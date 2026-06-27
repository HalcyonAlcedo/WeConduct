from __future__ import annotations

import base64
import binascii
from dataclasses import dataclass, field
import csv
from datetime import datetime, timezone
import ipaddress
import json
import os
from pathlib import Path
import re
import ast
import socket
import subprocess
import tempfile
from time import monotonic
import urllib.error
import urllib.parse
import urllib.request
from zipfile import BadZipFile
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from playwright.sync_api import Browser, Frame, Page, Playwright, sync_playwright
from weconduct.runtime.captcha_ocr import (
    DEFAULT_CAPTCHA_OCR_MODEL,
    CaptchaOcrRuntimeUnavailable,
    create_captcha_ocr_recognizer,
)


@dataclass
class RuntimeContext:
    variables: dict = field(default_factory=dict)
    node_outputs: dict[str, Any] = field(default_factory=dict)
    browser_runtime: dict[str, Any] = field(default_factory=dict)
    flow_runtime: dict[str, Any] = field(default_factory=dict)
    project_directory: Path | None = None
    workspace_root: Path | None = None
    embedded_resource_paths: dict[str, str] = field(default_factory=dict)
    allowed_path_roots: tuple[Path, ...] = field(default_factory=tuple)
    runtime_settings: dict[str, Any] = field(default_factory=dict)

    def close(self) -> None:
        browser_context = self.browser_runtime.get("browser_context")
        browser = self.browser_runtime.get("browser")
        playwright = self.browser_runtime.get("playwright")
        if browser_context is not None:
            browser_context.close()
        if browser is not None:
            browser.close()
        if playwright is not None:
            playwright.stop()
        self.browser_runtime.clear()


class RuntimeExecutorRegistry:
    def __init__(self, *, runtime_settings: dict | None = None) -> None:
        self._runtime_settings = runtime_settings or {}
        self._executors = {
            "flow.start": self._execute_flow_start,
            "http.request": self._execute_http_request,
            "browser.navigate": self._execute_browser_navigate,
            "browser.fill": self._execute_browser_fill,
            "browser.check": self._execute_browser_check,
            "browser.uncheck": self._execute_browser_uncheck,
            "browser.set_input_files": self._execute_browser_set_input_files,
            "browser.click": self._execute_browser_click,
            "browser.hover": self._execute_browser_hover,
            "browser.select_option": self._execute_browser_select_option,
            "browser.wait_for_element": self._execute_browser_wait_for_element,
            "browser.wait_for_navigation": self._execute_browser_wait_for_navigation,
            "browser.wait_for_timeout": self._execute_browser_wait_for_timeout,
            "browser.go_back": self._execute_browser_go_back,
            "browser.go_forward": self._execute_browser_go_forward,
            "browser.refresh": self._execute_browser_refresh,
            "browser.refresh_no_cache": self._execute_browser_refresh_no_cache,
            "browser.screenshot": self._execute_browser_screenshot,
            "browser.recognize_captcha": self._execute_browser_recognize_captcha,
            "browser.extract_web_table": self._execute_browser_extract_web_table,
            "browser.extract_web_table_to_excel": self._execute_browser_extract_web_table_to_excel,
            "browser.inject_js": self._execute_browser_inject_js,
            "browser.run_js": self._execute_browser_run_js,
            "browser.get_local_storage": self._execute_browser_get_local_storage,
            "browser.wait_for_text": self._execute_browser_wait_for_text,
            "browser.wait_for_attribute": self._execute_browser_wait_for_attribute,
            "browser.wait_for_value": self._execute_browser_wait_for_value,
            "browser.wait_for_request": self._execute_browser_wait_for_request,
            "browser.wait_for_response": self._execute_browser_wait_for_response,
            "browser.wait_for_popup": self._execute_browser_wait_for_popup,
            "browser.get_cookie": self._execute_browser_get_cookie,
            "browser.set_cookie": self._execute_browser_set_cookie,
            "browser.delete_cookie": self._execute_browser_delete_cookie,
            "browser.list_cookies": self._execute_browser_list_cookies,
            "browser.set_local_storage": self._execute_browser_set_local_storage,
            "browser.remove_local_storage": self._execute_browser_remove_local_storage,
            "browser.clear_local_storage": self._execute_browser_clear_local_storage,
            "browser.get_session_storage": self._execute_browser_get_session_storage,
            "browser.set_session_storage": self._execute_browser_set_session_storage,
            "browser.press_key": self._execute_browser_press_key,
            "browser.keyboard_type": self._execute_browser_keyboard_type,
            "browser.hotkey": self._execute_browser_hotkey,
            "browser.scroll_to_element": self._execute_browser_scroll_to_element,
            "browser.scroll_page": self._execute_browser_scroll_page,
            "browser.drag_and_drop": self._execute_browser_drag_and_drop,
            "browser.element_screenshot": self._execute_browser_element_screenshot,
            "browser.open_tab": self._execute_browser_open_tab,
            "browser.switch_tab": self._execute_browser_switch_tab,
            "browser.close_tab": self._execute_browser_close_tab,
            "browser.exists": self._execute_browser_exists,
            "browser.is_visible": self._execute_browser_is_visible,
            "browser.is_enabled": self._execute_browser_is_enabled,
            "browser.is_checked": self._execute_browser_is_checked,
            "browser.get_html": self._execute_browser_get_html,
            "browser.get_inner_html": self._execute_browser_get_inner_html,
            "browser.download_file": self._execute_browser_download_file,
            "browser.wait_for_download": self._execute_browser_wait_for_download,
            "browser.set_user_agent": self._execute_browser_set_user_agent,
            "browser.set_extra_headers": self._execute_browser_set_extra_headers,
            "browser.wait_for_url_change": self._execute_browser_wait_for_url_change,
            "browser.switch_to_frame": self._execute_browser_switch_to_frame,
            "browser.switch_to_parent_frame": self._execute_browser_switch_to_parent_frame,
            "browser.switch_to_default_content": self._execute_browser_switch_to_default_content,
            "browser.open_frame_page": self._execute_browser_open_frame_page,
            "data.map": self._execute_data_map,
            "data.set_variable": self._execute_set_variable,
            "data.get_variable": self._execute_get_variable,
            "data.get_text": self._execute_get_text,
            "data.get_attribute": self._execute_get_attribute,
            "data.get_value": self._execute_get_value,
            "data.get_element_count": self._execute_get_element_count,
            "data.set_variables_batch": self._execute_set_variables_batch,
            "data.increment_variable": self._execute_increment_variable,
            "data.decrement_variable": self._execute_decrement_variable,
            "data.convert_value": self._execute_convert_value,
            "data.create_list": self._execute_create_list,
            "data.list_append": self._execute_list_append,
            "data.list_extend": self._execute_list_extend,
            "data.list_get": self._execute_list_get,
            "data.list_set": self._execute_list_set,
            "data.list_index": self._execute_list_index,
            "data.list_length": self._execute_list_length,
            "data.list_insert": self._execute_list_insert,
            "data.list_remove": self._execute_list_remove,
            "data.list_slice": self._execute_list_slice,
            "data.list_sort": self._execute_list_sort,
            "data.list_reverse": self._execute_list_reverse,
            "data.evaluate_expression": self._execute_evaluate_expression,
            "data.regex_replace": self._execute_regex_replace,
            "time.get_current_time": self._execute_time_get_current_time,
            "file.write_text_file": self._execute_write_text_file,
            "file.read_text_file": self._execute_read_text_file,
            "file.read_csv_cell": self._execute_read_csv_cell,
            "file.read_csv_row": self._execute_read_csv_row,
            "file.read_csv_table": self._execute_read_csv_table,
            "excel.write_cell": self._execute_write_excel_cell,
            "excel.write_row": self._execute_write_excel_row,
            "excel.write_table": self._execute_write_excel_table,
            "excel.write_file": self._execute_write_excel_file,
            "excel.update_cells": self._execute_update_excel_cells,
            "excel.update_batch": self._execute_update_excel_batch,
            "excel.read_cell": self._execute_read_excel_cell,
            "excel.read_row": self._execute_read_excel_row,
            "excel.read_table": self._execute_read_excel_table,
            "python.run": self._execute_python_run,
            "control.foreach": self._execute_control_foreach,
            "control.jump_to_step": self._execute_control_jump_to_step,
            "control.end_foreach": self._execute_control_end_foreach,
            "control.foreach_continue": self._execute_control_foreach_continue,
            "control.foreach_break": self._execute_control_foreach_break,
            "session.apply_auth_session": self._execute_session_apply_auth_session,
            "dialog.switch_dialog_mode": self._execute_dialog_switch_dialog_mode,
            "dialog.watch_dialogs": self._execute_dialog_watch_dialogs,
            "dialog.handle_dialogs": self._execute_dialog_handle_dialogs,
            "dialog.set_agent_config": self._execute_dialog_set_agent_config,
        }

    def execute(self, node_kind: str | None, node: dict, context: RuntimeContext) -> dict:
        if not isinstance(node_kind, str) or node_kind not in self._executors:
            return {
                "status": "succeeded",
                "node_id": node["node_id"],
                "executor": "noop",
            }
        return self._executors[node_kind](node, context)

    def _execute_flow_start(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        initial_variables = node_config.get("initial_variables")
        if not isinstance(initial_variables, dict):
            initial_variables = {}
        for key, value in initial_variables.items():
            if isinstance(key, str) and key.strip():
                context.variables[key.strip()] = _resolve_value(value, context)
        browser_config = _resolve_browser_launch_config(node_config.get("browser_config"), context)
        if browser_config:
            context.browser_runtime["launch_options"] = dict(browser_config)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "initial_variable_count": len(initial_variables),
            "initial_variable_names": [
                key.strip()
                for key in initial_variables
                if isinstance(key, str) and key.strip()
            ],
            "browser_config": browser_config,
        }

    def _execute_http_request(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        method = str(_resolve_value(node_config.get("method", "GET"), context)).upper()
        url = _resolve_value(node_config.get("url"), context)
        if not isinstance(url, str) or not url.strip():
            return _failed_result(node, "http.url_required", "http.request requires node_config.url")
        try:
            normalized_url = _validate_http_request_url(
                url.strip(),
                allow_local_network_access=self._is_local_network_access_allowed(),
                allow_remote_network_access=self._is_remote_network_access_allowed(),
            )
        except ValueError as exc:
            message = str(exc)
            if "local network access is disabled" in message:
                return _failed_result(node, "http.local_network_disabled", message)
            if "remote network access is disabled" in message:
                return _failed_result(node, "http.remote_network_disabled", message)
            return _failed_result(node, "http.url_blocked", message)

        headers = _resolve_value(node_config.get("headers", {}), context)
        if not isinstance(headers, dict):
            return _failed_result(node, "http.headers_invalid", "node_config.headers must be an object")

        timeout_value = _resolve_value(node_config.get("timeout", 30), context)
        try:
            timeout = float(timeout_value)
        except (TypeError, ValueError):
            return _failed_result(node, "http.timeout_invalid", "node_config.timeout must be numeric")

        body_value = _resolve_value(node_config.get("body"), context)
        request_body = None
        request_headers = {str(key): str(value) for key, value in headers.items()}
        if body_value is not None:
            if isinstance(body_value, (dict, list)):
                request_body = json.dumps(body_value).encode("utf-8")
                request_headers.setdefault("Content-Type", "application/json")
            elif isinstance(body_value, bytes):
                request_body = body_value
            else:
                request_body = str(body_value).encode("utf-8")

        try:
            request = urllib.request.Request(
                normalized_url,
                data=request_body,
                headers=request_headers,
                method=method,
            )
            with urllib.request.urlopen(request, timeout=timeout) as response:
                raw_body = response.read()
                response_headers = {
                    key.lower(): value for key, value in response.headers.items()
                }
                response_body = _decode_http_body(raw_body, response_headers)
                result = {
                    "status": "succeeded",
                    "node_id": node["node_id"],
                    "method": method,
                    "url": normalized_url,
                    "status_code": response.status,
                    "headers": response_headers,
                    "body": response_body,
                    "body_text": raw_body.decode("utf-8", errors="replace"),
                }
        except urllib.error.HTTPError as exc:
            raw_body = exc.read()
            result = {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "http.status_error",
                "message": f"http request failed with status {exc.code}",
                "method": method,
                "url": normalized_url,
                "status_code": exc.code,
                "headers": {key.lower(): value for key, value in exc.headers.items()},
                "body": _decode_http_body(raw_body, dict(exc.headers.items())),
                "body_text": raw_body.decode("utf-8", errors="replace"),
            }
        except (urllib.error.URLError, TimeoutError, OSError, ValueError) as exc:
            result = {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "http.request_failed",
                "message": str(exc),
                "method": method,
                "url": normalized_url,
            }
        context.variables["last_http_response"] = result
        return result

    def _execute_browser_navigate(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_browser_executor_allowed():
            return _failed_result(node, "browser.executor_disabled", "browser executor is disabled")
        url = _resolve_value(node_config.get("url"), context)
        if not isinstance(url, str) or not url.strip():
            return _failed_result(node, "browser.url_required", "browser.navigate requires node_config.url")
        page = _require_browser_page(context)
        page.goto(url, wait_until="domcontentloaded")
        _reset_browser_frame_context(context)
        title = page.title()
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": page.url,
            "title": title,
        }
        context.variables["last_browser_url"] = page.url
        return result

    def _execute_browser_fill(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        value = _resolve_value(node_config.get("value"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        if not isinstance(value, str):
            value = "" if value is None else str(value)
        target = _require_browser_target(context)
        target.locator(selector).fill(value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "value": value,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_check(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        target = _require_browser_target(context)
        target.locator(selector).check()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_uncheck(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        target = _require_browser_target(context)
        target.locator(selector).uncheck()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_set_input_files(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_browser_uploads_allowed():
            return _failed_result(node, "browser.upload_disabled", "browser uploads are disabled")
        selector = _resolve_value(node_config.get("selector"), context)
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "browser.path_required", "path is required")
        resolved_path = _resolve_runtime_path(path_value, context)
        if not resolved_path.exists():
            return _failed_result(node, "browser.file_not_found", f"input file does not exist: {resolved_path}")
        target = _require_browser_target(context)
        target.locator(selector).set_input_files(str(resolved_path))
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "path": str(resolved_path.resolve()),
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_click(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        page = _require_browser_page(context)
        target = _require_browser_target(context)
        locator = target.locator(selector)
        try:
            locator.click()
        except Exception as exc:
            if not self._should_use_legacy_webcontrol_click_fallback(context=context, error=exc):
                raise
            locator.first.click()
        page.wait_for_load_state("domcontentloaded")
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "page_url": page.url,
        }

    def _execute_browser_hover(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        target = _require_browser_target(context)
        target.locator(selector).hover()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_select_option(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        value = _resolve_value(node_config.get("value"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        if not isinstance(value, str) or not value.strip():
            return _failed_result(node, "browser.value_required", "select_option requires value")
        target = _require_browser_target(context)
        selected = target.locator(selector).select_option(value=value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "value": value,
            "selected_values": selected,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_wait_for_element(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        target = _require_browser_target(context)
        target.locator(selector).wait_for(state="visible", timeout=timeout_ms)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "timeout_ms": timeout_ms,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_wait_for_navigation(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        url_pattern = _resolve_value(node_config.get("url_pattern"), context)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=15000)
        if not isinstance(url_pattern, str) or not url_pattern.strip():
            return _failed_result(node, "browser.url_pattern_required", "url_pattern is required")
        target = _require_browser_target(context)
        matched_url = _wait_for_browser_url(target, url_pattern.strip(), timeout_ms)
        if matched_url is None:
            return _failed_result(node, "browser.navigation_timeout", "browser.wait_for_navigation timed out")
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "url_pattern": url_pattern.strip(),
            "matched_url": matched_url,
            "timeout_ms": timeout_ms,
        }

    def _execute_browser_wait_for_timeout(self, node: dict, context: RuntimeContext) -> dict:
        timeout_ms = _resolve_int(_resolve_value(_node_config(node).get("timeout"), context), default=0)
        _require_browser_page(context).wait_for_timeout(timeout_ms)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "timeout_ms": timeout_ms,
        }

    def _execute_browser_go_back(self, node: dict, context: RuntimeContext) -> dict:
        page = _require_browser_page(context)
        response = page.go_back(wait_until="domcontentloaded")
        _reset_browser_frame_context(context)
        context.variables["last_browser_url"] = page.url
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": page.url,
            "title": page.title(),
            "navigated": response is not None,
        }

    def _execute_browser_go_forward(self, node: dict, context: RuntimeContext) -> dict:
        page = _require_browser_page(context)
        response = page.go_forward(wait_until="domcontentloaded")
        _reset_browser_frame_context(context)
        context.variables["last_browser_url"] = page.url
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": page.url,
            "title": page.title(),
            "navigated": response is not None,
        }

    def _execute_browser_refresh(self, node: dict, context: RuntimeContext) -> dict:
        page = _require_browser_page(context)
        page.reload(wait_until="domcontentloaded")
        _reset_browser_frame_context(context)
        context.variables["last_browser_url"] = page.url
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": page.url,
            "title": page.title(),
            "bypass_cache": False,
        }

    def _execute_browser_refresh_no_cache(self, node: dict, context: RuntimeContext) -> dict:
        page = _require_browser_page(context)
        page.evaluate(
            """
            () => {
              const currentUrl = window.location.href;
              const marker = currentUrl.includes('?') ? '&' : '?';
              window.location.replace(`${currentUrl}${marker}weconduct_no_cache=${Date.now()}`);
            }
            """
        )
        page.wait_for_load_state("domcontentloaded")
        _reset_browser_frame_context(context)
        context.variables["last_browser_url"] = page.url
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": page.url,
            "title": "",
            "bypass_cache": True,
        }

    def _execute_browser_screenshot(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_browser_screenshots_allowed():
            return _failed_result(node, "browser.screenshot_disabled", "browser screenshots are disabled")
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "browser.screenshot_path_required", "screenshot path is required")
        path = _resolve_runtime_path(path_value, context)
        path.parent.mkdir(parents=True, exist_ok=True)
        page = _require_browser_page(context)
        page.screenshot(path=str(path))
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(path.resolve()),
            "bytes_written": path.stat().st_size,
            "page_url": page.url,
        }

    def _execute_browser_inject_js(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_js_injection_allowed():
            return _failed_result(node, "browser.js_injection_disabled", "JavaScript injection is disabled")
        script = _resolve_value(node_config.get("script") or node_config.get("code"), context)
        if not isinstance(script, str) or not script.strip():
            return _failed_result(node, "browser.script_required", "JavaScript script is required")
        target = _require_browser_target(context)
        value = target.evaluate(script)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "script_length": len(script),
            "value": value,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_run_js(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_js_evaluation_allowed():
            return _failed_result(node, "browser.js_evaluation_disabled", "JavaScript evaluation is disabled")
        script = _resolve_value(node_config.get("script") or node_config.get("code"), context)
        if not isinstance(script, str) or not script.strip():
            return _failed_result(node, "browser.script_required", "JavaScript script is required")
        target = _require_browser_target(context)
        value = target.evaluate(script)
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "script_length": len(script),
            "value": value,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_get_local_storage(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        key = _resolve_value(node_config.get("key"), context)
        if not isinstance(key, str) or not key.strip():
            return _failed_result(
                node,
                "browser.local_storage_key_required",
                "localStorage key is required",
            )
        target = _require_browser_target(context)
        value = target.evaluate(
            "(storageKey) => window.localStorage.getItem(storageKey)",
            key.strip(),
        )
        if value is None:
            value = _resolve_value(node_config.get("default_value"), context)
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "key": key.strip(),
            "value": value,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_wait_for_text(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _require_selector(node, context)
        if isinstance(selector, dict):
            return selector
        expected_text = _resolve_value(node_config.get("text"), context)
        if not isinstance(expected_text, str):
            expected_text = "" if expected_text is None else str(expected_text)
        match_mode = str(_resolve_value(node_config.get("match_mode", "contains"), context) or "contains").strip().lower()
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        target = _require_browser_target(context)
        locator = target.locator(selector)
        _wait_until(
            timeout_ms,
            lambda: _match_text(str(locator.inner_text()), expected_text, match_mode),
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        matched_text = str(locator.inner_text())
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "matched_text": matched_text,
            "match_mode": match_mode,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_wait_for_attribute(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _require_selector(node, context)
        if isinstance(selector, dict):
            return selector
        attribute = _resolve_value(node_config.get("attribute"), context)
        if not isinstance(attribute, str) or not attribute.strip():
            return _failed_result(node, "browser.attribute_required", "attribute is required")
        expected_value = _resolve_value(node_config.get("value"), context)
        if not isinstance(expected_value, str):
            expected_value = "" if expected_value is None else str(expected_value)
        match_mode = str(_resolve_value(node_config.get("match_mode", "equals"), context) or "equals").strip().lower()
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        target = _require_browser_target(context)
        locator = target.locator(selector)
        _wait_until(
            timeout_ms,
            lambda: _match_text(locator.get_attribute(attribute.strip()) or "", expected_value, match_mode),
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        value = locator.get_attribute(attribute.strip()) or ""
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "attribute": attribute.strip(),
            "value": value,
            "match_mode": match_mode,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_wait_for_value(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _require_selector(node, context)
        if isinstance(selector, dict):
            return selector
        expected_value = _resolve_value(node_config.get("value"), context)
        if not isinstance(expected_value, str):
            expected_value = "" if expected_value is None else str(expected_value)
        match_mode = str(_resolve_value(node_config.get("match_mode", "equals"), context) or "equals").strip().lower()
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        target = _require_browser_target(context)
        locator = target.locator(selector)
        _wait_until(
            timeout_ms,
            lambda: _match_text(locator.input_value(), expected_value, match_mode),
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        value = locator.input_value()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector,
            "value": value,
            "match_mode": match_mode,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_wait_for_request(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        url_pattern = _resolve_value(node_config.get("url_pattern"), context)
        method = _resolve_value(node_config.get("method"), context)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        if not isinstance(url_pattern, str) or not url_pattern.strip():
            return _failed_result(node, "browser.url_pattern_required", "url_pattern is required")
        normalized_method = method.strip().upper() if isinstance(method, str) and method.strip() else None
        record = _wait_for_record(
            context.browser_runtime.setdefault("request_records", []),
            timeout_ms,
            lambda item: _request_record_matches(item, url_pattern.strip(), normalized_method),
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "url_pattern": url_pattern.strip(),
            "method": record["method"],
            "url": record["url"],
            "headers": record["headers"],
        }

    def _execute_browser_wait_for_response(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        url_pattern = _resolve_value(node_config.get("url_pattern"), context)
        status_code = _resolve_value(node_config.get("status_code"), context)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        if not isinstance(url_pattern, str) or not url_pattern.strip():
            return _failed_result(node, "browser.url_pattern_required", "url_pattern is required")
        expected_status = _resolve_int(status_code, default=-1) if status_code is not None else None
        record = _wait_for_record(
            context.browser_runtime.setdefault("response_records", []),
            timeout_ms,
            lambda item: _response_record_matches(item, url_pattern.strip(), expected_status),
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "url_pattern": url_pattern.strip(),
            "url": record["url"],
            "status_code": record["status_code"],
            "headers": record["headers"],
            "body_text": record["body_text"],
        }

    def _execute_browser_wait_for_popup(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        activate = bool(_resolve_value(node_config.get("activate", True), context))
        record = _wait_for_record(
            context.browser_runtime.setdefault("popup_records", []),
            timeout_ms,
            lambda item: isinstance(item, dict) and isinstance(item.get("page"), Page),
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        popup_page = record["page"]
        if activate:
            page_index = _set_active_browser_page(context, popup_page)
        else:
            page_index = record.get("page_index")
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": popup_page.url,
            "page_index": page_index,
            "label": _page_label(context, popup_page),
            "activated": activate,
        }
        _store_optional_variable(node_config, context, {key: value for key, value in result.items() if key != "node_id" and key != "status"})
        return result

    def _execute_browser_get_cookie(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_cookie_manipulation_allowed():
            return _failed_result(
                node,
                "browser.cookie_manipulation_disabled",
                "cookie manipulation is disabled",
            )
        name = _resolve_value(node_config.get("name"), context)
        if not isinstance(name, str) or not name.strip():
            return _failed_result(node, "browser.cookie_name_required", "cookie name is required")
        cookies = _require_browser_page(context).context.cookies()
        value = None
        selected = None
        for cookie in cookies:
            if cookie.get("name") == name.strip():
                selected = cookie
                value = cookie.get("value")
                break
        if value is None:
            value = _resolve_value(node_config.get("default_value"), context)
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "name": name.strip(),
            "value": value,
            "cookie": selected,
        }

    def _execute_browser_set_cookie(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_cookie_manipulation_allowed():
            return _failed_result(
                node,
                "browser.cookie_manipulation_disabled",
                "cookie manipulation is disabled",
            )
        name = _resolve_value(node_config.get("name"), context)
        if not isinstance(name, str) or not name.strip():
            return _failed_result(node, "browser.cookie_name_required", "cookie name is required")
        browser_context = _require_browser_page(context).context
        cookie = {
            "name": name.strip(),
            "value": "" if _resolve_value(node_config.get("value"), context) is None else str(_resolve_value(node_config.get("value"), context)),
        }
        for field_name in ("url", "domain"):
            field_value = _resolve_value(node_config.get(field_name), context)
            if isinstance(field_value, str) and field_value.strip():
                cookie[field_name] = field_value.strip()
        if "url" not in cookie and "domain" not in cookie:
            cookie["url"] = _require_browser_page(context).url
        if "url" not in cookie:
            cookie["path"] = str(_resolve_value(node_config.get("path", "/"), context) or "/")
        for field_name in ("http_only", "secure"):
            cookie[field_name] = bool(_resolve_value(node_config.get(field_name, False), context))
        same_site = _resolve_value(node_config.get("same_site", "Lax"), context)
        if isinstance(same_site, str) and same_site.strip():
            cookie["sameSite"] = same_site.strip()
        expires_value = _resolve_value(node_config.get("expires"), context)
        if expires_value is not None:
            cookie["expires"] = float(expires_value)
        browser_context.add_cookies([cookie])
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "cookie_name": cookie["name"],
            "url": cookie.get("url"),
            "domain": cookie.get("domain"),
        }

    def _execute_browser_delete_cookie(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_cookie_manipulation_allowed():
            return _failed_result(
                node,
                "browser.cookie_manipulation_disabled",
                "cookie manipulation is disabled",
            )
        name = _resolve_value(node_config.get("name"), context)
        if not isinstance(name, str) or not name.strip():
            return _failed_result(node, "browser.cookie_name_required", "cookie name is required")
        browser_context = _require_browser_page(context).context
        kept = []
        removed_count = 0
        for cookie in browser_context.cookies():
            if cookie.get("name") == name.strip():
                removed_count += 1
                continue
            kept.append(cookie)
        browser_context.clear_cookies()
        normalized = _normalize_browser_context_cookies(kept)
        if normalized:
            browser_context.add_cookies(normalized)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "cookie_name": name.strip(),
            "removed_count": removed_count,
        }

    def _execute_browser_list_cookies(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_cookie_manipulation_allowed():
            return _failed_result(
                node,
                "browser.cookie_manipulation_disabled",
                "cookie manipulation is disabled",
            )
        cookies = _require_browser_page(context).context.cookies()
        _store_optional_variable(node_config, context, cookies)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "cookie_count": len(cookies),
            "cookies": cookies,
        }

    def _execute_browser_set_local_storage(self, node: dict, context: RuntimeContext) -> dict:
        return _set_browser_storage_item(node=node, context=context, storage_name="localStorage")

    def _execute_browser_remove_local_storage(self, node: dict, context: RuntimeContext) -> dict:
        return _remove_browser_storage_item(node=node, context=context, storage_name="localStorage")

    def _execute_browser_clear_local_storage(self, node: dict, context: RuntimeContext) -> dict:
        return _clear_browser_storage(node=node, context=context, storage_name="localStorage")

    def _execute_browser_get_session_storage(self, node: dict, context: RuntimeContext) -> dict:
        return _get_browser_storage_item(node=node, context=context, storage_name="sessionStorage")

    def _execute_browser_set_session_storage(self, node: dict, context: RuntimeContext) -> dict:
        return _set_browser_storage_item(node=node, context=context, storage_name="sessionStorage")

    def _execute_browser_press_key(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        key = _resolve_value(node_config.get("key"), context)
        if not isinstance(key, str) or not key.strip():
            return _failed_result(node, "browser.key_required", "key is required")
        selector = _resolve_value(node_config.get("selector"), context)
        target = _require_browser_target(context)
        if isinstance(selector, str) and selector.strip():
            target.locator(selector.strip()).press(key.strip())
        else:
            _require_browser_page(context).keyboard.press(key.strip())
        return {"status": "succeeded", "node_id": node["node_id"], "key": key.strip(), "page_url": _require_browser_page(context).url}

    def _execute_browser_keyboard_type(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        text = _resolve_value(node_config.get("text"), context)
        if not isinstance(text, str):
            text = "" if text is None else str(text)
        delay_ms = _resolve_int(_resolve_value(node_config.get("delay_ms"), context), default=0)
        selector = _resolve_value(node_config.get("selector"), context)
        target = _require_browser_target(context)
        if isinstance(selector, str) and selector.strip():
            target.locator(selector.strip()).click()
        _require_browser_page(context).keyboard.type(text, delay=delay_ms)
        return {"status": "succeeded", "node_id": node["node_id"], "text": text, "delay_ms": delay_ms, "page_url": _require_browser_page(context).url}

    def _execute_browser_hotkey(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        combo = _resolve_value(node_config.get("combo"), context)
        if not isinstance(combo, str) or not combo.strip():
            return _failed_result(node, "browser.hotkey_required", "combo is required")
        selector = _resolve_value(node_config.get("selector"), context)
        target = _require_browser_target(context)
        if isinstance(selector, str) and selector.strip():
            target.locator(selector.strip()).click()
        _require_browser_page(context).keyboard.press(combo.strip())
        return {"status": "succeeded", "node_id": node["node_id"], "combo": combo.strip(), "page_url": _require_browser_page(context).url}

    def _execute_browser_scroll_to_element(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _require_selector(node, context)
        if isinstance(selector, dict):
            return selector
        block = str(_resolve_value(node_config.get("block", "center"), context) or "center")
        inline = str(_resolve_value(node_config.get("inline", "nearest"), context) or "nearest")
        target = _require_browser_target(context)
        target.locator(selector).scroll_into_view_if_needed()
        target.locator(selector).evaluate(
            "(el, args) => el.scrollIntoView({ block: args.block, inline: args.inline })",
            {"block": block, "inline": inline},
        )
        return {"status": "succeeded", "node_id": node["node_id"], "selector": selector, "block": block, "inline": inline, "page_url": _require_browser_page(context).url}

    def _execute_browser_scroll_page(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        x = _resolve_int(_resolve_value(node_config.get("x"), context), default=0)
        y = _resolve_int(_resolve_value(node_config.get("y"), context), default=0)
        mode = str(_resolve_value(node_config.get("mode", "by"), context) or "by").strip().lower()
        page = _require_browser_page(context)
        if mode == "to":
            page.evaluate("(coords) => window.scrollTo(coords.x, coords.y)", {"x": x, "y": y})
        else:
            mode = "by"
            page.evaluate("(coords) => window.scrollBy(coords.x, coords.y)", {"x": x, "y": y})
        return {"status": "succeeded", "node_id": node["node_id"], "x": x, "y": y, "mode": mode, "page_url": page.url}

    def _execute_browser_drag_and_drop(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        source_selector = _resolve_value(node_config.get("source_selector"), context)
        target_selector = _resolve_value(node_config.get("target_selector"), context)
        if not isinstance(source_selector, str) or not source_selector.strip():
            return _failed_result(node, "browser.source_selector_required", "source_selector is required")
        if not isinstance(target_selector, str) or not target_selector.strip():
            return _failed_result(node, "browser.target_selector_required", "target_selector is required")
        target = _require_browser_target(context)
        target.locator(source_selector.strip()).drag_to(target.locator(target_selector.strip()))
        return {"status": "succeeded", "node_id": node["node_id"], "source_selector": source_selector.strip(), "target_selector": target_selector.strip(), "page_url": _require_browser_page(context).url}

    def _execute_browser_element_screenshot(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_browser_screenshots_allowed():
            return _failed_result(node, "browser.screenshot_disabled", "browser screenshots are disabled")
        selector = _require_selector(node, context)
        if isinstance(selector, dict):
            return selector
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "browser.screenshot_path_required", "screenshot path is required")
        path = _resolve_runtime_path(path_value, context)
        path.parent.mkdir(parents=True, exist_ok=True)
        _require_browser_target(context).locator(selector).screenshot(path=str(path))
        return {"status": "succeeded", "node_id": node["node_id"], "selector": selector, "path": str(path.resolve()), "bytes_written": path.stat().st_size, "page_url": _require_browser_page(context).url}

    def _execute_browser_open_tab(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_new_browser_windows_allowed():
            return _failed_result(node, "browser.new_window_disabled", "new browser windows are disabled")
        url = _resolve_value(node_config.get("url"), context)
        if not isinstance(url, str) or not url.strip():
            return _failed_result(node, "browser.url_required", "browser.open_tab requires node_config.url")
        browser_context = _require_browser_page(context).context
        context.browser_runtime["suppress_next_page_record"] = int(context.browser_runtime.get("suppress_next_page_record", 0)) + 1
        page = browser_context.new_page()
        page.goto(url.strip(), wait_until="domcontentloaded")
        page_index = _register_browser_page(context, page)
        label = _resolve_value(node_config.get("label"), context)
        if isinstance(label, str) and label.strip():
            context.browser_runtime.setdefault("page_labels", {})[label.strip()] = page
        activate = bool(_resolve_value(node_config.get("activate", True), context))
        if activate:
            _set_active_browser_page(context, page)
        return {"status": "succeeded", "node_id": node["node_id"], "page_url": page.url, "page_index": page_index, "label": label.strip() if isinstance(label, str) and label.strip() else None, "activated": activate}

    def _execute_browser_switch_tab(self, node: dict, context: RuntimeContext) -> dict:
        page = _resolve_browser_page_reference(_node_config(node), context)
        if page is None:
            return _failed_result(node, "browser.tab_not_found", "target tab was not found")
        page_index = _set_active_browser_page(context, page)
        return {"status": "succeeded", "node_id": node["node_id"], "page_url": page.url, "page_index": page_index, "label": _page_label(context, page)}

    def _execute_browser_close_tab(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        page = _require_browser_page(context) if bool(_resolve_value(node_config.get("current", False), context)) else _resolve_browser_page_reference(node_config, context)
        if page is None:
            return _failed_result(node, "browser.tab_not_found", "target tab was not found")
        page.close()
        pages = _browser_pages(context)
        if pages:
            _set_active_browser_page(context, pages[min(len(pages) - 1, int(context.browser_runtime.get("active_page_index", 0)))])
        return {"status": "succeeded", "node_id": node["node_id"], "closed": True}

    def _execute_browser_exists(self, node: dict, context: RuntimeContext) -> dict:
        return _probe_browser_locator(node=node, context=context, probe="exists")

    def _execute_browser_is_visible(self, node: dict, context: RuntimeContext) -> dict:
        return _probe_browser_locator(node=node, context=context, probe="visible")

    def _execute_browser_is_enabled(self, node: dict, context: RuntimeContext) -> dict:
        return _probe_browser_locator(node=node, context=context, probe="enabled")

    def _execute_browser_is_checked(self, node: dict, context: RuntimeContext) -> dict:
        return _probe_browser_locator(node=node, context=context, probe="checked")

    def _execute_browser_get_html(self, node: dict, context: RuntimeContext) -> dict:
        return _read_browser_html(node=node, context=context, mode="outer")

    def _execute_browser_get_inner_html(self, node: dict, context: RuntimeContext) -> dict:
        return _read_browser_html(node=node, context=context, mode="inner")

    def _execute_browser_download_file(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_browser_downloads_allowed():
            return _failed_result(node, "browser.download_disabled", "browser downloads are disabled")
        url = _resolve_value(node_config.get("url"), context)
        if not isinstance(url, str) or not url.strip():
            return _failed_result(node, "browser.url_required", "browser.download_file requires node_config.url")
        try:
            url = _validate_http_request_url(
                url,
                allow_local_network_access=_is_local_network_access_allowed(context),
                allow_remote_network_access=_is_remote_network_access_allowed(context),
            )
        except ValueError as exc:
            return _failed_result(node, "runtime.executor_exception", str(exc))
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "browser.download_path_required", "download path is required")
        path = _resolve_runtime_path(path_value, context)
        path.parent.mkdir(parents=True, exist_ok=True)
        with urllib.request.urlopen(url.strip()) as response, path.open("wb") as handle:
            handle.write(response.read())
        return {"status": "succeeded", "node_id": node["node_id"], "url": url.strip(), "path": str(path.resolve()), "bytes_written": path.stat().st_size}

    def _execute_browser_wait_for_download(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_browser_downloads_allowed():
            return _failed_result(node, "browser.download_disabled", "browser downloads are disabled")
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        record = _wait_for_record(
            context.browser_runtime.setdefault("download_records", []),
            timeout_ms,
            lambda item: isinstance(item, dict) and item.get("download") is not None,
            on_poll=lambda: _pump_browser_events(context.browser_runtime.get("page")),
        )
        path_value = _resolve_value(node_config.get("path"), context)
        if isinstance(path_value, str) and path_value.strip():
            path = _resolve_runtime_path(path_value, context)
        else:
            path = _resolve_runtime_path(record["suggested_filename"], context)
        path.parent.mkdir(parents=True, exist_ok=True)
        record["download"].save_as(str(path))
        result = {"status": "succeeded", "node_id": node["node_id"], "url": record["url"], "path": str(path.resolve()), "suggested_filename": record["suggested_filename"], "page_index": record.get("page_index")}
        _store_optional_variable(node_config, context, {key: value for key, value in result.items() if key not in {"status", "node_id"}})
        return result

    def _execute_browser_set_user_agent(self, node: dict, context: RuntimeContext) -> dict:
        user_agent = _resolve_value(_node_config(node).get("user_agent"), context)
        if not isinstance(user_agent, str) or not user_agent.strip():
            return _failed_result(node, "browser.user_agent_required", "user_agent is required")
        launch_options = context.browser_runtime.setdefault("launch_options", {})
        launch_options["user_agent"] = user_agent.strip()
        return {"status": "succeeded", "node_id": node["node_id"], "user_agent": user_agent.strip()}

    def _execute_browser_set_extra_headers(self, node: dict, context: RuntimeContext) -> dict:
        headers = _resolve_value(_node_config(node).get("headers"), context)
        if not isinstance(headers, dict):
            return _failed_result(node, "browser.headers_invalid", "headers must be an object")
        launch_options = context.browser_runtime.setdefault("launch_options", {})
        launch_options["extra_http_headers"] = {str(key): "" if value is None else str(value) for key, value in headers.items()}
        return {"status": "succeeded", "node_id": node["node_id"], "header_count": len(headers)}

    def _execute_browser_wait_for_url_change(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        from_url = _resolve_value(node_config.get("from_url"), context)
        url_pattern = _resolve_value(node_config.get("url_pattern"), context)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=10000)
        page = _require_browser_page(context)
        if not isinstance(from_url, str) or not from_url.strip():
            from_url = page.url
        matched_url = _wait_for_url_change(page=page, from_url=from_url, url_pattern=url_pattern if isinstance(url_pattern, str) else "", timeout_ms=timeout_ms)
        return {"status": "succeeded", "node_id": node["node_id"], "from_url": from_url, "matched_url": matched_url, "page_url": page.url}

    def _execute_time_get_current_time(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        format_name = _resolve_value(node_config.get("format") or "iso", context)
        timezone_name = _resolve_value(node_config.get("timezone") or "utc", context)
        normalized_timezone = (
            timezone_name.strip().lower()
            if isinstance(timezone_name, str) and timezone_name.strip()
            else "utc"
        )
        tzinfo = timezone.utc if normalized_timezone == "utc" else datetime.now().astimezone().tzinfo
        now = datetime.now(tzinfo)
        if format_name == "timestamp":
            value: Any = int(now.timestamp())
        elif format_name == "timestamp_ms":
            value = int(now.timestamp() * 1000)
        else:
            format_name = "iso"
            value = now.isoformat()
        _store_optional_variable(_node_config(node), context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "value": value,
            "format": format_name,
            "timezone": normalized_timezone,
        }

    def _execute_browser_recognize_captcha(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        target_variable = (
            node_config.get("target_variable")
            or node_config.get("variable_name")
            or node_config.get("name")
        )
        if not isinstance(target_variable, str) or not target_variable.strip():
            return _failed_result(
                node,
                "browser.captcha_target_variable_required",
                "target_variable is required",
            )
        model_name = (
            node_config.get("model_name")
            or node_config.get("model")
            or node_config.get("captcha_model")
            or DEFAULT_CAPTCHA_OCR_MODEL
        )
        if not isinstance(model_name, str) or not model_name.strip():
            model_name = DEFAULT_CAPTCHA_OCR_MODEL

        try:
            image_bytes = self._resolve_captcha_image_bytes(node, context)
        except ValueError as exc:
            return _failed_result(node, "browser.captcha_image_invalid", str(exc))

        runtime_root = node_config.get("runtime_root")
        try:
            recognizer = create_captcha_ocr_recognizer(
                model_name=model_name.strip(),
                runtime_root=runtime_root,
            )
            try:
                text = recognizer.recognize_from_bytes(image_bytes)
            finally:
                close = getattr(recognizer, "close", None)
                if callable(close):
                    close()
        except (CaptchaOcrRuntimeUnavailable, RuntimeError) as exc:
            return _failed_result(node, "browser.captcha_ocr_unavailable", str(exc))

        if not text:
            return _failed_result(
                node,
                "browser.captcha_empty_result",
                "captcha_ocr returned an empty result",
            )

        target_variable = target_variable.strip()
        context.variables[target_variable] = text
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "text": text,
            "target_variable": target_variable,
            "model_name": model_name.strip(),
            "backend": "captcha_ocr",
        }

    def _resolve_captcha_image_bytes(self, node: dict, context: RuntimeContext) -> bytes:
        node_config = _node_config(node)
        encoded = _resolve_value(node_config.get("image_bytes_base64"), context)
        if isinstance(encoded, str) and encoded.strip():
            raw_encoded = encoded.split(",", 1)[1] if encoded.startswith("data:") else encoded
            try:
                return base64.b64decode(raw_encoded, validate=True)
            except (binascii.Error, ValueError) as exc:
                raise ValueError("image_bytes_base64 must be valid base64") from exc

        selector = _resolve_value(node_config.get("selector"), context)
        if not isinstance(selector, str) or not selector.strip():
            raise ValueError("selector or image_bytes_base64 is required")
        target = _require_browser_target(context)
        return target.locator(selector.strip()).screenshot()

    def _execute_browser_extract_web_table(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector = _resolve_value(node_config.get("selector"), context)
        if not isinstance(selector, str) or not selector.strip():
            return _failed_result(node, "browser.selector_required", "selector is required")
        target = _require_browser_target(context)
        try:
            table = target.locator(selector.strip()).evaluate(_WEB_TABLE_EXTRACT_SCRIPT)
        except Exception as exc:  # Playwright wraps selector/evaluate failures.
            return _failed_result(node, "browser.web_table_extract_failed", str(exc))
        if not isinstance(table, dict):
            return _failed_result(node, "browser.web_table_invalid", "web table extractor returned invalid data")
        rows = table.get("rows")
        headers = table.get("headers")
        if not isinstance(rows, list) or not isinstance(headers, list):
            return _failed_result(node, "browser.web_table_invalid", "web table rows or headers are invalid")
        _store_optional_variable(node_config, context, rows)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector.strip(),
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        }

    def _execute_browser_extract_web_table_to_excel(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        table_result = self._execute_browser_extract_web_table(node, context)
        if table_result.get("status") == "failed":
            return table_result
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "excel.path_required", "Excel file path is required")
        sheet_name = _resolve_value(node_config.get("sheet_name", "Sheet1"), context)
        if not isinstance(sheet_name, str) or not sheet_name.strip():
            return _failed_result(node, "excel.sheet_name_required", "Excel sheet_name is required")
        path = _resolve_runtime_path(path_value, context)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name.strip()
        headers = table_result["headers"]
        rows = table_result["rows"]
        if headers:
            for column_index, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=column_index, value=header)
            start_row = 2
        else:
            start_row = 1
        for row_offset, row in enumerate(rows, start=start_row):
            values = [row.get(str(header)) for header in headers] if isinstance(row, dict) else list(row)
            for column_index, value in enumerate(values, start=1):
                worksheet.cell(row=row_offset, column=column_index, value=value)
        workbook.save(path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(path.resolve()),
            "sheet_name": sheet_name.strip(),
            "headers": headers,
            "row_count": len(rows),
        }

    def _execute_browser_switch_to_frame(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        frame = _resolve_browser_frame(node_config, context)
        if frame is None:
            return _failed_result(node, "browser.frame_not_found", "target frame was not found")
        frame_stack = context.browser_runtime.setdefault("frame_stack", [])
        frame_stack.append(frame)
        context.browser_runtime["current_frame"] = frame
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "frame_name": frame.name,
            "frame_url": frame.url,
            "frame_depth": len(frame_stack),
        }

    def _execute_browser_switch_to_parent_frame(self, node: dict, context: RuntimeContext) -> dict:
        frame_stack = context.browser_runtime.setdefault("frame_stack", [])
        if frame_stack:
            frame_stack.pop()
        if frame_stack:
            context.browser_runtime["current_frame"] = frame_stack[-1]
        else:
            context.browser_runtime.pop("current_frame", None)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "frame_depth": len(frame_stack),
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_switch_to_default_content(self, node: dict, context: RuntimeContext) -> dict:
        _reset_browser_frame_context(context)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "frame_depth": 0,
            "page_url": _require_browser_page(context).url,
        }

    def _execute_browser_open_frame_page(self, node: dict, context: RuntimeContext) -> dict:
        frame = _resolve_browser_frame(_node_config(node), context)
        if frame is None:
            current_frame = context.browser_runtime.get("current_frame")
            frame = current_frame if isinstance(current_frame, Frame) else None
        if frame is None or not frame.url:
            return _failed_result(node, "browser.frame_not_found", "target frame was not found")
        page = _require_browser_page(context)
        previous_frame_url = frame.url
        page.goto(previous_frame_url, wait_until="domcontentloaded")
        _reset_browser_frame_context(context)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "page_url": page.url,
            "previous_frame_url": previous_frame_url,
        }

    def _execute_data_map(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        source_value = (
            _resolve_value(node_config["source"], context)
            if "source" in node_config
            else None
        )
        upstream = _last_node_output(context)
        if source_value is None:
            source_value = upstream
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "value": source_value,
            "mapped_from_node_id": upstream.get("node_id") if isinstance(upstream, dict) else None,
            "response_status_code": upstream.get("status_code") if isinstance(upstream, dict) else None,
        }
        variable_name = node_config.get("variable_name")
        if isinstance(variable_name, str) and variable_name.strip():
            context.variables[variable_name.strip()] = source_value
        context.variables["last_data_map_result"] = result
        return result

    def _execute_set_variable(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        name = node_config.get("name") or node_config.get("variable_name")
        if not isinstance(name, str) or not name.strip():
            return _failed_result(node, "data.variable_name_required", "variable name is required")
        value = _resolve_value(node_config.get("value"), context)
        context.variables[name.strip()] = value
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": name.strip(),
            "value": value,
        }

    def _execute_get_variable(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        name = node_config.get("name") or node_config.get("variable_name")
        if not isinstance(name, str) or not name.strip():
            return _failed_result(node, "data.variable_name_required", "variable name is required")
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": name.strip(),
            "value": context.variables.get(name.strip()),
        }

    def _execute_get_text(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector_result = _require_selector(node, context)
        if isinstance(selector_result, dict):
            return selector_result
        value = _require_browser_target(context).locator(selector_result).inner_text()
        target_type = node_config.get("target_type")
        if target_type is not None:
            try:
                value = _convert_runtime_value(value, target_type)
            except ValueError as exc:
                return _failed_result(node, "data.get_text_type_invalid", str(exc))
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector_result,
            "value": value,
        }

    def _execute_get_attribute(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        selector_result = _require_selector(node, context)
        if isinstance(selector_result, dict):
            return selector_result
        attribute = _resolve_value(node_config.get("attribute") or node_config.get("attribute_name"), context)
        if not isinstance(attribute, str) or not attribute.strip():
            return _failed_result(node, "data.attribute_required", "attribute is required")
        value = _require_browser_target(context).locator(selector_result).get_attribute(attribute.strip())
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector_result,
            "attribute": attribute.strip(),
            "value": value,
        }

    def _execute_get_value(self, node: dict, context: RuntimeContext) -> dict:
        selector_result = _require_selector(node, context)
        if isinstance(selector_result, dict):
            return selector_result
        value = _require_browser_target(context).locator(selector_result).input_value()
        _store_optional_variable(_node_config(node), context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector_result,
            "value": value,
        }

    def _execute_get_element_count(self, node: dict, context: RuntimeContext) -> dict:
        selector_result = _require_selector(node, context)
        if isinstance(selector_result, dict):
            return selector_result
        value = _require_browser_target(context).locator(selector_result).count()
        _store_optional_variable(_node_config(node), context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "selector": selector_result,
            "value": value,
        }

    def _execute_set_variables_batch(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        variables = _resolve_value(node_config.get("variables"), context)
        if not isinstance(variables, dict):
            return _failed_result(node, "data.variables_invalid", "variables must be an object")
        variable_names: list[str] = []
        for name, value in variables.items():
            if isinstance(name, str) and name.strip():
                context.variables[name.strip()] = value
                variable_names.append(name.strip())
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_names": variable_names,
            "count": len(variable_names),
        }

    def _execute_increment_variable(self, node: dict, context: RuntimeContext) -> dict:
        return self._execute_numeric_variable_delta(node, context, multiplier=1)

    def _execute_decrement_variable(self, node: dict, context: RuntimeContext) -> dict:
        return self._execute_numeric_variable_delta(node, context, multiplier=-1)

    def _execute_convert_value(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        target_type = _resolve_value(node_config.get("target_type", "string"), context)
        source_value = _resolve_value(node_config.get("source_value"), context)
        if source_value is None and "value" in node_config:
            source_value = _resolve_value(node_config.get("value"), context)
        try:
            converted = _convert_runtime_value(source_value, target_type)
        except ValueError as exc:
            return _failed_result(node, "data.convert_value_invalid", str(exc))

        in_place = bool(_resolve_value(node_config.get("in_place", False), context))
        source_variable_name = _resolve_value(node_config.get("source_variable_name"), context)
        output_variable_name = _resolve_value(node_config.get("variable_name"), context)
        target_variable_name = None
        if in_place and isinstance(source_variable_name, str) and source_variable_name.strip():
            target_variable_name = source_variable_name.strip()
        elif isinstance(output_variable_name, str) and output_variable_name.strip():
            target_variable_name = output_variable_name.strip()
        if target_variable_name:
            context.variables[target_variable_name] = converted

        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "target_type": _normalize_runtime_target_type(target_type),
            "value": converted,
            "variable_name": target_variable_name,
            "in_place": in_place,
        }

    def _execute_numeric_variable_delta(self, node: dict, context: RuntimeContext, *, multiplier: int) -> dict:
        node_config = _node_config(node)
        variable_name = node_config.get("variable_name") or node_config.get("name")
        if not isinstance(variable_name, str) or not variable_name.strip():
            return _failed_result(node, "data.variable_name_required", "variable name is required")
        step_value = _resolve_value(node_config.get("step", 1), context)
        current_value = context.variables.get(variable_name.strip(), 0)
        try:
            step = float(step_value)
            current = float(current_value)
        except (TypeError, ValueError):
            return _failed_result(node, "data.numeric_variable_invalid", "current value and step must be numeric")
        value: int | float = current + (step * multiplier)
        if value.is_integer():
            value = int(value)
        context.variables[variable_name.strip()] = value
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name.strip(),
            "step": step,
            "value": value,
        }

    def _execute_create_list(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        variable_name = node_config.get("variable_name") or node_config.get("name")
        if not isinstance(variable_name, str) or not variable_name.strip():
            return _failed_result(node, "data.variable_name_required", "variable name is required")
        items = _resolve_value(node_config.get("items", []), context)
        if not isinstance(items, list):
            return _failed_result(node, "data.list_items_invalid", "items must be a list")
        list_value = list(items)
        context.variables[variable_name.strip()] = list_value
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name.strip(),
            "items": list_value,
        }

    def _execute_list_append(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        variable_name = node_config.get("variable_name") or node_config.get("name")
        if not isinstance(variable_name, str) or not variable_name.strip():
            return _failed_result(node, "data.variable_name_required", "variable name is required")
        current = context.variables.get(variable_name.strip())
        if current is None:
            current = []
        if not isinstance(current, list):
            return _failed_result(node, "data.list_target_invalid", "target variable must be a list")
        value = _resolve_value(node_config.get("value"), context)
        updated = list(current)
        updated.append(value)
        context.variables[variable_name.strip()] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name.strip(),
            "value": value,
            "items": updated,
        }

    def _execute_list_extend(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        variable_name = node_config.get("variable_name") or node_config.get("name")
        if not isinstance(variable_name, str) or not variable_name.strip():
            return _failed_result(node, "data.variable_name_required", "variable name is required")
        current = context.variables.get(variable_name.strip())
        if current is None:
            current = []
        if not isinstance(current, list):
            return _failed_result(node, "data.list_target_invalid", "target variable must be a list")
        items = _resolve_value(node_config.get("items", []), context)
        if not isinstance(items, list):
            return _failed_result(node, "data.list_items_invalid", "items must be a list")
        updated = list(current)
        updated.extend(items)
        context.variables[variable_name.strip()] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name.strip(),
            "items": updated,
        }

    def _execute_list_get(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        index = _resolve_int(_resolve_value(node_config.get("index"), context), default=0)
        try:
            value = list_value[index]
        except IndexError:
            return _failed_result(node, "data.list_index_out_of_range", "list index is out of range")
        _store_optional_variable_name(node_config.get("output_variable_name"), context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "index": index,
            "value": value,
        }

    def _execute_list_set(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        index = _resolve_int(_resolve_value(node_config.get("index"), context), default=0)
        if index < 0 or index >= len(list_value):
            return _failed_result(node, "data.list_index_out_of_range", "list index is out of range")
        value = _resolve_value(node_config.get("value"), context)
        updated = list(list_value)
        updated[index] = value
        context.variables[variable_name] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "index": index,
            "value": value,
            "items": updated,
        }

    def _execute_list_index(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        value = _resolve_value(node_config.get("value"), context)
        try:
            index = list_value.index(value)
        except ValueError:
            return _failed_result(node, "data.list_value_not_found", "list value not found")
        _store_optional_variable_name(node_config.get("output_variable_name"), context, index)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "searched_value": value,
            "value": index,
        }

    def _execute_list_length(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        value = len(list_value)
        _store_optional_variable_name(node_config.get("output_variable_name"), context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "value": value,
        }

    def _execute_list_insert(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        index = _resolve_int(_resolve_value(node_config.get("index"), context), default=0)
        value = _resolve_value(node_config.get("value"), context)
        updated = list(list_value)
        if index < 0:
            index = 0
        if index > len(updated):
            index = len(updated)
        updated.insert(index, value)
        context.variables[variable_name] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "index": index,
            "value": value,
            "items": updated,
        }

    def _execute_list_remove(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        updated = list(list_value)
        if "index" in node_config:
            index = _resolve_int(_resolve_value(node_config.get("index"), context), default=0)
            if index < 0 or index >= len(updated):
                return _failed_result(node, "data.list_index_out_of_range", "list index is out of range")
            removed = updated.pop(index)
        else:
            value = _resolve_value(node_config.get("value"), context)
            try:
                updated.remove(value)
                removed = value
            except ValueError:
                return _failed_result(node, "data.list_value_not_found", "list value not found")
        context.variables[variable_name] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "value": removed,
            "items": updated,
        }

    def _execute_list_slice(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        start = _resolve_int(_resolve_value(node_config.get("start"), context), default=0)
        end_value = node_config.get("end")
        end = _resolve_int(_resolve_value(end_value, context), default=len(list_value)) if end_value is not None else len(list_value)
        sliced = list_value[start:end]
        _store_optional_variable_name(node_config.get("output_variable_name"), context, sliced)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "start": start,
            "end": end,
            "value": sliced,
        }

    def _execute_list_sort(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        updated = list(list_value)
        updated.sort()
        context.variables[variable_name] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "items": updated,
        }

    def _execute_list_reverse(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        list_value, variable_name = _require_runtime_list(node, context)
        if isinstance(list_value, dict):
            return list_value
        updated = list(list_value)
        updated.reverse()
        context.variables[variable_name] = updated
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name,
            "items": updated,
        }

    def _execute_evaluate_expression(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        expression = _resolve_value(node_config.get("expression"), context)
        if not isinstance(expression, str) or not expression.strip():
            return _failed_result(node, "data.expression_required", "expression is required")
        try:
            value = _safe_eval_expression(expression, context.variables)
        except ValueError as exc:
            return _failed_result(node, "data.expression_invalid", str(exc))
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "expression": expression,
            "value": value,
        }

    def _execute_regex_replace(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        text = _resolve_value(node_config.get("text", ""), context)
        pattern = _resolve_value(node_config.get("pattern"), context)
        replacement = _resolve_value(node_config.get("replacement", ""), context)
        if not isinstance(text, str):
            return _failed_result(node, "data.regex_text_invalid", "text must be a string")
        if not isinstance(pattern, str) or not pattern:
            return _failed_result(node, "data.regex_pattern_required", "pattern is required")
        if not isinstance(replacement, str):
            replacement = str(replacement)
        try:
            value = re.sub(pattern, replacement, text)
        except re.error as exc:
            return _failed_result(node, "data.regex_pattern_invalid", str(exc))
        _store_optional_variable(node_config, context, value)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "value": value,
        }

    def _execute_write_text_file(self, node: dict, context: RuntimeContext) -> dict:
        if not self._is_file_access_allowed():
            return _failed_result(node, "file.access_denied", "file access is disabled")
        node_config = _node_config(node)
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "file.path_required", "file path is required")
        encoding = str(_resolve_value(node_config.get("encoding", "utf-8"), context))
        content = _resolve_value(node_config.get("content", ""), context)
        text = content if isinstance(content, str) else json.dumps(content, ensure_ascii=False)
        path = _resolve_runtime_path(path_value, context)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text, encoding=encoding)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(path.resolve()),
            "encoding": encoding,
            "bytes_written": len(text.encode(encoding)),
        }

    def _execute_read_text_file(self, node: dict, context: RuntimeContext) -> dict:
        if not self._is_file_access_allowed():
            return _failed_result(node, "file.access_denied", "file access is disabled")
        node_config = _node_config(node)
        path_value = _resolve_value(node_config.get("path"), context)
        if not isinstance(path_value, str) or not path_value.strip():
            return _failed_result(node, "file.path_required", "file path is required")
        encoding = str(_resolve_value(node_config.get("encoding", "utf-8"), context))
        path = _resolve_runtime_path(path_value, context)
        content = path.read_text(encoding=encoding)
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(path.resolve()),
            "encoding": encoding,
            "content": content,
        }
        variable_name = node_config.get("variable_name")
        if isinstance(variable_name, str) and variable_name.strip():
            context.variables[variable_name.strip()] = content
        return result

    def _execute_read_csv_cell(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        table_result = _read_csv_table(node, context)
        if table_result.get("status") == "failed":
            return table_result
        row_index = _resolve_int(node_config.get("row_index"), default=0)
        column = _resolve_value(node_config.get("column"), context)
        rows = table_result["rows"]
        try:
            row = rows[row_index]
        except IndexError:
            return _failed_result(node, "file.csv_row_out_of_range", "CSV row_index is out of range")
        if isinstance(row, dict):
            if column is None:
                return _failed_result(node, "file.csv_column_required", "CSV column is required")
            value = row.get(str(column))
        else:
            column_index = _resolve_int(column, default=0)
            try:
                value = row[column_index]
            except IndexError:
                return _failed_result(node, "file.csv_column_out_of_range", "CSV column is out of range")
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": table_result["path"],
            "row_index": row_index,
            "column": column,
            "value": value,
        }
        _store_optional_variable(node_config, context, value)
        return result

    def _execute_read_csv_row(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        table_result = _read_csv_table(node, context)
        if table_result.get("status") == "failed":
            return table_result
        row_index = _resolve_int(node_config.get("row_index"), default=0)
        rows = table_result["rows"]
        try:
            row = rows[row_index]
        except IndexError:
            return _failed_result(node, "file.csv_row_out_of_range", "CSV row_index is out of range")
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": table_result["path"],
            "row_index": row_index,
            "row": row,
        }
        _store_optional_variable(node_config, context, row)
        return result

    def _execute_read_csv_table(self, node: dict, context: RuntimeContext) -> dict:
        table_result = _read_csv_table(node, context)
        if table_result.get("status") == "failed":
            return table_result
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": table_result["path"],
            "encoding": table_result["encoding"],
            "has_header": table_result["has_header"],
            "headers": table_result["headers"],
            "rows": table_result["rows"],
            "row_count": len(table_result["rows"]),
        }
        _store_optional_variable(_node_config(node), context, table_result["rows"])
        return result

    def _execute_write_excel_cell(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, workbook, worksheet = _load_excel_sheet_for_write(node, context)
        if worksheet is None:
            return workbook_path
        cell = _resolve_value(node_config.get("cell"), context)
        if not isinstance(cell, str) or not cell.strip():
            workbook.close()
            return _failed_result(node, "excel.cell_required", "Excel cell is required")
        value = _resolve_value(node_config.get("value"), context)
        worksheet[cell.strip()] = value
        workbook.save(workbook_path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(workbook_path.resolve()),
            "sheet_name": worksheet.title,
            "cell": cell.strip(),
            "value": value,
        }

    def _execute_write_excel_row(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, workbook, worksheet = _load_excel_sheet_for_write(node, context)
        if worksheet is None:
            return workbook_path
        row_index = _resolve_int(_resolve_value(node_config.get("row_index"), context), default=1)
        data = _resolve_value(node_config.get("data"), context)
        if row_index < 1:
            workbook.close()
            return _failed_result(node, "excel.row_out_of_range", "Excel row_index must be >= 1")
        values = _normalize_excel_row_payload(worksheet, data)
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)
        workbook.save(workbook_path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(workbook_path.resolve()),
            "sheet_name": worksheet.title,
            "row_index": row_index,
            "row": values,
        }

    def _execute_write_excel_table(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, workbook, worksheet = _load_excel_sheet_for_write(node, context)
        if worksheet is None:
            return workbook_path
        data = _resolve_value(node_config.get("data"), context)
        has_header = bool(_resolve_value(node_config.get("has_header", True), context))
        headers, rows = _normalize_excel_table_payload(data, has_header=has_header)
        _clear_worksheet(worksheet)
        current_row = 1
        if has_header and headers:
            for column_index, header in enumerate(headers, start=1):
                worksheet.cell(row=current_row, column=column_index, value=header)
            current_row += 1
        for row in rows:
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row=current_row, column=column_index, value=value)
            current_row += 1
        workbook.save(workbook_path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(workbook_path.resolve()),
            "sheet_name": worksheet.title,
            "has_header": has_header,
            "row_count": len(rows),
        }

    def _execute_write_excel_file(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        mode = str(_resolve_value(node_config.get("mode", "create"), context)).lower()
        workbook_path, workbook, worksheet = _load_excel_sheet_for_write(
            node,
            context,
            create_mode=(mode == "create"),
        )
        if worksheet is None:
            return workbook_path
        headers = _resolve_value(node_config.get("headers", []), context)
        rows = _resolve_value(node_config.get("rows", []), context)
        if not isinstance(headers, list):
            workbook.close()
            return _failed_result(node, "excel.headers_invalid", "Excel headers must be a list")
        if not isinstance(rows, list):
            workbook.close()
            return _failed_result(node, "excel.rows_invalid", "Excel rows must be a list")
        if mode == "create":
            _clear_worksheet(worksheet)
            current_row = 1
            if headers:
                for column_index, header in enumerate(headers, start=1):
                    worksheet.cell(row=current_row, column=column_index, value=header)
                current_row += 1
        else:
            current_row = worksheet.max_row + 1 if worksheet.max_row > 1 or worksheet["A1"].value is not None else 1
            if current_row == 1 and headers:
                for column_index, header in enumerate(headers, start=1):
                    worksheet.cell(row=current_row, column=column_index, value=header)
                current_row += 1
        for row in rows:
            row_values = list(row.values()) if isinstance(row, dict) else list(row)
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=current_row, column=column_index, value=value)
            current_row += 1
        workbook.save(workbook_path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(workbook_path.resolve()),
            "sheet_name": worksheet.title,
            "mode": mode,
            "row_count": len(rows),
        }

    def _execute_update_excel_cells(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, workbook, worksheet = _load_excel_sheet_for_write(node, context, create_if_not_exists=False)
        if worksheet is None:
            return workbook_path
        updates = _resolve_value(node_config.get("updates"), context)
        if not isinstance(updates, list):
            workbook.close()
            return _failed_result(node, "excel.updates_invalid", "Excel updates must be a list")
        updated_count = 0
        headers = [cell.value for cell in worksheet[1]]
        for update in updates:
            if not isinstance(update, dict):
                workbook.close()
                return _failed_result(node, "excel.update_invalid", "each Excel update must be an object")
            row_index = _resolve_int(update.get("row_index"), default=0)
            if row_index < 1:
                workbook.close()
                return _failed_result(node, "excel.row_out_of_range", "Excel row_index must be >= 1")
            if "column_name" in update:
                try:
                    column_index = headers.index(update.get("column_name")) + 1
                except ValueError:
                    workbook.close()
                    return _failed_result(node, "excel.column_not_found", "Excel column_name was not found")
            else:
                column_index = _resolve_int(update.get("column_index"), default=0)
                if column_index < 1:
                    workbook.close()
                    return _failed_result(node, "excel.column_out_of_range", "Excel column_index must be >= 1")
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=_resolve_value(update.get("value"), context),
            )
            updated_count += 1
        workbook.save(workbook_path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(workbook_path.resolve()),
            "sheet_name": worksheet.title,
            "updated_count": updated_count,
        }

    def _execute_update_excel_batch(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, workbook, worksheet = _load_excel_sheet_for_write(node, context, create_if_not_exists=False)
        if worksheet is None:
            return workbook_path
        condition = _resolve_value(node_config.get("condition"), context)
        updates = _resolve_value(node_config.get("updates"), context)
        if not isinstance(condition, str) or not condition.strip():
            workbook.close()
            return _failed_result(node, "excel.condition_required", "Excel update_batch requires condition")
        if not isinstance(updates, dict):
            workbook.close()
            return _failed_result(node, "excel.updates_invalid", "Excel update_batch updates must be an object")
        headers = [cell.value for cell in worksheet[1]]
        updated_count = 0
        for row_index in range(2, worksheet.max_row + 1):
            row_payload = {
                str(headers[column_index - 1]): worksheet.cell(row=row_index, column=column_index).value
                for column_index in range(1, len(headers) + 1)
            }
            try:
                matched = _evaluate_excel_batch_condition(condition, row_payload)
            except Exception as exc:
                workbook.close()
                return _failed_result(node, "excel.condition_invalid", str(exc))
            if not matched:
                continue
            for column_name, value in updates.items():
                try:
                    column_index = headers.index(column_name) + 1
                except ValueError:
                    workbook.close()
                    return _failed_result(node, "excel.column_not_found", "Excel column_name was not found")
                worksheet.cell(row=row_index, column=column_index, value=_resolve_value(value, context))
            updated_count += 1
        workbook.save(workbook_path)
        workbook.close()
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": str(workbook_path.resolve()),
            "sheet_name": worksheet.title,
            "updated_count": updated_count,
        }

    def _execute_read_excel_cell(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, worksheet = _require_excel_sheet(node, context)
        if worksheet is None:
            return workbook_path
        cell = _resolve_value(node_config.get("cell"), context)
        if not isinstance(cell, str) or not cell.strip():
            worksheet.parent.close()
            return _failed_result(node, "excel.cell_required", "Excel cell is required")
        value = worksheet[cell.strip()].value
        resolved_path = str(workbook_path.resolve())
        sheet_name = worksheet.title
        worksheet.parent.close()
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": resolved_path,
            "sheet_name": sheet_name,
            "cell": cell.strip(),
            "value": value,
        }
        _store_optional_variable(node_config, context, value)
        return result

    def _execute_read_excel_row(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        workbook_path, worksheet = _require_excel_sheet(node, context)
        if worksheet is None:
            return workbook_path
        row_index = _resolve_int(_resolve_value(node_config.get("row_index"), context), default=1)
        if row_index < 1:
            worksheet.parent.close()
            return _failed_result(node, "excel.row_out_of_range", "Excel row_index must be >= 1")
        try:
            row = [cell.value for cell in worksheet[row_index]]
        except (IndexError, ValueError):
            worksheet.parent.close()
            return _failed_result(node, "excel.row_out_of_range", "Excel row_index is out of range")
        resolved_path = str(workbook_path.resolve())
        sheet_name = worksheet.title
        worksheet.parent.close()
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": resolved_path,
            "sheet_name": sheet_name,
            "row_index": row_index,
            "row": row,
        }
        _store_optional_variable(node_config, context, row)
        return result

    def _execute_read_excel_table(self, node: dict, context: RuntimeContext) -> dict:
        table_result = _read_excel_table(node, context)
        if table_result.get("status") == "failed":
            return table_result
        result = {
            "status": "succeeded",
            "node_id": node["node_id"],
            "path": table_result["path"],
            "sheet_name": table_result["sheet_name"],
            "has_header": table_result["has_header"],
            "headers": table_result["headers"],
            "rows": table_result["rows"],
            "row_count": len(table_result["rows"]),
        }
        _store_optional_variable(_node_config(node), context, table_result["rows"])
        return result

    def _execute_python_run(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        if not self._is_python_execution_allowed():
            return _failed_result(node, "python.execution_disabled", "python execution is disabled")
        project_runtime_prepare_error = self._runtime_settings.get(
            "python_project_runtime_prepare_error"
        )
        if isinstance(project_runtime_prepare_error, str) and project_runtime_prepare_error.strip():
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.runtime_prepare_failed",
                "message": project_runtime_prepare_error.strip(),
                "stdout": "",
                "stderr": "",
            }
        code = _resolve_value(node_config.get("code"), context)
        if not isinstance(code, str) or not code.strip():
            return _failed_result(node, "python.code_required", "python.run requires node_config.code")
        default_variable_name = node_config.get("variable_name")
        python_executable_path = self._runtime_settings.get("python_executable_path")
        python_runtime_root = self._runtime_settings.get("python_project_runtime_root")
        python_runtime_source = (
            "project_runtime"
            if self._runtime_settings.get("python_project_runtime_enabled", False)
            else "preferences"
        )
        if self._runtime_settings.get("python_project_runtime_enabled", False):
            return self._execute_python_run_in_project_runtime(
                node=node,
                context=context,
                code=code,
                default_variable_name=default_variable_name,
                python_executable_path=python_executable_path,
                python_runtime_root=python_runtime_root,
                python_runtime_source=python_runtime_source,
            )
        return _failed_result(
            node,
            "python.runtime_disabled",
            "python.run requires project python runtime to be enabled",
        )

    def _execute_python_run_in_project_runtime(
        self,
        *,
        node: dict,
        context: RuntimeContext,
        code: str,
        default_variable_name: object,
        python_executable_path: object,
        python_runtime_root: object,
        python_runtime_source: str,
    ) -> dict:
        if not isinstance(python_executable_path, str) or not python_executable_path.strip():
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.runtime_prepare_failed",
                "message": "python runtime executable path is unavailable",
                "stdout": "",
                "stderr": "",
            }
        python_executable = Path(python_executable_path)
        capture_stdout_stderr = self._should_capture_stdout_stderr()
        payload = {
            "code": code,
            "variables": _make_python_run_json_safe(context.variables),
            "result_variable": default_variable_name,
            "allowed_imports": sorted(_PYTHON_ALLOWED_IMPORTS),
        }
        command_timeout = _resolve_int(self._runtime_settings.get("python_timeout_seconds"), default=60)
        if command_timeout <= 0:
            command_timeout = 60
        working_directory = context.project_directory or context.workspace_root
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        env.pop("PYTHONPATH", None)
        try:
            _validate_python_run_code(code)
            with tempfile.TemporaryDirectory(prefix="weconduct-python-run-") as temp_dir:
                temp_root = Path(temp_dir)
                input_path = temp_root / "input.json"
                output_path = temp_root / "output.json"
                runner_path = temp_root / "runner.py"
                input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
                runner_path.write_text(_PYTHON_RUN_CHILD_SCRIPT, encoding="utf-8")
                process = subprocess.run(
                    [str(python_executable), str(runner_path), str(input_path), str(output_path)],
                    cwd=str(working_directory) if working_directory is not None else None,
                    env=env,
                    capture_output=True,
                    text=True,
                    timeout=command_timeout,
                    check=False,
                )
                child_result = self._read_python_run_child_result(output_path)
        except PythonCodeRejected as exc:
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.code_rejected",
                "message": str(exc),
                "exception_type": type(exc).__name__,
                "stdout": "",
                "stderr": "",
            }
        except subprocess.TimeoutExpired:
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.execution_timeout",
                "message": f"python.run timed out after {command_timeout} seconds",
                "stdout": "",
                "stderr": "",
            }
        except OSError as exc:
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.execution_failed",
                "message": str(exc),
                "exception_type": type(exc).__name__,
                "stdout": "",
                "stderr": "",
            }
        except ValueError as exc:
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.execution_failed",
                "message": str(exc),
                "exception_type": type(exc).__name__,
                "stdout": "",
                "stderr": "",
            }
        child_stdout = child_result.get("stdout", "")
        child_stderr = child_result.get("stderr", "")
        if process.returncode != 0 and child_result.get("status") != "failed":
            detail = process.stderr.strip() or process.stdout.strip() or f"exit code {process.returncode}"
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.execution_failed",
                "message": detail,
                "stdout": child_stdout if capture_stdout_stderr else "",
                "stderr": child_stderr if capture_stdout_stderr else "",
            }
        if child_result.get("status") == "failed":
            return {
                "status": "failed",
                "node_id": node["node_id"],
                "error_code": "python.execution_failed",
                "message": str(child_result.get("message") or "python child execution failed"),
                "exception_type": child_result.get("exception_type"),
                "stdout": child_stdout if capture_stdout_stderr else "",
                "stderr": child_stderr if capture_stdout_stderr else "",
            }
        child_variables = child_result.get("variables")
        if isinstance(child_variables, dict):
            context.variables.clear()
            context.variables.update(child_variables)
        return self._finalize_python_run_result(
            node=node,
            context=context,
            result=child_result.get("result"),
            result_variable=child_result.get("result_variable"),
            python_runtime_source=python_runtime_source,
            python_executable_path=python_executable_path,
            python_runtime_root=python_runtime_root,
            stdout=child_stdout if capture_stdout_stderr else "",
            stderr=child_stderr if capture_stdout_stderr else "",
        )

    def _read_python_run_child_result(self, output_path: Path) -> dict:
        if not output_path.exists():
            raise ValueError(f"python.run child result file missing: {output_path}")
        try:
            payload = json.loads(output_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ValueError(f"python.run child result is unreadable: {output_path}") from exc
        if not isinstance(payload, dict):
            raise ValueError(f"python.run child result must be object: {output_path}")
        return payload

    def _finalize_python_run_result(
        self,
        *,
        node: dict,
        context: RuntimeContext,
        result: Any,
        result_variable: Any,
        python_runtime_source: str,
        python_executable_path: object,
        python_runtime_root: object,
        stdout: str,
        stderr: str,
    ) -> dict:
        if isinstance(result_variable, str) and result_variable.strip():
            context.variables[result_variable.strip()] = result
            result_variable = result_variable.strip()
        else:
            result_variable = None
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "result": result,
            "result_variable": result_variable,
            "python_runtime_source": python_runtime_source,
            "python_executable_path": python_executable_path,
            "python_runtime_root": python_runtime_root,
            "stdout": stdout,
            "stderr": stderr,
        }

    def _is_file_access_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_file_access", True))

    def _is_python_execution_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_python_execution", False))

    def _is_browser_executor_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_browser_executor", False))

    def _is_local_network_access_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_local_network_access", False))

    def _is_remote_network_access_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_remote_network_access", False))

    def _is_browser_screenshots_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_browser_screenshots", True))

    def _is_cookie_manipulation_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_cookie_manipulation", True))

    def _is_browser_storage_manipulation_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_browser_storage_manipulation", True))

    def _is_browser_uploads_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_browser_uploads", True))

    def _is_browser_downloads_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_browser_downloads", False))

    def _is_new_browser_windows_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_new_browser_windows", True))

    def _is_js_injection_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_js_injection", False))

    def _is_js_evaluation_allowed(self) -> bool:
        return bool(self._runtime_settings.get("allow_js_evaluation", False))

    def _should_capture_stdout_stderr(self) -> bool:
        return bool(self._runtime_settings.get("capture_stdout_stderr", True))

    def _should_use_legacy_webcontrol_click_fallback(
        self,
        *,
        context: RuntimeContext,
        error: Exception,
    ) -> bool:
        if "strict mode violation" not in str(error):
            return False
        root_metadata = context.flow_runtime.get("graph_root_metadata")
        if not isinstance(root_metadata, dict):
            return False
        return root_metadata.get("source_kind") == "webcontrol_main_flow"

    def _execute_session_apply_auth_session(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        page = _require_browser_page(context)
        cookies = _resolve_value(node_config.get("cookies", []), context)
        if cookies is None:
            cookies = []
        if not isinstance(cookies, list):
            return _failed_result(node, "session.cookies_invalid", "cookies must be a list")
        normalized_cookies = []
        for cookie in cookies:
            if not isinstance(cookie, dict):
                return _failed_result(node, "session.cookie_invalid", "each cookie must be an object")
            name = cookie.get("name")
            value = cookie.get("value")
            if not isinstance(name, str) or not name.strip():
                return _failed_result(node, "session.cookie_name_required", "cookie name is required")
            normalized = dict(cookie)
            normalized["name"] = name.strip()
            normalized["value"] = "" if value is None else str(value)
            normalized_cookies.append(normalized)
        if normalized_cookies:
            page.context.add_cookies(normalized_cookies)

        local_storage = _resolve_value(node_config.get("local_storage", {}), context)
        storage_origin_count = _apply_local_storage(page, local_storage)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "cookie_count": len(normalized_cookies),
            "local_storage_origin_count": storage_origin_count,
        }

    def _execute_dialog_set_agent_config(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        config = context.browser_runtime.setdefault("dialog_config", {})
        default_action = _resolve_value(node_config.get("default_action", node_config.get("action", "accept")), context)
        prompt_text = _resolve_value(node_config.get("prompt_text", ""), context)
        if not isinstance(default_action, str) or default_action.strip() not in {"accept", "dismiss"}:
            return _failed_result(node, "dialog.action_invalid", "default_action must be accept or dismiss")
        config["default_action"] = default_action.strip()
        config["prompt_text"] = "" if prompt_text is None else str(prompt_text)
        _ensure_dialog_handler(context)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "default_action": config["default_action"],
            "prompt_text": config["prompt_text"],
        }

    def _execute_dialog_switch_dialog_mode(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        mode = _resolve_value(node_config.get("mode", "auto"), context)
        if not isinstance(mode, str) or not mode.strip():
            return _failed_result(node, "dialog.mode_required", "dialog mode is required")
        normalized_mode = mode.strip()
        if normalized_mode not in {"auto", "manual", "accept", "dismiss"}:
            return _failed_result(node, "dialog.mode_invalid", "dialog mode must be auto, manual, accept, or dismiss")
        config = context.browser_runtime.setdefault("dialog_config", {})
        config["mode"] = normalized_mode
        if normalized_mode in {"accept", "dismiss"}:
            config["default_action"] = normalized_mode
        _ensure_dialog_handler(context)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "mode": normalized_mode,
            "default_action": config.get("default_action", "accept"),
        }

    def _execute_dialog_watch_dialogs(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        _ensure_dialog_handler(context)
        timeout_ms = _resolve_int(_resolve_value(node_config.get("timeout"), context), default=0)
        if timeout_ms > 0:
            _require_browser_page(context).wait_for_timeout(timeout_ms)
        records = list(context.browser_runtime.get("dialog_records", []))
        _store_optional_variable(node_config, context, records)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "dialog_count": len(records),
            "dialogs": records,
        }

    def _execute_dialog_handle_dialogs(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        _ensure_dialog_handler(context)
        records = list(context.browser_runtime.get("dialog_records", []))
        if bool(_resolve_value(node_config.get("clear_after", False), context)):
            context.browser_runtime["dialog_records"] = []
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "handled_count": len(records),
            "dialogs": records,
            "cleared": bool(node_config.get("clear_after", False)),
        }

    def _execute_control_foreach(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        variable_name = node_config.get("variable") or node_config.get("variable_name")
        if not isinstance(variable_name, str) or not variable_name.strip():
            return _failed_result(node, "control.variable_required", "variable is required")
        items = context.variables.get(variable_name.strip())
        if not isinstance(items, list):
            return _failed_result(node, "control.foreach_items_invalid", "variable must be a list")
        item_var = node_config.get("item_var")
        if not isinstance(item_var, str) or not item_var.strip():
            item_var = "item"
        index_var = node_config.get("index_var")
        if not isinstance(index_var, str) or not index_var.strip():
            index_var = "index"
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "variable_name": variable_name.strip(),
            "item_var": item_var.strip(),
            "index_var": index_var.strip(),
            "items": list(items),
            "iteration_count": len(items),
        }

    def _execute_control_jump_to_step(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        target_node_id = node_config.get("target_node_id")
        target_step = node_config.get("target_step")
        condition = _resolve_value(node_config.get("condition", "true"), context)
        max_jumps = _resolve_int(_resolve_value(node_config.get("max_jumps", -1), context), default=-1)
        if not isinstance(condition, str):
            condition = str(condition)
        should_jump = _evaluate_jump_condition(condition, context.variables)
        if not should_jump:
            return {
                "status": "succeeded",
                "node_id": node["node_id"],
                "jump_executed": False,
                "jump_count": 0,
                "target_node_id": target_node_id,
                "target_step": target_step,
            }
        flow_runtime = context.flow_runtime.setdefault("jump_runtime", {})
        jump_key = node["node_id"]
        jump_count = int(flow_runtime.get(jump_key, 0))
        if max_jumps >= 0 and jump_count >= max_jumps:
            return {
                "status": "succeeded",
                "node_id": node["node_id"],
                "jump_executed": False,
                "jump_count": jump_count,
                "stopped_by_max_jumps": True,
                "target_node_id": target_node_id,
                "target_step": target_step,
            }
        flow_runtime[jump_key] = jump_count + 1
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "jump_executed": True,
            "jump_count": jump_count + 1,
            "stopped_by_max_jumps": False,
            "target_node_id": target_node_id,
            "target_step": target_step,
        }

    def _execute_control_end_foreach(self, node: dict, context: RuntimeContext) -> dict:
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "end_marker": True,
        }

    def _execute_control_foreach_continue(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        condition = _resolve_value(node_config.get("condition", "true"), context)
        if not isinstance(condition, str):
            condition = str(condition)
        level = _resolve_int(_resolve_value(node_config.get("level", 1), context), default=1)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "continue_triggered": _evaluate_jump_condition(condition, context.variables),
            "level": level if level >= 1 else 1,
        }

    def _execute_control_foreach_break(self, node: dict, context: RuntimeContext) -> dict:
        node_config = _node_config(node)
        condition = _resolve_value(node_config.get("condition", "true"), context)
        if not isinstance(condition, str):
            condition = str(condition)
        level = _resolve_int(_resolve_value(node_config.get("level", 1), context), default=1)
        return {
            "status": "succeeded",
            "node_id": node["node_id"],
            "break_triggered": _evaluate_jump_condition(condition, context.variables),
            "level": level if level >= 1 else 1,
        }

    def execute_component_graph(
        self,
        *,
        graph_model: dict,
        inputs: dict[str, Any] | None = None,
    ) -> dict:
        graph_nodes = graph_model.get("nodes") if isinstance(graph_model, dict) else None
        graph_edges = graph_model.get("edges") if isinstance(graph_model, dict) else None
        if not isinstance(graph_nodes, list) or not isinstance(graph_edges, list):
            return {
                "status": "failed",
                "error_code": "component.graph_invalid",
                "message": "component graph is invalid",
            }
        context = RuntimeContext(runtime_settings=dict(self._runtime_settings))
        context.variables.update(inputs or {})
        node_lookup = {node.get("node_id"): node for node in graph_nodes if isinstance(node, dict)}
        if len(node_lookup) != len(graph_nodes):
            return {
                "status": "failed",
                "error_code": "component.graph_invalid",
                "message": "component graph node ids are invalid",
            }
        control_edges_by_source = _build_control_edges_by_source(graph_edges)
        node_outputs: dict[str, Any] = {}
        executed_node_ids: list[str] = []
        index = 0
        max_steps = max(1, len(graph_nodes) * 20)
        step_count = 0
        try:
            while 0 <= index < len(graph_nodes):
                if step_count >= max_steps:
                    return {
                        "status": "failed",
                        "error_code": "component.execution_step_limit_exceeded",
                        "message": "component execution step limit exceeded",
                        "executed_node_ids": executed_node_ids,
                        "outputs": dict(node_outputs),
                        "variables": dict(context.variables),
                    }
                step_count += 1
                node = graph_nodes[index]
                if not isinstance(node, dict):
                    return {
                        "status": "failed",
                        "error_code": "component.graph_invalid",
                        "message": "component node is invalid",
                    }
                output = execute_runtime_node(node, context, self)
                node_outputs[node["node_id"]] = output
                executed_node_ids.append(node["node_id"])
                if isinstance(output, dict) and output.get("status") == "failed":
                    return {
                        "status": "failed",
                        "error_code": output.get("error_code") or "runtime.node_failed",
                        "message": output.get("message", "runtime node failed"),
                        "executed_node_ids": executed_node_ids,
                        "outputs": dict(node_outputs),
                        "variables": dict(context.variables),
                    }
                next_index = index + 1
                if node.get("node_kind") == "control.foreach":
                    loop_body_end, loop_exit = _resolve_foreach_targets(node["node_id"], control_edges_by_source, node_lookup)
                    iterable = output.get("items", []) if isinstance(output, dict) else []
                    if not isinstance(iterable, list):
                        return {
                            "status": "failed",
                            "error_code": "control.foreach_items_invalid",
                            "message": "foreach items must be a list",
                            "executed_node_ids": executed_node_ids,
                            "outputs": dict(node_outputs),
                            "variables": dict(context.variables),
                        }
                    if not iterable:
                        if loop_exit is not None:
                            index = loop_exit
                            continue
                        next_index = index + 1
                    else:
                        item_var = output.get("item_var", "item")
                        index_var = output.get("index_var", "index")
                        body_start = _find_node_index(graph_nodes, loop_body_end) if loop_body_end is not None else None
                        exit_index = _find_node_index(graph_nodes, loop_exit) if loop_exit is not None else None
                        if body_start is None:
                            return {
                                "status": "failed",
                                "error_code": "control.foreach_body_missing",
                                "message": "foreach body node is missing",
                                "executed_node_ids": executed_node_ids,
                                "outputs": dict(node_outputs),
                                "variables": dict(context.variables),
                            }
                        for item_index, item_value in enumerate(iterable):
                            context.variables[item_var] = item_value
                            context.variables[index_var] = item_index
                            body_cursor = body_start
                            while True:
                                if step_count >= max_steps:
                                    return {
                                        "status": "failed",
                                        "error_code": "component.execution_step_limit_exceeded",
                                        "message": "component execution step limit exceeded",
                                        "executed_node_ids": executed_node_ids,
                                        "outputs": dict(node_outputs),
                                        "variables": dict(context.variables),
                                    }
                                step_count += 1
                                body_node = graph_nodes[body_cursor]
                                body_output = execute_runtime_node(body_node, context, self)
                                node_outputs[body_node["node_id"]] = body_output
                                executed_node_ids.append(body_node["node_id"])
                                if isinstance(body_output, dict) and body_output.get("status") == "failed":
                                    return {
                                        "status": "failed",
                                        "error_code": body_output.get("error_code") or "runtime.node_failed",
                                        "message": body_output.get("message", "runtime node failed"),
                                        "executed_node_ids": executed_node_ids,
                                        "outputs": dict(node_outputs),
                                        "variables": dict(context.variables),
                                    }
                                if body_node["node_id"] == loop_body_end:
                                    break
                                body_cursor += 1
                            if exit_index is not None:
                                context.flow_runtime["foreach_exit_index"] = exit_index
                        if loop_exit is not None:
                            index = loop_exit
                            continue
                        next_index = index + 1
                elif node.get("node_kind") == "control.jump_to_step":
                    jump_result = output
                    if isinstance(jump_result, dict) and jump_result.get("jump_executed") is True:
                        target_node_id = jump_result.get("target_node_id")
                        target_index = None
                        if isinstance(target_node_id, str) and target_node_id in node_lookup:
                            target_index = _find_node_index(graph_nodes, target_node_id)
                        elif jump_result.get("target_step") is not None:
                            target_index = _resolve_step_target_index(graph_nodes, jump_result.get("target_step"))
                        if target_index is None:
                            return {
                                "status": "failed",
                                "error_code": "control.jump_target_missing",
                                "message": "jump target node was not found",
                                "executed_node_ids": executed_node_ids,
                                "outputs": dict(node_outputs),
                                "variables": dict(context.variables),
                            }
                        next_index = target_index
                        continue
                index = next_index
            return {
                "status": "succeeded",
                "executed_node_ids": executed_node_ids,
                "outputs": node_outputs,
                "variables": dict(context.variables),
            }
        finally:
            context.close()


def execute_runtime_node(node: dict, context: RuntimeContext, registry: RuntimeExecutorRegistry) -> dict:
    output = registry.execute(node.get("node_kind"), node, context)
    context.node_outputs[node["node_id"]] = output
    return output


def _node_config(node: dict) -> dict:
    node_config = node.get("node_config")
    base_config = dict(node_config) if isinstance(node_config, dict) else {}
    runtime_input_overrides = node.get("__runtime_input_overrides__")
    if not isinstance(runtime_input_overrides, dict) or not runtime_input_overrides:
        return base_config
    merged_config = dict(base_config)
    for raw_key, value in runtime_input_overrides.items():
        if not isinstance(raw_key, str):
            continue
        normalized_key = raw_key.strip()
        if not normalized_key:
            continue
        if "." not in normalized_key:
            merged_config[normalized_key] = value
            continue
        segments = [segment.strip() for segment in normalized_key.split(".") if segment.strip()]
        if not segments:
            continue
        cursor = merged_config
        for segment in segments[:-1]:
            current_value = cursor.get(segment)
            if not isinstance(current_value, dict):
                current_value = {}
                cursor[segment] = current_value
            cursor = current_value
        cursor[segments[-1]] = value
    return merged_config


def _failed_result(node: dict, error_code: str, message: str) -> dict:
    return {
        "status": "failed",
        "node_id": node["node_id"],
        "error_code": error_code,
        "message": message,
    }


def _last_node_output(context: RuntimeContext) -> Any:
    upstream = None
    for output in context.node_outputs.values():
        upstream = output
    return upstream


def _build_control_edges_by_source(edges: list[Any]) -> dict[str, list[dict]]:
    result: dict[str, list[dict]] = {}
    for edge in edges:
        if not isinstance(edge, dict):
            continue
        if edge.get("relation_layer") != "control":
            continue
        source_id = edge.get("from_node_id")
        if not isinstance(source_id, str):
            continue
        result.setdefault(source_id, []).append(edge)
    return result


def _resolve_foreach_targets(
    source_node_id: str,
    control_edges_by_source: dict[str, list[dict]],
    node_lookup: dict[str, dict],
) -> tuple[str | None, str | None]:
    edges = control_edges_by_source.get(source_node_id, [])
    if not edges:
        return None, None
    ordered_targets = [
        edge.get("to_node_id")
        for edge in edges
        if isinstance(edge.get("to_node_id"), str) and edge.get("to_node_id") in node_lookup
    ]
    if not ordered_targets:
        return None, None
    loop_body = ordered_targets[0]
    loop_exit = ordered_targets[1] if len(ordered_targets) > 1 else None
    return loop_body, loop_exit


def _find_node_index(graph_nodes: list[dict], node_id: str) -> int | None:
    for index, node in enumerate(graph_nodes):
        if isinstance(node, dict) and node.get("node_id") == node_id:
            return index
    return None


def _resolve_step_target_index(graph_nodes: list[dict], target_step: Any) -> int | None:
    if isinstance(target_step, int):
        index = target_step - 1
        return index if 0 <= index < len(graph_nodes) else None
    if isinstance(target_step, str) and target_step.strip().isdigit():
        index = int(target_step.strip()) - 1
        return index if 0 <= index < len(graph_nodes) else None
    return None


def _evaluate_jump_condition(condition: str, variables: dict[str, Any]) -> bool:
    normalized = condition.strip().lower()
    if normalized in {"true", "1", "yes", "on"}:
        return True
    if normalized in {"false", "0", "no", "off", ""}:
        return False
    try:
        value = _safe_eval_expression(condition, variables)
        return bool(value)
    except Exception:
        return False


def _decode_http_body(raw_body: bytes, headers: dict) -> Any:
    body_text = raw_body.decode("utf-8", errors="replace")
    content_type = str(headers.get("content-type") or headers.get("Content-Type") or "")
    if "application/json" in content_type.lower():
        try:
            return json.loads(body_text)
        except json.JSONDecodeError:
            return body_text
    return body_text


_VARIABLE_PATTERN = re.compile(r"\$\{([^}]+)\}")

_WEB_TABLE_EXTRACT_SCRIPT = """
(table) => {
  const rows = Array.from(table.querySelectorAll('tr'));
  const headerCells = Array.from(table.querySelectorAll('thead tr:first-child th, thead tr:first-child td'));
  const headers = headerCells.map((cell) => cell.textContent.trim());
  const bodyRows = Array.from(table.querySelectorAll('tbody tr'));
  const sourceRows = bodyRows.length > 0 ? bodyRows : rows.slice(headers.length > 0 ? 1 : 0);
  const extractedRows = sourceRows.map((row) => {
    const values = Array.from(row.querySelectorAll('th, td')).map((cell) => cell.textContent.trim());
    if (headers.length > 0) {
      const item = {};
      headers.forEach((header, index) => {
        item[header || String(index)] = values[index] || '';
      });
      return item;
    }
    return values;
  });
  return { headers, rows: extractedRows };
}
"""

class PythonCodeRejected(ValueError):
    pass


_PYTHON_DANGEROUS_ATTRIBUTES = frozenset(
    {
        "__bases__",
        "__builtins__",
        "__class__",
        "__closure__",
        "__code__",
        "__dict__",
        "__func__",
        "__globals__",
        "__import__",
        "__mro__",
        "__self__",
        "__subclasses__",
    }
)

_PYTHON_ALLOWED_IMPORTS = frozenset(
    {
        "csv",
        "datetime",
        "json",
        "math",
        "openpyxl",
        "pathlib",
        "re",
        "samplepkg",
        "statistics",
    }
)


def _python_safe_import(
    name: str,
    globals_dict: dict | None = None,
    locals_dict: dict | None = None,
    fromlist: tuple | list = (),
    level: int = 0,
):
    if level != 0:
        raise ImportError("relative imports are not allowed")
    root_name = str(name or "").split(".", 1)[0]
    if root_name not in _PYTHON_ALLOWED_IMPORTS:
        raise ImportError(f"import not allowed: {name}")
    return __import__(name, globals_dict, locals_dict, fromlist, level)


_PYTHON_SAFE_EXEC_BUILTINS = {
    "abs": abs,
    "all": all,
    "any": any,
    "bool": bool,
    "dict": dict,
    "enumerate": enumerate,
    "float": float,
    "int": int,
    "isinstance": isinstance,
    "issubclass": issubclass,
    "len": len,
    "list": list,
    "map": map,
    "max": max,
    "min": min,
    "print": print,
    "range": range,
    "repr": repr,
    "reversed": reversed,
    "round": round,
    "set": set,
    "slice": slice,
    "sorted": sorted,
    "str": str,
    "sum": sum,
    "tuple": tuple,
    "type": type,
    "zip": zip,
    "__build_class__": __build_class__,
    "__import__": _python_safe_import,
}

_PYTHON_RUN_CHILD_SCRIPT = """
from __future__ import annotations

import io
import json
import sys
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path

_PYTHON_ALLOWED_IMPORTS = set()


def _python_safe_import(name, globals_dict=None, locals_dict=None, fromlist=(), level=0):
    if level != 0:
        raise ImportError("relative imports are not allowed")
    root_name = str(name or "").split(".", 1)[0]
    if root_name not in _PYTHON_ALLOWED_IMPORTS:
        raise ImportError(f"import not allowed: {name}")
    return __import__(name, globals_dict, locals_dict, fromlist, level)


_PYTHON_SAFE_EXEC_BUILTINS = {
    "abs": abs,
    "all": all,
    "any": any,
    "bool": bool,
    "dict": dict,
    "enumerate": enumerate,
    "float": float,
    "int": int,
    "isinstance": isinstance,
    "issubclass": issubclass,
    "len": len,
    "list": list,
    "map": map,
    "max": max,
    "min": min,
    "print": print,
    "range": range,
    "repr": repr,
    "reversed": reversed,
    "round": round,
    "set": set,
    "slice": slice,
    "sorted": sorted,
    "str": str,
    "sum": sum,
    "tuple": tuple,
    "type": type,
    "zip": zip,
    "__build_class__": __build_class__,
    "__import__": _python_safe_import,
}


def _json_safe(value):
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    return repr(value)


def main() -> int:
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    global _PYTHON_ALLOWED_IMPORTS
    _PYTHON_ALLOWED_IMPORTS = set(payload.get("allowed_imports", ()))
    scope = {
        "variables": payload.get("variables", {}),
        "page": None,
        "browser": None,
        "result": None,
        "result_variable": payload.get("result_variable"),
    }
    stdout_buffer = io.StringIO()
    stderr_buffer = io.StringIO()
    response = {}
    try:
        compiled = compile(payload["code"], "<weconduct-python.run>", "exec")
        with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
            exec(compiled, {"__builtins__": _PYTHON_SAFE_EXEC_BUILTINS}, scope)
        response = {
            "status": "succeeded",
            "result": _json_safe(scope.get("result")),
            "result_variable": _json_safe(scope.get("result_variable")),
            "variables": _json_safe(scope.get("variables", {})),
        }
    except Exception as exc:
        response = {
            "status": "failed",
            "message": str(exc),
            "exception_type": type(exc).__name__,
        }
    response["stdout"] = stdout_buffer.getvalue()
    response["stderr"] = stderr_buffer.getvalue()
    output_path.write_text(json.dumps(response, ensure_ascii=False), encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
""".strip()


def _make_python_run_json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, list):
        return [_make_python_run_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_make_python_run_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _make_python_run_json_safe(item) for key, item in value.items()}
    return repr(value)


_HTTP_BLOCKED_SCHEMES = frozenset({"file", "ftp", "gopher", "dict", "ldap"})
_HTTP_METADATA_BLOCKED_IPS = frozenset(
    {
        ipaddress.ip_address("169.254.169.254"),
        ipaddress.ip_address("fd00:ec2::254"),
    }
)


def _validate_python_run_code(code: str) -> None:
    try:
        tree = ast.parse(code, mode="exec")
    except SyntaxError as exc:
        raise PythonCodeRejected(f"syntax error: {exc.msg}") from exc
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                root_name = alias.name.split(".", 1)[0]
                if root_name not in _PYTHON_ALLOWED_IMPORTS:
                    raise PythonCodeRejected(f"import not allowed: {alias.name}")
            continue
        if isinstance(node, ast.ImportFrom):
            if node.module is None:
                raise PythonCodeRejected("relative imports are not allowed")
            root_name = node.module.split(".", 1)[0]
            if root_name not in _PYTHON_ALLOWED_IMPORTS:
                raise PythonCodeRejected(f"import not allowed: {node.module}")
            continue
        if isinstance(node, ast.Attribute) and node.attr in _PYTHON_DANGEROUS_ATTRIBUTES:
            raise PythonCodeRejected(f"access to attribute is not allowed: {node.attr}")
        if isinstance(node, ast.Name) and node.id in _PYTHON_DANGEROUS_ATTRIBUTES:
            raise PythonCodeRejected(f"access to name is not allowed: {node.id}")


def _validate_http_request_url(
    url: str,
    *,
    allow_local_network_access: bool,
    allow_remote_network_access: bool,
) -> str:
    try:
        parsed = urllib.parse.urlparse(url)
    except ValueError as exc:
        raise ValueError(f"http.request url is invalid: {url}") from exc
    scheme = parsed.scheme.lower()
    if scheme not in {"http", "https"}:
        raise ValueError(f"http.request blocked unsupported url scheme: {scheme or '<empty>'}")
    hostname = parsed.hostname
    if not hostname:
        raise ValueError("http.request blocked url without hostname")
    _validate_http_request_host(
        hostname,
        allow_local_network_access=allow_local_network_access,
        allow_remote_network_access=allow_remote_network_access,
    )
    return url


def _validate_http_request_host(
    hostname: str,
    *,
    allow_local_network_access: bool,
    allow_remote_network_access: bool,
) -> None:
    try:
        ip = ipaddress.ip_address(hostname)
        _validate_http_request_ip(
            ip,
            allow_local_network_access=allow_local_network_access,
            allow_remote_network_access=allow_remote_network_access,
        )
        return
    except ValueError:
        pass
    try:
        records = socket.getaddrinfo(hostname, None, 0, socket.SOCK_STREAM)
    except OSError as exc:
        raise ValueError(f"http.request hostname resolution failed: {hostname}") from exc
    for record in records:
        sockaddr = record[4]
        if not sockaddr:
            continue
        address_text = sockaddr[0]
        try:
            ip = ipaddress.ip_address(address_text)
        except ValueError:
            continue
        _validate_http_request_ip(
            ip,
            allow_local_network_access=allow_local_network_access,
            allow_remote_network_access=allow_remote_network_access,
        )


def _validate_http_request_ip(
    ip: ipaddress._BaseAddress,
    *,
    allow_local_network_access: bool,
    allow_remote_network_access: bool,
) -> None:
    if ip in _HTTP_METADATA_BLOCKED_IPS:
        raise ValueError(f"http.request blocked metadata address: {ip}")
    if ip.is_loopback or ip.is_private or ip.is_link_local:
        if not allow_local_network_access:
            raise ValueError(f"local network access is disabled for address: {ip}")
        return
    if not allow_remote_network_access:
        raise ValueError(f"remote network access is disabled for address: {ip}")


def _resolve_value(value: Any, context: RuntimeContext) -> Any:
    if isinstance(value, str):
        return _resolve_string(value, context)
    if isinstance(value, list):
        return [_resolve_value(item, context) for item in value]
    if isinstance(value, dict):
        return {key: _resolve_value(item, context) for key, item in value.items()}
    return value


def _resolve_string(value: str, context: RuntimeContext) -> Any:
    matches = list(_VARIABLE_PATTERN.finditer(value))
    if len(matches) == 1 and matches[0].span() == (0, len(value)):
        return _lookup_runtime_reference_expression(matches[0].group(1), context)

    def replace_match(match: re.Match) -> str:
        resolved = _lookup_runtime_reference_expression(match.group(1), context)
        if isinstance(resolved, (dict, list)):
            return json.dumps(resolved, ensure_ascii=False)
        return "" if resolved is None else str(resolved)

    return _VARIABLE_PATTERN.sub(replace_match, value)


def _lookup_runtime_reference_expression(reference_expression: str, context: RuntimeContext) -> Any:
    reference, target_type = _split_runtime_reference_expression(reference_expression)
    resolved = _lookup_runtime_reference(reference, context)
    if target_type is None:
        return resolved
    return _convert_runtime_value(resolved, target_type)


def _split_runtime_reference_expression(reference_expression: str) -> tuple[str, str | None]:
    reference = str(reference_expression or "").strip()
    if not reference:
        return "", None
    reference_body, separator, target_type = reference.rpartition("|")
    if not separator:
        return reference, None
    normalized_target_type = _normalize_runtime_target_type(target_type)
    if normalized_target_type is None:
        return reference, None
    return reference_body.strip(), normalized_target_type


def _lookup_runtime_reference(reference: str, context: RuntimeContext) -> Any:
    if reference.startswith("node."):
        parts = reference.split(".")
        if len(parts) < 3:
            return None
        node_id = parts[1]
        current: Any = context.node_outputs.get(node_id)
        for part in parts[2:]:
            if isinstance(current, dict):
                current = current.get(part)
            elif isinstance(current, list):
                try:
                    current = current[int(part)]
                except (ValueError, IndexError):
                    return None
            else:
                return None
        return current
    return context.variables.get(reference)


def _normalize_runtime_target_type(target_type: Any) -> str | None:
    if not isinstance(target_type, str):
        return None
    normalized = target_type.strip().lower().replace("-", "_")
    aliases = {
        "str": "string",
        "text": "string",
        "string": "string",
        "int": "int",
        "integer": "int",
        "float": "float",
        "number": "float",
        "bool": "bool",
        "boolean": "bool",
        "json": "json",
        "object": "json",
        "array": "json",
    }
    return aliases.get(normalized)


def _convert_runtime_value(value: Any, target_type: Any) -> Any:
    normalized_target_type = _normalize_runtime_target_type(target_type)
    if normalized_target_type is None:
        raise ValueError("target_type must be one of: string, int, float, bool, json")
    if normalized_target_type == "string":
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)
    if normalized_target_type == "int":
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                raise ValueError("cannot convert empty string to int")
            return int(float(stripped)) if any(marker in stripped for marker in (".", "e", "E")) else int(stripped)
        raise ValueError(f"cannot convert {type(value).__name__} to int")
    if normalized_target_type == "float":
        if isinstance(value, bool):
            return float(int(value))
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                raise ValueError("cannot convert empty string to float")
            return float(stripped)
        raise ValueError(f"cannot convert {type(value).__name__} to float")
    if normalized_target_type == "bool":
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            stripped = value.strip().lower()
            truthy = {"true", "1", "yes", "y", "on"}
            falsy = {"false", "0", "no", "n", "off", ""}
            if stripped in truthy:
                return True
            if stripped in falsy:
                return False
            raise ValueError(f"cannot convert string '{value}' to bool")
        return bool(value)
    if normalized_target_type == "json":
        if value is None:
            return None
        if isinstance(value, (dict, list, int, float, bool)):
            return value
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return None
            try:
                return json.loads(stripped)
            except json.JSONDecodeError as exc:
                raise ValueError(f"cannot parse json: {exc.msg}") from exc
        raise ValueError(f"cannot convert {type(value).__name__} to json")
    raise ValueError("unsupported target_type")


def _read_csv_table(node: dict, context: RuntimeContext) -> dict:
    node_config = _node_config(node)
    path_value = _resolve_value(node_config.get("path"), context)
    if not isinstance(path_value, str) or not path_value.strip():
        return _failed_result(node, "file.path_required", "file path is required")
    encoding = str(_resolve_value(node_config.get("encoding", "utf-8"), context))
    has_header = bool(_resolve_value(node_config.get("has_header", True), context))
    path = _resolve_runtime_path(path_value, context)
    with path.open("r", encoding=encoding, newline="") as csv_file:
        if has_header:
            reader = csv.DictReader(csv_file)
            rows = [dict(row) for row in reader]
            headers = list(reader.fieldnames or [])
        else:
            reader = csv.reader(csv_file)
            rows = [list(row) for row in reader]
            headers = []
    return {
        "status": "succeeded",
        "node_id": node["node_id"],
        "path": str(path.resolve()),
        "encoding": encoding,
        "has_header": has_header,
        "headers": headers,
        "rows": rows,
    }


def _require_excel_sheet(node: dict, context: RuntimeContext) -> tuple[Path | dict, Any | None]:
    node_config = _node_config(node)
    path_value = _resolve_value(node_config.get("path"), context)
    if not isinstance(path_value, str) or not path_value.strip():
        return _failed_result(node, "excel.path_required", "Excel file path is required"), None
    sheet_name = _resolve_value(node_config.get("sheet_name"), context)
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        return _failed_result(node, "excel.sheet_name_required", "Excel sheet_name is required"), None
    path = _resolve_runtime_path(path_value, context)
    if not path.exists():
        return _failed_result(node, "excel.path_missing", "Excel file path does not exist"), None
    try:
        workbook = load_workbook(path)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        return _failed_result(node, "excel.workbook_invalid", str(exc)), None
    if sheet_name.strip() not in workbook.sheetnames:
        workbook.close()
        return _failed_result(node, "excel.sheet_missing", "Excel sheet_name does not exist"), None
    worksheet = workbook[sheet_name.strip()]
    return path, worksheet


def _load_excel_sheet_for_write(
    node: dict,
    context: RuntimeContext,
    *,
    create_if_not_exists: bool = True,
    create_mode: bool = False,
) -> tuple[Path | dict, Workbook | None, Any | None]:
    node_config = _node_config(node)
    path_value = _resolve_value(node_config.get("path"), context)
    if not isinstance(path_value, str) or not path_value.strip():
        return _failed_result(node, "excel.path_required", "Excel file path is required"), None, None
    sheet_name = _resolve_value(node_config.get("sheet_name", "Sheet1"), context)
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        return _failed_result(node, "excel.sheet_name_required", "Excel sheet_name is required"), None, None
    path = _resolve_runtime_path(path_value, context)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and not create_mode:
        try:
            workbook = load_workbook(path)
        except (BadZipFile, InvalidFileException, OSError) as exc:
            return _failed_result(node, "excel.workbook_invalid", str(exc)), None, None
    elif create_if_not_exists or create_mode:
        workbook = Workbook()
        workbook.active.title = sheet_name.strip()
    else:
        return _failed_result(node, "excel.path_missing", "Excel file path does not exist"), None, None
    worksheet = workbook[sheet_name.strip()] if sheet_name.strip() in workbook.sheetnames else workbook.create_sheet(sheet_name.strip())
    return path, workbook, worksheet


def _read_excel_table(node: dict, context: RuntimeContext) -> dict:
    node_config = _node_config(node)
    workbook_path, worksheet = _require_excel_sheet(node, context)
    if worksheet is None:
        return workbook_path
    has_header = bool(_resolve_value(node_config.get("has_header", False), context))
    rows_raw = list(worksheet.iter_rows(values_only=True))
    headers: list[Any] = []
    rows: list[Any]
    if has_header and rows_raw:
        headers = list(rows_raw[0])
        rows = [
            {str(headers[index]): value for index, value in enumerate(row)}
            for row in rows_raw[1:]
        ]
    else:
        rows = [list(row) for row in rows_raw]
    result = {
        "status": "succeeded",
        "node_id": node["node_id"],
        "path": str(workbook_path.resolve()),
        "sheet_name": worksheet.title,
        "has_header": has_header,
        "headers": headers,
        "rows": rows,
    }
    worksheet.parent.close()
    return result


def _clear_worksheet(worksheet: Any) -> None:
    if worksheet.max_row >= 1:
        worksheet.delete_rows(1, worksheet.max_row)


def _normalize_excel_row_payload(worksheet: Any, data: Any) -> list[Any]:
    if isinstance(data, dict):
        headers = [cell.value for cell in worksheet[1] if cell.value is not None]
        if headers:
            return [data.get(str(header)) for header in headers]
        return list(data.values())
    if isinstance(data, list):
        return list(data)
    raise ValueError("excel row data must be a list or object")


def _normalize_excel_table_payload(data: Any, *, has_header: bool) -> tuple[list[Any], list[list[Any]]]:
    if not isinstance(data, list):
        raise ValueError("excel table data must be a list")
    if not data:
        return [], []
    first = data[0]
    if isinstance(first, dict):
        headers = list(first.keys())
        rows = [[item.get(header) for header in headers] for item in data]
        return (headers if has_header else []), rows
    return [], [list(item) for item in data]


def _evaluate_excel_batch_condition(condition: str, row: dict[str, Any]) -> bool:
    if not condition or not condition.strip():
        return True
    result = _safe_eval_expression(condition, {"row": row})
    return bool(result)


def _resolve_int(value: Any, *, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _store_optional_variable(node_config: dict, context: RuntimeContext, value: Any) -> None:
    variable_name = node_config.get("variable_name")
    if isinstance(variable_name, str) and variable_name.strip():
        context.variables[variable_name.strip()] = value


def _store_optional_variable_name(variable_name: Any, context: RuntimeContext, value: Any) -> None:
    if isinstance(variable_name, str) and variable_name.strip():
        context.variables[variable_name.strip()] = value


def _resolve_runtime_path(path_value: str, context: RuntimeContext) -> Path:
    normalized_literal = path_value.strip()
    if normalized_literal:
        embedded_match = context.embedded_resource_paths.get(normalized_literal)
        if embedded_match is None:
            embedded_match = context.embedded_resource_paths.get(normalized_literal.replace("\\", "/"))
        if embedded_match is None:
            embedded_match = context.embedded_resource_paths.get(normalized_literal.replace("/", "\\"))
        if embedded_match is not None:
            return Path(embedded_match).expanduser().resolve()
    path = Path(path_value).expanduser()
    if context.runtime_settings.get("file_access_require_absolute_path", False) and not path.is_absolute():
        raise ValueError(f"file path must be absolute: {path_value}")
    if path.is_absolute():
        resolved = path.resolve()
        _validate_path_within_allowed_roots(resolved, context)
        return resolved
    base_directory = context.project_directory or context.workspace_root
    if base_directory is None:
        resolved = path.resolve()
        _validate_path_within_allowed_roots(resolved, context)
        return resolved
    resolved = (base_directory / path).resolve()
    _validate_path_within_allowed_roots(resolved, context)
    return resolved


def _resolve_allowed_path_roots(context: RuntimeContext) -> tuple[Path, ...]:
    roots: list[Path] = []
    for configured_root in context.allowed_path_roots:
        resolved_root = Path(configured_root).resolve()
        if resolved_root not in roots:
            roots.append(resolved_root)
    if context.project_directory is not None:
        resolved_project_directory = Path(context.project_directory).resolve()
        if resolved_project_directory not in roots:
            roots.append(resolved_project_directory)
    if context.workspace_root is not None:
        workspace_root = Path(context.workspace_root).resolve()
        if workspace_root not in roots:
            roots.append(workspace_root)
    downloads = Path.home() / "Downloads"
    if downloads.exists():
        resolved_downloads = downloads.resolve()
        if resolved_downloads not in roots:
            roots.append(resolved_downloads)
    return tuple(roots)


def _validate_path_within_allowed_roots(path: Path, context: RuntimeContext) -> None:
    file_access_scope = context.runtime_settings.get("file_access_scope", "restricted")
    resolved_path = path.resolve()
    blocked_roots = [
        Path(item).expanduser().resolve()
        for item in context.runtime_settings.get("file_access_blocked_roots", [])
        if isinstance(item, str) and item.strip()
    ]
    for root in blocked_roots:
        try:
            resolved_path.relative_to(root)
            raise ValueError(
                "file path is inside blocked directory: "
                f"{resolved_path}; blocked roots: {[str(item) for item in blocked_roots]}"
            )
        except ValueError as exc:
            if str(exc).startswith("file path is inside blocked directory:"):
                raise
            continue
    suffix = path.suffix.strip().lower()
    blocked_extensions = {
        str(item).strip().lower()
        for item in context.runtime_settings.get("file_access_blocked_extensions", [])
        if isinstance(item, str) and str(item).strip()
    }
    if suffix and suffix in blocked_extensions:
        raise ValueError(f"file path uses blocked extension: {suffix}")
    allowed_extensions = [
        str(item).strip().lower()
        for item in context.runtime_settings.get("file_access_allowed_extensions", [])
        if isinstance(item, str) and str(item).strip()
    ]
    if allowed_extensions and suffix not in allowed_extensions:
        raise ValueError(
            "file path extension is not allowed: "
            f"{suffix or '<none>'}; allowed extensions: {allowed_extensions}"
        )
    if file_access_scope == "allow_all":
        return
    allowed_roots = _resolve_allowed_path_roots(context)
    for root in allowed_roots:
        try:
            resolved_path.relative_to(root)
            return
        except ValueError:
            continue
    raise ValueError(
        "file path is outside allowed directories: "
        f"{resolved_path}; allowed roots: {[str(root) for root in allowed_roots]}"
    )


def _require_selector(node: dict, context: RuntimeContext) -> str | dict:
    selector = _resolve_value(_node_config(node).get("selector"), context)
    if not isinstance(selector, str) or not selector.strip():
        return _failed_result(node, "browser.selector_required", "selector is required")
    return selector.strip()


def _require_runtime_list(node: dict, context: RuntimeContext) -> tuple[list[Any] | dict, str]:
    node_config = _node_config(node)
    variable_name = node_config.get("variable_name") or node_config.get("name")
    if not isinstance(variable_name, str) or not variable_name.strip():
        return _failed_result(node, "data.variable_name_required", "variable name is required"), ""
    current = context.variables.get(variable_name.strip())
    if not isinstance(current, list):
        return (
            _failed_result(node, "data.list_target_invalid", "target variable must be a list"),
            variable_name.strip(),
        )
    return list(current), variable_name.strip()


def _require_browser_page(context: RuntimeContext) -> Page:
    page = context.browser_runtime.get("page")
    if page is not None:
        return page
    launch_options = context.browser_runtime.get("launch_options")
    if not isinstance(launch_options, dict):
        launch_options = {}
    chromium_launch_kwargs: dict[str, Any] = {
        "headless": bool(launch_options.get("headless", True))
    }
    if "slow_mo_ms" in launch_options:
        chromium_launch_kwargs["slow_mo"] = _resolve_int(launch_options.get("slow_mo_ms"), default=0)
    playwright: Playwright = sync_playwright().start()
    chromium_launch_kwargs["channel"] = "msedge"
    browser: Browser = playwright.chromium.launch(**chromium_launch_kwargs)
    context_kwargs: dict[str, Any] = {"accept_downloads": True}
    user_agent = launch_options.get("user_agent")
    if isinstance(user_agent, str) and user_agent.strip():
        context_kwargs["user_agent"] = user_agent.strip()
    extra_http_headers = launch_options.get("extra_http_headers")
    if isinstance(extra_http_headers, dict) and extra_http_headers:
        context_kwargs["extra_http_headers"] = {
            str(key): "" if value is None else str(value)
            for key, value in extra_http_headers.items()
        }
    browser_context = browser.new_context(**context_kwargs)
    page = browser_context.new_page()
    context.browser_runtime["playwright"] = playwright
    context.browser_runtime["browser"] = browser
    context.browser_runtime["browser_context"] = browser_context
    context.browser_runtime["page"] = page
    context.browser_runtime["pages"] = []
    context.browser_runtime["active_page_index"] = 0
    context.browser_runtime["page_labels"] = {}
    context.browser_runtime.setdefault("request_records", [])
    context.browser_runtime.setdefault("response_records", [])
    context.browser_runtime.setdefault("popup_records", [])
    context.browser_runtime.setdefault("download_records", [])
    _ensure_browser_context_handlers(context)
    return page


def _resolve_browser_launch_config(value: Any, context: RuntimeContext) -> dict[str, Any]:
    resolved = _resolve_value(value, context)
    if not isinstance(resolved, dict):
        return {}
    browser_config: dict[str, Any] = {}
    if "headless" in resolved:
        browser_config["headless"] = bool(resolved.get("headless"))
    if "slow_mo_ms" in resolved:
        browser_config["slow_mo_ms"] = _resolve_int(resolved.get("slow_mo_ms"), default=0)
    if isinstance(resolved.get("user_agent"), str) and resolved.get("user_agent").strip():
        browser_config["user_agent"] = resolved.get("user_agent").strip()
    extra_http_headers = resolved.get("extra_http_headers")
    if isinstance(extra_http_headers, dict):
        browser_config["extra_http_headers"] = dict(extra_http_headers)
    return browser_config


def _ensure_browser_context_handlers(context: RuntimeContext) -> None:
    browser_context = context.browser_runtime.get("browser_context")
    if browser_context is None or context.browser_runtime.get("context_handlers_installed") is True:
        return

    def handle_request(request: Any) -> None:
        context.browser_runtime.setdefault("request_records", []).append(
            {
                "url": request.url,
                "method": request.method,
                "headers": dict(request.headers),
                "resource_type": request.resource_type,
            }
        )

    def handle_response(response: Any) -> None:
        try:
            body_text = response.text()
        except Exception:
            body_text = None
        context.browser_runtime.setdefault("response_records", []).append(
            {
                "url": response.url,
                "status_code": response.status,
                "headers": dict(response.headers),
                "body_text": body_text,
                "ok": response.ok,
            }
        )

    def handle_page(page: Page) -> None:
        page_index = _register_browser_page(context, page)
        suppressed_count = int(context.browser_runtime.get("suppress_next_page_record", 0))
        if suppressed_count > 0:
            context.browser_runtime["suppress_next_page_record"] = suppressed_count - 1
            return
        context.browser_runtime.setdefault("popup_records", []).append(
            {
                "page": page,
                "page_url": page.url,
                "page_index": page_index,
            }
        )

    browser_context.on("request", handle_request)
    browser_context.on("response", handle_response)
    browser_context.on("page", handle_page)
    context.browser_runtime["context_handlers_installed"] = True
    _register_browser_page(context, context.browser_runtime["page"])


def _browser_pages(context: RuntimeContext) -> list[Page]:
    pages = context.browser_runtime.get("pages")
    if not isinstance(pages, list):
        page = _require_browser_page(context)
        pages = [page]
        context.browser_runtime["pages"] = pages
    alive_pages = [page for page in pages if not page.is_closed()]
    if alive_pages != pages:
        context.browser_runtime["pages"] = alive_pages
    return context.browser_runtime["pages"]


def _set_active_browser_page(context: RuntimeContext, page: Page) -> int:
    pages = _browser_pages(context)
    if page not in pages:
        pages.append(page)
    index = pages.index(page)
    context.browser_runtime["page"] = page
    context.browser_runtime["active_page_index"] = index
    _reset_browser_frame_context(context)
    return index


def _page_label(context: RuntimeContext, page: Page) -> str | None:
    labels = context.browser_runtime.get("page_labels")
    if not isinstance(labels, dict):
        return None
    for label, labeled_page in labels.items():
        if labeled_page is page:
            return label
    return None


def _register_browser_page(context: RuntimeContext, page: Page) -> int:
    pages = context.browser_runtime.setdefault("pages", [])
    if page not in pages:
        pages.append(page)
    registered_page_ids = context.browser_runtime.setdefault("download_handler_page_ids", set())
    if id(page) not in registered_page_ids:
        def handle_download(download: Any) -> None:
            context.browser_runtime.setdefault("download_records", []).append(
                {
                    "download": download,
                    "url": download.url,
                    "suggested_filename": download.suggested_filename,
                    "page": page,
                    "page_index": _browser_pages(context).index(page) if page in _browser_pages(context) else None,
                }
            )

        page.on("download", handle_download)
        registered_page_ids.add(id(page))
    return pages.index(page)


def _normalize_browser_context_cookies(cookies: list[dict]) -> list[dict]:
    normalized: list[dict] = []
    for cookie in cookies:
        if not isinstance(cookie, dict):
            continue
        name = cookie.get("name")
        if not isinstance(name, str) or not name.strip():
            continue
        item = dict(cookie)
        item["name"] = name.strip()
        if item.get("value") is None:
            item["value"] = ""
        normalized.append(item)
    return normalized


def _pump_browser_events(page: Page | None, delay_ms: int = 25) -> None:
    if page is None:
        return
    try:
        page.wait_for_timeout(delay_ms)
    except Exception:
        return


def _wait_until(timeout_ms: int, predicate: Any, *, on_poll: Any | None = None) -> None:
    started_at = monotonic()
    while (monotonic() - started_at) * 1000 <= timeout_ms:
        if predicate():
            return
        if on_poll is not None:
            on_poll()
    raise TimeoutError(f"condition was not met within {timeout_ms} ms")


def _match_text(actual: str, expected: str, match_mode: str) -> bool:
    if match_mode == "equals":
        return actual == expected
    if match_mode == "starts_with":
        return actual.startswith(expected)
    if match_mode == "ends_with":
        return actual.endswith(expected)
    return expected in actual


def _wait_for_record(records: list[dict], timeout_ms: int, predicate: Any, *, on_poll: Any | None = None) -> dict:
    started_at = monotonic()
    while (monotonic() - started_at) * 1000 <= timeout_ms:
        for record in records:
            if predicate(record):
                return record
        if on_poll is not None:
            on_poll()
    raise TimeoutError(f"record was not observed within {timeout_ms} ms")


def _request_record_matches(record: dict, url_pattern: str, method: str | None) -> bool:
    if not isinstance(record, dict):
        return False
    if url_pattern not in str(record.get("url", "")):
        return False
    if method is not None and str(record.get("method", "")).upper() != method:
        return False
    return True


def _response_record_matches(record: dict, url_pattern: str, expected_status: int | None) -> bool:
    if not isinstance(record, dict):
        return False
    if url_pattern not in str(record.get("url", "")):
        return False
    if expected_status is not None and expected_status >= 0 and int(record.get("status_code", -1)) != expected_status:
        return False
    return True


def _get_browser_storage_item(*, node: dict, context: RuntimeContext, storage_name: str) -> dict:
    node_config = _node_config(node)
    key = _resolve_value(node_config.get("key"), context)
    if not isinstance(key, str) or not key.strip():
        return _failed_result(node, "browser.storage_key_required", f"{storage_name} key is required")
    target = _require_browser_target(context)
    script = (
        "(storageKey) => window.sessionStorage.getItem(storageKey)"
        if storage_name == "sessionStorage"
        else "(storageKey) => window.localStorage.getItem(storageKey)"
    )
    value = target.evaluate(script, key.strip())
    if value is None:
        value = _resolve_value(node_config.get("default_value"), context)
    _store_optional_variable(node_config, context, value)
    return {"status": "succeeded", "node_id": node["node_id"], "key": key.strip(), "value": value, "storage_name": storage_name}


def _set_browser_storage_item(*, node: dict, context: RuntimeContext, storage_name: str) -> dict:
    node_config = _node_config(node)
    key = _resolve_value(node_config.get("key"), context)
    if not isinstance(key, str) or not key.strip():
        return _failed_result(node, "browser.storage_key_required", f"{storage_name} key is required")
    value = _resolve_value(node_config.get("value"), context)
    target = _require_browser_target(context)
    script = (
        "(payload) => window.sessionStorage.setItem(payload.key, String(payload.value))"
        if storage_name == "sessionStorage"
        else "(payload) => window.localStorage.setItem(payload.key, String(payload.value))"
    )
    target.evaluate(script, {"key": key.strip(), "value": "" if value is None else value})
    return {"status": "succeeded", "node_id": node["node_id"], "key": key.strip(), "value": value, "storage_name": storage_name}


def _remove_browser_storage_item(*, node: dict, context: RuntimeContext, storage_name: str) -> dict:
    key = _resolve_value(_node_config(node).get("key"), context)
    if not isinstance(key, str) or not key.strip():
        return _failed_result(node, "browser.storage_key_required", f"{storage_name} key is required")
    target = _require_browser_target(context)
    script = (
        "(storageKey) => window.sessionStorage.removeItem(storageKey)"
        if storage_name == "sessionStorage"
        else "(storageKey) => window.localStorage.removeItem(storageKey)"
    )
    target.evaluate(script, key.strip())
    return {"status": "succeeded", "node_id": node["node_id"], "key": key.strip(), "storage_name": storage_name}


def _clear_browser_storage(*, node: dict, context: RuntimeContext, storage_name: str) -> dict:
    target = _require_browser_target(context)
    script = "() => window.sessionStorage.clear()" if storage_name == "sessionStorage" else "() => window.localStorage.clear()"
    target.evaluate(script)
    return {"status": "succeeded", "node_id": node["node_id"], "storage_name": storage_name}


def _probe_browser_locator(*, node: dict, context: RuntimeContext, probe: str) -> dict:
    selector = _require_selector(node, context)
    if isinstance(selector, dict):
        return selector
    locator = _require_browser_target(context).locator(selector)
    if probe == "visible":
        value = locator.is_visible()
    elif probe == "enabled":
        value = locator.is_enabled()
    elif probe == "checked":
        value = locator.is_checked()
    else:
        value = locator.count() > 0
    _store_optional_variable(_node_config(node), context, value)
    return {"status": "succeeded", "node_id": node["node_id"], "selector": selector, "value": value, "probe": probe}


def _read_browser_html(*, node: dict, context: RuntimeContext, mode: str) -> dict:
    selector = _require_selector(node, context)
    if isinstance(selector, dict):
        return selector
    locator = _require_browser_target(context).locator(selector)
    value = locator.evaluate("(el) => el.innerHTML" if mode == "inner" else "(el) => el.outerHTML")
    _store_optional_variable(_node_config(node), context, value)
    return {"status": "succeeded", "node_id": node["node_id"], "selector": selector, "value": value, "mode": mode}


def _resolve_browser_page_reference(node_config: dict, context: RuntimeContext) -> Page | None:
    pages = _browser_pages(context)
    label = _resolve_value(node_config.get("label"), context)
    labels = context.browser_runtime.get("page_labels")
    if isinstance(label, str) and label.strip() and isinstance(labels, dict):
        labeled_page = labels.get(label.strip())
        if isinstance(labeled_page, Page) and not labeled_page.is_closed():
            return labeled_page
    index_value = _resolve_value(node_config.get("index"), context)
    if index_value is not None:
        index = _resolve_int(index_value, default=-1)
        if 0 <= index < len(pages):
            return pages[index]
    url_pattern = _resolve_value(node_config.get("url_pattern"), context)
    if isinstance(url_pattern, str) and url_pattern.strip():
        for page in pages:
            if url_pattern.strip() in page.url:
                return page
    return None


def _wait_for_url_change(*, page: Page, from_url: str, url_pattern: str, timeout_ms: int) -> str:
    started_at = monotonic()
    while (monotonic() - started_at) * 1000 <= timeout_ms:
        current_url = page.url
        if current_url != from_url and (not url_pattern or url_pattern in current_url):
            return current_url
        _pump_browser_events(page)
    raise TimeoutError(f"url did not change within {timeout_ms} ms")


def _apply_local_storage(page: Page, local_storage: Any) -> int:
    if local_storage in (None, {}, []):
        return 0
    origin_items: list[tuple[str, dict[str, Any]]] = []
    if isinstance(local_storage, dict):
        for origin, items in local_storage.items():
            if isinstance(origin, str) and origin.strip() and isinstance(items, dict):
                origin_items.append((origin.strip(), items))
    elif isinstance(local_storage, list):
        for item in local_storage:
            if not isinstance(item, dict):
                continue
            origin = item.get("origin")
            items = item.get("items")
            if isinstance(origin, str) and origin.strip() and isinstance(items, dict):
                origin_items.append((origin.strip(), items))
    else:
        raise ValueError("local_storage must be an object or list")

    applied_count = 0
    current_origin = _page_origin(page.url)
    for origin, items in origin_items:
        target_page = page
        opened_temp_page = None
        if current_origin != origin:
            opened_temp_page = page.context.new_page()
            target_page = opened_temp_page
            target_page.goto(origin, wait_until="domcontentloaded")
        target_page.evaluate(
            """
            (items) => {
              for (const [key, value] of Object.entries(items)) {
                window.localStorage.setItem(key, String(value));
              }
            }
            """,
            items,
        )
        if opened_temp_page is not None:
            opened_temp_page.close()
        applied_count += 1
    return applied_count


def _page_origin(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        return ""
    return f"{parsed.scheme}://{parsed.netloc}"


def _ensure_dialog_handler(context: RuntimeContext) -> None:
    page = _require_browser_page(context)
    if context.browser_runtime.get("dialog_handler_installed") is True:
        return
    context.browser_runtime.setdefault("dialog_records", [])
    context.browser_runtime.setdefault(
        "dialog_config",
        {
            "mode": "auto",
            "default_action": "accept",
            "prompt_text": "",
        },
    )

    def handle_dialog(dialog: Any) -> None:
        config = context.browser_runtime.setdefault("dialog_config", {})
        mode = str(config.get("mode", "auto"))
        action = str(config.get("default_action", "accept"))
        if mode in {"accept", "dismiss"}:
            action = mode
        record = {
            "type": dialog.type,
            "message": dialog.message,
            "default_value": dialog.default_value,
            "action": action,
        }
        try:
            if action == "dismiss":
                dialog.dismiss()
            else:
                dialog.accept(str(config.get("prompt_text", "")))
        except Exception as exc:
            record["error"] = str(exc)
        context.browser_runtime.setdefault("dialog_records", []).append(record)

    page.on("dialog", handle_dialog)
    context.browser_runtime["dialog_handler_installed"] = True


def _require_browser_target(context: RuntimeContext) -> Page | Frame:
    current_frame = context.browser_runtime.get("current_frame")
    if isinstance(current_frame, Frame):
        if current_frame.is_detached():
            context.browser_runtime.pop("current_frame", None)
            context.browser_runtime["frame_depth"] = 0
            return _require_browser_page(context)
        return current_frame
    return _require_browser_page(context)


def _reset_browser_frame_context(context: RuntimeContext) -> None:
    context.browser_runtime.pop("current_frame", None)
    context.browser_runtime["frame_stack"] = []


def _resolve_browser_frame(node_config: dict, context: RuntimeContext) -> Frame | None:
    page = _require_browser_page(context)
    selector = _resolve_value(node_config.get("selector"), context)
    name = _resolve_value(node_config.get("name"), context)
    url_contains = _resolve_value(node_config.get("url_contains"), context)
    index = _resolve_int(_resolve_value(node_config.get("index"), context), default=-1)
    if isinstance(selector, str) and selector.strip():
        frame_handle = page.locator(selector.strip()).element_handle()
        return frame_handle.content_frame() if frame_handle is not None else None
    for frame in page.frames:
        if isinstance(name, str) and name.strip() and frame.name == name.strip():
            return frame
        if isinstance(url_contains, str) and url_contains.strip() and url_contains.strip() in frame.url:
            return frame
    if index >= 0:
        frames = page.frames
        return frames[index] if index < len(frames) else None
    return None


def _wait_for_browser_url(target: Page | Frame, url_pattern: str, timeout_ms: int) -> str | None:
    started_at = monotonic()
    page = target.page if isinstance(target, Frame) else target
    while (monotonic() - started_at) * 1000 <= timeout_ms:
        current_url = target.url
        if url_pattern in current_url:
            return current_url
        page.wait_for_timeout(50)
    return None


def _safe_eval_expression(expression: str, variables: dict[str, Any]) -> Any:
    tree = ast.parse(expression, mode="eval")
    return _eval_ast_node(tree.body, variables)


def _eval_ast_node(node: ast.AST, variables: dict[str, Any]) -> Any:
    if isinstance(node, ast.Constant):
        return node.value
    if isinstance(node, ast.Name):
        return variables.get(node.id)
    if isinstance(node, ast.BoolOp):
        if isinstance(node.op, ast.And):
            return all(bool(_eval_ast_node(value, variables)) for value in node.values)
        if isinstance(node.op, ast.Or):
            return any(bool(_eval_ast_node(value, variables)) for value in node.values)
        raise ValueError("unsupported boolean operator")
    if isinstance(node, ast.List):
        return [_eval_ast_node(item, variables) for item in node.elts]
    if isinstance(node, ast.Tuple):
        return tuple(_eval_ast_node(item, variables) for item in node.elts)
    if isinstance(node, ast.BinOp):
        left = _eval_ast_node(node.left, variables)
        right = _eval_ast_node(node.right, variables)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return left / right
        if isinstance(node.op, ast.FloorDiv):
            return left // right
        raise ValueError("unsupported binary operator")
    if isinstance(node, ast.UnaryOp):
        operand = _eval_ast_node(node.operand, variables)
        if isinstance(node.op, ast.USub):
            return -operand
        if isinstance(node.op, ast.UAdd):
            return +operand
        raise ValueError("unsupported unary operator")
    if isinstance(node, ast.Call):
        if isinstance(node.func, ast.Name):
            if node.func.id == "len" and len(node.args) == 1:
                return len(_eval_ast_node(node.args[0], variables))
            raise ValueError("unsupported function call")
        if (
            isinstance(node.func, ast.Attribute)
            and node.func.attr == "get"
            and len(node.args) in {1, 2}
        ):
            target = _eval_ast_node(node.func.value, variables)
            if not isinstance(target, dict):
                raise ValueError("dict.get target must be an object")
            key = _eval_ast_node(node.args[0], variables)
            default = _eval_ast_node(node.args[1], variables) if len(node.args) == 2 else None
            return target.get(key, default)
        raise ValueError("unsupported function call")
    if isinstance(node, ast.Subscript):
        target = _eval_ast_node(node.value, variables)
        if isinstance(node.slice, ast.Slice):
            raise ValueError("unsupported subscript slice")
        key = _eval_ast_node(node.slice, variables)
        return target[key]
    if isinstance(node, ast.Compare):
        left = _eval_ast_node(node.left, variables)
        current = left
        for operator, comparator in zip(node.ops, node.comparators, strict=False):
            right = _eval_ast_node(comparator, variables)
            if isinstance(operator, ast.Lt):
                matched = current < right
            elif isinstance(operator, ast.LtE):
                matched = current <= right
            elif isinstance(operator, ast.Gt):
                matched = current > right
            elif isinstance(operator, ast.GtE):
                matched = current >= right
            elif isinstance(operator, ast.Eq):
                matched = current == right
            elif isinstance(operator, ast.NotEq):
                matched = current != right
            else:
                raise ValueError("unsupported compare operator")
            if not matched:
                return False
            current = right
        return True
    raise ValueError("unsupported expression syntax")
