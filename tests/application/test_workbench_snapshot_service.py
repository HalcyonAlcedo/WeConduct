import json
from pathlib import Path
import threading
import base64
from http.server import BaseHTTPRequestHandler
from socketserver import TCPServer
import urllib.parse
from openpyxl import Workbook, load_workbook

from weconduct.application import CompilationWorkbenchService
from weconduct.application.workspace_state_store import InMemoryWorkspaceStateStore
from weconduct.application.workspace_state_store import FileWorkspaceStateStore
from weconduct.builtin_components.registry import BUILTIN_COMPONENT_DEFINITIONS
from weconduct.contracts import GraphEdge, GraphModel, GraphNode, GraphPort, GraphPosition
from weconduct.runtime.engine import (
    RuntimeContext,
    RuntimeExecutorRegistry,
    execute_runtime_node,
    _require_browser_page,
)


class _RuntimeHttpHandler(BaseHTTPRequestHandler):
    def do_POST(self) -> None:  # noqa: N802
        content_length = int(self.headers.get("Content-Length", "0"))
        request_body = self.rfile.read(content_length).decode("utf-8")
        payload = {
            "ok": True,
            "path": self.path,
            "method": "POST",
            "body": json.loads(request_body),
            "header_value": self.headers.get("X-WeConduct-Test"),
        }
        body = json.dumps(payload).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("X-WeConduct-Reply", "runtime")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


class _RuntimeHttpServer(TCPServer):
    allow_reuse_address = True


def _start_runtime_http_server() -> tuple[_RuntimeHttpServer, threading.Thread]:
    server = _RuntimeHttpServer(("127.0.0.1", 0), _RuntimeHttpHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server, thread


class _BrowserMockSiteHandler(BaseHTTPRequestHandler):
    clicked = False
    last_form_value = ""

    def do_GET(self) -> None:  # noqa: N802
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        if path == "/dashboard":
            body = b"""
<!doctype html>
<html>
  <body>
    <div id="dashboard-status">dashboard-ready</div>
    <button id="open-popup" type="button" onclick="window.open('/popup', 'phase14Popup')">Open Popup</button>
    <a id="download-link" href="/download">Download</a>
  </body>
</html>
""".strip()
        elif path == "/popup":
            body = b"""
<!doctype html>
<html>
  <body>
    <div id="popup-status">popup-ready</div>
  </body>
</html>
""".strip()
        elif path == "/headers":
            payload = {
                "user_agent": self.headers.get("User-Agent"),
                "x_weconduct_extra": self.headers.get("X-WeConduct-Extra"),
            }
            body = f"""
<!doctype html>
<html>
  <body>
    <pre id="header-json">{json.dumps(payload, ensure_ascii=False)}</pre>
  </body>
</html>
""".strip().encode("utf-8")
        elif path == "/download":
            body = b"phase14-download"
            self.send_response(200)
            self.send_header("Content-Type", "application/octet-stream")
            self.send_header("Content-Disposition", 'attachment; filename="phase14.txt"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        elif path == "/frame":
            body = b"""
<!doctype html>
<html>
  <body>
    <div id="frame-status">frame-ready</div>
    <button id="frame-submit" type="button" onclick="window.location='/frame-details'">Open Frame Details</button>
  </body>
</html>
""".strip()
        elif path == "/frame-details":
            body = b"""
<!doctype html>
<html>
  <body>
    <div id="frame-details">frame-details-ready</div>
  </body>
</html>
""".strip()
        else:
            body = f"""
<!doctype html>
<html>
  <body>
    <form method="post" action="/submit">
      <input id="name" name="name" value="{self.last_form_value}">
      <label><input id="agree" type="checkbox" name="agree"> Agree</label>
      <input id="upload-file" type="file" name="upload_file">
      <button id="submit" type="submit">Submit</button>
    </form>
    <select id="city">
      <option value="">pick</option>
      <option value="beijing">Beijing</option>
      <option value="shanghai">Shanghai</option>
    </select>
    <input
      id="key-input"
      value=""
      oninput="document.getElementById('key-output').textContent = this.value"
      onkeydown="if (event.key === 'Enter') document.getElementById('key-enter-status').textContent = 'enter'; if (event.ctrlKey && event.key.toLowerCase() === 'a') document.getElementById('key-hotkey-status').textContent = 'ctrl+a';"
    >
    <div id="key-output"></div>
    <div id="key-enter-status"></div>
    <div id="key-hotkey-status"></div>
    <div id="hover-target" onmouseover="document.getElementById('hover-result').style.display='block'">hover-me</div>
    <div id="hover-result" style="display:none">hovered</div>
    <button id="go-dashboard" type="button" onclick="window.location='/dashboard'">Go Dashboard</button>
    <button id="alert-button" type="button" onclick="alert('hello-dialog')">Alert</button>
    <button id="open-popup" type="button" onclick="window.open('/popup', 'phase14Popup')">Open Popup</button>
    <button
      id="fetch-button"
      type="button"
      onclick="fetch('/api/ping?token=abc', {{ method: 'POST', headers: {{ 'Content-Type': 'application/json', 'X-Client-Action': 'fetch' }}, body: JSON.stringify({{ kind: 'ping' }}) }}).then((response) => response.json()).then((payload) => {{ const el = document.getElementById('fetch-status'); el.textContent = payload.message; el.setAttribute('data-response', payload.status); }});"
    >
      Fetch
    </button>
    <div id="fetch-status" data-response="idle"></div>
    <button id="disabled-button" type="button" disabled>Disabled</button>
    <div
      id="drag-source"
      draggable="true"
      ondragstart="event.dataTransfer.setData('text/plain', 'drag-payload')"
    >
      drag-source
    </div>
    <div
      id="drop-target"
      ondragover="event.preventDefault()"
      ondrop="event.preventDefault(); document.getElementById('drop-status').textContent = event.dataTransfer.getData('text/plain') || 'dropped';"
      style="width:120px;height:60px;border:1px solid #333;"
    >
      drop-target
    </div>
    <div id="drop-status"></div>
    <div id="scroll-target" style="margin-top: 2200px;">scroll-target</div>
    <a id="download-link" href="/download">Download</a>
    <table id="sample-table">
      <thead>
        <tr><th>Name</th><th>Score</th></tr>
      </thead>
      <tbody>
        <tr><td>Alice</td><td>95</td></tr>
        <tr><td>Bob</td><td>88</td></tr>
      </tbody>
    </table>
    <iframe id="content-frame" name="contentFrame" src="/frame"></iframe>
    <div id="status">{"clicked" if self.clicked else "ready"}</div>
    <script>
      window.sessionStorage.setItem('boot', 'ready');
    </script>
  </body>
</html>
""".strip().encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self) -> None:  # noqa: N802
        content_length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(content_length).decode("utf-8")
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/ping":
            payload = json.loads(body or "{}")
            response_body = json.dumps(
                {
                    "status": "ok",
                    "message": "pong",
                    "query": urllib.parse.parse_qs(parsed.query),
                    "payload": payload,
                }
            ).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(response_body)))
            self.end_headers()
            self.wfile.write(response_body)
            return
        form_value = ""
        for part in body.split("&"):
            if part.startswith("name="):
                form_value = part.split("=", 1)[1].replace("+", " ")
        type(self).clicked = True
        type(self).last_form_value = form_value
        self.send_response(303)
        self.send_header("Location", "/")
        self.end_headers()

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


class _BrowserMockSiteServer(TCPServer):
    allow_reuse_address = True


def _start_browser_mock_site() -> tuple[_BrowserMockSiteServer, threading.Thread]:
    _BrowserMockSiteHandler.clicked = False
    _BrowserMockSiteHandler.last_form_value = ""
    server = _BrowserMockSiteServer(("127.0.0.1", 0), _BrowserMockSiteHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server, thread


def test_workbench_snapshot_exposes_ui_read_model() -> None:
    service = CompilationWorkbenchService()

    snapshot = service.get_workbench_snapshot()

    assert isinstance(snapshot["workbench"]["workspace_session_id"], str)
    assert snapshot["workbench"]["workspace_session_id"].startswith("ws-")
    assert isinstance(snapshot["workbench"]["service_started_at"], str)
    assert snapshot["workbench"]["service_started_at"]
    assert snapshot["workbench"]["compile_counter"] == 0
    assert snapshot["project"]["loaded"] is True
    assert snapshot["project"]["project_id"] == "weconduct-workspace"
    assert snapshot["project"]["project_name"] == "WeConduct Workspace"
    assert snapshot["project"]["project_status"] == "ready"
    assert snapshot["project"]["project_schema_version"] == "project-v1"
    assert snapshot["project"]["source_of_truth"] == "graph_document"
    assert snapshot["project"]["main_graph_document_id"] == "graph:workspace"
    assert snapshot["project"]["resource_registry_revision"] == 0
    assert snapshot["project"]["workspace_root"] == str(Path(__file__).resolve().parents[2])
    assert snapshot["project"]["has_persisted_workspace_state"] is False
    assert snapshot["project"]["last_compile_status"] is None
    assert snapshot["project"]["last_runtime_status"] is None
    assert snapshot["project"]["last_debug_status"] is None
    assert snapshot["capabilities"]["compiler_available"] is True
    assert snapshot["entrypoints"]["compile_action"] == "/api/workbench/compile"
    assert snapshot["entrypoints"]["graph_source_projection"] == "/api/workbench/graph/source-projection"
    assert snapshot["workbench"]["host_mode"] == "python_core"
    assert snapshot["workbench"]["api_version"] == "0.4.0"
    assert snapshot["compiler"]["available_source_kinds"] == [
        "graph_workspace",
        "native_flow",
        "webcontrol_main_flow",
        "webcontrol_blueprint",
    ]
    assert snapshot["compiler"]["supported_stage_names"] == [
        "parse",
        "bind",
        "validate",
        "normalize",
        "lower",
        "emit",
    ]
    assert snapshot["compiler"]["compile_statuses"] == [
        "succeeded",
        "failed",
        "unsupported",
    ]
    assert snapshot["compiler"]["compile_history_limit"] == 5
    assert snapshot["compiler"]["diagnostic_severities"] == [
        "info",
        "warning",
        "degraded",
        "error",
        "fatal",
    ]
    assert snapshot["compiler"]["default_source_kind"] == "graph_workspace"
    assert snapshot["compiler"]["source_templates"]["graph_workspace"]["entry_document"] == "graph:workspace"
    assert (
        snapshot["compiler"]["source_templates"]["graph_workspace"]["source_text"]
        == '{"graph_model_id":"graph:workspace","compilation_id":null,"graph_schema_version":"graph-v1","nodes":[{"node_id":"node-1","lowered_kind":"execution","source_anchor_ref":"n1","expansion_role":"action:request","display_name":"HTTP Request","node_kind":"http.request","position":{"x":120,"y":80},"ports":[{"port_id":"out-main","direction":"output","relation_layer":"data","semantic_slot":"out.result"}],"node_config":{"method":"GET"}},{"node_id":"node-2","lowered_kind":"execution","source_anchor_ref":"n2","expansion_role":"transform:map","display_name":"Map Result","node_kind":"data.map","position":{"x":360,"y":80},"ports":[{"port_id":"in-main","direction":"input","relation_layer":"data","semantic_slot":"in.default"}],"node_config":{"mode":"map"}}],"edges":[{"edge_id":"edge-1","relation_layer":"data","from_node_id":"node-1","to_node_id":"node-2","from_port_id":"out-main","to_port_id":"in-main","edge_state":"draft"}],"viewport":{"x":0,"y":0,"zoom":1.1},"graph_effective_diagnostic_anchor_refs":[]}'
    )
    assert snapshot["compiler"]["source_templates"]["native_flow"]["entry_document"] == "examples/native-flow.json"
    assert (
        snapshot["compiler"]["source_templates"]["native_flow"]["source_text"]
        == '{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}'
    )
    assert (
        snapshot["compiler"]["source_templates"]["webcontrol_main_flow"]["entry_document"]
        == "examples/webcontrol-main-flow.json"
    )
    assert (
        snapshot["compiler"]["source_templates"]["webcontrol_main_flow"]["source_text"]
        == '{"project_info":{"name":"demo"},"automation_steps":[{"step_id":"step-1","action":"open_url"}]}'
    )
    assert (
        snapshot["compiler"]["source_templates"]["webcontrol_blueprint"]["entry_document"]
        == "examples/webcontrol-blueprint.json"
    )
    assert (
        snapshot["compiler"]["source_templates"]["webcontrol_blueprint"]["source_text"]
        == '{"blueprint_info":{"id":"blueprint-demo","name":"Demo Blueprint"},"input_schema":{"username":{"type":"string"}},"output_schema":{"logged_in":{"type":"boolean"}},"automation_steps":[{"step_id":"step-1","action":"open_url"}]}'
    )
    assert snapshot["last_compile"] is None
    assert snapshot["compile_history"] == []


def test_workbench_snapshot_and_project_document_expose_project_execution_overview() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "snapshot_value", "value": "ok"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-set",
                    "from_port_id": "out-control",
                    "to_port_id": None,
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    runtime_started = service.start_runtime_session(None)
    service.run_runtime_session(session_id=runtime_started["runtime_session"]["session_id"])
    debug_started = service.start_debug_session(None)

    snapshot = service.get_workbench_snapshot()
    project_document = service.get_project_document()

    assert snapshot["project"]["last_runtime_status"] == "completed"
    assert snapshot["project"]["last_runtime_session_id"] == runtime_started["runtime_session"]["session_id"]
    assert snapshot["project"]["last_debug_status"] == "prepared"
    assert snapshot["project"]["last_debug_session_id"] == debug_started["debug_session"]["session_id"]
    assert snapshot["project"]["execution_overview"]["runtime_run_count"] == 1
    assert snapshot["project"]["execution_overview"]["debug_session_count"] == 1
    assert snapshot["project"]["execution_overview"]["runtime_status_counts"]["completed"] == 1
    assert project_document["project"]["execution_overview"]["debug_status_counts"]["prepared"] == 1


def test_service_can_create_new_project_and_reset_workspace_state() -> None:
    store = InMemoryWorkspaceStateStore()
    service = CompilationWorkbenchService(state_store=store)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.compile_graph_document(None)
    previous_snapshot = service.get_workbench_snapshot()

    result = service.create_project(project_name="Demo Project")
    current_snapshot = service.get_workbench_snapshot()

    assert result["status"] == "created"
    assert result["project"]["project_name"] == "Demo Project"
    assert result["project"]["project_id"] != previous_snapshot["project"]["project_id"]
    assert result["project"]["project_schema_version"] == "project-v1"
    assert result["graph_document"].graph_model_id == "graph:workspace"
    assert result["graph_document"].nodes == []
    assert current_snapshot["project"]["project_name"] == "Demo Project"
    assert current_snapshot["project"]["project_id"] == result["project"]["project_id"]
    assert current_snapshot["workbench"]["compile_counter"] == 0
    assert current_snapshot["last_compile"] is None
    assert current_snapshot["compile_history"] == []


def test_service_snapshot_exposes_preferences_document() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "en-US",
                    "theme": "dark",
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 7,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": False,
                    "allow_file_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(
        state_store=InMemoryWorkspaceStateStore(),
        preferences_service=preferences_service,
    )

    snapshot = service.get_workbench_snapshot()

    assert snapshot["preferences"]["program_settings"]["language"] == "en-US"
    assert snapshot["preferences"]["program_settings"]["recent_project_limit"] == 7
    assert snapshot["preferences"]["graph_settings"]["auto_sync_mode"] == "responsive"


def test_service_create_project_uses_default_preferences_directory_when_missing(tmp_path) -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    default_project_directory = tmp_path / "preferred-project-home"
    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "theme": "light",
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": str(default_project_directory.resolve()),
                    "recent_project_limit": 10,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": False,
                    "allow_file_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(
        state_store=InMemoryWorkspaceStateStore(),
        preferences_service=preferences_service,
    )

    result = service.create_project(project_name="Preferred Home Project")

    expected_project_path = default_project_directory / "Preferred Home Project.weconduct.json"
    assert result["project"]["project_file_path"] == str(expected_project_path.resolve())
    assert result["project"]["workspace_root"] == str(default_project_directory.resolve())
    assert expected_project_path.exists() is True


def test_service_create_project_can_persist_into_selected_directory(tmp_path) -> None:
    store = InMemoryWorkspaceStateStore()
    service = CompilationWorkbenchService(state_store=store)
    project_directory = tmp_path / "custom-project-home" / "demo-project"

    result = service.create_project(
        project_name="Demo Project",
        project_directory=project_directory,
    )
    current_snapshot = service.get_workbench_snapshot()
    expected_project_path = project_directory / "Demo Project.weconduct.json"

    assert result["status"] == "created"
    assert result["project"]["project_name"] == "Demo Project"
    assert result["project"]["project_file_path"] == str(expected_project_path.resolve())
    assert result["project"]["workspace_root"] == str(project_directory.resolve())
    assert result["project"]["is_dirty"] is False
    assert expected_project_path.exists() is True
    assert current_snapshot["project"]["project_file_path"] == str(expected_project_path.resolve())
    assert current_snapshot["project"]["workspace_root"] == str(project_directory.resolve())
    assert current_snapshot["project"]["recent_project_count"] == 1
    assert current_snapshot["project"]["recent_projects"][0]["project_path"] == str(
        expected_project_path.resolve()
    )


def test_service_recent_projects_respects_preferences_limit(tmp_path) -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "theme": "light",
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 1,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": False,
                    "allow_file_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(
        state_store=InMemoryWorkspaceStateStore(),
        preferences_service=preferences_service,
    )

    service.create_project(project_name="First Project", project_directory=tmp_path / "first")
    service.create_project(project_name="Second Project", project_directory=tmp_path / "second")

    recent_projects = service.get_recent_projects_document()
    assert len(recent_projects["recent_projects"]) == 1
    assert recent_projects["recent_projects"][0]["project_name"] == "Second Project"


def test_service_can_save_project_as_track_dirty_state_and_open_recent_project(tmp_path) -> None:
    store = InMemoryWorkspaceStateStore()
    service = CompilationWorkbenchService(state_store=store)
    project_a_path = tmp_path / "project-a.weconduct.json"
    project_b_path = tmp_path / "project-b.weconduct.json"

    created_project = service.create_project(project_name="Project A")
    created_snapshot = service.get_workbench_snapshot()

    assert created_project["status"] == "created"
    assert created_snapshot["project"]["project_name"] == "Project A"
    assert created_snapshot["project"]["project_file_path"] is None
    assert created_snapshot["project"]["is_dirty"] is True
    assert created_snapshot["project"]["recent_project_count"] == 0

    save_as_result = service.save_project_as(project_path=project_a_path)
    saved_snapshot = service.get_workbench_snapshot()

    assert save_as_result["status"] == "saved"
    assert save_as_result["project"]["project_file_path"] == str(project_a_path.resolve())
    assert save_as_result["project"]["is_dirty"] is False
    assert project_a_path.exists() is True
    assert saved_snapshot["project"]["project_file_name"] == "project-a.weconduct.json"
    assert saved_snapshot["project"]["recent_project_count"] == 1
    assert saved_snapshot["project"]["recent_projects"][0]["project_path"] == str(
        project_a_path.resolve()
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    dirty_snapshot = service.get_workbench_snapshot()

    assert dirty_snapshot["project"]["is_dirty"] is True

    save_result = service.save_project()
    clean_snapshot = service.get_workbench_snapshot()

    assert save_result["status"] == "saved"
    assert save_result["project"]["project_file_path"] == str(project_a_path.resolve())
    assert clean_snapshot["project"]["is_dirty"] is False
    assert clean_snapshot["project"]["recent_project_count"] == 1

    other_service = CompilationWorkbenchService(state_store=store)
    other_service.create_project(project_name="Project B")
    other_service.save_project_as(project_path=project_b_path)

    opened_result = service.open_project(project_path=project_b_path)
    opened_snapshot = service.get_workbench_snapshot()

    assert opened_result["status"] == "opened"
    assert opened_result["project"]["project_name"] == "Project B"
    assert opened_result["project"]["project_file_path"] == str(project_b_path.resolve())
    assert opened_result["graph_document"].nodes == []
    assert opened_snapshot["project"]["is_dirty"] is False
    assert opened_snapshot["project"]["recent_project_count"] == 2
    assert opened_snapshot["project"]["recent_projects"][0]["project_path"] == str(
        project_b_path.resolve()
    )
    assert opened_snapshot["project"]["recent_projects"][1]["project_path"] == str(
        project_a_path.resolve()
    )


def test_service_rejects_save_project_when_project_file_path_is_missing() -> None:
    service = CompilationWorkbenchService()

    try:
        service.save_project()
    except ValueError as exc:
        assert str(exc) == "project_file_path is not set; use save_project_as first"
    else:
        raise AssertionError("expected ValueError when saving project without file path")


def test_service_reports_save_project_requires_save_as_when_project_file_path_is_missing() -> None:
    service = CompilationWorkbenchService()

    try:
        service.save_project()
    except ValueError as exc:
        assert getattr(exc, "error_code", None) == "project.needs_save_as"
        assert getattr(exc, "recovery_action", None) == "save_as"
        assert str(exc) == "project_file_path is not set; use save_project_as first"
    else:
        raise AssertionError("expected save-as requirement when saving project without file path")


def test_service_exposes_recent_projects_document(tmp_path) -> None:
    service = CompilationWorkbenchService()
    first_path = tmp_path / "first.weconduct.json"
    second_path = tmp_path / "second.weconduct.json"

    service.create_project(project_name="First Project")
    service.save_project_as(project_path=first_path)
    service.create_project(project_name="Second Project")
    service.save_project_as(project_path=second_path)

    recent_projects = service.get_recent_projects_document()

    assert recent_projects["recent_projects"][0]["project_name"] == "Second Project"
    assert recent_projects["recent_projects"][0]["project_path"] == str(second_path.resolve())
    assert recent_projects["recent_projects"][1]["project_name"] == "First Project"
    assert recent_projects["recent_projects"][1]["project_path"] == str(first_path.resolve())


def test_service_component_library_and_resource_registry_expose_i18n_fields() -> None:
    service = CompilationWorkbenchService()

    component_library = service.get_component_library_document(
        query="captcha",
        tags=["domain:browser"],
    )
    full_component_library = service.get_component_library_document(
        tags=["origin:builtin"],
        enabled=True,
    )
    resource_registry = service.get_resource_registry_document(
        query="navigate",
        tags=["origin:builtin"],
        enabled=True,
    )
    full_registry = service.get_resource_registry_document(
        tags=["origin:builtin"],
        enabled=True,
    )

    captcha_item = next(
        item
        for item in component_library["items"]
        if item["resource_key"] == "browser.recognize_captcha"
    )
    navigate_resource = next(
        item
        for item in resource_registry["resources"]
        if item["resource_key"] == "browser.navigate"
    )
    control_if_resource = next(
        item
        for item in full_component_library["items"]
        if item["resource_key"] == "control.if"
    )
    excel_write_cell_resource = next(
        item
        for item in full_registry["resources"]
        if item["resource_key"] == "excel.write_cell"
    )

    assert captcha_item["display_name_i18n"]["zh-CN"] == "识别验证码"
    assert captcha_item["display_name_i18n"]["en-US"] == "Recognize Captcha"
    assert captcha_item["description_i18n"]["zh-CN"] == "使用 captcha_ocr 识别验证码图片。"
    assert captcha_item["description_i18n"]["en-US"] == "Recognize captcha image with captcha_ocr."
    assert navigate_resource["display_name_i18n"]["zh-CN"] == "导航"
    assert navigate_resource["display_name_i18n"]["en-US"] == "Navigate"
    assert navigate_resource["description_i18n"]["zh-CN"] == "导航浏览器到目标 URL。"
    assert navigate_resource["description_i18n"]["en-US"] == "Navigate browser to target URL."
    assert control_if_resource["display_name_i18n"]["zh-CN"] == "条件分支"
    assert control_if_resource["display_name_i18n"]["en-US"] == "If"
    assert control_if_resource["description_i18n"]["zh-CN"] == "根据布尔条件分支控制流。"
    assert control_if_resource["description_i18n"]["en-US"] == "Branch control flow based on a boolean condition."
    assert excel_write_cell_resource["display_name_i18n"]["zh-CN"] == "写入 Excel 单元格"
    assert excel_write_cell_resource["display_name_i18n"]["en-US"] == "Write Excel Cell"
    assert excel_write_cell_resource["description_i18n"]["zh-CN"] == "写入 Excel 单元格的值。"
    assert excel_write_cell_resource["description_i18n"]["en-US"] == "Write Excel cell value."


def test_service_exposes_resource_registry_and_can_save_user_component_and_toggle_resource(
    tmp_path,
) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "resource-project.weconduct.json"

    initial_registry = service.get_resource_registry_document()

    assert initial_registry["registry_revision"] == 0
    assert initial_registry["summary"]["builtin_resource_count"] >= 2
    assert initial_registry["summary"]["user_resource_count"] == 0
    assert initial_registry["resources"][0]["resource_type"] == "builtin_component"
    assert initial_registry["resources"][0]["enabled"] is True
    assert all(item["resource_key"] != "data.create_list" for item in initial_registry["resources"])
    assert any(
        item["resource_key"] == "data.create_list"
        and item["node_taxonomy"] == "logic_expression"
        and item["component_library_visible"] is True
        and item["resource_manager_visible"] is False
        for item in service.get_component_library_document()["items"]
    )
    assert all(
        item["resource_key"] != "graph.call_subgraph"
        for item in service.get_component_library_document()["items"]
    )
    assert all(
        item["resource_key"] != "graph.call_subgraph"
        for item in initial_registry["resources"]
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    saved_resource = service.save_user_component_resource(resource_name="My Component")
    after_create_registry = service.get_resource_registry_document()

    assert saved_resource["status"] == "saved"
    assert saved_resource["resource"]["resource_type"] == "user_component"
    assert saved_resource["resource"]["display_name"] == "My Component"
    assert saved_resource["resource"]["source_graph_document_id"] == "graph:workspace"
    assert after_create_registry["registry_revision"] == 1
    assert after_create_registry["summary"]["user_resource_count"] == 1

    builtin_resource_id = after_create_registry["resources"][0]["resource_id"]
    toggled_resource = service.set_resource_enabled(resource_id=builtin_resource_id, enabled=False)
    after_toggle_snapshot = service.get_workbench_snapshot()

    assert toggled_resource["status"] == "updated"
    assert toggled_resource["resource"]["resource_id"] == builtin_resource_id
    assert toggled_resource["resource"]["enabled"] is False
    assert after_toggle_snapshot["project"]["resource_registry_revision"] == 2

    service.save_project_as(project_path=project_path)
    reopened_service = CompilationWorkbenchService()
    reopened_result = reopened_service.open_project(project_path=project_path)
    reopened_registry = reopened_service.get_resource_registry_document()

    assert reopened_result["status"] == "opened"
    assert reopened_registry["registry_revision"] == 2
    assert reopened_registry["summary"]["user_resource_count"] == 1
    assert any(
        item["resource_type"] == "user_component" and item["display_name"] == "My Component"
        for item in reopened_registry["resources"]
    )
    assert any(
        item["resource_id"] == builtin_resource_id and item["enabled"] is False
        for item in reopened_registry["resources"]
    )


def test_service_can_save_subgraph_resource_and_expose_it_in_registry() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "msg", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    saved = service.save_subgraph_resource(resource_name="Greeting Subgraph")

    assert saved["resource"]["resource_type"] == "custom_node_graph"
    assert saved["resource"]["display_name"] == "Greeting Subgraph"

    registry = service.get_resource_registry_document()
    assert any(
        item["resource_type"] == "custom_node_graph"
        and item["display_name"] == "Greeting Subgraph"
        for item in registry["resources"]
    )


def test_service_can_save_custom_node_graph_resource_and_expose_it_in_registry() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:custom-node-graph",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "text": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-node-1",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Value",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    saved = service.save_custom_node_graph_resource(resource_name="Greeting Graph")

    assert saved["status"] == "saved"
    assert saved["resource"]["resource_type"] == "custom_node_graph"
    assert saved["resource"]["display_name"] == "Greeting Graph"
    assert saved["resource"]["input_schema"] == {
        "text": {"type": "string", "required": True},
    }
    assert saved["resource"]["output_schema"] == {
        "message": {"type": "string"},
    }

    registry = service.get_resource_registry_document()["resources"]
    assert any(
        item["resource_type"] == "custom_node_graph"
        and item["display_name"] == "Greeting Graph"
        for item in registry
    )


def test_service_validates_custom_node_graph_instance_and_reports_missing_resource() -> None:
    service = CompilationWorkbenchService()
    graph_document = {
        "graph_model_id": "graph:validation",
        "compilation_id": None,
        "nodes": [
            {
                "node_id": "node-1",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-node-1",
                "expansion_role": "action:custom_node_graph",
                "display_name": "Missing Custom Graph",
                "node_kind": "custom_node_graph:missing",
                "node_config": {},
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    result = service.validate_graph_document(graph_document)
    categories = [item["category"] for item in result["diagnostics"]]

    assert "custom_node_graph.missing" in categories


def test_service_open_project_migrates_legacy_subgraph_resource_to_custom_node_graph(
    tmp_path,
) -> None:
    project_path = tmp_path / "legacy-project.weconduct.json"
    project_path.write_text(
        json.dumps(
            {
                "project_file_schema_version": 1,
                "saved_at": "2026-06-16T12:00:00+00:00",
                "project": {
                    "project_id": "legacy-project",
                    "project_name": "Legacy Project",
                    "project_schema_version": "project-v1",
                    "project_status": "ready",
                    "workspace_root": str(tmp_path),
                    "source_of_truth": "graph_document",
                    "main_graph_document_id": "graph:main",
                    "resource_registry_revision": 1,
                },
                "resource_registry": [
                    {
                        "resource_id": "subgraph_resource:legacy",
                        "resource_type": "subgraph_resource",
                        "display_name": "Legacy Subgraph",
                        "resource_key": "subgraph_resource:legacy",
                        "enabled": True,
                        "source_graph_document_id": "graph:legacy",
                        "source_graph_document_save_revision": 1,
                        "source_graph_document": {
                            "graph_model_id": "graph:legacy",
                            "compilation_id": None,
                            "nodes": [],
                            "edges": [],
                            "graph_effective_diagnostic_anchor_refs": [],
                        },
                    }
                ],
                "graph_document": {
                    "graph_model_id": "graph:main",
                    "compilation_id": None,
                    "nodes": [],
                    "edges": [],
                    "graph_effective_diagnostic_anchor_refs": [],
                },
                "graph_document_meta": {
                    "save_revision": 1,
                    "saved_at": "2026-06-16T12:00:00+00:00",
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    service = CompilationWorkbenchService()
    service.open_project(project_path=project_path)
    resources = service.get_resource_registry_document()["resources"]
    migrated = next(item for item in resources if item["display_name"] == "Legacy Subgraph")

    assert migrated["resource_type"] == "custom_node_graph"
    assert migrated["resource_key"] == "custom_node_graph:legacy"


def test_service_save_subgraph_resource_inherits_input_output_schema_from_graph_root_metadata() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                    "optional_value": {"type": "string", "required": False},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "msg", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    saved = service.save_subgraph_resource(resource_name="Schema Subgraph")

    assert saved["resource"]["input_schema"] == {
        "incoming": {"type": "string", "required": True},
        "optional_value": {"type": "string", "required": False},
    }
    assert saved["resource"]["output_schema"] == {
        "message": {"type": "string"},
    }


def test_service_rejects_resource_toggle_for_unknown_resource_id() -> None:
    service = CompilationWorkbenchService()

    try:
        service.set_resource_enabled(resource_id="missing-resource", enabled=False)
    except ValueError as exc:
        assert str(exc) == "resource not found: missing-resource"
    else:
        raise AssertionError("expected ValueError for missing resource")


def test_service_can_remove_recent_project_and_track_editor_history(tmp_path) -> None:
    service = CompilationWorkbenchService()
    first_path = tmp_path / "first.weconduct.json"
    second_path = tmp_path / "second.weconduct.json"

    service.create_project(project_name="First Project")
    service.save_project_as(project_path=first_path)
    service.create_project(project_name="Second Project")
    service.save_project_as(project_path=second_path)

    removed_recent = service.remove_recent_project(project_path=first_path)
    history_before = service.get_editor_history_document()

    service.record_editor_operation(
        operation_kind="graph.node.added",
        label="Add HTTP Request",
        payload={"node_id": "node-1"},
    )
    history_after = service.get_editor_history_document()

    assert removed_recent["status"] == "removed"
    assert removed_recent["recent_projects"][0]["project_path"] == str(second_path.resolve())
    assert len(removed_recent["recent_projects"]) == 1
    assert history_before["undo_depth"] == 0
    assert history_before["redo_depth"] == 0
    assert history_after["undo_depth"] == 1
    assert history_after["undo_stack"][0]["operation_kind"] == "graph.node.added"
    assert history_after["redo_stack"] == []


def test_service_can_export_import_and_replace_user_component_resource(tmp_path) -> None:
    service = CompilationWorkbenchService()
    export_path = tmp_path / "exported-resource.wecresource.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    first_saved = service.save_user_component_resource(resource_name="Reusable Component")
    exported = service.export_resource(resource_id=first_saved["resource"]["resource_id"], export_path=export_path)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "node_kind": "data.map",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    replaced = service.save_user_component_resource(
        resource_name="Reusable Component",
        replace_existing_resource_id=first_saved["resource"]["resource_id"],
    )
    imported = service.import_resource(import_path=export_path, replace_existing=True)
    registry = service.get_resource_registry_document()

    assert exported["status"] == "exported"
    assert export_path.exists() is True
    assert replaced["status"] == "saved"
    assert replaced["resource"]["resource_id"] == first_saved["resource"]["resource_id"]
    assert replaced["resource"]["source_graph_document_save_revision"] == 2
    assert imported["status"] == "imported"
    assert imported["resource"]["display_name"] == "Reusable Component"
    assert registry["summary"]["user_resource_count"] == 1
    assert registry["registry_revision"] >= 3


def test_service_can_export_import_subgraph_resource_with_schema(tmp_path) -> None:
    service = CompilationWorkbenchService()
    export_path = tmp_path / "exported-subgraph-resource.wecresource.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    saved = service.save_subgraph_resource(resource_name="Schema Export Subgraph")
    exported = service.export_resource(
        resource_id=saved["resource"]["resource_id"],
        export_path=export_path,
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.save_subgraph_resource(
        resource_name="Schema Export Subgraph",
        replace_existing_resource_id=saved["resource"]["resource_id"],
    )
    imported = service.import_resource(import_path=export_path, replace_existing=True)

    assert exported["status"] == "exported"
    assert imported["status"] == "imported"
    assert imported["resource"]["input_schema"] == {
        "incoming": {"type": "string", "required": True},
    }
    assert imported["resource"]["output_schema"] == {
        "message": {"type": "string"},
    }


def test_service_can_import_legacy_webcontrol_blueprint_as_custom_node_graph(
    tmp_path,
) -> None:
    service = CompilationWorkbenchService()
    import_path = tmp_path / "legacy-blueprint.json"
    import_path.write_text(
        json.dumps(
            {
                "blueprint_info": {
                    "id": "bp-login",
                    "name": "Login Blueprint",
                },
                "input_schema": {
                    "username": {"type": "string"},
                },
                "output_schema": {
                    "logged_in": {"type": "boolean"},
                },
                "automation_steps": [
                    {
                        "step_id": "step-1",
                        "action": "open_url",
                        "url": "https://example.com/login",
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    imported = service.import_resource(import_path=import_path)

    assert imported["status"] == "imported"
    assert imported["resource"]["resource_type"] == "custom_node_graph"
    assert imported["resource"]["resource_id"] == "custom_node_graph:bp-login"
    assert imported["resource"]["display_name"] == "Login Blueprint"
    assert imported["resource"]["input_schema"] == {
        "username": {"type": "string"},
    }
    assert imported["resource"]["output_schema"] == {
        "logged_in": {"type": "boolean"},
    }
    assert imported["resource"]["source_graph_document"]["root_metadata"]["source_kind"] == (
        "webcontrol_blueprint"
    )


def test_service_can_convert_legacy_webcontrol_project_into_split_project_layout(
    tmp_path: Path,
) -> None:
    service = CompilationWorkbenchService()
    source_path = tmp_path / "legacy-main.yaml"
    blueprint_path = tmp_path / "bp-login.yaml"
    output_project_path = tmp_path / "converted" / "converted-demo.weconduct.json"
    source_path.write_text(
        """
project_info:
  name: 转换后的演示项目
browser_config:
  headless: true
initial_variables:
  username: alice
automation_steps:
  - step: 1
    action: open_url
    url: "https://example.com/login"
  - step: 2
    action: call_blueprint
    blueprint_id: "bp-login"
    inputs:
      username: "${username}"
    outputs:
      logged_in: "logged_in"
""".strip(),
        encoding="utf-8",
    )
    blueprint_path.write_text(
        """
blueprint_info:
  id: bp-login
  name: 登录蓝图
input_schema:
  username:
    type: string
output_schema:
  logged_in:
    type: boolean
automation_steps:
  - step: 1
    action: open_url
    url: "https://example.com/form"
""".strip(),
        encoding="utf-8",
    )

    result = service.convert_webcontrol_project(
        source_path=source_path,
        blueprint_paths=[blueprint_path],
        output_project_path=output_project_path,
        auto_open_project=False,
        preserve_legacy_metadata=True,
        write_conversion_report=True,
    )

    report_path = output_project_path.parent / f"{output_project_path.stem}.data" / "conversion-report.json"
    reopened = CompilationWorkbenchService()
    opened = reopened.open_project(project_path=output_project_path)
    registry = reopened.get_resource_registry_document()["resources"]

    assert result["status"] == "converted"
    assert result["output_project_path"] == str(output_project_path.resolve())
    assert result["report"]["source_kind"] == "webcontrol_main_flow"
    assert result["report"]["imported_blueprint_count"] == 1
    assert result["report"]["generated_resource_count"] == 1
    assert result["report"]["errors"] == []
    assert output_project_path.exists() is True
    assert report_path.exists() is True
    assert opened["status"] == "opened"
    assert opened["project"]["project_name"] == "转换后的演示项目"
    assert opened["graph_document"].root_metadata["source_kind"] == "webcontrol_main_flow"
    assert opened["graph_document"].root_metadata["legacy_webcontrol_source"]["source_kind"] == (
        "webcontrol_main_flow"
    )
    assert any(item["resource_id"] == "custom_node_graph:bp-login" for item in registry)


def test_converted_legacy_webcontrol_project_becomes_editable_graph_workspace(
    tmp_path: Path,
) -> None:
    service = CompilationWorkbenchService()
    source_path = tmp_path / "legacy-main-full.yaml"
    output_project_path = tmp_path / "converted" / "editable-demo.weconduct.json"
    source_path.write_text(
        """
project_info:
  name: 可编辑转换项目
browser_config:
  headless: false
  slow_mo: 150
initial_variables:
  base_url: https://example.com/login
  username: alice
automation_steps:
  - step: 1
    action: open_url
    url: "${base_url}"
  - step: 2
    action: screenshot
    filename: "${base_url}/shot.png"
  - step: 3
    action: call_blueprint
    blueprint_id: "bp-login"
    inputs:
      username: "${username}"
    outputs:
      logged_in: "logged_in"
""".strip(),
        encoding="utf-8",
    )
    blueprint_path = tmp_path / "bp-login.yaml"
    blueprint_path.write_text(
        """
blueprint_info:
  id: bp-login
  name: 登录蓝图
input_schema:
  username:
    type: string
output_schema:
  logged_in:
    type: boolean
automation_steps:
  - step: 1
    action: open_url
    url: "https://example.com/form"
""".strip(),
        encoding="utf-8",
    )

    service.convert_webcontrol_project(
        source_path=source_path,
        blueprint_paths=[blueprint_path],
        output_project_path=output_project_path,
        auto_open_project=False,
        preserve_legacy_metadata=True,
        write_conversion_report=True,
    )

    reopened = CompilationWorkbenchService()
    opened = reopened.open_project(project_path=output_project_path)
    graph_document = opened["graph_document"]
    graph_payload = graph_document.model_dump(mode="json")
    nodes_by_kind = {}
    for node in graph_payload["nodes"]:
        nodes_by_kind.setdefault(node["node_kind"], []).append(node)

    assert "flow.start" in nodes_by_kind
    assert nodes_by_kind["flow.start"][0]["lowered_kind"] == "control"
    assert nodes_by_kind["flow.start"][0]["node_config"]["initial_variables"] == {
        "base_url": "https://example.com/login",
        "username": "alice",
    }
    assert nodes_by_kind["flow.start"][0]["node_config"]["browser_config"] == {
        "headless": False,
        "slow_mo_ms": 150,
    }

    assert "browser.navigate" in nodes_by_kind
    assert nodes_by_kind["browser.navigate"][0]["node_config"]["url"] == "${base_url}"
    assert "browser.screenshot" in nodes_by_kind
    assert nodes_by_kind["browser.screenshot"][0]["node_config"]["path"] == "${base_url}/shot.png"
    assert "call_blueprint" in nodes_by_kind
    assert nodes_by_kind["call_blueprint"][0]["node_config"]["blueprint_id"] == "bp-login"
    assert nodes_by_kind["call_blueprint"][0]["node_config"]["inputs"] == {
        "username": "${username}",
    }
    assert nodes_by_kind["call_blueprint"][0]["node_config"]["outputs"] == {
        "logged_in": "logged_in",
    }

    control_edges = [edge for edge in graph_payload["edges"] if edge["relation_layer"] == "control"]
    assert len(control_edges) == 3
    assert graph_payload["root_metadata"]["source_kind"] == "webcontrol_main_flow"

    compile_result = reopened.compile_graph_document(None)

    assert compile_result["status"] == "succeeded"
    assert compile_result["view"]["status"] == "succeeded"


def test_service_save_and_open_project_preserves_subgraph_resource_schema(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "subgraph-schema-project.weconduct.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema Project Subgraph")
    service.save_project_as(project_path=project_path)

    reopened_service = CompilationWorkbenchService()
    reopened_service.open_project(project_path=project_path)
    reopened_registry = reopened_service.get_resource_registry_document()
    reopened_resource = next(
        item
        for item in reopened_registry["resources"]
        if item["resource_id"] == saved["resource"]["resource_id"]
    )

    assert reopened_resource["input_schema"] == {
        "incoming": {"type": "string", "required": True},
    }
    assert reopened_resource["output_schema"] == {
        "message": {"type": "string"},
    }


def test_service_persists_execution_history_and_user_component_runtime_bridge(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "execution-history.weconduct.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_component = service.save_user_component_resource(resource_name="Reusable HTTP Block")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-user-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "user1",
                    "expansion_role": "module:user-component",
                    "node_kind": saved_component["resource"]["resource_id"],
                    "node_config": {
                        "outputs": {"message": "greeting"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started_runtime = service.start_runtime_session(None)
    runtime_session = service.run_runtime_session(
        session_id=started_runtime["runtime_session"]["session_id"]
    )
    started_debug = service.start_debug_session(None)
    execution_history = service.get_execution_history_document()
    registry = service.get_resource_registry_document()

    assert started_runtime["runtime_plan"]["executable_nodes"][0]["resolved_resource_id"] == (
        saved_component["resource"]["resource_id"]
    )
    assert started_runtime["runtime_plan"]["executable_nodes"][0]["resource_type"] == "user_component"
    assert started_runtime["runtime_plan"]["executable_nodes"][0]["resource_origin"] == "project"
    assert started_runtime["runtime_plan"]["executable_nodes"][0]["component_source_graph_document_id"] == (
        "graph:workspace"
    )
    assert (
        started_runtime["runtime_plan"]["executable_nodes"][0]["component_source_graph_document"]["nodes"][0]["node_id"]
        == "node-1"
    )
    assert runtime_session["status"] == "completed"
    assert runtime_session["result"]["variables"]["greeting"] == "hello"
    assert execution_history["summary"]["runtime_run_count"] == 1
    assert execution_history["summary"]["debug_session_count"] == 1
    assert execution_history["runtime_runs"][0]["status"] == "completed"
    assert execution_history["runtime_runs"][0]["graph_model_id"] == "graph:workspace"
    assert execution_history["debug_sessions"][0]["status"] == "prepared"
    assert execution_history["debug_sessions"][0]["session_id"] == (
        started_debug["debug_session"]["session_id"]
    )
    user_component = next(
        item
        for item in registry["resources"]
        if item["resource_id"] == saved_component["resource"]["resource_id"]
    )
    assert user_component["source_graph_document"]["nodes"][0]["node_id"] == "node-1"

    service.save_project_as(project_path=project_path)
    reopened_service = CompilationWorkbenchService()
    reopened_service.open_project(project_path=project_path)
    reopened_history = reopened_service.get_execution_history_document()
    reopened_registry = reopened_service.get_resource_registry_document()

    assert reopened_history["summary"]["runtime_run_count"] == 1
    assert reopened_history["summary"]["debug_session_count"] == 1


def test_service_execution_history_exposes_summary_and_filters_by_status() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "phase11_filter", "value": "ok"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "control",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out-control",
                    "to_port_id": None,
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    runtime_started = service.start_runtime_session(None)
    service.run_runtime_session(session_id=runtime_started["runtime_session"]["session_id"])
    service.start_debug_session(None)

    full_history = service.get_execution_history_document()
    completed_history = service.get_execution_history_document(runtime_status="completed")
    prepared_debug_history = service.get_execution_history_document(debug_status="prepared")

    assert full_history["summary"]["runtime_run_count"] == 1
    assert full_history["summary"]["debug_session_count"] == 1
    assert full_history["summary"]["runtime_status_counts"]["completed"] == 1
    assert full_history["summary"]["debug_status_counts"]["prepared"] == 1
    assert completed_history["summary"]["runtime_run_count"] == 1
    assert completed_history["runtime_runs"][0]["status"] == "completed"
    assert prepared_debug_history["summary"]["debug_session_count"] == 1
    assert prepared_debug_history["debug_sessions"][0]["status"] == "prepared"


def test_service_exposes_project_documents_index_and_component_library_view(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "documents-project.weconduct.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.save_user_component_resource(resource_name="Library Component")
    service.save_project_as(project_path=project_path)

    documents = service.get_project_documents_document()
    component_library = service.get_component_library_document()

    assert documents["main_graph_document_id"] == "graph:workspace"
    assert documents["documents"][0]["document_id"] == "graph:workspace"
    assert documents["documents"][0]["document_role"] == "main_graph"
    assert component_library["summary"]["available_resource_count"] >= 3
    assert any(item["display_name"] == "Library Component" for item in component_library["items"])


def test_service_exposes_node_taxonomy_visibility_for_resources_and_component_library() -> None:
    service = CompilationWorkbenchService()

    registry = service.get_resource_registry_document()
    component_library = service.get_component_library_document()

    registry_by_key = {item["resource_key"]: item for item in registry["resources"]}
    library_by_key = {item["resource_key"]: item for item in component_library["items"]}

    assert registry_by_key["browser.navigate"]["node_taxonomy"] == "builtin_component"
    assert registry_by_key["browser.navigate"]["resource_manager_visible"] is True
    assert registry_by_key["browser.navigate"]["component_library_visible"] is True
    assert registry_by_key["browser.navigate"]["user_creatable"] is True
    assert registry_by_key["browser.recognize_captcha"]["node_taxonomy"] == "builtin_component"
    assert registry_by_key["browser.recognize_captcha"]["resource_manager_visible"] is True

    assert "control.foreach" not in registry_by_key
    assert library_by_key["control.foreach"]["node_taxonomy"] == "control_structure"
    assert library_by_key["control.foreach"]["resource_manager_visible"] is False
    assert library_by_key["control.foreach"]["component_library_visible"] is True
    assert library_by_key["control.foreach"]["user_creatable"] is True

    assert "control.jump_to_step" not in registry_by_key
    assert "control.jump_to_step" not in library_by_key


def test_service_exposes_system_tags_for_builtin_and_project_resources() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_custom_node_graph_resource(resource_name="My Tagged Graph")

    registry = service.get_resource_registry_document()["resources"]
    library = service.get_component_library_document()["items"]

    browser_click = next(item for item in registry if item["resource_key"] == "browser.click")
    custom_graph = next(
        item for item in registry if item["resource_id"] == saved["resource"]["resource_id"]
    )
    custom_graph_library = next(
        item for item in library if item["resource_id"] == saved["resource"]["resource_id"]
    )

    assert "domain:browser" in browser_click["tags"]
    assert "type:builtin_component" in browser_click["tags"]
    assert "origin:builtin" in browser_click["tags"]

    assert "type:custom_node_graph" in custom_graph["tags"]
    assert "origin:project" in custom_graph["tags"]
    assert "taxonomy:user_component" in custom_graph["tags"]
    assert custom_graph_library["tags"] == custom_graph["tags"]


def test_service_filters_component_library_and_resource_registry_by_query_and_tags() -> None:
    service = CompilationWorkbenchService()

    component_library = service.get_component_library_document(
        query="captcha",
        tags=["domain:browser"],
    )
    assert any(item["resource_key"] == "browser.recognize_captcha" for item in component_library["items"])
    assert all("domain:browser" in item["tags"] for item in component_library["items"])
    assert all(
        "captcha" in json.dumps(item, ensure_ascii=False).lower()
        for item in component_library["items"]
    )

    registry = service.get_resource_registry_document(
        query="dialog",
        tags=["origin:builtin"],
        enabled=True,
    )
    assert registry["resources"]
    assert all("origin:builtin" in item["tags"] for item in registry["resources"])
    assert all(item["enabled"] is True for item in registry["resources"])


def test_component_library_exposes_new_flow_control_nodes_with_control_tags() -> None:
    service = CompilationWorkbenchService()
    library = service.get_component_library_document(tags=["domain:control"])
    by_key = {item["resource_key"]: item for item in library["items"]}

    for resource_key in [
        "control.if",
        "control.switch",
        "control.parallel_fork",
        "control.join",
        "control.while",
        "control.retry",
        "control.failover",
    ]:
        assert resource_key in by_key
        assert "domain:control" in by_key[resource_key]["tags"]


def test_service_builds_graph_node_drafts_for_builtin_resources() -> None:
    service = CompilationWorkbenchService()

    flow_start = service.build_graph_node_draft(resource_key="flow.start")
    control_if = service.build_graph_node_draft(resource_key="control.if")
    browser_fill = service.build_graph_node_draft(resource_key="browser.fill")
    data_set_variable = service.build_graph_node_draft(resource_key="data.set_variable")
    http_request = service.build_graph_node_draft(resource_key="http.request")
    browser_click = service.build_graph_node_draft(resource_key="browser.click")
    browser_wait_for_element = service.build_graph_node_draft(
        resource_key="browser.wait_for_element"
    )
    browser_screenshot = service.build_graph_node_draft(resource_key="browser.screenshot")
    data_get_variable = service.build_graph_node_draft(resource_key="data.get_variable")
    data_get_text = service.build_graph_node_draft(resource_key="data.get_text")
    file_read_text = service.build_graph_node_draft(resource_key="file.read_text_file")
    excel_write_file = service.build_graph_node_draft(resource_key="excel.write_file")
    python_run = service.build_graph_node_draft(resource_key="python.run")
    graph_call_subgraph = service.build_graph_node_draft(resource_key="graph.call_subgraph")
    control_foreach = service.build_graph_node_draft(resource_key="control.foreach")

    assert flow_start["resource"]["resource_key"] == "flow.start"
    assert flow_start["node"]["lowered_kind"] == "control"
    assert flow_start["node"]["display_name"] == "开始"
    assert flow_start["node"]["expansion_role"] == "flow:start"
    assert flow_start["node"]["ports"] == [
        {
            "port_id": "out",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.control",
        },
        {
            "port_id": "out:variables",
            "direction": "output",
            "relation_layer": "data",
            "semantic_slot": "out.variables",
        }
    ]
    assert flow_start["node"]["node_config"] == {
        "initial_variables": {},
        "browser_config": {
            "headless": True,
            "slow_mo_ms": 0,
        },
    }

    assert control_if["node"]["lowered_kind"] == "control"
    assert control_if["node"]["display_name"] == "条件分支"
    assert control_if["node"]["expansion_role"] == "control:if"
    assert control_if["node"]["ports"] == [
        {
            "port_id": "in",
            "direction": "input",
            "relation_layer": "control",
            "semantic_slot": "in.control",
        },
        {
            "port_id": "condition",
            "direction": "input",
            "relation_layer": "data",
            "semantic_slot": "in.condition",
        },
        {
            "port_id": "true",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.true",
        },
        {
            "port_id": "false",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.false",
        },
    ]
    assert control_if["node"]["node_config"] == {"expression": ""}

    assert browser_fill["node"]["lowered_kind"] == "execution"
    assert browser_fill["node"]["display_name"] == "填入"
    assert browser_fill["node"]["ports"] == [
        {
            "port_id": "in",
            "direction": "input",
            "relation_layer": "control",
            "semantic_slot": "in.control",
        },
        {
            "port_id": "in:selector",
            "direction": "input",
            "relation_layer": "data",
            "semantic_slot": "in.selector",
        },
        {
            "port_id": "in:value",
            "direction": "input",
            "relation_layer": "data",
            "semantic_slot": "in.value",
        },
        {
            "port_id": "out",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.control",
        },
    ]
    assert browser_fill["node"]["node_config"] == {"selector": "", "value": ""}

    assert data_set_variable["node"]["lowered_kind"] == "execution"
    assert data_set_variable["node"]["display_name"] == "写入变量"
    assert data_set_variable["node"]["ports"] == [
        {
            "port_id": "in",
            "direction": "input",
            "relation_layer": "control",
            "semantic_slot": "in.control",
        },
        {
            "port_id": "in:value",
            "direction": "input",
            "relation_layer": "data",
            "semantic_slot": "in.value",
        },
        {
            "port_id": "out",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.control",
        },
        {
            "port_id": "out:value",
            "direction": "output",
            "relation_layer": "data",
            "semantic_slot": "out.value",
        },
    ]
    assert data_set_variable["node"]["node_config"] == {"name": "", "value": None}

    assert http_request["node"]["ports"] == [
        {
            "port_id": "in",
            "direction": "input",
            "relation_layer": "control",
            "semantic_slot": "in.control",
        },
        {
            "port_id": "in:url",
            "direction": "input",
            "relation_layer": "data",
            "semantic_slot": "in.url",
        },
        {
            "port_id": "out",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.control",
        },
        {
            "port_id": "out:body",
            "direction": "output",
            "relation_layer": "data",
            "semantic_slot": "out.body",
        },
    ]
    assert http_request["node"]["node_config"] == {
        "method": "GET",
        "url": "",
        "headers": {},
        "timeout": 30,
        "body": None,
    }

    assert browser_click["node"]["node_config"] == {"selector": ""}
    assert browser_wait_for_element["node"]["node_config"] == {
        "selector": "",
        "timeout": 10000,
    }
    assert browser_screenshot["node"]["node_config"] == {"path": ""}

    assert data_get_variable["node"]["node_config"] == {"name": ""}
    assert data_get_text["node"]["node_config"] == {
        "selector": "",
        "variable_name": "",
        "target_type": "string",
    }

    assert file_read_text["node"]["node_config"] == {"path": "", "encoding": "utf-8"}
    assert excel_write_file["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "rows": [],
    }
    assert python_run["node"]["node_config"] == {"code": ""}
    assert graph_call_subgraph["node"]["node_config"] == {
        "subgraph_id": "",
        "inputs": {},
        "outputs": {},
    }
    assert control_foreach["node"]["node_config"] == {
        "variable": "",
        "item_var": "item",
        "index_var": "index",
    }


def test_service_builds_graph_node_drafts_for_extended_builtin_resources() -> None:
    service = CompilationWorkbenchService()

    expected_resource_keys = [
        "browser.hover",
        "browser.check",
        "browser.uncheck",
        "browser.set_input_files",
        "browser.select_option",
        "browser.wait_for_navigation",
        "browser.wait_for_timeout",
        "browser.recognize_captcha",
        "browser.switch_to_frame",
        "browser.switch_to_parent_frame",
        "browser.switch_to_default_content",
        "browser.open_frame_page",
        "data.map",
        "data.get_attribute",
        "data.get_value",
        "data.get_element_count",
        "data.set_variables_batch",
        "data.increment_variable",
        "data.decrement_variable",
        "data.convert_value",
        "data.evaluate_expression",
        "data.regex_replace",
        "data.create_list",
        "data.list_append",
        "data.list_extend",
        "data.list_get",
        "data.list_set",
        "data.list_index",
        "data.list_length",
        "data.list_insert",
        "data.list_remove",
        "data.list_slice",
        "data.list_sort",
        "data.list_reverse",
        "file.write_text_file",
        "file.read_csv_cell",
        "file.read_csv_row",
        "file.read_csv_table",
        "excel.read_cell",
        "excel.read_row",
        "excel.read_table",
        "excel.write_cell",
        "excel.write_row",
        "excel.write_table",
        "excel.update_cells",
        "excel.update_batch",
        "browser.extract_web_table",
        "browser.extract_web_table_to_excel",
        "browser.inject_js",
        "browser.run_js",
        "browser.get_local_storage",
        "browser.go_back",
        "browser.go_forward",
        "browser.refresh",
        "browser.refresh_no_cache",
        "time.get_current_time",
        "control.jump_to_step",
        "control.end_foreach",
        "control.foreach_continue",
        "control.foreach_break",
        "call_blueprint",
        "session.apply_auth_session",
        "dialog.switch_dialog_mode",
        "dialog.watch_dialogs",
        "dialog.handle_dialogs",
        "dialog.set_agent_config",
    ]

    drafts = {
        resource_key: service.build_graph_node_draft(resource_key=resource_key)
        for resource_key in expected_resource_keys
    }

    for resource_key, draft in drafts.items():
        assert draft["resource"]["resource_key"] == resource_key
        assert draft["node"]["node_kind"] == resource_key
        assert isinstance(draft["node"]["display_name"], str) and draft["node"]["display_name"]
        assert isinstance(draft["node"]["ports"], list) and draft["node"]["ports"]
        assert isinstance(draft["node"]["node_config"], dict)

    assert drafts["browser.hover"]["node"]["node_config"] == {"selector": ""}
    assert drafts["browser.check"]["node"]["node_config"] == {"selector": ""}
    assert drafts["browser.uncheck"]["node"]["node_config"] == {"selector": ""}
    assert drafts["browser.set_input_files"]["node"]["node_config"] == {
        "selector": "",
        "path": "",
    }
    assert drafts["browser.select_option"]["node"]["node_config"] == {
        "selector": "",
        "value": "",
    }
    assert drafts["browser.wait_for_navigation"]["node"]["node_config"] == {
        "url_pattern": "",
        "timeout": 15000,
    }
    assert drafts["browser.wait_for_timeout"]["node"]["node_config"] == {"timeout": 0}
    assert drafts["browser.recognize_captcha"]["node"]["node_config"] == {
        "selector": "",
        "image_bytes_base64": "",
        "target_variable": "",
        "model_name": "",
        "runtime_root": "",
    }
    assert drafts["browser.switch_to_frame"]["node"]["node_config"] == {
        "selector": "",
        "name": "",
        "url_contains": "",
        "index": -1,
    }
    assert drafts["browser.switch_to_parent_frame"]["node"]["node_config"] == {}
    assert drafts["browser.switch_to_default_content"]["node"]["node_config"] == {}
    assert drafts["browser.open_frame_page"]["node"]["node_config"] == {
        "selector": "",
        "name": "",
        "url_contains": "",
        "index": -1,
    }

    assert drafts["data.map"]["node"]["node_config"] == {
        "source": None,
        "variable_name": "",
        "mode": "map",
    }
    assert drafts["data.get_attribute"]["node"]["node_config"] == {
        "selector": "",
        "attribute": "",
        "variable_name": "",
    }
    assert drafts["data.get_value"]["node"]["node_config"] == {
        "selector": "",
        "variable_name": "",
    }
    assert drafts["data.get_element_count"]["node"]["node_config"] == {
        "selector": "",
        "variable_name": "",
    }
    assert drafts["data.set_variables_batch"]["node"]["node_config"] == {"variables": {}}
    assert drafts["data.increment_variable"]["node"]["node_config"] == {
        "variable_name": "",
        "step": 1,
    }
    assert drafts["data.decrement_variable"]["node"]["node_config"] == {
        "variable_name": "",
        "step": 1,
    }
    assert drafts["data.convert_value"]["node"]["node_config"] == {
        "source_value": None,
        "target_type": "string",
        "variable_name": "",
        "in_place": False,
        "source_variable_name": "",
    }
    assert drafts["data.evaluate_expression"]["node"]["node_config"] == {
        "expression": "",
        "variable_name": "",
    }
    assert drafts["data.regex_replace"]["node"]["node_config"] == {
        "text": "",
        "pattern": "",
        "replacement": "",
        "variable_name": "",
    }
    assert drafts["data.create_list"]["node"]["node_config"] == {
        "variable_name": "",
        "items": [],
    }
    assert drafts["data.list_append"]["node"]["node_config"] == {
        "variable_name": "",
        "value": None,
    }
    assert drafts["data.list_extend"]["node"]["node_config"] == {
        "variable_name": "",
        "items": [],
    }
    assert drafts["data.list_get"]["node"]["node_config"] == {
        "variable_name": "",
        "index": 0,
        "output_variable_name": "",
    }
    assert drafts["data.list_set"]["node"]["node_config"] == {
        "variable_name": "",
        "index": 0,
        "value": None,
    }
    assert drafts["data.list_index"]["node"]["node_config"] == {
        "variable_name": "",
        "value": None,
        "output_variable_name": "",
    }
    assert drafts["data.list_length"]["node"]["node_config"] == {
        "variable_name": "",
        "output_variable_name": "",
    }
    assert drafts["data.list_insert"]["node"]["node_config"] == {
        "variable_name": "",
        "index": 0,
        "value": None,
    }
    assert drafts["data.list_remove"]["node"]["node_config"] == {
        "variable_name": "",
        "index": None,
        "value": None,
    }
    assert drafts["data.list_slice"]["node"]["node_config"] == {
        "variable_name": "",
        "start": 0,
        "end": None,
        "output_variable_name": "",
    }
    assert drafts["data.list_sort"]["node"]["node_config"] == {"variable_name": ""}
    assert drafts["data.list_reverse"]["node"]["node_config"] == {"variable_name": ""}

    assert drafts["file.write_text_file"]["node"]["node_config"] == {
        "path": "",
        "encoding": "utf-8",
        "content": "",
    }
    assert drafts["file.read_csv_cell"]["node"]["node_config"] == {
        "path": "",
        "encoding": "utf-8",
        "has_header": True,
        "row_index": 0,
        "column": None,
        "variable_name": "",
    }
    assert drafts["file.read_csv_row"]["node"]["node_config"] == {
        "path": "",
        "encoding": "utf-8",
        "has_header": True,
        "row_index": 0,
        "variable_name": "",
    }
    assert drafts["file.read_csv_table"]["node"]["node_config"] == {
        "path": "",
        "encoding": "utf-8",
        "has_header": True,
        "variable_name": "",
    }

    assert drafts["excel.read_cell"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "cell": "",
        "variable_name": "",
    }
    assert drafts["excel.read_row"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "row_index": 1,
        "variable_name": "",
    }
    assert drafts["excel.read_table"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "has_header": True,
        "variable_name": "",
    }
    assert drafts["excel.write_cell"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "cell": "",
        "value": None,
    }
    assert drafts["excel.write_row"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "row_index": 1,
        "data": [],
    }
    assert drafts["excel.write_table"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "data": [],
        "has_header": True,
    }
    assert drafts["excel.update_cells"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "updates": [],
    }
    assert drafts["excel.update_batch"]["node"]["node_config"] == {
        "path": "",
        "sheet_name": "Sheet1",
        "condition": "",
        "updates": {},
    }

    assert drafts["browser.extract_web_table"]["node"]["node_config"] == {
        "selector": "",
        "variable_name": "",
    }
    assert drafts["browser.extract_web_table_to_excel"]["node"]["node_config"] == {
        "selector": "",
        "path": "",
        "sheet_name": "Sheet1",
    }
    assert drafts["browser.inject_js"]["node"]["node_config"] == {"script": ""}
    assert drafts["browser.run_js"]["node"]["node_config"] == {
        "script": "",
        "variable_name": "",
    }
    assert drafts["browser.get_local_storage"]["node"]["node_config"] == {
        "key": "",
        "variable_name": "",
        "default_value": None,
    }
    assert drafts["browser.go_back"]["node"]["node_config"] == {}
    assert drafts["browser.go_forward"]["node"]["node_config"] == {}
    assert drafts["browser.refresh"]["node"]["node_config"] == {}
    assert drafts["browser.refresh_no_cache"]["node"]["node_config"] == {}
    assert drafts["time.get_current_time"]["node"]["node_config"] == {
        "variable_name": "",
        "format": "iso",
        "timezone": "utc",
    }

    assert drafts["control.jump_to_step"]["node"]["node_config"] == {
        "target_node_id": "",
        "target_step": None,
        "condition": "true",
        "max_jumps": -1,
    }
    assert drafts["control.end_foreach"]["node"]["node_config"] == {}
    assert drafts["control.foreach_continue"]["node"]["node_config"] == {
        "condition": "true",
        "level": 1,
    }
    assert drafts["control.foreach_break"]["node"]["node_config"] == {
        "condition": "true",
        "level": 1,
    }

    assert drafts["call_blueprint"]["node"]["node_config"] == {
        "blueprint_id": "",
        "inputs": {},
        "outputs": {},
    }
    assert drafts["session.apply_auth_session"]["node"]["node_config"] == {
        "cookies": [],
        "local_storage": {},
    }
    assert drafts["dialog.switch_dialog_mode"]["node"]["node_config"] == {"mode": "auto"}
    assert drafts["dialog.watch_dialogs"]["node"]["node_config"] == {
        "timeout": 0,
        "variable_name": "",
    }
    assert drafts["dialog.handle_dialogs"]["node"]["node_config"] == {"clear_after": False}
    assert drafts["dialog.set_agent_config"]["node"]["node_config"] == {
        "default_action": "accept",
        "prompt_text": "",
    }


def test_service_builds_graph_node_drafts_for_phase14_browser_resources() -> None:
    service = CompilationWorkbenchService()

    expected_resource_keys = [
        "browser.wait_for_text",
        "browser.wait_for_attribute",
        "browser.wait_for_value",
        "browser.wait_for_request",
        "browser.wait_for_response",
        "browser.wait_for_popup",
        "browser.get_cookie",
        "browser.set_cookie",
        "browser.delete_cookie",
        "browser.list_cookies",
        "browser.set_local_storage",
        "browser.remove_local_storage",
        "browser.clear_local_storage",
        "browser.get_session_storage",
        "browser.set_session_storage",
        "browser.press_key",
        "browser.keyboard_type",
        "browser.hotkey",
        "browser.scroll_to_element",
        "browser.scroll_page",
        "browser.drag_and_drop",
        "browser.element_screenshot",
        "browser.open_tab",
        "browser.switch_tab",
        "browser.close_tab",
        "browser.exists",
        "browser.is_visible",
        "browser.is_enabled",
        "browser.is_checked",
        "browser.get_html",
        "browser.get_inner_html",
        "browser.download_file",
        "browser.wait_for_download",
        "browser.set_user_agent",
        "browser.set_extra_headers",
        "browser.wait_for_url_change",
    ]

    drafts = {
        resource_key: service.build_graph_node_draft(resource_key=resource_key)
        for resource_key in expected_resource_keys
    }

    for resource_key, draft in drafts.items():
        assert draft["resource"]["resource_key"] == resource_key
        assert draft["node"]["node_kind"] == resource_key
        assert isinstance(draft["node"]["ports"], list) and draft["node"]["ports"]
        assert isinstance(draft["node"]["node_config"], dict)

    assert drafts["browser.wait_for_text"]["node"]["node_config"] == {
        "selector": "",
        "text": "",
        "match_mode": "contains",
        "timeout": 10000,
    }
    assert drafts["browser.set_cookie"]["node"]["node_config"] == {
        "name": "",
        "value": "",
        "url": "",
        "domain": "",
        "path": "/",
        "http_only": False,
        "secure": False,
        "same_site": "Lax",
        "expires": None,
    }
    assert drafts["browser.element_screenshot"]["node"]["node_config"] == {
        "selector": "",
        "path": "",
    }
    assert drafts["browser.download_file"]["node"]["node_config"] == {
        "url": "",
        "path": "",
    }
    assert drafts["browser.set_extra_headers"]["node"]["node_config"] == {
        "headers": {},
    }


def test_service_graph_node_draft_exposes_parameter_metadata_and_enhanced_flow_control_drafts() -> None:
    service = CompilationWorkbenchService()

    flow_start = service.build_graph_node_draft(resource_key="flow.start")
    browser_screenshot = service.build_graph_node_draft(resource_key="browser.screenshot")
    parallel_fork = service.build_graph_node_draft(resource_key="control.parallel_fork")
    control_join = service.build_graph_node_draft(resource_key="control.join")
    control_while = service.build_graph_node_draft(resource_key="control.while")

    assert flow_start["node"]["ports"] == [
        {
            "port_id": "out",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.control",
        },
        {
            "port_id": "out:variables",
            "direction": "output",
            "relation_layer": "data",
            "semantic_slot": "out.variables",
        },
    ]
    assert flow_start["parameter_schema"]["initial_variables"] == {
        "type": "object",
        "required": False,
        "editor_kind": "key_value_map",
        "path_kind": None,
    }
    assert flow_start["parameter_schema"]["browser_config"] == {
        "type": "object",
        "required": False,
        "editor_kind": "object",
        "path_kind": None,
    }

    assert browser_screenshot["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "save_file",
    }

    assert parallel_fork["node"]["node_config"] == {
        "branches": [
            {"key": "left", "label": "Left"},
            {"key": "right", "label": "Right"},
        ]
    }
    assert parallel_fork["parameter_schema"]["branches"] == {
        "type": "array",
        "required": True,
        "editor_kind": "branch_list",
        "path_kind": None,
    }

    assert control_join["node"]["node_config"] == {
        "branches": [
            {"key": "left", "label": "Left"},
            {"key": "right", "label": "Right"},
        ],
        "mode": "all",
        "quorum": None,
    }
    assert control_join["parameter_schema"]["branches"] == {
        "type": "array",
        "required": True,
        "editor_kind": "branch_list",
        "path_kind": None,
    }
    assert control_join["parameter_schema"]["quorum"] == {
        "type": "integer",
        "required": False,
        "editor_kind": "number",
        "path_kind": None,
    }
    assert control_while["node"]["ports"] == [
        {
            "port_id": "in",
            "direction": "input",
            "relation_layer": "control",
            "semantic_slot": "in.control",
        },
        {
            "port_id": "repeat",
            "direction": "input",
            "relation_layer": "control",
            "semantic_slot": "in.repeat",
        },
        {
            "port_id": "condition",
            "direction": "input",
            "relation_layer": "data",
            "semantic_slot": "in.condition",
        },
        {
            "port_id": "loop",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.loop",
        },
        {
            "port_id": "done",
            "direction": "output",
            "relation_layer": "control",
            "semantic_slot": "out.done",
        },
    ]


def test_service_graph_node_draft_exposes_path_parameter_metadata_for_file_and_excel_nodes() -> None:
    service = CompilationWorkbenchService()

    file_read_text = service.build_graph_node_draft(resource_key="file.read_text_file")
    file_write_text = service.build_graph_node_draft(resource_key="file.write_text_file")
    file_read_csv_table = service.build_graph_node_draft(resource_key="file.read_csv_table")
    excel_read_cell = service.build_graph_node_draft(resource_key="excel.read_cell")
    excel_write_cell = service.build_graph_node_draft(resource_key="excel.write_cell")
    excel_write_file = service.build_graph_node_draft(resource_key="excel.write_file")
    excel_update_cells = service.build_graph_node_draft(resource_key="excel.update_cells")

    assert file_read_text["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "open_file",
    }
    assert file_write_text["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "save_file",
    }
    assert file_read_csv_table["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "open_file",
    }
    assert excel_read_cell["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "open_file",
    }
    assert excel_write_cell["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "open_file",
    }
    assert excel_write_file["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "save_file",
    }
    assert excel_update_cells["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "open_file",
    }


def test_service_graph_node_draft_exposes_phase14_browser_parameter_metadata() -> None:
    service = CompilationWorkbenchService()

    element_screenshot = service.build_graph_node_draft(resource_key="browser.element_screenshot")
    download_file = service.build_graph_node_draft(resource_key="browser.download_file")
    wait_for_download = service.build_graph_node_draft(resource_key="browser.wait_for_download")
    set_cookie = service.build_graph_node_draft(resource_key="browser.set_cookie")
    set_extra_headers = service.build_graph_node_draft(resource_key="browser.set_extra_headers")

    assert element_screenshot["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "save_file",
    }
    assert download_file["parameter_schema"]["path"] == {
        "type": "string",
        "required": True,
        "editor_kind": "path",
        "path_kind": "save_file",
    }
    assert wait_for_download["parameter_schema"]["path"] == {
        "type": "string",
        "required": False,
        "editor_kind": "path",
        "path_kind": "save_file",
    }
    assert set_cookie["parameter_schema"]["same_site"] == {
        "type": "string",
        "required": False,
        "editor_kind": "select",
        "path_kind": None,
    }
    assert set_extra_headers["parameter_schema"]["headers"] == {
        "type": "object",
        "required": True,
        "editor_kind": "object",
        "path_kind": None,
    }


def test_service_exposes_category_paths_and_facets_for_resources_and_component_library() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "resource_tags": ["team:qa", "folder:imported"],
            },
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_custom_node_graph_resource(resource_name="Imported Graph")

    registry_document = service.get_resource_registry_document()
    library_document = service.get_component_library_document()

    custom_resource = next(
        item
        for item in registry_document["resources"]
        if item["resource_id"] == saved["resource"]["resource_id"]
    )
    browser_click = next(
        item for item in registry_document["resources"] if item["resource_key"] == "browser.click"
    )
    control_if = next(
        item for item in library_document["items"] if item["resource_key"] == "control.if"
    )
    browser_click_library = next(
        item for item in library_document["items"] if item["resource_key"] == "browser.click"
    )

    assert browser_click["category_path"] == ["builtin", "browser", "action"]
    assert control_if["category_path"] == ["builtin", "control_structure", "if"]
    assert custom_resource["category_path"] == ["project", "custom_node_graph", "imported"]
    assert browser_click["category_group_path"] == ["builtin", "browser"]
    assert browser_click["category_group_label"] == "浏览器"
    assert browser_click_library["category_group_path"] == ["builtin", "browser"]
    assert browser_click_library["category_group_label"] == "浏览器"
    assert control_if["category_group_path"] == ["builtin", "control_structure"]
    assert control_if["category_group_label"] == "流程控制"
    assert custom_resource["category_group_path"] == ["project", "custom_node_graph"]
    assert custom_resource["category_group_label"] == "用户组件"
    assert "display:imported graph" in custom_resource["search_tokens"]
    assert "tag:team:qa" in custom_resource["search_tokens"]

    resource_paths = {
        tuple(item["path"])
        for item in registry_document["facets"]["category_paths"]
    }
    library_paths = {
        tuple(item["path"])
        for item in library_document["facets"]["category_paths"]
    }

    assert ("builtin", "browser", "action") in resource_paths
    assert ("project", "custom_node_graph", "imported") in resource_paths
    assert ("builtin", "control_structure", "if") in library_paths
    assert {
        tuple(item["path"])
        for item in registry_document["facets"]["category_groups"]
    } >= {
        ("builtin", "browser"),
        ("project", "custom_node_graph"),
    }
    assert {
        tuple(item["path"])
        for item in library_document["facets"]["category_groups"]
    } >= {
        ("builtin", "browser"),
        ("builtin", "control_structure"),
    }
    assert "team:qa" in registry_document["facets"]["user_tags"]


def test_component_library_and_registry_expose_category_group_protocol() -> None:
    service = CompilationWorkbenchService()

    registry_document = service.get_resource_registry_document()
    library_document = service.get_component_library_document()

    browser_resource = next(
        item for item in registry_document["resources"] if item["resource_key"] == "browser.click"
    )
    control_resource = next(
        item for item in library_document["items"] if item["resource_key"] == "control.retry"
    )

    assert browser_resource["category_group_path"] == ["builtin", "browser"]
    assert browser_resource["category_group_label"] == "浏览器"
    assert control_resource["category_group_path"] == ["builtin", "control_structure"]
    assert control_resource["category_group_label"] == "流程控制"
    assert any(
        item["path"] == ["builtin", "browser"] and item["label"] == "浏览器"
        for item in registry_document["facets"]["category_groups"]
    )
    assert any(
        item["path"] == ["builtin", "control_structure"] and item["label"] == "流程控制"
        for item in library_document["facets"]["category_groups"]
    )


def test_service_can_update_project_resource_tags_and_search_by_category_path_token() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "hello"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_user_component_resource(resource_name="Team Shared Component")

    updated = service.update_resource_tags(
        resource_id=saved["resource"]["resource_id"],
        tags=["team:alpha", "folder:shared", " team:alpha ", ""],
    )

    assert updated["resource"]["tags"].count("team:alpha") == 1
    assert "folder:shared" in updated["resource"]["tags"]
    assert updated["resource"]["category_path"] == ["project", "user_component", "shared"]

    filtered = service.get_resource_registry_document(query="shared")
    assert any(
        item["resource_id"] == saved["resource"]["resource_id"]
        for item in filtered["resources"]
    )

def test_builtin_registry_covers_p9_webcontrol_gap_components() -> None:
    component_keys = {
        item["resource_key"]
        for item in BUILTIN_COMPONENT_DEFINITIONS
    }

    assert {
        "data.get_text",
        "data.get_attribute",
        "data.get_value",
        "data.get_element_count",
        "data.set_variables_batch",
        "data.increment_variable",
        "data.decrement_variable",
        "data.list_index",
        "browser.extract_web_table",
        "browser.extract_web_table_to_excel",
        "browser.inject_js",
        "browser.run_js",
        "session.apply_auth_session",
        "dialog.switch_dialog_mode",
        "dialog.watch_dialogs",
        "dialog.handle_dialogs",
        "dialog.set_agent_config",
    }.issubset(component_keys)


def test_builtin_registry_covers_phase14_browser_components() -> None:
    component_keys = {
        item["resource_key"]
        for item in BUILTIN_COMPONENT_DEFINITIONS
    }

    assert {
        "browser.wait_for_text",
        "browser.wait_for_attribute",
        "browser.wait_for_value",
        "browser.wait_for_request",
        "browser.wait_for_response",
        "browser.wait_for_popup",
        "browser.get_cookie",
        "browser.set_cookie",
        "browser.delete_cookie",
        "browser.list_cookies",
        "browser.set_local_storage",
        "browser.remove_local_storage",
        "browser.clear_local_storage",
        "browser.get_session_storage",
        "browser.set_session_storage",
        "browser.press_key",
        "browser.keyboard_type",
        "browser.hotkey",
        "browser.scroll_to_element",
        "browser.scroll_page",
        "browser.drag_and_drop",
        "browser.element_screenshot",
        "browser.open_tab",
        "browser.switch_tab",
        "browser.close_tab",
        "browser.exists",
        "browser.is_visible",
        "browser.is_enabled",
        "browser.is_checked",
        "browser.get_html",
        "browser.get_inner_html",
        "browser.download_file",
        "browser.wait_for_download",
        "browser.set_user_agent",
        "browser.set_extra_headers",
        "browser.wait_for_url_change",
    }.issubset(component_keys)


def test_builtin_registry_covers_all_legacy_webcontrol_action_names() -> None:
    legacy_actions_path = (
        Path(__file__).resolve().parents[3]
        / "I:/WebControl/docs/api-reference/actions.md"
    )
    legacy_actions = []
    for raw_line in legacy_actions_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if line.startswith("- [`"):
            legacy_actions.append(line.split("`", 2)[1])

    resolved_names = set()
    for item in BUILTIN_COMPONENT_DEFINITIONS:
        resolved_names.add(item["resource_key"])
        for alias in item.get("compatibility_aliases", []):
            resolved_names.add(alias)

    missing_actions = sorted(action for action in legacy_actions if action not in resolved_names)

    assert missing_actions == []


def test_service_exposes_graph_source_projection_for_source_input_subview() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "display_name": "HTTP Request",
                    "node_kind": "http.request",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    projection = service.get_graph_source_projection_document(
        target_source_kind="native_flow",
    )

    assert projection["status"] == "ready"
    assert projection["source_kind"] == "native_flow"
    assert projection["request_origin"] == "saved_graph_document"
    assert projection["graph_model_id"] == "graph:workspace"
    assert projection["graph_document_save_revision"] == 1
    assert projection["entry_document"] == "graph:workspace.native-flow.json"
    assert projection["diagnostics"] == []
    assert projection["source_text"] == (
        '{"nodes":[{"id":"n1","role":"action","capability_domain":"http",'
        '"action_kind":"request"}],"edges":[]}'
    )


def test_service_exposes_builtin_registry_baseline_for_phase8() -> None:
    service = CompilationWorkbenchService()

    registry = service.get_resource_registry_document()
    resource_ids = {item["resource_id"] for item in registry["resources"]}

    assert "builtin:browser.navigate" in resource_ids
    assert "builtin:excel.read_cell" in resource_ids
    assert "builtin:python.run" in resource_ids
    assert registry["summary"]["builtin_resource_count"] >= 20


def test_service_can_project_unsaved_graph_document_payload_for_source_input_subview() -> None:
    service = CompilationWorkbenchService()

    payload = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-1",
                "lowered_kind": "execution",
                "source_anchor_ref": "n1",
                "expansion_role": "action:request",
                "display_name": "HTTP Request",
                "node_kind": "http.request",
                "ports": [],
            },
            {
                "node_id": "node-2",
                "lowered_kind": "execution",
                "source_anchor_ref": "n2",
                "expansion_role": "transform:map",
                "display_name": "Map Result",
                "node_kind": "data.map",
                "ports": [],
            },
        ],
        "edges": [
            {
                "edge_id": "edge-1",
                "relation_layer": "data",
                "from_node_id": "node-1",
                "to_node_id": "node-2",
            }
        ],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    projection = service.get_graph_source_projection_document(
        target_source_kind="native_flow",
        graph_document_payload=payload,
    )

    assert projection["status"] == "ready"
    assert projection["request_origin"] == "graph_document_payload"
    assert projection["graph_model_id"] == "graph:workspace"
    assert projection["graph_document_save_revision"] is None
    assert projection["diagnostics"] == []
    assert projection["source_text"] == (
        '{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"},'
        '{"id":"n2","role":"transform","capability_domain":"data","action_kind":"map"}],'
        '"edges":[{"id":"edge-1","from":"n1","to":"n2","relation_layer":"data"}]}'
    )


def test_service_projects_browser_excel_python_node_kinds_to_native_flow() -> None:
    service = CompilationWorkbenchService()

    payload = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-1",
                "lowered_kind": "execution",
                "source_anchor_ref": "n1",
                "expansion_role": "action:navigate",
                "display_name": "Navigate",
                "node_kind": "browser.navigate",
                "ports": [],
            },
            {
                "node_id": "node-2",
                "lowered_kind": "execution",
                "source_anchor_ref": "n2",
                "expansion_role": "action:read_cell",
                "display_name": "Read Cell",
                "node_kind": "excel.read_cell",
                "ports": [],
            },
            {
                "node_id": "node-3",
                "lowered_kind": "execution",
                "source_anchor_ref": "n3",
                "expansion_role": "action:run",
                "display_name": "Run Python",
                "node_kind": "python.run",
                "ports": [],
            },
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    projection = service.get_graph_source_projection_document(
        target_source_kind="native_flow",
        graph_document_payload=payload,
    )

    assert projection["status"] == "ready"
    assert projection["diagnostics"] == []
    assert projection["source_text"] == (
        '{"nodes":[{"id":"n1","role":"action","capability_domain":"browser","action_kind":"navigate"},'
        '{"id":"n2","role":"action","capability_domain":"excel","action_kind":"read_cell"},'
        '{"id":"n3","role":"action","capability_domain":"python","action_kind":"run"}],"edges":[]}'
    )


def test_service_falls_back_to_graph_workspace_source_projection_when_native_flow_mapping_is_unsupported() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "action:start",
                    "display_name": "Flow Start",
                    "node_kind": "flow.start",
                    "node_config": {
                        "initial_variables": {"status": "ready"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    projection = service.get_graph_source_projection_document(
        target_source_kind="native_flow",
    )

    assert projection["status"] == "ready"
    assert projection["source_kind"] == "graph_workspace"
    assert projection["request_origin"] == "saved_graph_document"
    assert projection["graph_model_id"] == "graph:workspace"
    assert projection["graph_document_save_revision"] == 1
    assert len(projection["diagnostics"]) == 1
    assert projection["diagnostics"][0]["category"] == "graph.compile.mapping_error"
    projected_graph = json.loads(projection["source_text"])
    assert projected_graph["nodes"][0]["node_kind"] == "flow.start"
    assert projected_graph["nodes"][0]["node_config"]["initial_variables"] == {"status": "ready"}


def test_service_can_start_and_run_runtime_session_with_execution_log_and_result() -> None:
    service = CompilationWorkbenchService()
    server, thread = _start_runtime_http_server()

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-1",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n1",
                        "expansion_role": "action:request",
                        "display_name": "HTTP Request",
                        "node_kind": "http.request",
                        "node_config": {
                            "method": "POST",
                            "url": f"http://127.0.0.1:{server.server_address[1]}/api/echo",
                            "body": {"ok": True},
                        },
                        "ports": [
                            {
                                "port_id": "out",
                                "direction": "output",
                                "relation_layer": "data",
                                "semantic_slot": "out.default",
                            }
                        ],
                    },
                    {
                        "node_id": "node-2",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n2",
                        "expansion_role": "transform:map",
                        "display_name": "Map Result",
                        "node_kind": "data.map",
                        "ports": [
                            {
                                "port_id": "in",
                                "direction": "input",
                                "relation_layer": "data",
                                "semantic_slot": "in.default",
                            }
                        ],
                    },
                ],
                "edges": [
                    {
                        "edge_id": "edge-1",
                        "relation_layer": "data",
                        "from_node_id": "node-1",
                        "to_node_id": "node-2",
                        "from_port_id": "out",
                        "to_port_id": "in",
                    }
                ],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])
        runtime_session = service.get_runtime_session(session_id=started["runtime_session"]["session_id"])
        runtime_sessions = service.list_runtime_sessions()
    finally:
        server.shutdown()
        server.server_close()

    assert started["status"] == "started"
    assert started["runtime_session"]["status"] == "running"
    assert started["runtime_session"]["execution_supported"] is True
    assert started["runtime_plan"]["start_node_ids"] == ["node-1"]
    assert session["status"] == "completed"
    assert session["runtime_session"]["status"] == "completed"
    assert session["runtime_session"]["completed_node_count"] == 2
    assert session["result"]["status"] == "succeeded"
    assert session["result"]["completed_node_ids"] == ["node-1", "node-2"]
    assert session["result"]["outputs"]["node-1"]["status_code"] == 200
    assert session["result"]["outputs"]["node-1"]["body"]["ok"] is True
    assert session["result"]["outputs"]["node-2"]["mapped_from_node_id"] == "node-1"
    assert session["result"]["outputs"]["node-2"]["response_status_code"] == 200
    assert session["event_log"][0]["event_kind"] == "session.started"
    assert session["event_log"][-1]["event_kind"] == "session.completed"
    assert runtime_session["runtime_session"]["status"] == "completed"
    assert runtime_session["node_states"][0]["node_status"] == "completed"
    assert runtime_session["node_states"][1]["node_status"] == "completed"
    assert runtime_session["node_states"][0]["output"]["status_code"] == 200
    assert runtime_session["node_states"][1]["output"]["response_status_code"] == 200
    assert runtime_session["debug_snapshot"]["scheduler_mode"] == "legacy_sequence"
    assert runtime_session["debug_snapshot"]["token_queue"] == []
    assert runtime_session["debug_snapshot"]["queued_node_ids"] == []
    assert runtime_session["debug_snapshot"]["executed_node_ids"] == ["node-1", "node-2"]
    assert runtime_session["debug_snapshot"]["join_buffers"] == {}
    assert runtime_session["debug_snapshot"]["retry_states"] == {}
    assert runtime_session["diagnostic_events"] == []
    assert runtime_sessions["sessions"][0]["session_id"] == started["runtime_session"]["session_id"]


def test_service_runtime_session_exposes_execution_summary_and_diagnostic_events() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "phase11_summary", "value": "ok"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "control",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out-control",
                    "to_port_id": None,
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])
    listed = service.list_runtime_sessions()

    assert session["runtime_session"]["status"] == "completed"
    assert session["runtime_session"]["completed_node_count"] == 2
    assert session["runtime_session"]["completed_at"] is not None
    assert session["runtime_session"]["execution_supported"] is True
    assert isinstance(session["diagnostic_events"], list)
    assert session["diagnostic_events"] == []
    assert session["execution_summary"]["status"] == "succeeded"
    assert session["execution_summary"]["completed_node_count"] == 2
    assert session["execution_summary"]["failed_node_count"] == 0
    assert session["execution_summary"]["event_count"] == len(session["event_log"])
    assert session["execution_summary"]["diagnostic_event_count"] == 0
    assert listed["sessions"][0]["completed_node_count"] == 2
    assert listed["sessions"][0]["event_count"] == len(session["event_log"])
    assert session["execution_summary"]["node_status_counts"]["completed"] == 2
    assert session["execution_summary"]["latest_event_kind"] == "session.completed"


def test_service_runtime_data_edge_outputs_do_not_create_persisted_state_cycle(tmp_path: Path) -> None:
    service = CompilationWorkbenchService(
        state_store=FileWorkspaceStateStore(tmp_path / "workspace-state.json")
    )
    server, thread = _start_runtime_http_server()

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-1",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n1",
                        "expansion_role": "action:request",
                        "display_name": "HTTP Request",
                        "node_kind": "http.request",
                        "node_config": {
                            "method": "POST",
                            "url": f"http://127.0.0.1:{server.server_address[1]}/api/echo",
                            "body": {"ok": True},
                        },
                        "ports": [
                            {
                                "port_id": "out",
                                "direction": "output",
                                "relation_layer": "data",
                                "semantic_slot": "out.default",
                            }
                        ],
                    },
                    {
                        "node_id": "node-2",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n2",
                        "expansion_role": "transform:map",
                        "display_name": "Map Result",
                        "node_kind": "data.map",
                        "ports": [
                            {
                                "port_id": "in",
                                "direction": "input",
                                "relation_layer": "data",
                                "semantic_slot": "in.default",
                            }
                        ],
                    },
                ],
                "edges": [
                    {
                        "edge_id": "edge-1",
                        "relation_layer": "data",
                        "from_node_id": "node-1",
                        "to_node_id": "node-2",
                        "from_port_id": "out",
                        "to_port_id": "in",
                    }
                ],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])
    finally:
        server.shutdown()
        server.server_close()

    assert session["status"] == "completed"
    assert session["runtime_session"]["status"] == "completed"
    assert session["node_states"][0]["output"]["status_code"] == 200
    assert session["node_states"][1]["output"]["mapped_from_node_id"] == "node-1"


def test_service_runtime_marks_executor_failure_as_failed_session() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-http",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-http",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "node_config": {"method": "GET"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["runtime_session"]["status"] == "failed"
    assert session["runtime_session"]["failed_node_count"] == 1
    assert session["node_states"][0]["node_status"] == "failed"
    assert session["node_states"][0]["error"]["error_code"] == "http.url_required"
    assert session["result"]["failure_reason"] == "http.url_required"


def test_service_runtime_converts_unexpected_executor_exception_to_failed_session(
    monkeypatch,
) -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-browser",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-browser",
                    "expansion_role": "browser.navigate",
                    "display_name": "Navigate",
                    "node_kind": "browser.navigate",
                    "node_config": {"url": "http://127.0.0.1:1"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    def fail_executor(**_kwargs):
        raise RuntimeError("browser launch failed")

    monkeypatch.setattr(service, "_execute_runtime_plan_node", fail_executor)

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["runtime_session"]["status"] == "failed"
    assert session["result"]["status"] == "failed"
    assert session["result"]["failure_reason"] == "runtime.executor_exception"
    assert session["result"]["failed_node_ids"] == ["node-browser"]
    assert session["node_states"][0]["node_status"] == "failed"
    assert session["node_states"][0]["error"] == {
        "error_code": "runtime.executor_exception",
        "message": "browser launch failed",
        "exception_type": "RuntimeError",
    }
    assert session["node_states"][0]["output"] == {
        "status": "failed",
        "node_id": "node-browser",
        "error_code": "runtime.executor_exception",
        "message": "browser launch failed",
        "exception_type": "RuntimeError",
    }


def test_service_runtime_marks_invalid_http_url_as_failed_session() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-http",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-http",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "node_config": {"method": "GET", "url": "__P8_1_HTTP_URL__"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["runtime_session"]["status"] == "failed"
    assert session["runtime_session"]["failed_node_count"] == 1
    assert session["node_states"][0]["node_status"] == "failed"
    assert session["node_states"][0]["error"]["error_code"] == "http.request_failed"
    assert "unknown url type" in session["node_states"][0]["error"]["message"]
    assert session["result"]["failure_reason"] == "http.request_failed"


def test_service_runtime_executes_http_data_and_text_file_actions(tmp_path) -> None:
    service = CompilationWorkbenchService()
    output_path = tmp_path / "runtime-output.txt"
    server, thread = _start_runtime_http_server()

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-var",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-var",
                        "expansion_role": "transform:set_variable",
                        "display_name": "Set Token",
                        "node_kind": "data.set_variable",
                        "node_config": {
                            "name": "token",
                            "value": "runtime-token",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-http",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-http",
                        "expansion_role": "action:request",
                        "display_name": "POST Request",
                        "node_kind": "http.request",
                        "node_config": {
                            "method": "POST",
                            "url": f"http://127.0.0.1:{server.server_address[1]}/api/echo",
                            "headers": {
                                "X-WeConduct-Test": "${token}",
                            },
                            "body": {
                                "token": "${token}",
                                "source": "runtime-test",
                            },
                            "timeout": 5,
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-map",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-map",
                        "expansion_role": "transform:map",
                        "display_name": "Map Response",
                        "node_kind": "data.map",
                        "node_config": {
                            "variable_name": "mapped_status",
                            "source": "${node.node-http.body.body.token}",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-write",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-write",
                        "expansion_role": "action:write_text_file",
                        "display_name": "Write Text",
                        "node_kind": "file.write_text_file",
                        "node_config": {
                            "path": str(output_path),
                            "content": "status=${node.node-http.status_code}; token=${mapped_status}",
                            "encoding": "utf-8",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-read",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-read",
                        "expansion_role": "action:read_text_file",
                        "display_name": "Read Text",
                        "node_kind": "file.read_text_file",
                        "node_config": {
                            "path": str(output_path),
                            "encoding": "utf-8",
                        },
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(
            session_id=started["runtime_session"]["session_id"],
        )

        assert session["status"] == "completed"
        assert session["result"]["status"] == "succeeded"
        assert session["result"]["outputs"]["node-http"]["status_code"] == 200
        assert session["result"]["outputs"]["node-http"]["body"]["body"]["token"] == "runtime-token"
        assert session["result"]["outputs"]["node-http"]["headers"]["x-weconduct-reply"] == "runtime"
        assert session["result"]["outputs"]["node-map"]["value"] == "runtime-token"
        assert session["result"]["variables"]["mapped_status"] == "runtime-token"
        assert session["result"]["outputs"]["node-write"]["bytes_written"] == len(
            "status=200; token=runtime-token".encode("utf-8")
        )
        assert session["result"]["outputs"]["node-read"]["content"] == "status=200; token=runtime-token"
        assert output_path.read_text(encoding="utf-8") == "status=200; token=runtime-token"
    finally:
        server.shutdown()
        server.server_close()


def test_service_runtime_reads_csv_cell_row_and_table(tmp_path) -> None:
    service = CompilationWorkbenchService()
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "name,score\nalice,10\nbob,20\n",
        encoding="utf-8",
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-cell",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-cell",
                    "expansion_role": "action:read_csv_cell",
                    "node_kind": "file.read_csv_cell",
                        "node_config": {
                            "path": str(csv_path),
                            "row_index": 0,
                            "column": "score",
                            "has_header": True,
                            "variable_name": "alice_score",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-row",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-row",
                    "expansion_role": "action:read_csv_row",
                    "node_kind": "file.read_csv_row",
                        "node_config": {
                            "path": str(csv_path),
                            "row_index": 1,
                            "has_header": True,
                            "variable_name": "bob_row",
                        },
                    "ports": [],
                },
                {
                    "node_id": "node-table",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-table",
                    "expansion_role": "action:read_csv_table",
                    "node_kind": "file.read_csv_table",
                    "node_config": {
                        "path": str(csv_path),
                        "has_header": True,
                        "variable_name": "score_table",
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-cell"]["value"] == "10"
    assert session["result"]["outputs"]["node-row"]["row"] == {"name": "bob", "score": "20"}
    assert session["result"]["outputs"]["node-table"]["rows"] == [
        {"name": "alice", "score": "10"},
        {"name": "bob", "score": "20"},
    ]
    assert session["result"]["variables"]["alice_score"] == "10"
    assert session["result"]["variables"]["bob_row"] == {"name": "bob", "score": "20"}
    assert session["result"]["variables"]["score_table"][1]["score"] == "20"


def test_service_runtime_reads_and_writes_excel_cells_rows_and_tables(tmp_path) -> None:
    service = CompilationWorkbenchService()
    workbook_path = tmp_path / "input.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SheetA"
    worksheet["A1"] = "name"
    worksheet["B1"] = "score"
    worksheet["A2"] = "alice"
    worksheet["B2"] = 10
    worksheet["A3"] = "bob"
    worksheet["B3"] = 20
    workbook.save(workbook_path)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-write",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write",
                    "expansion_role": "action:write_excel_cell",
                    "node_kind": "excel.write_cell",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "SheetA",
                        "cell": "B3",
                        "value": 25,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-cell",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-cell",
                    "expansion_role": "action:read_excel_cell",
                    "node_kind": "excel.read_cell",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "SheetA",
                        "cell": "B3",
                        "variable_name": "bob_score",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-row",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-row",
                    "expansion_role": "action:read_excel_row",
                    "node_kind": "excel.read_row",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "SheetA",
                        "row_index": 2,
                        "variable_name": "alice_row",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-table",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-table",
                    "expansion_role": "action:read_excel_table",
                    "node_kind": "excel.read_table",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "SheetA",
                        "has_header": True,
                        "variable_name": "score_table",
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-write"]["cell"] == "B3"
    assert session["result"]["outputs"]["node-write"]["value"] == 25
    assert session["result"]["outputs"]["node-cell"]["value"] == 25
    assert session["result"]["outputs"]["node-row"]["row"] == ["alice", 10]
    assert session["result"]["outputs"]["node-table"]["rows"] == [
        {"name": "alice", "score": 10},
        {"name": "bob", "score": 25},
    ]
    assert session["result"]["variables"]["bob_score"] == 25
    assert session["result"]["variables"]["alice_row"] == ["alice", 10]
    assert session["result"]["variables"]["score_table"][1]["score"] == 25


def test_service_runtime_blocks_file_actions_when_file_access_is_disabled(tmp_path) -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    output_path = tmp_path / "blocked.txt"
    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "theme": "light",
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": False,
                    "allow_file_access": False,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-write",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write",
                    "expansion_role": "action:write_text_file",
                    "node_kind": "file.write_text_file",
                    "node_config": {"path": str(output_path), "content": "blocked"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["outputs"]["node-write"]["error_code"] == "file.access_denied"
    assert output_path.exists() is False


from http.server import BaseHTTPRequestHandler
from socketserver import TCPServer
import threading

class _HttpOkHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:  # noqa: N802
        body = b"ok"
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return

class _HttpOkServer(TCPServer):
    allow_reuse_address = True


def test_service_runtime_executes_http_request_and_browser_navigation_with_defaults() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    server = _HttpOkServer(("127.0.0.1", 0), _HttpOkHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        preferences_service = PreferencesService(
            preferences_store=InMemoryPreferencesStore(
                {
                    "preferences_file_version": 1,
                    "program_settings": {
                        "language": "zh-CN",
                        "resource_language": "zh-CN",
                        "theme": "light",
                        "default_window_size": {"width": 1440, "height": 900},
                        "startup_action": "restore_last_workspace",
                        "default_project_directory": None,
                        "recent_project_limit": 10,
                        "preferences_auto_save": True,
                        "font_scale": 100,
                    },
                    "compile_settings": {
                        "default_source_kind": "graph_workspace",
                        "diagnostic_level": "error",
                        "block_on_disabled_components": True,
                        "allow_degraded_compile": True,
                        "stop_on_first_error": True,
                        "emit_runtime_plan": True,
                        "emit_debug_plan": True,
                    },
                    "security_settings": {
                        "confirm_high_risk_actions": True,
                        "allow_external_programs": False,
                        "allow_file_access": True,
                        "allow_browser_executor": True,
                        "allow_local_network_access": True,
                    },
                    "python_runtime_settings": {
                        "python_executable_path": None,
                        "timeout_seconds": 60,
                        "sandbox_mode": "restricted",
                        "capture_stdout_stderr": True,
                    },
                    "graph_settings": {
                        "auto_sync_mode": "responsive",
                        "show_node_id_on_node": True,
                        "show_disabled_resource_badge": True,
                        "snap_to_grid": True,
                        "grid_enabled": True,
                        "auto_open_node_on_drop": True,
                        "confirm_delete_node": True,
                        "show_inline_config_summary": True,
                    },
                    "other_settings": {
                        "workspace_draft_recovery_enabled": True,
                        "workspace_draft_recovery_ttl_minutes": 30,
                    },
                }
            )
        )
        service = CompilationWorkbenchService(preferences_service=preferences_service)

        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-http",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-http",
                        "expansion_role": "action:http_request",
                        "node_kind": "http.request",
                        "node_config": {"url": f"http://127.0.0.1:{server.server_address[1]}/"},
                        "ports": [],
                    }
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-http"]["status"] == "succeeded"
        assert session["result"]["outputs"]["node-http"]["status_code"] == 200
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def test_service_runtime_executes_python_run_and_writes_overridden_result_variable() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-a",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-a",
                    "expansion_role": "transform:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "A", "value": 7},
                    "ports": [],
                },
                {
                    "node_id": "node-b",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-b",
                    "expansion_role": "transform:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "B", "value": 5},
                    "ports": [],
                },
                {
                    "node_id": "node-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-python",
                    "expansion_role": "action:run_python",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "total = variables.get('A', 0) + variables.get('B', 0)\n"
                            "result = {'sum': total, 'page_available': page is not None}\n"
                            "result_variable = 'python_summary'\n"
                        ),
                        "variable_name": "default_summary",
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-python"]["result"] == {
        "sum": 12,
        "page_available": False,
    }
    assert session["result"]["outputs"]["node-python"]["result_variable"] == "python_summary"
    assert session["result"]["variables"]["python_summary"] == {
        "sum": 12,
        "page_available": False,
    }
    assert "default_summary" not in session["result"]["variables"]


def test_service_runtime_hides_python_stdout_stderr_when_capture_is_disabled() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": False,
                    "allow_local_network_access": False,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": False,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-python",
                    "expansion_role": "action:run_python",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": "print('visible'); result = 1",
                        "variable_name": "python_result",
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-python"]["stdout"] == ""
    assert session["result"]["outputs"]["node-python"]["stderr"] == ""


def test_service_runtime_blocks_python_run_when_external_programs_are_disabled() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "theme": "light",
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": False,
                    "allow_file_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {"code": "result = 1"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["outputs"]["node-python"]["error_code"] == "python.execution_disabled"


def test_service_runtime_python_run_captures_output_and_allows_safe_imports() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "import math\n"
                            "print('hello stdout')\n"
                            "result = math.sqrt(16)\n"
                            "result_variable = 'sqrt_value'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-python"]["stdout"] == "hello stdout\n"
    assert session["result"]["outputs"]["node-python"]["stderr"] == ""
    assert session["result"]["variables"]["sqrt_value"] == 4.0


def test_service_runtime_python_run_blocks_disallowed_imports() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {"code": "import os\nresult = os.getcwd()"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["failure_reason"] == "python.import_not_allowed"
    assert session["result"]["outputs"]["node-python"]["error_code"] == "python.import_not_allowed"


def test_service_runtime_recognizes_captcha_with_captcha_ocr(monkeypatch) -> None:
    service = CompilationWorkbenchService()

    class FakeCaptchaRecognizer:
        def __init__(self, *, model_name: str, runtime_root=None):
            self.model_name = model_name
            self.runtime_root = runtime_root

        def recognize_from_bytes(self, image_bytes: bytes) -> str:
            assert image_bytes == b"captcha-bytes"
            return "A1B2"

    monkeypatch.setattr(
        "weconduct.runtime.engine.create_captcha_ocr_recognizer",
        lambda **kwargs: FakeCaptchaRecognizer(**kwargs),
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-captcha",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-captcha",
                    "expansion_role": "action:recognize_captcha",
                    "node_kind": "browser.recognize_captcha",
                    "node_config": {
                        "image_bytes_base64": base64.b64encode(b"captcha-bytes").decode("ascii"),
                        "target_variable": "captcha_code",
                        "model_name": "common_old.onnx",
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-captcha"]["text"] == "A1B2"
    assert session["result"]["outputs"]["node-captcha"]["target_variable"] == "captcha_code"
    assert session["result"]["variables"]["captcha_code"] == "A1B2"


def test_service_runtime_reports_captcha_ocr_unavailable(monkeypatch) -> None:
    service = CompilationWorkbenchService()

    def raise_unavailable(**kwargs):
        raise RuntimeError("captcha_ocr runtime not found")

    monkeypatch.setattr(
        "weconduct.runtime.engine.create_captcha_ocr_recognizer",
        raise_unavailable,
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-captcha",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-captcha",
                    "expansion_role": "action:recognize_captcha",
                    "node_kind": "browser.recognize_captcha",
                    "node_config": {
                        "image_bytes_base64": base64.b64encode(b"captcha-bytes").decode("ascii"),
                        "target_variable": "captcha_code",
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["failure_reason"] == "browser.captcha_ocr_unavailable"
    assert session["result"]["status"] == "failed"
    assert session["result"]["outputs"]["node-captcha"]["error_code"] == "browser.captcha_ocr_unavailable"
    assert "captcha_ocr runtime not found" in session["result"]["outputs"]["node-captcha"]["message"]


def test_service_runtime_executes_extended_browser_atomic_components(tmp_path) -> None:
    service = CompilationWorkbenchService()
    site_server, site_thread = _start_browser_mock_site()
    screenshot_path = tmp_path / "browser-extended-shot.png"

    try:
        base_url = f"http://127.0.0.1:{site_server.server_address[1]}"
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {"url": f"{base_url}/"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-hover",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-hover",
                        "expansion_role": "action:hover",
                        "node_kind": "browser.hover",
                        "node_config": {"selector": "#hover-target"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-wait-hover",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-wait-hover",
                        "expansion_role": "action:wait_for_element",
                        "node_kind": "browser.wait_for_element",
                        "node_config": {"selector": "#hover-result", "timeout": 3000},
                        "ports": [],
                    },
                    {
                        "node_id": "node-select",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-select",
                        "expansion_role": "action:select_option",
                        "node_kind": "browser.select_option",
                        "node_config": {"selector": "#city", "value": "beijing"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-click-go",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-click-go",
                        "expansion_role": "action:click",
                        "node_kind": "browser.click",
                        "node_config": {"selector": "#go-dashboard"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-wait-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-wait-nav",
                        "expansion_role": "action:wait_for_navigation",
                        "node_kind": "browser.wait_for_navigation",
                        "node_config": {"url_pattern": "/dashboard", "timeout": 3000},
                        "ports": [],
                    },
                    {
                        "node_id": "node-back",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-back",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {"url": f"{base_url}/"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-frame",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-frame",
                        "expansion_role": "action:switch_to_frame",
                        "node_kind": "browser.switch_to_frame",
                        "node_config": {"selector": "#content-frame"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-frame-wait",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-frame-wait",
                        "expansion_role": "action:wait_for_element",
                        "node_kind": "browser.wait_for_element",
                        "node_config": {"selector": "#frame-status", "timeout": 3000},
                        "ports": [],
                    },
                    {
                        "node_id": "node-parent",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-parent",
                        "expansion_role": "action:switch_to_parent_frame",
                        "node_kind": "browser.switch_to_parent_frame",
                        "node_config": {},
                        "ports": [],
                    },
                    {
                        "node_id": "node-default",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-default",
                        "expansion_role": "action:switch_to_default_content",
                        "node_kind": "browser.switch_to_default_content",
                        "node_config": {},
                        "ports": [],
                    },
                    {
                        "node_id": "node-open-frame-page",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-open-frame-page",
                        "expansion_role": "action:open_frame_page",
                        "node_kind": "browser.open_frame_page",
                        "node_config": {"selector": "#content-frame"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-timeout",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-timeout",
                        "expansion_role": "action:wait_for_timeout",
                        "node_kind": "browser.wait_for_timeout",
                        "node_config": {"timeout": 10},
                        "ports": [],
                    },
                    {
                        "node_id": "node-shot",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-shot",
                        "expansion_role": "action:screenshot",
                        "node_kind": "browser.screenshot",
                        "node_config": {"path": str(screenshot_path)},
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-hover"]["selector"] == "#hover-target"
        assert session["result"]["outputs"]["node-wait-hover"]["selector"] == "#hover-result"
        assert session["result"]["outputs"]["node-select"]["value"] == "beijing"
        assert session["result"]["outputs"]["node-wait-nav"]["matched_url"].endswith("/dashboard")
        assert session["result"]["outputs"]["node-frame"]["frame_url"].endswith("/frame")
        assert session["result"]["outputs"]["node-frame-wait"]["selector"] == "#frame-status"
        assert session["result"]["outputs"]["node-parent"]["frame_depth"] == 0
        assert session["result"]["outputs"]["node-default"]["frame_depth"] == 0
        assert session["result"]["outputs"]["node-open-frame-page"]["page_url"].endswith("/frame")
        assert session["result"]["outputs"]["node-timeout"]["timeout_ms"] == 10
        assert session["result"]["outputs"]["node-shot"]["path"] == str(screenshot_path.resolve())
        assert screenshot_path.exists() is True
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_flow_start_configures_browser_launch_options(monkeypatch) -> None:
    service = CompilationWorkbenchService()
    captured_launch_kwargs: dict[str, object] = {}

    class FakePage:
        url = "about:blank"

        def goto(self, url: str, wait_until: str | None = None) -> None:
            self.url = url

        def title(self) -> str:
            return "Fake Browser Page"

    class FakeBrowser:
        def new_page(self):
            return FakePage()

        def close(self) -> None:
            return None

    class FakeChromium:
        def launch(self, **kwargs):
            captured_launch_kwargs.update(kwargs)
            return FakeBrowser()

    class FakePlaywrightInstance:
        chromium = FakeChromium()

        def stop(self) -> None:
            return None

    class FakeSyncPlaywright:
        def start(self):
            return FakePlaywrightInstance()

    monkeypatch.setattr(
        "weconduct.runtime.engine.sync_playwright",
        lambda: FakeSyncPlaywright(),
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {
                        "initial_variables": {"base_url": "http://example.test"},
                        "browser_config": {
                            "headless": False,
                            "slow_mo_ms": 150,
                        },
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-nav",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-nav",
                    "expansion_role": "action:navigate",
                    "node_kind": "browser.navigate",
                    "node_config": {"url": "${base_url}"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-nav",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-nav",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert captured_launch_kwargs == {"headless": False, "slow_mo": 150, "channel": "msedge"}
    assert session["result"]["outputs"]["node-start"]["browser_config"] == {
        "headless": False,
        "slow_mo_ms": 150,
    }
    assert session["result"]["outputs"]["node-nav"]["page_url"] == "http://example.test"


def test_service_runtime_executes_extended_excel_write_actions(tmp_path) -> None:
    service = CompilationWorkbenchService()
    workbook_path = tmp_path / "output.xlsx"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-write-file",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write-file",
                    "expansion_role": "action:write_excel_file",
                    "node_kind": "excel.write_file",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Users",
                        "headers": ["name", "score"],
                        "rows": [["alice", 10], ["bob", 20]],
                        "mode": "create",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-write-row",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write-row",
                    "expansion_role": "action:write_excel_row",
                    "node_kind": "excel.write_row",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Users",
                        "row_index": 4,
                        "data": ["carol", 30],
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-write-table",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write-table",
                    "expansion_role": "action:write_excel_table",
                    "node_kind": "excel.write_table",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Summary",
                        "has_header": True,
                        "data": [
                            {"kind": "passed", "count": 2},
                            {"kind": "failed", "count": 1},
                        ],
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-update-cells",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-update-cells",
                    "expansion_role": "action:update_excel_cells",
                    "node_kind": "excel.update_cells",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Users",
                        "updates": [
                            {"row_index": 2, "column_name": "score", "value": 15},
                            {"row_index": 3, "column_index": 2, "value": 25},
                        ],
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-update-batch",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-update-batch",
                    "expansion_role": "action:update_excel_batch",
                    "node_kind": "excel.update_batch",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Users",
                        "condition": "row.get('score', 0) >= 25",
                        "updates": {"score": 99},
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read-users",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read-users",
                    "expansion_role": "action:read_excel_table",
                    "node_kind": "excel.read_table",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Users",
                        "has_header": True,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read-summary",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read-summary",
                    "expansion_role": "action:read_excel_table",
                    "node_kind": "excel.read_table",
                    "node_config": {
                        "path": str(workbook_path),
                        "sheet_name": "Summary",
                        "has_header": True,
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-write-file"]["row_count"] == 2
    assert session["result"]["outputs"]["node-write-row"]["row_index"] == 4
    assert session["result"]["outputs"]["node-write-table"]["sheet_name"] == "Summary"
    assert session["result"]["outputs"]["node-update-cells"]["updated_count"] == 2
    assert session["result"]["outputs"]["node-update-batch"]["updated_count"] == 2
    assert session["result"]["outputs"]["node-read-users"]["rows"] == [
        {"name": "alice", "score": 15},
        {"name": "bob", "score": 99},
        {"name": "carol", "score": 99},
    ]
    assert session["result"]["outputs"]["node-read-summary"]["rows"][0] == {
        "kind": "passed",
        "count": 2,
    }


def test_service_runtime_excel_read_missing_file_returns_structured_error(tmp_path) -> None:
    service = CompilationWorkbenchService()
    missing_path = tmp_path / "missing.xlsx"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-read-missing",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read-missing",
                    "expansion_role": "action:read_excel_cell",
                    "node_kind": "excel.read_cell",
                    "node_config": {
                        "path": str(missing_path),
                        "sheet_name": "Sheet1",
                        "cell": "A1",
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["failure_reason"] == "excel.path_missing"
    assert session["result"]["outputs"]["node-read-missing"]["error_code"] == "excel.path_missing"
    diagnostic_events = [
        item for item in session["event_log"] if item.get("event_kind") == "diagnostic.raised"
    ]
    assert diagnostic_events
    assert diagnostic_events[0]["node_id"] == "node-read-missing"
    assert diagnostic_events[0]["severity"] == "error"
    assert diagnostic_events[0]["error_code"] == "excel.path_missing"


def test_service_runtime_resolves_relative_file_paths_against_project_directory(tmp_path) -> None:
    project_dir = tmp_path / "demo-project"
    project_path = project_dir / "demo.weconduct.json"
    service = CompilationWorkbenchService()
    service.create_project(project_name="Demo", project_directory=project_dir)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-write-text",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write-text",
                    "expansion_role": "action:write_text_file",
                    "node_kind": "file.write_text_file",
                    "node_config": {
                        "path": r"artifacts\result.txt",
                        "content": "hello project",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-write-excel",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-write-excel",
                    "expansion_role": "action:write_excel_file",
                    "node_kind": "excel.write_file",
                    "node_config": {
                        "path": r"artifacts\report.xlsx",
                        "sheet_name": "Users",
                        "headers": ["name", "score"],
                        "rows": [["alice", 10]],
                        "mode": "create",
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.save_project_as(project_path=project_path)

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    expected_text_path = project_dir / "artifacts" / "result.txt"
    expected_excel_path = project_dir / "artifacts" / "report.xlsx"

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-write-text"]["path"] == str(expected_text_path.resolve())
    assert session["result"]["outputs"]["node-write-excel"]["path"] == str(expected_excel_path.resolve())
    assert expected_text_path.read_text(encoding="utf-8") == "hello project"
    assert expected_excel_path.exists() is True


def test_service_runtime_resolves_relative_browser_and_text_paths_against_project_directory(tmp_path) -> None:
    project_dir = tmp_path / "browser-project"
    project_path = project_dir / "browser.weconduct.json"
    service = CompilationWorkbenchService()
    service.create_project(project_name="Browser", project_directory=project_dir)
    site_server, site_thread = _start_browser_mock_site()
    try:
        screenshot_rel_path = r"artifacts\shot.png"
        text_rel_path = r"artifacts\notes.txt"
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {"url": f"http://127.0.0.1:{site_server.server_address[1]}/"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-shot",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-shot",
                        "expansion_role": "action:screenshot",
                        "node_kind": "browser.screenshot",
                        "node_config": {"path": screenshot_rel_path},
                        "ports": [],
                    },
                    {
                        "node_id": "node-text",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-text",
                        "expansion_role": "action:write_text_file",
                        "node_kind": "file.write_text_file",
                        "node_config": {"path": text_rel_path, "content": "project notes"},
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )
        service.save_project_as(project_path=project_path)

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        expected_screenshot_path = project_dir / "artifacts" / "shot.png"
        expected_text_path = project_dir / "artifacts" / "notes.txt"

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-shot"]["path"] == str(expected_screenshot_path.resolve())
        assert session["result"]["outputs"]["node-text"]["path"] == str(expected_text_path.resolve())
        assert expected_screenshot_path.exists() is True
        assert expected_text_path.read_text(encoding="utf-8") == "project notes"
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_scheduler_foreach_uses_control_edges_and_populates_item_and_index_vars() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-list",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-list",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "items", "items": [10, 20, 30]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {
                        "variable": "items",
                        "item_var": "item",
                        "index_var": "idx",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {
                        "variable_name": "results",
                        "value": "${item}",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-last-index",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-last-index",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {
                        "name": "last_index",
                        "value": "${idx}",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read-results",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "results"},
                    "ports": [],
                },
                {
                    "node_id": "node-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-loop-body",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-loop-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-read-results",
                },
                {
                    "edge_id": "edge-body-tail",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-last-index",
                },
                {
                    "edge_id": "edge-tail-end",
                    "relation_layer": "control",
                    "from_node_id": "node-last-index",
                    "to_node_id": "node-end",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["results"] == [10, 20, 30]
    assert session["result"]["variables"]["last_index"] == 2
    assert session["result"]["outputs"]["node-read-results"]["value"] == [10, 20, 30]
    assert session["result"]["outputs"]["node-foreach"]["iteration_count"] == 3
    assert session["result"]["skipped_node_ids"] == []


def test_service_runtime_scheduler_jump_to_step_repeats_by_condition_and_stops_at_max_jumps() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-init",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-init",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "retries", "value": 0},
                    "ports": [],
                },
                {
                    "node_id": "node-inc",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inc",
                    "expansion_role": "action:evaluate_expression",
                    "node_kind": "data.evaluate_expression",
                    "node_config": {
                        "expression": "retries + 1",
                        "variable_name": "retries",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-jump",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-jump",
                    "expansion_role": "action:jump_to_step",
                    "node_kind": "control.jump_to_step",
                    "node_config": {
                        "target_node_id": "node-inc",
                        "condition": "retries < 10",
                        "max_jumps": 2,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "retries"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["retries"] == 3
    assert session["result"]["outputs"]["node-read"]["value"] == 3
    assert session["result"]["outputs"]["node-jump"]["jump_executed"] is False
    assert session["result"]["outputs"]["node-jump"]["jump_count"] == 2
    assert session["result"]["outputs"]["node-jump"]["stopped_by_max_jumps"] is True


def test_service_runtime_flow_start_uses_control_edges_instead_of_node_array_order() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"status": "from-start"}},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "status"},
                    "ports": [],
                },
                {
                    "node_id": "node-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "from-control-edge"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-set",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-set",
                },
                {
                    "edge_id": "edge-set-read",
                    "relation_layer": "control",
                    "from_node_id": "node-set",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["runtime_plan"]["entry_node_ids"] == ["node-start"]
    assert session["runtime_plan"]["scheduler_mode"] == "flow_graph"
    assert session["result"]["completed_node_ids"] == [
        "node-start",
        "node-set",
        "node-read",
    ]
    assert session["result"]["variables"]["status"] == "from-control-edge"
    assert session["result"]["outputs"]["node-start"]["initial_variable_count"] == 1
    assert session["result"]["outputs"]["node-read"]["value"] == "from-control-edge"


def test_service_runtime_flow_start_does_not_execute_unreachable_business_nodes() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"status": "started"}},
                    "ports": [],
                },
                {
                    "node_id": "node-unreachable",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-unreachable",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "unreachable"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["runtime_plan"]["entry_node_ids"] == ["node-start"]
    assert session["result"]["completed_node_ids"] == ["node-start"]
    assert session["node_states"][1]["node_status"] == "pending"
    assert "node-unreachable" not in session["result"]["outputs"]
    assert session["result"]["variables"]["status"] == "started"
    assert session["result"]["skipped_node_ids"] == ["node-unreachable"]
    assert session["result"]["unreachable_node_ids"] == ["node-unreachable"]
    assert session["event_log"][-1]["event_kind"] == "session.completed"
    assert session["event_log"][-1]["unreachable_node_ids"] == ["node-unreachable"]
    skipped_events = [
        item for item in session["event_log"] if item.get("event_kind") == "node.skipped"
    ]
    assert skipped_events
    assert skipped_events[0]["node_id"] == "node-unreachable"


def test_service_runtime_flow_graph_reports_multiple_unreachable_business_nodes() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"status": "started"}},
                    "ports": [],
                },
                {
                    "node_id": "node-a",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-a",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "branch-a"},
                    "ports": [],
                },
                {
                    "node_id": "node-dead-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-dead-1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "dead-1"},
                    "ports": [],
                },
                {
                    "node_id": "node-dead-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-dead-2",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "dead-2"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-a",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-a",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["completed_node_ids"] == ["node-start", "node-a"]
    assert session["result"]["unreachable_node_ids"] == ["node-dead-1", "node-dead-2"]
    assert session["result"]["skipped_node_ids"] == ["node-dead-1", "node-dead-2"]
    assert session["node_states"][2]["node_status"] == "pending"
    assert session["node_states"][3]["node_status"] == "pending"


def test_service_runtime_flow_graph_deduplicates_requeued_targets() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-a",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-a",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "a", "value": 1},
                    "ports": [],
                },
                {
                    "node_id": "node-b",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-b",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "b", "value": 2},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "b"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-a",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-a",
                },
                {
                    "edge_id": "edge-start-b",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-b",
                },
                {
                    "edge_id": "edge-a-read",
                    "relation_layer": "control",
                    "from_node_id": "node-a",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-b-read",
                    "relation_layer": "control",
                    "from_node_id": "node-b",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["completed_node_ids"] == [
        "node-start",
        "node-a",
        "node-b",
        "node-read",
    ]
    assert session["result"]["completed_node_ids"].count("node-read") == 1
    assert session["result"]["outputs"]["node-read"]["value"] == 2


def test_service_runtime_flow_graph_prefers_control_edges_over_node_order() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch", "value": "left"},
                    "ports": [],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch", "value": "right"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "branch"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-right",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-right",
                },
                {
                    "edge_id": "edge-right-read",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-start-left",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-left",
                },
                {
                    "edge_id": "edge-left-read",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["runtime_plan"]["scheduler_mode"] == "flow_graph"
    assert session["result"]["completed_node_ids"] == [
        "node-start",
        "node-right",
        "node-left",
        "node-read",
    ]
    assert session["result"]["outputs"]["node-read"]["value"] == "left"


def test_service_runtime_flow_graph_executes_foreach_and_honors_loop_exit() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "items", "items": [1, 2, 3]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "items", "item_var": "item"},
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {"variable_name": "results", "value": "${item}"},
                    "ports": [],
                },
                {
                    "node_id": "node-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "results"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-items",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-items",
                },
                {
                    "edge_id": "edge-items-results",
                    "relation_layer": "control",
                    "from_node_id": "node-items",
                    "to_node_id": "node-results",
                },
                {
                    "edge_id": "edge-results-foreach",
                    "relation_layer": "control",
                    "from_node_id": "node-results",
                    "to_node_id": "node-foreach",
                },
                {
                    "edge_id": "edge-foreach-body",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-foreach-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-append-end",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-end",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["runtime_plan"]["scheduler_mode"] == "flow_graph"
    assert session["result"]["variables"]["results"] == [1, 2, 3]
    assert session["result"]["outputs"]["node-foreach"]["iteration_count"] == 3
    assert session["result"]["outputs"]["node-read"]["value"] == [1, 2, 3]
    foreach_body_started_events = [
        item
        for item in session["event_log"]
        if item.get("event_kind") == "node.started" and item.get("node_id") == "node-append"
    ]
    foreach_body_completed_events = [
        item
        for item in session["event_log"]
        if item.get("event_kind") == "node.completed" and item.get("node_id") == "node-append"
    ]
    assert len(foreach_body_started_events) == 3
    assert len(foreach_body_completed_events) == 3


def test_service_runtime_flow_graph_executes_nodes_after_end_foreach_exit_marker() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "items", "items": ["a", "b", "c"]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-foreach",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-foreach",
                    "expansion_role": "control:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "items", "item_var": "item"},
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {"variable_name": "results", "value": "${item}"},
                    "ports": [],
                },
                {
                    "node_id": "node-end",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-end",
                    "expansion_role": "control:end_foreach",
                    "node_kind": "control.end_foreach",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-after",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-after",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "after_loop", "value": "done"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "after_loop"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-items",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-items",
                },
                {
                    "edge_id": "edge-items-results",
                    "relation_layer": "control",
                    "from_node_id": "node-items",
                    "to_node_id": "node-results",
                },
                {
                    "edge_id": "edge-results-foreach",
                    "relation_layer": "control",
                    "from_node_id": "node-results",
                    "to_node_id": "node-foreach",
                },
                {
                    "edge_id": "edge-foreach-body",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-append-end",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-end",
                },
                {
                    "edge_id": "edge-end-after",
                    "relation_layer": "control",
                    "from_node_id": "node-end",
                    "to_node_id": "node-after",
                },
                {
                    "edge_id": "edge-after-read",
                    "relation_layer": "control",
                    "from_node_id": "node-after",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["runtime_plan"]["scheduler_mode"] == "flow_graph"
    assert session["result"]["variables"]["results"] == ["a", "b", "c"]
    assert session["result"]["variables"]["after_loop"] == "done"
    assert session["result"]["outputs"]["node-read"]["value"] == "done"
    assert "node-after" in session["result"]["completed_node_ids"]
    assert "node-read" in session["result"]["completed_node_ids"]
    assert "node-after" not in session["result"]["skipped_node_ids"]
    assert "node-read" not in session["result"]["skipped_node_ids"]


def test_service_runtime_flow_graph_jump_to_step_requeues_target() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-init",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-init",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "count", "value": 0},
                    "ports": [],
                },
                {
                    "node_id": "node-inc",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inc",
                    "expansion_role": "action:evaluate_expression",
                    "node_kind": "data.evaluate_expression",
                    "node_config": {
                        "expression": "count + 1",
                        "variable_name": "count",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-jump",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-jump",
                    "expansion_role": "action:jump_to_step",
                    "node_kind": "control.jump_to_step",
                    "node_config": {
                        "target_node_id": "node-inc",
                        "condition": "count < 3",
                        "max_jumps": 5,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "count"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-init",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-init",
                },
                {
                    "edge_id": "edge-init-inc",
                    "relation_layer": "control",
                    "from_node_id": "node-init",
                    "to_node_id": "node-inc",
                },
                {
                    "edge_id": "edge-inc-jump",
                    "relation_layer": "control",
                    "from_node_id": "node-inc",
                    "to_node_id": "node-jump",
                },
                {
                    "edge_id": "edge-jump-read",
                    "relation_layer": "control",
                    "from_node_id": "node-jump",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["runtime_plan"]["scheduler_mode"] == "flow_graph"
    assert session["result"]["variables"]["count"] == 3
    assert session["result"]["outputs"]["node-read"]["value"] == 3
    assert session["result"]["completed_node_ids"].count("node-inc") == 3
    assert session["result"]["completed_node_ids"].count("node-jump") == 3


def test_service_runtime_plan_reports_data_dependencies_and_potential_write_conflicts() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "shared", "value": "left"},
                    "ports": [
                        {
                            "port_id": "out:value",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "value",
                        }
                    ],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "shared", "value": "right"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "shared"},
                    "ports": [
                        {
                            "port_id": "in:value",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "value",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-left",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-left",
                },
                {
                    "edge_id": "edge-start-right",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-right",
                },
                {
                    "edge_id": "edge-left-read",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-right-read",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-data-left-read",
                    "relation_layer": "data",
                    "from_node_id": "node-left",
                    "to_node_id": "node-read",
                    "from_port_id": "out:value",
                    "to_port_id": "in:value",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)

    assert started["status"] == "started"
    assert started["runtime_plan"]["scheduler_hints"]["data_dependency_edges"] == [
        {
            "edge_id": "edge-data-left-read",
            "from_node_id": "node-left",
            "to_node_id": "node-read",
            "from_port_id": "out:value",
            "to_port_id": "in:value",
        }
    ]
    assert started["runtime_plan"]["scheduler_hints"]["potential_write_conflicts"] == [
        {
            "variable_name": "shared",
            "writer_node_ids": ["node-left", "node-right"],
            "reachable_writer_node_ids": ["node-left", "node-right"],
        }
    ]


def test_service_runtime_flow_graph_if_routes_only_selected_branch() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"flag": True}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-if",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-if",
                    "expansion_role": "control:if",
                    "node_kind": "control.if",
                    "node_config": {"expression": "flag"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "true",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.true",
                        },
                        {
                            "port_id": "false",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.false",
                        },
                    ],
                },
                {
                    "node_id": "node-true",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-true",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch", "value": "true-branch"},
                    "ports": [],
                },
                {
                    "node_id": "node-false",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-false",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch", "value": "false-branch"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "branch"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-if",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-if",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-if-true",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-true",
                    "from_port_id": "true",
                },
                {
                    "edge_id": "edge-if-false",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-false",
                    "from_port_id": "false",
                },
                {
                    "edge_id": "edge-true-read",
                    "relation_layer": "control",
                    "from_node_id": "node-true",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-false-read",
                    "relation_layer": "control",
                    "from_node_id": "node-false",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == "true-branch"
    assert "node-true" in session["result"]["completed_node_ids"]
    assert "node-false" not in session["result"]["completed_node_ids"]
    branch_events = [
        item for item in session["event_log"] if item.get("event_kind") == "branch.selected"
    ]
    assert branch_events
    assert branch_events[0]["node_id"] == "node-if"
    assert branch_events[0]["selected_port_id"] == "true"


def test_service_runtime_flow_graph_if_allows_initial_entry_with_repeat_port_present() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"flag": True}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-if",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-if",
                    "expansion_role": "control:if",
                    "node_kind": "control.if",
                    "node_config": {"expression": "flag"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "repeat",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.repeat",
                        },
                        {
                            "port_id": "true",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.true",
                        },
                        {
                            "port_id": "false",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.false",
                        },
                    ],
                },
                {
                    "node_id": "node-true",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-true",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch", "value": "true-branch"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "branch"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-if",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-if",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-if-true",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-true",
                    "from_port_id": "true",
                },
                {
                    "edge_id": "edge-true-read",
                    "relation_layer": "control",
                    "from_node_id": "node-true",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["branch"] == "true-branch"
    assert "node-if" in session["result"]["completed_node_ids"]
    assert "node-true" in session["result"]["completed_node_ids"]
    join_waiting_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.waiting"
    ]
    assert all(item.get("node_id") != "node-if" for item in join_waiting_events)


def test_service_runtime_flow_graph_data_edge_binds_source_output_into_target_input() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-producer",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-producer",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "upstream_name", "value": "alice"},
                    "ports": [
                        {
                            "port_id": "out:value",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.value",
                        }
                    ],
                },
                {
                    "node_id": "node-consumer",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-consumer",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "consumed_name"},
                    "ports": [
                        {
                            "port_id": "in:value",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.value",
                        }
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "consumed_name"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-producer",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-producer",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-producer-consumer-control",
                    "relation_layer": "control",
                    "from_node_id": "node-producer",
                    "to_node_id": "node-consumer",
                },
                {
                    "edge_id": "edge-producer-consumer-data",
                    "relation_layer": "data",
                    "from_node_id": "node-producer",
                    "to_node_id": "node-consumer",
                    "from_port_id": "out:value",
                    "to_port_id": "in:value",
                },
                {
                    "edge_id": "edge-consumer-read",
                    "relation_layer": "control",
                    "from_node_id": "node-consumer",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == "alice"
    assert session["result"]["outputs"]["node-consumer"]["value"] == "alice"


def test_service_runtime_node_config_merges_runtime_input_overrides() -> None:
    context = RuntimeContext()
    registry = RuntimeExecutorRegistry()

    output = execute_runtime_node(
        {
            "node_id": "node-consumer",
            "node_kind": "data.set_variable",
            "node_config": {
                "name": "merged_name",
            },
            "__runtime_input_overrides__": {
                "value": "alice",
                "in:value": "ignored-shadow",
            },
        },
        context,
        registry,
    )

    assert output["status"] == "succeeded"
    assert output["value"] == "alice"
    assert context.variables["merged_name"] == "alice"


def test_service_runtime_node_config_merges_nested_runtime_input_overrides() -> None:
    context = RuntimeContext()
    registry = RuntimeExecutorRegistry()

    output = execute_runtime_node(
        {
            "node_id": "node-start",
            "node_kind": "flow.start",
            "node_config": {
                "browser_config": {
                    "headless": True,
                    "slow_mo_ms": 0,
                }
            },
            "__runtime_input_overrides__": {
                "browser_config.headless": False,
                "browser_config.slow_mo_ms": 150,
            },
        },
        context,
        registry,
    )

    assert output["status"] == "succeeded"
    assert output["browser_config"] == {
        "headless": False,
        "slow_mo_ms": 150,
    }
    assert context.browser_runtime["launch_options"] == {
        "headless": False,
        "slow_mo_ms": 150,
    }


def test_runtime_reference_supports_explicit_type_cast_suffixes() -> None:
    context = RuntimeContext(
        variables={
            "count": "3",
            "enabled": "true",
            "payload": '{"name":"alice"}',
        }
    )
    registry = RuntimeExecutorRegistry()

    set_result = execute_runtime_node(
        {
            "node_id": "node-set",
            "node_kind": "data.set_variable",
            "node_config": {
                "name": "resolved_count",
                "value": "${count|int}",
            },
        },
        context,
        registry,
    )
    expression_result = execute_runtime_node(
        {
            "node_id": "node-expr",
            "node_kind": "data.evaluate_expression",
            "node_config": {
                "expression": "count > 1 and enabled",
                "variable_name": "expr_result",
            },
            "__runtime_input_overrides__": {
                "expression": "${count|int} > 1 and ${enabled|bool}",
            },
        },
        context,
        registry,
    )
    payload_result = execute_runtime_node(
        {
            "node_id": "node-payload",
            "node_kind": "data.set_variable",
            "node_config": {
                "name": "resolved_payload",
                "value": "${payload|json}",
            },
        },
        context,
        registry,
    )

    assert set_result["status"] == "succeeded"
    assert set_result["value"] == 3
    assert isinstance(context.variables["resolved_count"], int)
    assert expression_result["status"] == "succeeded"
    assert expression_result["value"] is True
    assert context.variables["expr_result"] is True
    assert payload_result["status"] == "succeeded"
    assert payload_result["value"] == {"name": "alice"}
    assert context.variables["resolved_payload"] == {"name": "alice"}


def test_runtime_convert_value_node_supports_new_and_in_place_targets() -> None:
    context = RuntimeContext(
        variables={
            "count": "7",
            "enabled": "false",
        }
    )
    registry = RuntimeExecutorRegistry()

    first = execute_runtime_node(
        {
            "node_id": "node-convert-new",
            "node_kind": "data.convert_value",
            "node_config": {
                "source_value": "${count}",
                "target_type": "int",
                "variable_name": "count_number",
                "in_place": False,
                "source_variable_name": "count",
            },
        },
        context,
        registry,
    )
    second = execute_runtime_node(
        {
            "node_id": "node-convert-inplace",
            "node_kind": "data.convert_value",
            "node_config": {
                "source_value": "${enabled}",
                "target_type": "bool",
                "in_place": True,
                "source_variable_name": "enabled",
                "variable_name": "",
            },
        },
        context,
        registry,
    )

    assert first["status"] == "succeeded"
    assert first["value"] == 7
    assert first["variable_name"] == "count_number"
    assert context.variables["count_number"] == 7
    assert context.variables["count"] == "7"
    assert second["status"] == "succeeded"
    assert second["value"] is False
    assert second["variable_name"] == "enabled"
    assert context.variables["enabled"] is False


def test_service_runtime_browser_launch_uses_system_edge_channel(monkeypatch) -> None:
    captured_launch_kwargs = {}

    class FakePage:
        url = "about:blank"

        def goto(self, url: str, wait_until: str | None = None) -> None:
            self.url = url

        def title(self) -> str:
            return "Fake Browser Page"

    class FakeBrowser:
        def new_page(self):
            return FakePage()

        def close(self) -> None:
            return None

    class FakeChromium:
        def launch(self, **kwargs):
            captured_launch_kwargs.update(kwargs)
            return FakeBrowser()

    class FakePlaywrightInstance:
        chromium = FakeChromium()

        def stop(self) -> None:
            return None

    class FakeSyncPlaywright:
        def start(self):
            return FakePlaywrightInstance()

    monkeypatch.setattr(
        "weconduct.runtime.engine.sync_playwright",
        lambda: FakeSyncPlaywright(),
    )

    context = RuntimeContext()
    context.browser_runtime["launch_options"] = {"headless": True, "slow_mo_ms": 25}

    page = _require_browser_page(context)

    assert page.title() == "Fake Browser Page"
    assert captured_launch_kwargs == {
        "headless": True,
        "slow_mo": 25,
        "channel": "msedge",
    }


def test_service_runtime_flow_graph_user_component_output_data_edge_binds_child_output_into_downstream_input() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:component-output-edge",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "component-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "component-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = f\"hello {variables.get('name')}\"\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-component-start-python",
                    "relation_layer": "control",
                    "from_node_id": "component-start",
                    "to_node_id": "component-python",
                    "from_port_id": "out",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    component = service.save_user_component_resource(resource_name="Output Edge Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-call-component",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-component",
                    "expansion_role": "module:user-component",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": {"name": "Alice"},
                    },
                    "ports": [
                        {
                            "port_id": "result-port",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.message",
                        }
                    ],
                },
                {
                    "node_id": "node-consumer",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-consumer",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "component_message"},
                    "ports": [
                        {
                            "port_id": "input-port",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.value",
                        }
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "component_message"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-call",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-call-component",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-call-consumer-control",
                    "relation_layer": "control",
                    "from_node_id": "node-call-component",
                    "to_node_id": "node-consumer",
                },
                {
                    "edge_id": "edge-call-consumer-data",
                    "relation_layer": "data",
                    "from_node_id": "node-call-component",
                    "to_node_id": "node-consumer",
                    "from_port_id": "result-port",
                    "to_port_id": "input-port",
                },
                {
                    "edge_id": "edge-consumer-read",
                    "relation_layer": "control",
                    "from_node_id": "node-consumer",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-consumer"]["value"] == "hello Alice"
    assert session["result"]["outputs"]["node-read"]["value"] == "hello Alice"


def test_service_runtime_flow_graph_parallel_fork_and_join_wait_for_all_inputs() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch:left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "branch:right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_stage", "value": "pre-join"},
                    "ports": [],
                },
                {
                    "node_id": "node-right-final",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-final",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "node_config": {"mode": "all"},
                    "ports": [
                        {
                            "port_id": "in:left",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:left",
                        },
                        {
                            "port_id": "in:right",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:right",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "right_done"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-fork",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-fork",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "branch:left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right",
                    "from_port_id": "branch:right",
                },
                {
                    "edge_id": "edge-left-join",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-join",
                    "to_port_id": "in:left",
                },
                {
                    "edge_id": "edge-right-right-final",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-right-final",
                },
                {
                    "edge_id": "edge-right-final-join",
                    "relation_layer": "control",
                    "from_node_id": "node-right-final",
                    "to_node_id": "node-join",
                    "to_port_id": "in:right",
                },
                {
                    "edge_id": "edge-join-read",
                    "relation_layer": "control",
                    "from_node_id": "node-join",
                    "to_node_id": "node-read",
                    "from_port_id": "out",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] is True
    assert session["result"]["completed_node_ids"].count("node-join") == 1
    assert session["result"]["completed_node_ids"].index("node-right-final") < (
        session["result"]["completed_node_ids"].index("node-join")
    )
    assert session["result"]["completed_node_ids"][-1] == "node-read"
    join_waiting_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.waiting"
    ]
    join_released_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.released"
    ]
    assert join_waiting_events
    assert join_released_events
    assert join_waiting_events[0]["node_id"] == "node-join"
    assert join_released_events[0]["node_id"] == "node-join"
    token_enqueued_events = [
        item for item in session["event_log"] if item.get("event_kind") == "token.enqueued"
    ]
    token_dispatched_events = [
        item for item in session["event_log"] if item.get("event_kind") == "token.dispatched"
    ]
    node_ready_events = [
        item for item in session["event_log"] if item.get("event_kind") == "node.ready"
    ]
    assert token_enqueued_events
    assert token_dispatched_events
    assert node_ready_events
    assert any(item.get("target_node_id") == "node-fork" for item in token_enqueued_events)
    assert any(item.get("node_id") == "node-fork" for item in node_ready_events)


def test_service_runtime_flow_graph_implicit_join_waits_for_all_control_inputs() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                        ]
                    },
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch:left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "branch:right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right-final",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-final",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_final_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-collector",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-collector",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {
                        "name": "collector_ready",
                        "value": {"left": "${left_done}", "right": "${right_final_done}"},
                    },
                    "ports": [
                        {
                            "port_id": "in:left",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control:left",
                        },
                        {
                            "port_id": "in:right",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control:right",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "collector_ready"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-fork",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-fork",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "branch:left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right",
                    "from_port_id": "branch:right",
                },
                {
                    "edge_id": "edge-left-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in:left",
                },
                {
                    "edge_id": "edge-right-right-final",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-right-final",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-right-final-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-right-final",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in:right",
                },
                {
                    "edge_id": "edge-collector-read",
                    "relation_layer": "control",
                    "from_node_id": "node-collector",
                    "to_node_id": "node-read",
                    "from_port_id": "out",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == {"left": True, "right": True}
    assert session["result"]["completed_node_ids"].count("node-collector") == 1
    assert session["result"]["completed_node_ids"].index("node-right-final") < (
        session["result"]["completed_node_ids"].index("node-collector")
    )
    implicit_join_waiting_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.waiting"
    ]
    implicit_join_released_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.released"
    ]
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_waiting_events)
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_released_events)


def test_service_runtime_flow_graph_implicit_join_waits_for_all_control_edges_even_on_same_input_port() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                        ]
                    },
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch:left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "branch:right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_same_port_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_same_port_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right-final",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-final",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_same_port_final_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-collector",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-collector",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {
                        "name": "collector_same_port_ready",
                        "value": {"left": "${left_same_port_done}", "right": "${right_same_port_final_done}"},
                    },
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "collector_same_port_ready"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-fork",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-fork",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "branch:left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right",
                    "from_port_id": "branch:right",
                },
                {
                    "edge_id": "edge-left-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-right-right-final",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-right-final",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-right-final-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-right-final",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-collector-read",
                    "relation_layer": "control",
                    "from_node_id": "node-collector",
                    "to_node_id": "node-read",
                    "from_port_id": "out",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == {"left": True, "right": True}
    assert session["result"]["completed_node_ids"].count("node-collector") == 1
    assert session["result"]["completed_node_ids"].index("node-right-final") < (
        session["result"]["completed_node_ids"].index("node-collector")
    )
    implicit_join_waiting_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.waiting"
    ]
    implicit_join_released_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.released"
    ]
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_waiting_events)
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_released_events)


def test_service_runtime_flow_graph_implicit_join_waits_for_all_branches_from_regular_node_fanout() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-split",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-split",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "fanout_root_done", "value": True},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "regular_left_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-middle",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-middle",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "regular_middle_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right-step-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-step-1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "regular_right_step_1_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right-step-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-step-2",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "regular_right_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-collector",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-collector",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {
                        "name": "regular_fanout_collected",
                        "value": {
                            "left": "${regular_left_done}",
                            "middle": "${regular_middle_done}",
                            "right": "${regular_right_done}",
                        },
                    },
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "regular_fanout_collected"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-split",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-split",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-split-left",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-left",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-split-middle",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-middle",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-split-right",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-right-step-1",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-right-step-1-step-2",
                    "relation_layer": "control",
                    "from_node_id": "node-right-step-1",
                    "to_node_id": "node-right-step-2",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-left-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-middle-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-middle",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-right-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-right-step-2",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-collector-read",
                    "relation_layer": "control",
                    "from_node_id": "node-collector",
                    "to_node_id": "node-read",
                    "from_port_id": "out",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == {
        "left": True,
        "middle": True,
        "right": True,
    }
    assert session["result"]["completed_node_ids"].count("node-collector") == 1
    assert session["result"]["completed_node_ids"].index("node-right-step-2") < (
        session["result"]["completed_node_ids"].index("node-collector")
    )
    implicit_join_waiting_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.waiting"
    ]
    implicit_join_released_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.released"
    ]
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_waiting_events)
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_released_events)


def test_service_runtime_flow_graph_implicit_join_waits_when_one_branch_contains_internal_join() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-split",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-split",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "root_done", "value": True},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-branch-wait",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-branch-wait",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch_wait_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-branch-ready",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-branch-ready",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch_ready_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-branch-join",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-branch-join",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch_join_done", "value": True},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-branch-final",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-branch-final",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch_final_done", "value": True},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-collector",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-collector",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {
                        "name": "collector_nested_join_ready",
                        "value": {
                            "left": "${left_done}",
                            "right": "${right_done}",
                            "branch": "${branch_final_done}",
                        },
                    },
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "collector_nested_join_ready"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-split",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-split",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-split-left",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-left",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-split-right",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-right",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-split-branch-wait",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-branch-wait",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-split-branch-ready",
                    "relation_layer": "control",
                    "from_node_id": "node-split",
                    "to_node_id": "node-branch-ready",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-branch-wait-join",
                    "relation_layer": "control",
                    "from_node_id": "node-branch-wait",
                    "to_node_id": "node-branch-join",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-branch-ready-join",
                    "relation_layer": "control",
                    "from_node_id": "node-branch-ready",
                    "to_node_id": "node-branch-join",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-branch-join-final",
                    "relation_layer": "control",
                    "from_node_id": "node-branch-join",
                    "to_node_id": "node-branch-final",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-left-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-right-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-branch-final-collector",
                    "relation_layer": "control",
                    "from_node_id": "node-branch-final",
                    "to_node_id": "node-collector",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-collector-read",
                    "relation_layer": "control",
                    "from_node_id": "node-collector",
                    "to_node_id": "node-read",
                    "from_port_id": "out",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == {
        "left": True,
        "right": True,
        "branch": True,
    }
    assert session["result"]["completed_node_ids"].count("node-collector") == 1
    assert session["result"]["completed_node_ids"].index("node-branch-final") < (
        session["result"]["completed_node_ids"].index("node-collector")
    )
    implicit_join_waiting_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.waiting"
    ]
    implicit_join_released_events = [
        item for item in session["event_log"] if item.get("event_kind") == "join.released"
    ]
    assert any(item.get("node_id") == "node-branch-join" for item in implicit_join_waiting_events)
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_waiting_events)
    assert any(item.get("node_id") == "node-branch-join" for item in implicit_join_released_events)
    assert any(item.get("node_id") == "node-collector" for item in implicit_join_released_events)


def test_service_runtime_flow_graph_if_uses_semantic_slot_when_control_port_ids_are_renamed() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"flag": True}},
                    "ports": [
                        {
                            "port_id": "entry-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-if",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-if",
                    "expansion_role": "control:if",
                    "node_kind": "control.if",
                    "node_config": {"expression": "flag"},
                    "ports": [
                        {
                            "port_id": "control-in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.condition",
                        },
                        {
                            "port_id": "input-control",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch-yes",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.true",
                        },
                        {
                            "port_id": "branch-no",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.false",
                        },
                    ],
                },
                {
                    "node_id": "node-true",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-true",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch_message", "value": "true-branch"},
                    "ports": [],
                },
                {
                    "node_id": "node-false",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-false",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "branch_message", "value": "false-branch"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "branch_message"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-if",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-if",
                    "from_port_id": "entry-control",
                    "to_port_id": "input-control",
                },
                {
                    "edge_id": "edge-if-true",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-true",
                    "from_port_id": "branch-yes",
                },
                {
                    "edge_id": "edge-if-false",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-false",
                    "from_port_id": "branch-no",
                },
                {
                    "edge_id": "edge-true-read",
                    "relation_layer": "control",
                    "from_node_id": "node-true",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-false-read",
                    "relation_layer": "control",
                    "from_node_id": "node-false",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == "true-branch"
    assert "node-true" in session["result"]["completed_node_ids"]
    assert "node-false" not in session["result"]["completed_node_ids"]


def test_service_runtime_flow_graph_join_uses_semantic_slot_when_control_port_ids_are_renamed() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "entry-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "fork-input",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "fork-left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "fork-right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "node_config": {"mode": "all"},
                    "ports": [
                        {
                            "port_id": "join-left",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:left",
                        },
                        {
                            "port_id": "join-right",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:right",
                        },
                        {
                            "port_id": "join-output",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "right_done"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-fork",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-fork",
                    "from_port_id": "entry-control",
                    "to_port_id": "fork-input",
                },
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "fork-left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right",
                    "from_port_id": "fork-right",
                },
                {
                    "edge_id": "edge-left-join",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-join",
                    "to_port_id": "join-left",
                },
                {
                    "edge_id": "edge-right-join",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-join",
                    "to_port_id": "join-right",
                },
                {
                    "edge_id": "edge-join-read",
                    "relation_layer": "control",
                    "from_node_id": "node-join",
                    "to_node_id": "node-read",
                    "from_port_id": "join-output",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] is True
    assert session["result"]["completed_node_ids"].index("node-right") < (
        session["result"]["completed_node_ids"].index("node-join")
    )


def test_service_runtime_session_node_states_use_static_flow_order_before_execution() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "display_name": "Read",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "right_done"},
                    "ports": [],
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "display_name": "Join",
                    "node_kind": "control.join",
                    "node_config": {"mode": "all"},
                    "ports": [
                        {
                            "port_id": "join-left",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:left",
                        },
                        {
                            "port_id": "join-right",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:right",
                        },
                        {
                            "port_id": "join-output",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "display_name": "Left",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "display_name": "Start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {}},
                    "ports": [
                        {
                            "port_id": "entry-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "display_name": "Right",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "display_name": "Fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                        ]
                    },
                    "ports": [
                        {
                            "port_id": "fork-input",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "fork-left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "fork-right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-fork",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-fork",
                    "from_port_id": "entry-control",
                    "to_port_id": "fork-input",
                },
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "fork-left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right",
                    "from_port_id": "fork-right",
                },
                {
                    "edge_id": "edge-left-join",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-join",
                    "to_port_id": "join-left",
                },
                {
                    "edge_id": "edge-right-join",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-join",
                    "to_port_id": "join-right",
                },
                {
                    "edge_id": "edge-join-read",
                    "relation_layer": "control",
                    "from_node_id": "node-join",
                    "to_node_id": "node-read",
                    "from_port_id": "join-output",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)

    assert [item["node_id"] for item in started["node_states"]] == [
        "node-start",
        "node-fork",
        "node-left",
        "node-right",
        "node-join",
        "node-read",
    ]


def test_service_runtime_session_node_states_use_execution_order_after_run() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "display_name": "Read",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "right_done"},
                    "ports": [],
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "display_name": "Join",
                    "node_kind": "control.join",
                    "node_config": {"mode": "all"},
                    "ports": [
                        {
                            "port_id": "join-left",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:left",
                        },
                        {
                            "port_id": "join-right",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:right",
                        },
                        {
                            "port_id": "join-output",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "display_name": "Left",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "display_name": "Start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {}},
                    "ports": [
                        {
                            "port_id": "entry-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-right",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right",
                    "expansion_role": "action:set_variable",
                    "display_name": "Right",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "display_name": "Fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                        ]
                    },
                    "ports": [
                        {
                            "port_id": "fork-input",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "fork-left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "fork-right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-fork",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-fork",
                    "from_port_id": "entry-control",
                    "to_port_id": "fork-input",
                },
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "fork-left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right",
                    "from_port_id": "fork-right",
                },
                {
                    "edge_id": "edge-left-join",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-join",
                    "to_port_id": "join-left",
                },
                {
                    "edge_id": "edge-right-join",
                    "relation_layer": "control",
                    "from_node_id": "node-right",
                    "to_node_id": "node-join",
                    "to_port_id": "join-right",
                },
                {
                    "edge_id": "edge-join-read",
                    "relation_layer": "control",
                    "from_node_id": "node-join",
                    "to_node_id": "node-read",
                    "from_port_id": "join-output",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert [item["node_id"] for item in session["node_states"]] == [
        "node-start",
        "node-fork",
        "node-left",
        "node-right",
        "node-join",
        "node-read",
    ]


def test_service_runtime_flow_graph_switch_routes_selected_case_only() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"route": "beta"}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-switch",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-switch",
                    "expansion_role": "control:switch",
                    "node_kind": "control.switch",
                    "node_config": {"selector": "route"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "case:alpha",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.case:alpha",
                        },
                        {
                            "port_id": "case:beta",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.case:beta",
                        },
                        {
                            "port_id": "default",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.default",
                        },
                    ],
                },
                {
                    "node_id": "node-alpha",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-alpha",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "selected_route", "value": "alpha"},
                    "ports": [],
                },
                {
                    "node_id": "node-beta",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-beta",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "selected_route", "value": "beta"},
                    "ports": [],
                },
                {
                    "node_id": "node-default",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-default",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "selected_route", "value": "default"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "selected_route"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-switch",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-switch",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-switch-alpha",
                    "relation_layer": "control",
                    "from_node_id": "node-switch",
                    "to_node_id": "node-alpha",
                    "from_port_id": "case:alpha",
                },
                {
                    "edge_id": "edge-switch-beta",
                    "relation_layer": "control",
                    "from_node_id": "node-switch",
                    "to_node_id": "node-beta",
                    "from_port_id": "case:beta",
                },
                {
                    "edge_id": "edge-switch-default",
                    "relation_layer": "control",
                    "from_node_id": "node-switch",
                    "to_node_id": "node-default",
                    "from_port_id": "default",
                },
                {
                    "edge_id": "edge-alpha-read",
                    "relation_layer": "control",
                    "from_node_id": "node-alpha",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-beta-read",
                    "relation_layer": "control",
                    "from_node_id": "node-beta",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-default-read",
                    "relation_layer": "control",
                    "from_node_id": "node-default",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == "beta"
    assert "node-beta" in session["result"]["completed_node_ids"]
    assert "node-alpha" not in session["result"]["completed_node_ids"]
    assert "node-default" not in session["result"]["completed_node_ids"]


def test_service_runtime_flow_graph_while_routes_loop_then_done() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"count": 0}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-while",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-while",
                    "expansion_role": "control:while",
                    "node_kind": "control.while",
                    "node_config": {"expression": "count < 2"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "repeat",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.repeat",
                        },
                        {
                            "port_id": "loop",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.loop",
                        },
                        {
                            "port_id": "done",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.done",
                        },
                    ],
                },
                {
                    "node_id": "node-inc",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inc",
                    "expansion_role": "action:increment_variable",
                    "node_kind": "data.increment_variable",
                    "node_config": {"name": "count", "step": 1},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "count"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-while",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-while",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-while-loop",
                    "relation_layer": "control",
                    "from_node_id": "node-while",
                    "to_node_id": "node-inc",
                    "from_port_id": "loop",
                },
                {
                    "edge_id": "edge-inc-while",
                    "relation_layer": "control",
                    "from_node_id": "node-inc",
                    "to_node_id": "node-while",
                    "to_port_id": "repeat",
                },
                {
                    "edge_id": "edge-while-done",
                    "relation_layer": "control",
                    "from_node_id": "node-while",
                    "to_node_id": "node-read",
                    "from_port_id": "done",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["count"] == 2
    assert session["result"]["outputs"]["node-read"]["value"] == 2
    assert session["result"]["completed_node_ids"].count("node-inc") == 2
    assert session["result"]["completed_node_ids"].count("node-while") == 3
    loop_events = [
        item for item in session["event_log"] if item.get("event_kind") == "loop.iteration"
    ]
    assert len(loop_events) == 2
    assert [item["iteration_index"] for item in loop_events] == [1, 2]


def test_service_runtime_flow_graph_retry_requeues_attempt_until_success() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"attempt_counter": 0}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-retry",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-retry",
                    "expansion_role": "control:retry",
                    "node_kind": "control.retry",
                    "node_config": {"max_attempts": 3, "success_expression": "attempt_counter >= 2"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "attempt",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.attempt",
                        },
                        {
                            "port_id": "exhausted",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.exhausted",
                        },
                    ],
                },
                {
                    "node_id": "node-inc",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inc",
                    "expansion_role": "action:increment_variable",
                    "node_kind": "data.increment_variable",
                    "node_config": {"name": "attempt_counter", "step": 1},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "attempt_counter"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-retry",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-retry",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-retry-attempt",
                    "relation_layer": "control",
                    "from_node_id": "node-retry",
                    "to_node_id": "node-inc",
                    "from_port_id": "attempt",
                },
                {
                    "edge_id": "edge-inc-retry",
                    "relation_layer": "control",
                    "from_node_id": "node-inc",
                    "to_node_id": "node-retry",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-retry-exhausted",
                    "relation_layer": "control",
                    "from_node_id": "node-retry",
                    "to_node_id": "node-read",
                    "from_port_id": "exhausted",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["attempt_counter"] == 2
    assert session["result"]["outputs"]["node-read"]["value"] == 2
    assert session["result"]["completed_node_ids"].count("node-inc") == 2
    assert session["result"]["completed_node_ids"].count("node-retry") == 3
    retry_events = [
        item for item in session["event_log"] if item.get("event_kind") == "retry.scheduled"
    ]
    assert len(retry_events) == 2
    assert [item["attempt_index"] for item in retry_events] == [1, 2]


def test_service_runtime_flow_graph_failover_switches_to_fallback_then_finishes() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"use_fallback": True}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-failover",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-failover",
                    "expansion_role": "control:failover",
                    "node_kind": "control.failover",
                    "node_config": {"fallback_expression": "use_fallback"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "primary",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.primary",
                        },
                        {
                            "port_id": "fallback:backup",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.fallback:backup",
                        },
                        {
                            "port_id": "failed",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.failed",
                        },
                    ],
                },
                {
                    "node_id": "node-primary",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-primary",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "path", "value": "primary"},
                    "ports": [],
                },
                {
                    "node_id": "node-fallback",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-fallback",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "path", "value": "fallback"},
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "path"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-failover",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-failover",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-failover-primary",
                    "relation_layer": "control",
                    "from_node_id": "node-failover",
                    "to_node_id": "node-primary",
                    "from_port_id": "primary",
                },
                {
                    "edge_id": "edge-failover-fallback",
                    "relation_layer": "control",
                    "from_node_id": "node-failover",
                    "to_node_id": "node-fallback",
                    "from_port_id": "fallback:backup",
                },
                {
                    "edge_id": "edge-primary-read",
                    "relation_layer": "control",
                    "from_node_id": "node-primary",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-fallback-read",
                    "relation_layer": "control",
                    "from_node_id": "node-fallback",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-read"]["value"] == "fallback"
    assert "node-fallback" in session["result"]["completed_node_ids"]
    assert "node-primary" not in session["result"]["completed_node_ids"]
    failover_events = [
        item for item in session["event_log"] if item.get("event_kind") == "failover.switched"
    ]
    assert failover_events
    assert failover_events[0]["node_id"] == "node-failover"
    assert failover_events[0]["selected_port_id"] == "fallback:backup"


def test_service_runtime_scheduler_user_component_runs_child_graph_and_maps_outputs_to_parent_variables() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "component-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "component-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"component_status": "seed"}},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "variables['internal_only'] = 'hidden'\n"
                            "result = f\"hello {variables.get('name')}\"\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                },
                {
                    "node_id": "component-dead",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-dead",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "dead", "value": True},
                    "ports": [],
                }
            ],
            "edges": [
                {
                    "edge_id": "edge-component-start-python",
                    "relation_layer": "control",
                    "from_node_id": "component-start",
                    "to_node_id": "component-python",
                    "from_port_id": "out",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_component = service.save_user_component_resource(resource_name="Greeter Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-component",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-component",
                    "expansion_role": "module:user-component",
                    "node_kind": saved_component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": {"name": "Alice"},
                        "outputs": {"message": "greeting"},
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read-greeting",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-read-greeting",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "greeting"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["greeting"] == "hello Alice"
    assert "message" not in session["result"]["variables"]
    assert "internal_only" not in session["result"]["variables"]
    assert session["result"]["outputs"]["node-call-component"]["mapped_outputs"] == {
        "greeting": "hello Alice"
    }
    assert session["result"]["outputs"]["node-read-greeting"]["value"] == "hello Alice"
    component_result = session["result"]["outputs"]["node-call-component"]["component_result"]
    assert component_result["event_log"]
    assert component_result["skipped_node_ids"] == ["component-dead"]
    assert component_result["unreachable_node_ids"] == ["component-dead"]
    assert any(
        item.get("event_kind") == "token.enqueued"
        for item in component_result["event_log"]
    )
    assert any(
        item.get("event_kind") == "token.dispatched"
        for item in component_result["event_log"]
    )
    assert any(
        item.get("event_kind") == "node.ready"
        for item in component_result["event_log"]
    )
    assert any(
        item.get("event_kind") == "node.skipped" and item.get("node_id") == "component-dead"
        for item in component_result["event_log"]
    )


def test_service_runtime_scheduler_foreach_supports_continue_break_and_end_markers() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "items", "items": [1, 2, 3, 4]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "items", "item_var": "item"},
                    "ports": [],
                },
                {
                    "node_id": "node-continue",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-continue",
                    "expansion_role": "action:foreach_continue",
                    "node_kind": "control.foreach_continue",
                    "node_config": {"condition": "item == 2"},
                    "ports": [],
                },
                {
                    "node_id": "node-break",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-break",
                    "expansion_role": "action:foreach_break",
                    "node_kind": "control.foreach_break",
                    "node_config": {"condition": "item > 3"},
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {"variable_name": "results", "value": "${item}"},
                    "ports": [],
                },
                {
                    "node_id": "node-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "results"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-body",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-continue",
                },
                {
                    "edge_id": "edge-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-continue-break",
                    "relation_layer": "control",
                    "from_node_id": "node-continue",
                    "to_node_id": "node-break",
                },
                {
                    "edge_id": "edge-break-append",
                    "relation_layer": "control",
                    "from_node_id": "node-break",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-append-end",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-end",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["results"] == [1, 3]
    assert session["result"]["outputs"]["node-read"]["value"] == [1, 3]
    assert session["result"]["outputs"]["node-continue"]["continue_triggered"] is False
    assert session["result"]["outputs"]["node-break"]["break_triggered"] is True
    assert session["result"]["outputs"]["node-end"]["end_marker"] is True


def test_service_runtime_scheduler_foreach_continue_supports_multilevel_level() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-outer-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-outer-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "outer_items", "items": [1, 2]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-outer-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-outer-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "outer_items", "item_var": "outer_item"},
                    "ports": [],
                },
                {
                    "node_id": "node-inner-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inner-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "inner_items", "items": [10, 20]},
                    "ports": [],
                },
                {
                    "node_id": "node-inner-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inner-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "inner_items", "item_var": "inner_item"},
                    "ports": [],
                },
                {
                    "node_id": "node-continue-outer",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-continue-outer",
                    "expansion_role": "action:foreach_continue",
                    "node_kind": "control.foreach_continue",
                    "node_config": {
                        "condition": "outer_item == 1 and inner_item == 10",
                        "level": 2,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {
                        "variable_name": "results",
                        "value": "${outer_item}:${inner_item}",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-inner-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inner-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-outer-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-outer-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "results"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-outer-body",
                    "relation_layer": "control",
                    "from_node_id": "node-outer-foreach",
                    "to_node_id": "node-inner-items",
                },
                {
                    "edge_id": "edge-outer-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-outer-foreach",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-inner-items-foreach",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-items",
                    "to_node_id": "node-inner-foreach",
                },
                {
                    "edge_id": "edge-inner-body",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-foreach",
                    "to_node_id": "node-continue-outer",
                },
                {
                    "edge_id": "edge-inner-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-foreach",
                    "to_node_id": "node-outer-end",
                },
                {
                    "edge_id": "edge-continue-append",
                    "relation_layer": "control",
                    "from_node_id": "node-continue-outer",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-append-inner-end",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-inner-end",
                },
                {
                    "edge_id": "edge-inner-end-outer-end",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-end",
                    "to_node_id": "node-outer-end",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["results"] == ["2:10", "2:20"]
    assert session["result"]["outputs"]["node-read"]["value"] == ["2:10", "2:20"]
    assert session["result"]["outputs"]["node-continue-outer"]["level"] == 2


def test_service_runtime_scheduler_foreach_break_supports_multilevel_level() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-outer-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-outer-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "outer_items", "items": [1, 2, 3]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-outer-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-outer-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "outer_items", "item_var": "outer_item"},
                    "ports": [],
                },
                {
                    "node_id": "node-inner-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inner-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "inner_items", "items": [10, 20]},
                    "ports": [],
                },
                {
                    "node_id": "node-inner-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inner-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "inner_items", "item_var": "inner_item"},
                    "ports": [],
                },
                {
                    "node_id": "node-break-all",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-break-all",
                    "expansion_role": "action:foreach_break",
                    "node_kind": "control.foreach_break",
                    "node_config": {
                        "condition": "outer_item == 2 and inner_item == 20",
                        "level": 2,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {
                        "variable_name": "results",
                        "value": "${outer_item}:${inner_item}",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-inner-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inner-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-outer-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-outer-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "results"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-outer-body",
                    "relation_layer": "control",
                    "from_node_id": "node-outer-foreach",
                    "to_node_id": "node-inner-items",
                },
                {
                    "edge_id": "edge-outer-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-outer-foreach",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-inner-items-foreach",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-items",
                    "to_node_id": "node-inner-foreach",
                },
                {
                    "edge_id": "edge-inner-body",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-foreach",
                    "to_node_id": "node-break-all",
                },
                {
                    "edge_id": "edge-inner-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-foreach",
                    "to_node_id": "node-outer-end",
                },
                {
                    "edge_id": "edge-break-append",
                    "relation_layer": "control",
                    "from_node_id": "node-break-all",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-append-inner-end",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-inner-end",
                },
                {
                    "edge_id": "edge-inner-end-outer-end",
                    "relation_layer": "control",
                    "from_node_id": "node-inner-end",
                    "to_node_id": "node-outer-end",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["results"] == ["1:10", "1:20", "2:10"]
    assert session["result"]["outputs"]["node-read"]["value"] == ["1:10", "1:20", "2:10"]
    assert session["result"]["outputs"]["node-break-all"]["break_triggered"] is True


def test_service_runtime_scheduler_counts_foreach_body_steps_against_runtime_limit() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "items", "items": list(range(1001))},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "items", "item_var": "item"},
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {"variable_name": "results", "value": "${item}"},
                    "ports": [],
                },
                {
                    "node_id": "node-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-body",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-append",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["failure_reason"] == "runtime.execution_step_limit_exceeded"
    assert len(session["result"]["completed_node_ids"]) <= 1000


def test_service_runtime_scheduler_foreach_body_uses_control_edges_not_node_range() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-items",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-items",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "items", "items": [1, 2]},
                    "ports": [],
                },
                {
                    "node_id": "node-results",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-results",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "results", "items": []},
                    "ports": [],
                },
                {
                    "node_id": "node-foreach",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-foreach",
                    "expansion_role": "action:foreach",
                    "node_kind": "control.foreach",
                    "node_config": {"variable": "items", "item_var": "item"},
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {"variable_name": "results", "value": "${item}"},
                    "ports": [],
                },
                {
                    "node_id": "node-stray",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-stray",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "stray", "value": "should-not-run"},
                    "ports": [],
                },
                {
                    "node_id": "node-end",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-end",
                    "expansion_role": "action:end_foreach",
                    "node_kind": "control.end_foreach",
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "results"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-body",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-append",
                },
                {
                    "edge_id": "edge-exit",
                    "relation_layer": "control",
                    "from_node_id": "node-foreach",
                    "to_node_id": "node-read",
                },
                {
                    "edge_id": "edge-append-end",
                    "relation_layer": "control",
                    "from_node_id": "node-append",
                    "to_node_id": "node-end",
                },
                {
                    "edge_id": "edge-end-read",
                    "relation_layer": "control",
                    "from_node_id": "node-end",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["results"] == [1, 2]
    assert "stray" not in session["result"]["variables"]
    assert session["node_states"][4]["node_status"] == "pending"


def test_service_runtime_scheduler_call_blueprint_uses_component_resource_id() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = variables.get('name').upper()\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_component = service.save_user_component_resource(resource_name="Upper Name")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-blueprint",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-blueprint",
                    "expansion_role": "action:call_blueprint",
                    "node_kind": "call_blueprint",
                    "node_config": {
                        "blueprint_id": saved_component["resource"]["resource_id"],
                        "inputs": {"name": "alice"},
                        "outputs": {"message": "upper_name"},
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "upper_name"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["upper_name"] == "ALICE"
    assert session["result"]["outputs"]["node-call-blueprint"]["blueprint_id"] == (
        saved_component["resource"]["resource_id"]
    )
    assert session["result"]["outputs"]["node-call-blueprint"]["mapped_outputs"] == {
        "upper_name": "ALICE"
    }


def test_service_runtime_scheduler_call_blueprint_resolves_legacy_blueprint_info_id() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {"blueprint_info": {"id": "legacy-upper-blueprint"}},
            "nodes": [
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = variables.get('name').upper()\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.save_user_component_resource(resource_name="Legacy Upper Name")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-blueprint",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-blueprint",
                    "expansion_role": "action:call_blueprint",
                    "node_kind": "call_blueprint",
                    "node_config": {
                        "blueprint_id": "legacy-upper-blueprint",
                        "inputs": {"name": "alice"},
                        "outputs": {"message": "upper_name"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["upper_name"] == "ALICE"
    assert session["result"]["outputs"]["node-call-blueprint"]["blueprint_id"] == (
        "legacy-upper-blueprint"
    )


def test_service_runtime_scheduler_graph_call_subgraph_runs_child_graph_and_maps_outputs() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "${incoming}"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Echo Subgraph")

    parent_graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call-subgraph",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call-subgraph",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {"incoming": "hello"},
                    "outputs": {"message": "result_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(parent_graph)
    completed = service.run_runtime_session(
        session_id=started["runtime_session"]["session_id"]
    )

    assert completed["result"]["status"] == "succeeded"
    assert completed["result"]["variables"]["result_message"] == "hello"


def test_service_runtime_scheduler_user_component_accepts_data_edge_bound_inputs() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:component-input-edge",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = variables.get('name')\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    component = service.save_user_component_resource(resource_name="Edge Input Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-producer",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-producer",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "upstream_name", "value": "alice"},
                    "ports": [
                        {
                            "port_id": "out:value",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.value",
                        }
                    ],
                },
                {
                    "node_id": "node-call-component",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-component",
                    "expansion_role": "module:input-edge",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": {},
                        "outputs": {"message": "greeting"},
                    },
                    "ports": [
                        {
                            "port_id": "in:name",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.name",
                        }
                    ],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "greeting"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-producer",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-producer",
                    "from_port_id": "out",
                },
                {
                    "edge_id": "edge-producer-call-control",
                    "relation_layer": "control",
                    "from_node_id": "node-producer",
                    "to_node_id": "node-call-component",
                },
                {
                    "edge_id": "edge-producer-call-data",
                    "relation_layer": "data",
                    "from_node_id": "node-producer",
                    "to_node_id": "node-call-component",
                    "from_port_id": "out:value",
                    "to_port_id": "in:name",
                },
                {
                    "edge_id": "edge-call-read",
                    "relation_layer": "control",
                    "from_node_id": "node-call-component",
                    "to_node_id": "node-read",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["greeting"] == "alice"
    assert session["result"]["outputs"]["node-call-component"]["mapped_outputs"] == {
        "greeting": "alice"
    }


def test_service_runtime_scheduler_graph_call_subgraph_resolves_legacy_user_component_resource() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:legacy-subgraph",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "legacy-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "legacy-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "legacy"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    legacy = service.save_user_component_resource(resource_name="Legacy Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": legacy["resource"]["resource_id"],
                    "outputs": {"message": "legacy_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)
    result = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert result["result"]["status"] == "succeeded"
    assert result["result"]["variables"]["legacy_message"] == "legacy"


def test_service_runtime_scheduler_graph_call_subgraph_reports_subgraph_error_codes() -> None:
    service = CompilationWorkbenchService()

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {"subgraph_id": "missing-subgraph"},
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)

    assert started["status"] == "failed"
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["session_id"] is None
    assert started["diagnostics"]["entries"][0]["category"] == (
        "graph.call_subgraph.subgraph_missing"
    )
    assert started["diagnostics"]["entries"][0]["object_ref"] == "node-call"


def test_service_runtime_scheduler_user_component_internal_if_uses_selected_branch_only() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:component-if",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-if",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-if",
                    "expansion_role": "control:if",
                    "node_kind": "control.if",
                    "node_config": {"expression": "flag"},
                    "ports": [
                        {
                            "port_id": "true",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.true",
                        },
                        {
                            "port_id": "false",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.false",
                        },
                    ],
                },
                {
                    "node_id": "node-true",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-true",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "true-branch"},
                    "ports": [],
                },
                {
                    "node_id": "node-false",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-false",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "false-branch"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-if-true",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-true",
                    "from_port_id": "true",
                },
                {
                    "edge_id": "edge-if-false",
                    "relation_layer": "control",
                    "from_node_id": "node-if",
                    "to_node_id": "node-false",
                    "from_port_id": "false",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    component = service.save_user_component_resource(resource_name="Branching Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call",
                    "expansion_role": "module:user-component",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": {"flag": True},
                        "outputs": {"message": "selected_message"},
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-read",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "selected_message"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["selected_message"] == "true-branch"
    assert session["result"]["outputs"]["node-read"]["value"] == "true-branch"
    assert session["result"]["outputs"]["node-call"]["component_result"]["variables"]["message"] == (
        "true-branch"
    )


def test_service_runtime_scheduler_custom_node_graph_internal_while_repeats_until_done() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-while",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-while",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-while",
                    "expansion_role": "control:while",
                    "node_kind": "control.while",
                    "node_config": {"expression": "count < 2"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "repeat",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.repeat",
                        },
                        {
                            "port_id": "loop",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.loop",
                        },
                        {
                            "port_id": "done",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.done",
                        },
                    ],
                },
                {
                    "node_id": "node-inc",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inc",
                    "expansion_role": "action:increment_variable",
                    "node_kind": "data.increment_variable",
                    "node_config": {"name": "count", "step": 1},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-while-loop",
                    "relation_layer": "control",
                    "from_node_id": "node-while",
                    "to_node_id": "node-inc",
                    "from_port_id": "loop",
                },
                {
                    "edge_id": "edge-inc-while",
                    "relation_layer": "control",
                    "from_node_id": "node-inc",
                    "to_node_id": "node-while",
                    "to_port_id": "repeat",
                },
            ],
            "root_metadata": {
                "output_schema": {"count": {"type": "number"}},
            },
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    custom_graph = service.save_custom_node_graph_resource(resource_name="Counter Graph")

    parent_graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "flow-start",
                "lowered_kind": "control",
                "source_anchor_ref": "n-flow-start",
                "expansion_role": "flow:start",
                "node_kind": "flow.start",
                "node_config": {},
                "ports": [],
            },
            {
                "node_id": "call-custom",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call-custom",
                "expansion_role": "action:custom_node_graph",
                "node_kind": custom_graph["resource"]["resource_key"],
                "node_config": {
                    "inputs": {"count": 0},
                    "outputs": {"count": "final_count"},
                },
                "ports": [],
            },
            {
                "node_id": "node-read",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-read",
                "expansion_role": "action:get_variable",
                "node_kind": "data.get_variable",
                "node_config": {"name": "final_count"},
                "ports": [],
            },
        ],
        "edges": [
            {
                "edge_id": "edge-start-call",
                "relation_layer": "control",
                "from_node_id": "flow-start",
                "to_node_id": "call-custom",
            },
            {
                "edge_id": "edge-call-read",
                "relation_layer": "control",
                "from_node_id": "call-custom",
                "to_node_id": "node-read",
            },
        ],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(parent_graph)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["final_count"] == 2
    assert session["result"]["outputs"]["node-read"]["value"] == 2
    assert session["result"]["outputs"]["call-custom"]["component_result"]["variables"]["count"] == 2


def test_service_runtime_scheduler_graph_call_subgraph_internal_parallel_join_waits_for_all_branches() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-join",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "branch:left",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:left",
                        },
                        {
                            "port_id": "branch:right",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:right",
                        },
                    ],
                },
                {
                    "node_id": "node-left",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-left",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "left_done", "value": True},
                    "ports": [],
                },
                {
                    "node_id": "node-right-mid",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-mid",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_stage", "value": "pre-join"},
                    "ports": [],
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "node_config": {"mode": "all"},
                    "ports": [
                        {
                            "port_id": "in:left",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:left",
                        },
                        {
                            "port_id": "in:right",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:right",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-after-join",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-after-join",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "join_result", "value": "${right_done}"},
                    "ports": [],
                },
                {
                    "node_id": "node-right-final",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-right-final",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "right_done", "value": True},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-fork-left",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-left",
                    "from_port_id": "branch:left",
                },
                {
                    "edge_id": "edge-fork-right",
                    "relation_layer": "control",
                    "from_node_id": "node-fork",
                    "to_node_id": "node-right-mid",
                    "from_port_id": "branch:right",
                },
                {
                    "edge_id": "edge-left-join",
                    "relation_layer": "control",
                    "from_node_id": "node-left",
                    "to_node_id": "node-join",
                    "to_port_id": "in:left",
                },
                {
                    "edge_id": "edge-right-mid-right-final",
                    "relation_layer": "control",
                    "from_node_id": "node-right-mid",
                    "to_node_id": "node-right-final",
                },
                {
                    "edge_id": "edge-right-final-join",
                    "relation_layer": "control",
                    "from_node_id": "node-right-final",
                    "to_node_id": "node-join",
                    "to_port_id": "in:right",
                },
                {
                    "edge_id": "edge-join-after",
                    "relation_layer": "control",
                    "from_node_id": "node-join",
                    "to_node_id": "node-after-join",
                    "from_port_id": "out",
                },
            ],
            "root_metadata": {
                "output_schema": {"join_result": {"type": "boolean"}},
            },
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    subgraph = service.save_subgraph_resource(resource_name="Join Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": subgraph["resource"]["resource_id"],
                    "outputs": {"join_result": "subgraph_join_result"},
                },
                "ports": [],
            },
            {
                "node_id": "node-read",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-read",
                "expansion_role": "action:get_variable",
                "node_kind": "data.get_variable",
                "node_config": {"name": "subgraph_join_result"},
                "ports": [],
            },
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["subgraph_join_result"] is True
    assert session["result"]["outputs"]["node-read"]["value"] is True
    assert session["result"]["outputs"]["node-call"]["component_result"]["variables"]["join_result"] is (
        True
    )


def test_service_compile_graph_document_blocks_missing_graph_call_subgraph_resource() -> None:
    service = CompilationWorkbenchService()

    result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {"subgraph_id": "missing-subgraph"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert result["status"] == "failed"
    assert result["view"]["diagnostic_summary"]["total_count"] == 1
    assert result["view"]["diagnostic_groups"][0]["category"] == (
        "graph.call_subgraph.subgraph_missing"
    )


def test_service_runtime_scheduler_graph_call_subgraph_rejects_missing_required_schema_input() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "schema"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema Input Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {},
                    "outputs": {"message": "result_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)

    assert started["status"] == "failed"
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["session_id"] is None
    assert started["diagnostics"]["entries"][0]["category"] == (
        "graph.call_subgraph.input_mapping_missing_required"
    )


def test_service_runtime_scheduler_graph_call_subgraph_accepts_required_input_from_data_edge() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-edge",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "${incoming}"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema Edge Input Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-start",
                "lowered_kind": "control",
                "source_anchor_ref": "n-start",
                "expansion_role": "flow:start",
                "node_kind": "flow.start",
                "node_config": {},
                "ports": [
                    {
                        "port_id": "out",
                        "direction": "output",
                        "relation_layer": "control",
                        "semantic_slot": "out.control",
                    }
                ],
            },
            {
                "node_id": "node-producer",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-producer",
                "expansion_role": "action:set_variable",
                "node_kind": "data.set_variable",
                "node_config": {"name": "upstream_name", "value": "hello"},
                "ports": [
                    {
                        "port_id": "out:value",
                        "direction": "output",
                        "relation_layer": "data",
                        "semantic_slot": "out.value",
                    }
                ],
            },
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {},
                    "outputs": {"message": "result_message"},
                },
                "ports": [
                    {
                        "port_id": "in:incoming",
                        "direction": "input",
                        "relation_layer": "data",
                        "semantic_slot": "in.incoming",
                    }
                ],
            },
            {
                "node_id": "node-read",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-read",
                "expansion_role": "action:get_variable",
                "node_kind": "data.get_variable",
                "node_config": {"name": "result_message"},
                "ports": [],
            },
        ],
        "edges": [
            {
                "edge_id": "edge-start-producer",
                "relation_layer": "control",
                "from_node_id": "node-start",
                "to_node_id": "node-producer",
                "from_port_id": "out",
            },
            {
                "edge_id": "edge-producer-call-control",
                "relation_layer": "control",
                "from_node_id": "node-producer",
                "to_node_id": "node-call",
            },
            {
                "edge_id": "edge-producer-call-data",
                "relation_layer": "data",
                "from_node_id": "node-producer",
                "to_node_id": "node-call",
                "from_port_id": "out:value",
                "to_port_id": "in:incoming",
            },
            {
                "edge_id": "edge-call-read",
                "relation_layer": "control",
                "from_node_id": "node-call",
                "to_node_id": "node-read",
            },
        ],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)
    assert started["status"] == "started"

    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])
    assert session["status"] == "completed"
    assert session["result"]["variables"]["result_message"] == "hello"
    assert session["result"]["outputs"]["node-read"]["value"] == "hello"


def test_service_runtime_scheduler_graph_call_subgraph_rejects_unknown_schema_output_mapping() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-output",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": False},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "schema"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema Output Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {"incoming": "hello"},
                    "outputs": {"unknown_output": "result_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)

    assert started["status"] == "failed"
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["session_id"] is None
    assert started["diagnostics"]["entries"][0]["category"] == (
        "graph.call_subgraph.output_mapping_unknown_output"
    )


def test_service_runtime_scheduler_graph_call_subgraph_rejects_string_schema_input_type_mismatch() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-string",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "schema"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema String Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {"incoming": ["not-a-string"]},
                    "outputs": {"message": "result_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)

    assert started["status"] == "failed"
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["session_id"] is None
    assert started["diagnostics"]["entries"][0]["category"] == (
        "graph.call_subgraph.input_mapping_type_mismatch"
    )


def test_service_runtime_scheduler_graph_call_subgraph_rejects_number_schema_input_type_mismatch() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-number",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "count": {"type": "number", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "schema"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema Number Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {"count": "not-a-number"},
                    "outputs": {"message": "result_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)

    assert started["status"] == "failed"
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["session_id"] is None
    assert started["diagnostics"]["entries"][0]["category"] == (
        "graph.call_subgraph.input_mapping_type_mismatch"
    )


def test_service_runtime_scheduler_graph_call_subgraph_rejects_output_schema_type_mismatch() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-output-type",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": False},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [
                {
                    "node_id": "child-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-child-set",
                    "expansion_role": "action:set_variable",
                    "display_name": "Set Variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": 123},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_subgraph_resource(resource_name="Schema Output Type Subgraph")

    graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-call",
                "lowered_kind": "execution",
                "source_anchor_ref": "n-call",
                "expansion_role": "action:call_subgraph",
                "display_name": "Call Subgraph",
                "node_kind": "graph.call_subgraph",
                "node_config": {
                    "subgraph_id": saved["resource"]["resource_id"],
                    "inputs": {"incoming": "hello"},
                    "outputs": {"message": "result_message"},
                },
                "ports": [],
            }
        ],
        "edges": [],
        "graph_effective_diagnostic_anchor_refs": [],
    }

    started = service.start_runtime_session(graph)
    result = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert result["status"] == "failed"
    assert result["result"]["outputs"]["node-call"]["error_code"] == "subgraph.output_mapping_invalid"


def test_service_runtime_scheduler_graph_call_subgraph_reports_recursive_subgraph_error() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:recursive-subgraph",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-recursive",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-recursive",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Recursive Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {"subgraph_id": "subgraph_resource:recursive_subgraph"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    recursive_subgraph = service.save_subgraph_resource(resource_name="Recursive Subgraph")
    service.save_graph_document(
        {
            "graph_model_id": "graph:recursive-subgraph",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-recursive",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-recursive",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Recursive Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": recursive_subgraph["resource"]["resource_id"]
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.save_subgraph_resource(
        resource_name="Recursive Subgraph",
        replace_existing_resource_id=recursive_subgraph["resource"]["resource_id"],
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-recursive",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-recursive",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Recursive Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": recursive_subgraph["resource"]["resource_id"]
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    output = session["result"]["outputs"]["node-call-recursive"]
    assert output["error_code"] == "subgraph.recursive_call_detected"


def test_service_runtime_scheduler_graph_call_subgraph_reports_call_depth_exceeded() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "leaf-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "leaf-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "depth_result", "value": "ok"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    child_component = service.save_subgraph_resource(resource_name="Depth Leaf Subgraph")

    for depth in range(1, 10):
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": f"call-depth-{depth}",
                        "lowered_kind": "execution",
                        "source_anchor_ref": f"call-depth-{depth}",
                        "expansion_role": f"action:call_subgraph_depth_{depth}",
                        "display_name": f"Call Depth {depth}",
                        "node_kind": "graph.call_subgraph",
                        "node_config": {
                            "subgraph_id": child_component["resource"]["resource_id"]
                        },
                        "ports": [],
                    }
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )
        child_component = service.save_subgraph_resource(
            resource_name=f"Depth Layer Subgraph {depth}",
        )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-depth-chain",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-depth-chain",
                    "expansion_role": "action:call_subgraph_depth_chain",
                    "display_name": "Call Depth Chain",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {"subgraph_id": child_component["resource"]["resource_id"]},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    output = session["result"]["outputs"]["node-call-depth-chain"]
    assert output["error_code"] == "subgraph.call_depth_exceeded"


def test_phase9_1_subgraph_success_project_fixture_runs_successfully(tmp_path) -> None:
    fixture_path = (
        Path(__file__).resolve().parents[3]
        / "docs"
        / "dev"
        / "phase-9-1"
        / "fixtures"
        / "p9_1_subgraph_success_project.weconduct.json"
    )
    assert fixture_path.exists() is True

    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)
    opened = service.open_project(project_path=fixture_path)
    started = service.start_runtime_session(None)
    result = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert opened["project"]["project_name"] == "P9.1 Subgraph Success Project"
    assert result["result"]["status"] == "succeeded"
    assert result["result"]["variables"]["subgraph_result"] == "HELLO SUBGRAPH"


def test_phase9_1_subgraph_failure_project_fixture_reports_subgraph_error(tmp_path) -> None:
    fixture_path = (
        Path(__file__).resolve().parents[3]
        / "docs"
        / "dev"
        / "phase-9-1"
        / "fixtures"
        / "p9_1_subgraph_failure_project.weconduct.json"
    )
    assert fixture_path.exists() is True

    service = CompilationWorkbenchService()
    service.open_project(project_path=fixture_path)
    started = service.start_runtime_session(None)

    assert started["status"] == "failed"
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["session_id"] is None
    assert started["diagnostics"]["entries"][0]["category"] == "custom_node_graph.missing"


def test_service_runtime_scheduler_runs_nested_user_components_with_call_stack() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "base-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "base-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = f\"base {variables.get('name')}\"\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    base_component = service.save_user_component_resource(resource_name="Base Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "wrapper-call-base",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "wrapper-call-base",
                    "expansion_role": "module:base-component",
                    "node_kind": base_component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": {"name": "${outer_name}"},
                        "outputs": {"message": "wrapped_message"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    wrapper_component = service.save_user_component_resource(resource_name="Wrapper Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-wrapper",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-wrapper",
                    "expansion_role": "module:wrapper-component",
                    "node_kind": wrapper_component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": {"outer_name": "Alice"},
                        "outputs": {"wrapped_message": "final_message"},
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-read-final",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-read-final",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": "final_message"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["variables"]["final_message"] == "base Alice"
    wrapper_output = session["result"]["outputs"]["node-call-wrapper"]
    assert wrapper_output["component_call_stack"] == [
        wrapper_component["resource"]["resource_id"]
    ]
    assert wrapper_output["component_result"]["outputs"]["wrapper-call-base"]["mapped_outputs"] == {
        "wrapped_message": "base Alice"
    }


def test_service_runtime_scheduler_blocks_recursive_user_component_calls() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "placeholder",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "placeholder",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "placeholder", "value": "created"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    component = service.save_user_component_resource(resource_name="Recursive Component")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "self-call",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "self-call",
                    "expansion_role": "module:self",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    service.save_user_component_resource(
        resource_name="Recursive Component",
        replace_existing_resource_id=component["resource"]["resource_id"],
    )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-recursive",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-recursive",
                    "expansion_role": "module:recursive",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    output = session["result"]["outputs"]["node-call-recursive"]
    assert output["error_code"] == "component.recursive_call_detected"
    assert output["component_call_stack"] == [
        component["resource"]["resource_id"],
        component["resource"]["resource_id"],
    ]


def test_service_runtime_scheduler_blocks_user_component_call_depth_overflow() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "leaf-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "leaf-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "depth_result", "value": "ok"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    child_component = service.save_user_component_resource(resource_name="Depth Leaf")

    for depth in range(1, 10):
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": f"call-depth-{depth}",
                        "lowered_kind": "execution",
                        "source_anchor_ref": f"call-depth-{depth}",
                        "expansion_role": f"module:depth-{depth}",
                        "node_kind": child_component["resource"]["resource_id"],
                        "node_config": {},
                        "ports": [],
                    }
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )
        child_component = service.save_user_component_resource(
            resource_name=f"Depth Layer {depth}",
        )

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-depth-chain",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-depth-chain",
                    "expansion_role": "module:depth-chain",
                    "node_kind": child_component["resource"]["resource_id"],
                    "node_config": {},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    output = session["result"]["outputs"]["node-call-depth-chain"]
    assert output["error_code"] == "component.call_depth_exceeded"
    assert len(output["component_call_stack"]) == 8


def test_service_runtime_scheduler_rejects_invalid_component_input_mapping() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = variables.get('name')\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    component = service.save_user_component_resource(resource_name="Input Mapping Target")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-component",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-component",
                    "expansion_role": "module:input-invalid",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {
                        "inputs": [{"name": "alice"}],
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    output = session["result"]["outputs"]["node-call-component"]
    assert output["error_code"] == "component.input_mapping_invalid"


def test_service_runtime_scheduler_rejects_invalid_component_output_mapping() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "component-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "component-python",
                    "expansion_role": "action:python_run",
                    "node_kind": "python.run",
                    "node_config": {
                        "code": (
                            "result = 'ok'\n"
                            "result_variable = 'message'\n"
                        )
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    component = service.save_user_component_resource(resource_name="Output Mapping Target")

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-component",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "node-call-component",
                    "expansion_role": "module:output-invalid",
                    "node_kind": component["resource"]["resource_id"],
                    "node_config": {
                        "outputs": ["message", "target"],
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    output = session["result"]["outputs"]["node-call-component"]
    assert output["error_code"] == "component.output_mapping_invalid"


def test_service_runtime_executes_builtin_custom_data_components() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-list",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-list",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {
                        "variable_name": "numbers",
                        "items": [1, 2],
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-append",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-append",
                    "expansion_role": "action:list_append",
                    "node_kind": "data.list_append",
                    "node_config": {
                        "variable_name": "numbers",
                        "value": 3,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-eval",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-eval",
                    "expansion_role": "action:evaluate_expression",
                    "node_kind": "data.evaluate_expression",
                    "node_config": {
                        "expression": "len(numbers) * 10",
                        "variable_name": "score",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-regex",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-regex",
                    "expansion_role": "action:regex_replace",
                    "node_kind": "data.regex_replace",
                        "node_config": {
                            "text": "score=${score}",
                            "pattern": r"\d+",
                            "replacement": "done",
                            "variable_name": "message",
                        },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-list"]["items"] == [1, 2]
    assert session["result"]["outputs"]["node-append"]["items"] == [1, 2, 3]
    assert session["result"]["outputs"]["node-eval"]["value"] == 30
    assert session["result"]["outputs"]["node-regex"]["value"] == "score=done"
    assert session["result"]["variables"]["numbers"] == [1, 2, 3]
    assert session["result"]["variables"]["score"] == 30
    assert session["result"]["variables"]["message"] == "score=done"


def test_service_runtime_executes_extended_builtin_custom_list_components() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-list",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-list",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "numbers", "items": [1, 2]},
                    "ports": [],
                },
                {
                    "node_id": "node-extend",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-extend",
                    "expansion_role": "action:list_extend",
                    "node_kind": "data.list_extend",
                    "node_config": {"variable_name": "numbers", "items": [3, 4]},
                    "ports": [],
                },
                {
                    "node_id": "node-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set",
                    "expansion_role": "action:list_set",
                    "node_kind": "data.list_set",
                    "node_config": {"variable_name": "numbers", "index": 1, "value": 20},
                    "ports": [],
                },
                {
                    "node_id": "node-get",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-get",
                    "expansion_role": "action:list_get",
                    "node_kind": "data.list_get",
                    "node_config": {"variable_name": "numbers", "index": 2, "output_variable_name": "picked"},
                    "ports": [],
                },
                {
                    "node_id": "node-length",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-length",
                    "expansion_role": "action:list_length",
                    "node_kind": "data.list_length",
                    "node_config": {"variable_name": "numbers", "output_variable_name": "count"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-extend"]["items"] == [1, 2, 3, 4]
    assert session["result"]["outputs"]["node-set"]["items"] == [1, 20, 3, 4]
    assert session["result"]["outputs"]["node-get"]["value"] == 3
    assert session["result"]["outputs"]["node-length"]["value"] == 4
    assert session["result"]["variables"]["numbers"] == [1, 20, 3, 4]
    assert session["result"]["variables"]["picked"] == 3
    assert session["result"]["variables"]["count"] == 4


def test_service_runtime_executes_remaining_builtin_custom_list_components() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-list",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-list",
                    "expansion_role": "action:create_list",
                    "node_kind": "data.create_list",
                    "node_config": {"variable_name": "letters", "items": ["c", "a", "b"]},
                    "ports": [],
                },
                {
                    "node_id": "node-insert",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-insert",
                    "expansion_role": "action:list_insert",
                    "node_kind": "data.list_insert",
                    "node_config": {"variable_name": "letters", "index": 1, "value": "z"},
                    "ports": [],
                },
                {
                    "node_id": "node-remove",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-remove",
                    "expansion_role": "action:list_remove",
                    "node_kind": "data.list_remove",
                    "node_config": {"variable_name": "letters", "value": "a"},
                    "ports": [],
                },
                {
                    "node_id": "node-sort",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-sort",
                    "expansion_role": "action:list_sort",
                    "node_kind": "data.list_sort",
                    "node_config": {"variable_name": "letters"},
                    "ports": [],
                },
                {
                    "node_id": "node-slice",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-slice",
                    "expansion_role": "action:list_slice",
                    "node_kind": "data.list_slice",
                    "node_config": {"variable_name": "letters", "start": 1, "end": 3, "output_variable_name": "window"},
                    "ports": [],
                },
                {
                    "node_id": "node-reverse",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-reverse",
                    "expansion_role": "action:list_reverse",
                    "node_kind": "data.list_reverse",
                    "node_config": {"variable_name": "letters"},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-insert"]["items"] == ["c", "z", "a", "b"]
    assert session["result"]["outputs"]["node-remove"]["items"] == ["c", "z", "b"]
    assert session["result"]["outputs"]["node-sort"]["items"] == ["b", "c", "z"]
    assert session["result"]["outputs"]["node-slice"]["value"] == ["c", "z"]
    assert session["result"]["outputs"]["node-reverse"]["items"] == ["z", "c", "b"]
    assert session["result"]["variables"]["letters"] == ["z", "c", "b"]
    assert session["result"]["variables"]["window"] == ["c", "z"]


def test_service_runtime_executes_p9_data_batch_numeric_and_list_index_components() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-batch",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-batch",
                    "expansion_role": "action:set_variables_batch",
                    "node_kind": "data.set_variables_batch",
                    "node_config": {
                        "variables": {
                            "counter": 5,
                            "letters": ["a", "b", "c"],
                        },
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-increment",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-increment",
                    "expansion_role": "action:increment_variable",
                    "node_kind": "data.increment_variable",
                    "node_config": {
                        "variable_name": "counter",
                        "step": 2,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-decrement",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-decrement",
                    "expansion_role": "action:decrement_variable",
                    "node_kind": "data.decrement_variable",
                    "node_config": {
                        "variable_name": "counter",
                        "step": 1,
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-index",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-index",
                    "expansion_role": "action:list_index",
                    "node_kind": "data.list_index",
                    "node_config": {
                        "variable_name": "letters",
                        "value": "b",
                        "output_variable_name": "letter_index",
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert session["result"]["outputs"]["node-batch"]["variable_names"] == ["counter", "letters"]
    assert session["result"]["outputs"]["node-increment"]["value"] == 7
    assert session["result"]["outputs"]["node-decrement"]["value"] == 6
    assert session["result"]["outputs"]["node-index"]["value"] == 1
    assert session["result"]["variables"]["counter"] == 6
    assert session["result"]["variables"]["letter_index"] == 1


def test_service_runtime_executes_browser_atomic_components(tmp_path) -> None:
    service = CompilationWorkbenchService()
    site_server, site_thread = _start_browser_mock_site()
    screenshot_path = tmp_path / "browser-shot.png"
    upload_file_path = tmp_path / "upload.txt"
    upload_file_path.write_text("upload-content", encoding="utf-8")

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {
                            "url": f"http://127.0.0.1:{site_server.server_address[1]}/",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-fill",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-fill",
                        "expansion_role": "action:fill",
                        "node_kind": "browser.fill",
                        "node_config": {
                            "selector": "#name",
                            "value": "Alice",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-check",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-check",
                        "expansion_role": "action:check",
                        "node_kind": "browser.check",
                        "node_config": {
                            "selector": "#agree",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-uncheck",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-uncheck",
                        "expansion_role": "action:uncheck",
                        "node_kind": "browser.uncheck",
                        "node_config": {
                            "selector": "#agree",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-upload",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-upload",
                        "expansion_role": "action:set_input_files",
                        "node_kind": "browser.set_input_files",
                        "node_config": {
                            "selector": "#upload-file",
                            "path": str(upload_file_path),
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-upload-verify",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-upload-verify",
                        "expansion_role": "action:run_js",
                        "node_kind": "browser.run_js",
                        "node_config": {
                            "script": "(() => { const input = document.querySelector('#upload-file'); return JSON.stringify({ checked: document.querySelector('#agree').checked, fileName: input.files && input.files[0] ? input.files[0].name : '' }); })()",
                            "variable_name": "browser_atomic_state",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-click",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-click",
                        "expansion_role": "action:click",
                        "node_kind": "browser.click",
                        "node_config": {
                            "selector": "#submit",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-shot",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-shot",
                        "expansion_role": "action:screenshot",
                        "node_kind": "browser.screenshot",
                        "node_config": {
                            "path": str(screenshot_path),
                        },
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-nav"]["page_url"].startswith("http://127.0.0.1:")
        assert session["result"]["outputs"]["node-fill"]["value"] == "Alice"
        assert session["result"]["outputs"]["node-check"]["selector"] == "#agree"
        assert session["result"]["outputs"]["node-uncheck"]["selector"] == "#agree"
        assert session["result"]["outputs"]["node-upload"]["path"] == str(upload_file_path.resolve())
        assert json.loads(session["result"]["variables"]["browser_atomic_state"]) == {
            "checked": False,
            "fileName": "upload.txt",
        }
        assert session["result"]["outputs"]["node-click"]["page_url"].startswith("http://127.0.0.1:")
        assert session["result"]["outputs"]["node-shot"]["path"] == str(screenshot_path.resolve())
        assert session["result"]["outputs"]["node-shot"]["bytes_written"] > 0
        assert screenshot_path.exists() is True
        assert _BrowserMockSiteHandler.clicked is True
        assert _BrowserMockSiteHandler.last_form_value == "Alice"
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_executes_browser_get_local_storage_component() -> None:
    service = CompilationWorkbenchService()
    site_server, site_thread = _start_browser_mock_site()

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {
                            "url": f"http://127.0.0.1:{site_server.server_address[1]}/",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-seed-storage",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-seed-storage",
                        "expansion_role": "action:run_js",
                        "node_kind": "browser.run_js",
                        "node_config": {
                            "script": "(() => { window.localStorage.setItem('authToken', 'token-123'); return true; })()",
                            "variable_name": "seeded",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-read-storage",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-read-storage",
                        "expansion_role": "action:get_local_storage",
                        "node_kind": "browser.get_local_storage",
                        "node_config": {
                            "key": "authToken",
                            "variable_name": "auth_token",
                            "default_value": "missing",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-read-storage-default",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-read-storage-default",
                        "expansion_role": "action:get_local_storage",
                        "node_kind": "browser.get_local_storage",
                        "node_config": {
                            "key": "missingKey",
                            "variable_name": "missing_token",
                            "default_value": "fallback",
                        },
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-read-storage"]["value"] == "token-123"
        assert session["result"]["outputs"]["node-read-storage"]["key"] == "authToken"
        assert session["result"]["variables"]["auth_token"] == "token-123"
        assert session["result"]["outputs"]["node-read-storage-default"]["value"] == "fallback"
        assert session["result"]["variables"]["missing_token"] == "fallback"
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_executes_browser_history_refresh_and_current_time_components() -> None:
    service = CompilationWorkbenchService()
    site_server, site_thread = _start_browser_mock_site()

    try:
        base_url = f"http://127.0.0.1:{site_server.server_address[1]}"
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav-home",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav-home",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {"url": f"{base_url}/"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-click-dashboard",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-click-dashboard",
                        "expansion_role": "action:click",
                        "node_kind": "browser.click",
                        "node_config": {"selector": "#go-dashboard"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-wait-dashboard",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-wait-dashboard",
                        "expansion_role": "action:wait_for_navigation",
                        "node_kind": "browser.wait_for_navigation",
                        "node_config": {"url_pattern": "/dashboard", "timeout": 3000},
                        "ports": [],
                    },
                    {
                        "node_id": "node-go-back",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-go-back",
                        "expansion_role": "action:go_back",
                        "node_kind": "browser.go_back",
                        "node_config": {},
                        "ports": [],
                    },
                    {
                        "node_id": "node-go-forward",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-go-forward",
                        "expansion_role": "action:go_forward",
                        "node_kind": "browser.go_forward",
                        "node_config": {},
                        "ports": [],
                    },
                    {
                        "node_id": "node-refresh",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-refresh",
                        "expansion_role": "action:refresh",
                        "node_kind": "browser.refresh",
                        "node_config": {},
                        "ports": [],
                    },
                    {
                        "node_id": "node-refresh-no-cache",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-refresh-no-cache",
                        "expansion_role": "action:refresh_no_cache",
                        "node_kind": "browser.refresh_no_cache",
                        "node_config": {},
                        "ports": [],
                    },
                    {
                        "node_id": "node-now",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-now",
                        "expansion_role": "action:get_current_time",
                        "node_kind": "time.get_current_time",
                        "node_config": {
                            "variable_name": "current_time_value",
                            "format": "iso",
                            "timezone": "utc",
                        },
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-go-back"]["page_url"] == f"{base_url}/"
        assert session["result"]["outputs"]["node-go-forward"]["page_url"].endswith("/dashboard")
        assert session["result"]["outputs"]["node-refresh"]["page_url"].endswith("/dashboard")
        assert session["result"]["outputs"]["node-refresh-no-cache"]["page_url"].endswith("/dashboard")
        assert session["result"]["outputs"]["node-now"]["timezone"] == "utc"
        assert session["result"]["outputs"]["node-now"]["format"] == "iso"
        assert isinstance(session["result"]["outputs"]["node-now"]["value"], str)
        assert session["result"]["variables"]["current_time_value"] == session["result"]["outputs"]["node-now"]["value"]
        assert "T" in session["result"]["outputs"]["node-now"]["value"]
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_executes_phase14_browser_storage_wait_keyboard_and_probe_components(
    tmp_path,
) -> None:
    project_dir = tmp_path / "phase14-browser-project"
    project_path = project_dir / "phase14-browser.weconduct.json"
    service = CompilationWorkbenchService()
    service.create_project(project_name="Phase14Browser", project_directory=project_dir)
    site_server, site_thread = _start_browser_mock_site()
    element_shot_rel = r"artifacts\table-element.png"

    try:
        base_url = f"http://127.0.0.1:{site_server.server_address[1]}"
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {"node_id": "node-nav", "lowered_kind": "execution", "source_anchor_ref": "n-nav", "expansion_role": "action:navigate", "node_kind": "browser.navigate", "node_config": {"url": f"{base_url}/"}, "ports": []},
                    {"node_id": "node-set-cookie", "lowered_kind": "execution", "source_anchor_ref": "n-set-cookie", "expansion_role": "action:set_cookie", "node_kind": "browser.set_cookie", "node_config": {"name": "phase14", "value": "cookie-123", "url": f"{base_url}/"}, "ports": []},
                    {"node_id": "node-list-cookies", "lowered_kind": "execution", "source_anchor_ref": "n-list-cookies", "expansion_role": "action:list_cookies", "node_kind": "browser.list_cookies", "node_config": {"url": f"{base_url}/", "variable_name": "cookies_before_delete"}, "ports": []},
                    {"node_id": "node-get-cookie", "lowered_kind": "execution", "source_anchor_ref": "n-get-cookie", "expansion_role": "action:get_cookie", "node_kind": "browser.get_cookie", "node_config": {"name": "phase14", "url": f"{base_url}/", "variable_name": "phase14_cookie"}, "ports": []},
                    {"node_id": "node-set-local-storage", "lowered_kind": "execution", "source_anchor_ref": "n-set-local-storage", "expansion_role": "action:set_local_storage", "node_kind": "browser.set_local_storage", "node_config": {"key": "authToken", "value": "token-123"}, "ports": []},
                    {"node_id": "node-get-local-storage", "lowered_kind": "execution", "source_anchor_ref": "n-get-local-storage", "expansion_role": "action:get_local_storage", "node_kind": "browser.get_local_storage", "node_config": {"key": "authToken", "variable_name": "auth_token"}, "ports": []},
                    {"node_id": "node-set-session-storage", "lowered_kind": "execution", "source_anchor_ref": "n-set-session-storage", "expansion_role": "action:set_session_storage", "node_kind": "browser.set_session_storage", "node_config": {"key": "sessionToken", "value": "session-123"}, "ports": []},
                    {"node_id": "node-get-session-storage", "lowered_kind": "execution", "source_anchor_ref": "n-get-session-storage", "expansion_role": "action:get_session_storage", "node_kind": "browser.get_session_storage", "node_config": {"key": "sessionToken", "variable_name": "session_token"}, "ports": []},
                    {"node_id": "node-keyboard-type", "lowered_kind": "execution", "source_anchor_ref": "n-keyboard-type", "expansion_role": "action:keyboard_type", "node_kind": "browser.keyboard_type", "node_config": {"selector": "#key-input", "text": "Alice", "delay_ms": 0}, "ports": []},
                    {"node_id": "node-wait-value", "lowered_kind": "execution", "source_anchor_ref": "n-wait-value", "expansion_role": "action:wait_for_value", "node_kind": "browser.wait_for_value", "node_config": {"selector": "#key-input", "value": "Alice", "timeout": 3000}, "ports": []},
                    {"node_id": "node-press-enter", "lowered_kind": "execution", "source_anchor_ref": "n-press-enter", "expansion_role": "action:press_key", "node_kind": "browser.press_key", "node_config": {"selector": "#key-input", "key": "Enter"}, "ports": []},
                    {"node_id": "node-hotkey", "lowered_kind": "execution", "source_anchor_ref": "n-hotkey", "expansion_role": "action:hotkey", "node_kind": "browser.hotkey", "node_config": {"selector": "#key-input", "combo": "Control+A"}, "ports": []},
                    {"node_id": "node-hover", "lowered_kind": "execution", "source_anchor_ref": "n-hover", "expansion_role": "action:hover", "node_kind": "browser.hover", "node_config": {"selector": "#hover-target"}, "ports": []},
                    {"node_id": "node-wait-text", "lowered_kind": "execution", "source_anchor_ref": "n-wait-text", "expansion_role": "action:wait_for_text", "node_kind": "browser.wait_for_text", "node_config": {"selector": "#hover-result", "text": "hovered", "match_mode": "contains", "timeout": 3000}, "ports": []},
                    {"node_id": "node-check", "lowered_kind": "execution", "source_anchor_ref": "n-check", "expansion_role": "action:check", "node_kind": "browser.check", "node_config": {"selector": "#agree"}, "ports": []},
                    {"node_id": "node-is-checked", "lowered_kind": "execution", "source_anchor_ref": "n-is-checked", "expansion_role": "action:is_checked", "node_kind": "browser.is_checked", "node_config": {"selector": "#agree", "variable_name": "agree_checked"}, "ports": []},
                    {"node_id": "node-click-fetch", "lowered_kind": "execution", "source_anchor_ref": "n-click-fetch", "expansion_role": "action:click", "node_kind": "browser.click", "node_config": {"selector": "#fetch-button"}, "ports": []},
                    {"node_id": "node-wait-request", "lowered_kind": "execution", "source_anchor_ref": "n-wait-request", "expansion_role": "action:wait_for_request", "node_kind": "browser.wait_for_request", "node_config": {"url_pattern": "/api/ping", "method": "POST", "timeout": 3000}, "ports": []},
                    {"node_id": "node-wait-response", "lowered_kind": "execution", "source_anchor_ref": "n-wait-response", "expansion_role": "action:wait_for_response", "node_kind": "browser.wait_for_response", "node_config": {"url_pattern": "/api/ping", "status_code": 200, "timeout": 3000}, "ports": []},
                    {"node_id": "node-wait-attribute", "lowered_kind": "execution", "source_anchor_ref": "n-wait-attribute", "expansion_role": "action:wait_for_attribute", "node_kind": "browser.wait_for_attribute", "node_config": {"selector": "#fetch-status", "attribute": "data-response", "value": "ok", "timeout": 3000}, "ports": []},
                    {"node_id": "node-exists", "lowered_kind": "execution", "source_anchor_ref": "n-exists", "expansion_role": "action:exists", "node_kind": "browser.exists", "node_config": {"selector": "#hover-result", "variable_name": "hover_exists"}, "ports": []},
                    {"node_id": "node-is-visible", "lowered_kind": "execution", "source_anchor_ref": "n-is-visible", "expansion_role": "action:is_visible", "node_kind": "browser.is_visible", "node_config": {"selector": "#hover-result", "variable_name": "hover_visible"}, "ports": []},
                    {"node_id": "node-is-enabled", "lowered_kind": "execution", "source_anchor_ref": "n-is-enabled", "expansion_role": "action:is_enabled", "node_kind": "browser.is_enabled", "node_config": {"selector": "#submit", "variable_name": "submit_enabled"}, "ports": []},
                    {"node_id": "node-get-html", "lowered_kind": "execution", "source_anchor_ref": "n-get-html", "expansion_role": "action:get_html", "node_kind": "browser.get_html", "node_config": {"selector": "#sample-table", "variable_name": "table_html"}, "ports": []},
                    {"node_id": "node-get-inner-html", "lowered_kind": "execution", "source_anchor_ref": "n-get-inner-html", "expansion_role": "action:get_inner_html", "node_kind": "browser.get_inner_html", "node_config": {"selector": "#sample-table", "variable_name": "table_inner_html"}, "ports": []},
                    {"node_id": "node-drag-drop", "lowered_kind": "execution", "source_anchor_ref": "n-drag-drop", "expansion_role": "action:drag_and_drop", "node_kind": "browser.drag_and_drop", "node_config": {"source_selector": "#drag-source", "target_selector": "#drop-target"}, "ports": []},
                    {"node_id": "node-wait-drop", "lowered_kind": "execution", "source_anchor_ref": "n-wait-drop", "expansion_role": "action:wait_for_text", "node_kind": "browser.wait_for_text", "node_config": {"selector": "#drop-status", "text": "drag-payload", "match_mode": "contains", "timeout": 3000}, "ports": []},
                    {"node_id": "node-scroll-to", "lowered_kind": "execution", "source_anchor_ref": "n-scroll-to", "expansion_role": "action:scroll_to_element", "node_kind": "browser.scroll_to_element", "node_config": {"selector": "#scroll-target", "block": "center", "inline": "nearest"}, "ports": []},
                    {"node_id": "node-scroll-page", "lowered_kind": "execution", "source_anchor_ref": "n-scroll-page", "expansion_role": "action:scroll_page", "node_kind": "browser.scroll_page", "node_config": {"x": 0, "y": -120, "mode": "by"}, "ports": []},
                    {"node_id": "node-element-shot", "lowered_kind": "execution", "source_anchor_ref": "n-element-shot", "expansion_role": "action:element_screenshot", "node_kind": "browser.element_screenshot", "node_config": {"selector": "#sample-table", "path": element_shot_rel}, "ports": []},
                    {"node_id": "node-remove-local-storage", "lowered_kind": "execution", "source_anchor_ref": "n-remove-local-storage", "expansion_role": "action:remove_local_storage", "node_kind": "browser.remove_local_storage", "node_config": {"key": "authToken"}, "ports": []},
                    {"node_id": "node-clear-local-storage", "lowered_kind": "execution", "source_anchor_ref": "n-clear-local-storage", "expansion_role": "action:clear_local_storage", "node_kind": "browser.clear_local_storage", "node_config": {}, "ports": []},
                    {"node_id": "node-delete-cookie", "lowered_kind": "execution", "source_anchor_ref": "n-delete-cookie", "expansion_role": "action:delete_cookie", "node_kind": "browser.delete_cookie", "node_config": {"name": "phase14"}, "ports": []},
                    {"node_id": "node-list-cookies-after-delete", "lowered_kind": "execution", "source_anchor_ref": "n-list-cookies-after-delete", "expansion_role": "action:list_cookies", "node_kind": "browser.list_cookies", "node_config": {"url": f"{base_url}/", "variable_name": "cookies_after_delete"}, "ports": []},
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )
        service.save_project_as(project_path=project_path)

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        expected_element_path = project_dir / "artifacts" / "table-element.png"

        assert session["status"] == "completed"
        assert session["result"]["outputs"]["node-get-cookie"]["value"] == "cookie-123"
        assert session["result"]["variables"]["phase14_cookie"] == "cookie-123"
        assert session["result"]["outputs"]["node-get-local-storage"]["value"] == "token-123"
        assert session["result"]["variables"]["auth_token"] == "token-123"
        assert session["result"]["outputs"]["node-get-session-storage"]["value"] == "session-123"
        assert session["result"]["variables"]["session_token"] == "session-123"
        assert session["result"]["outputs"]["node-wait-value"]["value"] == "Alice"
        assert session["result"]["outputs"]["node-wait-text"]["matched_text"] == "hovered"
        assert session["result"]["outputs"]["node-is-checked"]["value"] is True
        assert session["result"]["outputs"]["node-wait-request"]["method"] == "POST"
        assert session["result"]["outputs"]["node-wait-response"]["status_code"] == 200
        assert session["result"]["outputs"]["node-wait-attribute"]["value"] == "ok"
        assert session["result"]["outputs"]["node-exists"]["value"] is True
        assert session["result"]["outputs"]["node-is-visible"]["value"] is True
        assert session["result"]["outputs"]["node-is-enabled"]["value"] is True
        assert "<table" in session["result"]["outputs"]["node-get-html"]["value"]
        assert "<tbody>" in session["result"]["outputs"]["node-get-inner-html"]["value"]
        assert session["result"]["outputs"]["node-wait-drop"]["matched_text"] == "drag-payload"
        assert session["result"]["outputs"]["node-element-shot"]["path"] == str(expected_element_path.resolve())
        assert expected_element_path.exists() is True
        assert session["result"]["outputs"]["node-list-cookies"]["cookie_count"] >= 1
        assert session["result"]["outputs"]["node-list-cookies-after-delete"]["cookie_count"] == 0
        assert session["result"]["variables"]["agree_checked"] is True
        assert session["result"]["variables"]["hover_exists"] is True
        assert session["result"]["variables"]["hover_visible"] is True
        assert session["result"]["variables"]["submit_enabled"] is True
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_executes_phase14_browser_tabs_popup_download_and_context_components(
    tmp_path,
) -> None:
    project_dir = tmp_path / "phase14-context-project"
    project_path = project_dir / "phase14-context.weconduct.json"
    service = CompilationWorkbenchService()
    service.create_project(project_name="Phase14Context", project_directory=project_dir)
    site_server, site_thread = _start_browser_mock_site()

    try:
        base_url = f"http://127.0.0.1:{site_server.server_address[1]}"
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {"node_id": "node-set-user-agent", "lowered_kind": "execution", "source_anchor_ref": "n-set-user-agent", "expansion_role": "action:set_user_agent", "node_kind": "browser.set_user_agent", "node_config": {"user_agent": "WeConductPhase14/1.0"}, "ports": []},
                    {"node_id": "node-set-extra-headers", "lowered_kind": "execution", "source_anchor_ref": "n-set-extra-headers", "expansion_role": "action:set_extra_headers", "node_kind": "browser.set_extra_headers", "node_config": {"headers": {"X-WeConduct-Extra": "phase14"}}, "ports": []},
                    {"node_id": "node-nav-headers", "lowered_kind": "execution", "source_anchor_ref": "n-nav-headers", "expansion_role": "action:navigate", "node_kind": "browser.navigate", "node_config": {"url": f"{base_url}/headers"}, "ports": []},
                    {"node_id": "node-read-headers", "lowered_kind": "execution", "source_anchor_ref": "n-read-headers", "expansion_role": "action:run_js", "node_kind": "browser.run_js", "node_config": {"script": "JSON.parse(document.getElementById('header-json').textContent)", "variable_name": "header_payload"}, "ports": []},
                    {"node_id": "node-nav-root", "lowered_kind": "execution", "source_anchor_ref": "n-nav-root", "expansion_role": "action:navigate", "node_kind": "browser.navigate", "node_config": {"url": f"{base_url}/"}, "ports": []},
                    {"node_id": "node-click-go", "lowered_kind": "execution", "source_anchor_ref": "n-click-go", "expansion_role": "action:click", "node_kind": "browser.click", "node_config": {"selector": "#go-dashboard"}, "ports": []},
                    {"node_id": "node-wait-url-change", "lowered_kind": "execution", "source_anchor_ref": "n-wait-url-change", "expansion_role": "action:wait_for_url_change", "node_kind": "browser.wait_for_url_change", "node_config": {"from_url": f"{base_url}/", "url_pattern": "/dashboard", "timeout": 3000}, "ports": []},
                    {"node_id": "node-open-tab", "lowered_kind": "execution", "source_anchor_ref": "n-open-tab", "expansion_role": "action:open_tab", "node_kind": "browser.open_tab", "node_config": {"url": f"{base_url}/dashboard", "label": "dashboard-tab", "activate": True}, "ports": []},
                    {"node_id": "node-switch-main", "lowered_kind": "execution", "source_anchor_ref": "n-switch-main", "expansion_role": "action:switch_tab", "node_kind": "browser.switch_tab", "node_config": {"index": 0}, "ports": []},
                    {"node_id": "node-click-popup", "lowered_kind": "execution", "source_anchor_ref": "n-click-popup", "expansion_role": "action:click", "node_kind": "browser.click", "node_config": {"selector": "#open-popup"}, "ports": []},
                    {"node_id": "node-wait-popup", "lowered_kind": "execution", "source_anchor_ref": "n-wait-popup", "expansion_role": "action:wait_for_popup", "node_kind": "browser.wait_for_popup", "node_config": {"timeout": 3000, "activate": True, "variable_name": "popup_record"}, "ports": []},
                    {"node_id": "node-close-tab", "lowered_kind": "execution", "source_anchor_ref": "n-close-tab", "expansion_role": "action:close_tab", "node_kind": "browser.close_tab", "node_config": {"current": True}, "ports": []},
                    {"node_id": "node-switch-dashboard", "lowered_kind": "execution", "source_anchor_ref": "n-switch-dashboard", "expansion_role": "action:switch_tab", "node_kind": "browser.switch_tab", "node_config": {"label": "dashboard-tab"}, "ports": []},
                    {"node_id": "node-switch-main-again", "lowered_kind": "execution", "source_anchor_ref": "n-switch-main-again", "expansion_role": "action:switch_tab", "node_kind": "browser.switch_tab", "node_config": {"index": 0}, "ports": []},
                    {"node_id": "node-click-download", "lowered_kind": "execution", "source_anchor_ref": "n-click-download", "expansion_role": "action:click", "node_kind": "browser.click", "node_config": {"selector": "#download-link"}, "ports": []},
                    {"node_id": "node-wait-download", "lowered_kind": "execution", "source_anchor_ref": "n-wait-download", "expansion_role": "action:wait_for_download", "node_kind": "browser.wait_for_download", "node_config": {"path": r"artifacts\browser-download.txt", "timeout": 3000, "variable_name": "download_record"}, "ports": []},
                    {"node_id": "node-download-file", "lowered_kind": "execution", "source_anchor_ref": "n-download-file", "expansion_role": "action:download_file", "node_kind": "browser.download_file", "node_config": {"url": f"{base_url}/download", "path": r"artifacts\direct-download.txt"}, "ports": []},
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )
        service.save_project_as(project_path=project_path)

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        wait_download_path = project_dir / "artifacts" / "browser-download.txt"
        direct_download_path = project_dir / "artifacts" / "direct-download.txt"

        assert session["status"] == "completed"
        assert session["result"]["variables"]["header_payload"]["user_agent"] == "WeConductPhase14/1.0"
        assert session["result"]["variables"]["header_payload"]["x_weconduct_extra"] == "phase14"
        assert session["result"]["outputs"]["node-wait-url-change"]["matched_url"].endswith("/dashboard")
        assert session["result"]["outputs"]["node-open-tab"]["page_url"].endswith("/dashboard")
        assert session["result"]["outputs"]["node-switch-main"]["page_index"] == 0
        assert session["result"]["outputs"]["node-wait-popup"]["page_url"].endswith("/popup")
        assert session["result"]["variables"]["popup_record"]["page_url"].endswith("/popup")
        assert session["result"]["outputs"]["node-close-tab"]["closed"] is True
        assert session["result"]["outputs"]["node-switch-dashboard"]["label"] == "dashboard-tab"
        assert session["result"]["outputs"]["node-wait-download"]["path"] == str(wait_download_path.resolve())
        assert session["result"]["outputs"]["node-download-file"]["path"] == str(direct_download_path.resolve())
        assert wait_download_path.read_text(encoding="utf-8") == "phase14-download"
        assert direct_download_path.read_text(encoding="utf-8") == "phase14-download"
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_executes_p9_browser_extraction_js_and_table_components(tmp_path) -> None:
    service = CompilationWorkbenchService()
    site_server, site_thread = _start_browser_mock_site()
    workbook_path = tmp_path / "web-table.xlsx"

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {"url": f"http://127.0.0.1:{site_server.server_address[1]}/"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-get-text",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-get-text",
                        "expansion_role": "action:get_text",
                        "node_kind": "data.get_text",
                        "node_config": {"selector": "#status", "variable_name": "status_text"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-get-attr",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-get-attr",
                        "expansion_role": "action:get_attribute",
                        "node_kind": "data.get_attribute",
                        "node_config": {"selector": "#name", "attribute": "name", "variable_name": "name_attr"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-get-value",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-get-value",
                        "expansion_role": "action:get_value",
                        "node_kind": "data.get_value",
                        "node_config": {"selector": "#name", "variable_name": "name_value"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-count",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-count",
                        "expansion_role": "action:get_element_count",
                        "node_kind": "data.get_element_count",
                        "node_config": {"selector": "#city option", "variable_name": "option_count"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-inject-js",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-inject-js",
                        "expansion_role": "action:inject_js",
                        "node_kind": "browser.inject_js",
                        "node_config": {"script": "window.weconductInjected = 'ok';"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-run-js",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-run-js",
                        "expansion_role": "action:run_js",
                        "node_kind": "browser.run_js",
                        "node_config": {
                            "script": "window.weconductInjected + ':' + document.querySelector('#status').textContent",
                            "variable_name": "js_value",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-web-table",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-web-table",
                        "expansion_role": "action:extract_web_table",
                        "node_kind": "browser.extract_web_table",
                        "node_config": {"selector": "#sample-table", "variable_name": "web_rows"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-web-table-excel",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-web-table-excel",
                        "expansion_role": "action:extract_web_table_to_excel",
                        "node_kind": "browser.extract_web_table_to_excel",
                        "node_config": {
                            "selector": "#sample-table",
                            "path": str(workbook_path),
                            "sheet_name": "WebTable",
                        },
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert session["result"]["variables"]["status_text"] == "ready"
        assert session["result"]["variables"]["name_attr"] == "name"
        assert session["result"]["variables"]["name_value"] == ""
        assert session["result"]["variables"]["option_count"] == 3
        assert session["result"]["variables"]["js_value"] == "ok:ready"
        assert session["result"]["outputs"]["node-web-table"]["headers"] == ["Name", "Score"]
        assert session["result"]["outputs"]["node-web-table"]["rows"] == [
            {"Name": "Alice", "Score": "95"},
            {"Name": "Bob", "Score": "88"},
        ]
        assert session["result"]["outputs"]["node-web-table-excel"]["row_count"] == 2

        workbook = load_workbook(workbook_path)
        try:
            worksheet = workbook["WebTable"]
            assert worksheet["A1"].value == "Name"
            assert worksheet["B2"].value == "95"
        finally:
            workbook.close()
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_executes_p9_session_apply_auth_session_and_dialog_components() -> None:
    service = CompilationWorkbenchService()
    site_server, site_thread = _start_browser_mock_site()
    origin = f"http://127.0.0.1:{site_server.server_address[1]}"

    try:
        service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [
                    {
                        "node_id": "node-nav",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-nav",
                        "expansion_role": "action:navigate",
                        "node_kind": "browser.navigate",
                        "node_config": {"url": origin + "/"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-auth",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-auth",
                        "expansion_role": "action:apply_auth_session",
                        "node_kind": "session.apply_auth_session",
                        "node_config": {
                            "cookies": [
                                {
                                    "name": "session_id",
                                    "value": "abc123",
                                    "url": origin + "/",
                                }
                            ],
                            "local_storage": {
                                origin: {
                                    "auth_token": "local-xyz",
                                }
                            },
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-auth-check",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-auth-check",
                        "expansion_role": "action:run_js",
                        "node_kind": "browser.run_js",
                        "node_config": {
                            "script": "document.cookie + '|' + window.localStorage.getItem('auth_token')",
                            "variable_name": "auth_state",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-dialog-config",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-dialog-config",
                        "expansion_role": "action:set_agent_config",
                        "node_kind": "dialog.set_agent_config",
                        "node_config": {
                            "default_action": "accept",
                            "prompt_text": "typed-by-runtime",
                        },
                        "ports": [],
                    },
                    {
                        "node_id": "node-dialog-mode",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-dialog-mode",
                        "expansion_role": "action:switch_dialog_mode",
                        "node_kind": "dialog.switch_dialog_mode",
                        "node_config": {"mode": "auto"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-click-alert",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-click-alert",
                        "expansion_role": "action:click",
                        "node_kind": "browser.click",
                        "node_config": {"selector": "#alert-button"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-watch-dialogs",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-watch-dialogs",
                        "expansion_role": "action:watch_dialogs",
                        "node_kind": "dialog.watch_dialogs",
                        "node_config": {"timeout": 50, "variable_name": "dialog_records"},
                        "ports": [],
                    },
                    {
                        "node_id": "node-handle-dialogs",
                        "lowered_kind": "execution",
                        "source_anchor_ref": "n-handle-dialogs",
                        "expansion_role": "action:handle_dialogs",
                        "node_kind": "dialog.handle_dialogs",
                        "node_config": {"clear_after": True},
                        "ports": [],
                    },
                ],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            }
        )

        started = service.start_runtime_session(None)
        session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

        assert session["status"] == "completed"
        assert "session_id=abc123" in session["result"]["variables"]["auth_state"]
        assert session["result"]["variables"]["auth_state"].endswith("|local-xyz")
        assert session["result"]["outputs"]["node-auth"]["cookie_count"] == 1
        assert session["result"]["outputs"]["node-dialog-config"]["default_action"] == "accept"
        assert session["result"]["outputs"]["node-dialog-mode"]["mode"] == "auto"
        assert session["result"]["outputs"]["node-watch-dialogs"]["dialog_count"] == 1
        assert session["result"]["variables"]["dialog_records"][0]["message"] == "hello-dialog"
        assert session["result"]["outputs"]["node-handle-dialogs"]["handled_count"] == 1
    finally:
        site_server.shutdown()
        site_server.server_close()


def test_service_runtime_closes_runtime_context_after_completed_session(monkeypatch) -> None:
    service = CompilationWorkbenchService()
    close_calls: list[str] = []

    def patched_close(self) -> None:
        close_calls.append("closed")

    monkeypatch.setattr("weconduct.application.compilation_workbench_service.RuntimeContext.close", patched_close)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-set",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "ok"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "completed"
    assert close_calls == ["closed"]


def test_service_runtime_closes_runtime_context_after_failed_session(monkeypatch) -> None:
    service = CompilationWorkbenchService()
    close_calls: list[str] = []

    def patched_close(self) -> None:
        close_calls.append("closed")

    monkeypatch.setattr("weconduct.application.compilation_workbench_service.RuntimeContext.close", patched_close)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-get",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-get",
                    "expansion_role": "action:get_variable",
                    "node_kind": "data.get_variable",
                    "node_config": {"name": ""},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    session = service.run_runtime_session(session_id=started["runtime_session"]["session_id"])

    assert session["status"] == "failed"
    assert session["result"]["failure_reason"] == "data.variable_name_required"
    assert close_calls == ["closed"]


def test_service_runtime_and_debug_start_fail_for_empty_workspace_graph_with_diagnostics() -> None:
    service = CompilationWorkbenchService()

    runtime_started = service.start_runtime_session(None)
    debug_started = service.start_debug_session(None)

    assert runtime_started["status"] == "failed"
    assert runtime_started["runtime_session"]["status"] == "diagnostic_blocked"
    assert runtime_started["runtime_session"]["execution_supported"] is False
    assert runtime_started["request"]["request_origin"] == "saved_graph_document"
    assert runtime_started["request"]["requested_graph_model_id"] == "graph:workspace"
    assert runtime_started["request"]["compile_status"] == "failed"
    assert runtime_started["diagnostics"]["total_count"] == 2
    assert runtime_started["diagnostics"]["highest_severity"] == "fatal"
    assert runtime_started["diagnostics"]["entries"][-1]["category"] == "source.empty"

    assert debug_started["status"] == "failed"
    assert debug_started["debug_session"]["status"] == "diagnostic_blocked"
    assert debug_started["request"]["request_origin"] == "saved_graph_document"
    assert debug_started["request"]["requested_graph_model_id"] == "graph:workspace"
    assert debug_started["request"]["compile_status"] == "failed"
    assert debug_started["stage_timeline"][2]["stage"] == "validate"
    assert debug_started["stage_timeline"][2]["status"] == "failed"
    assert debug_started["diagnostic_links"][-1]["category"] == "source.empty"


def test_service_rejects_runtime_run_for_unknown_session_id() -> None:
    service = CompilationWorkbenchService()

    try:
        service.run_runtime_session(session_id="runtime-session-missing")
    except ValueError as exc:
        assert str(exc) == "runtime session not found: runtime-session-missing"
    else:
        raise AssertionError("expected ValueError for missing runtime session")


def test_service_blocks_runtime_session_start_when_required_resource_is_disabled() -> None:
    service = CompilationWorkbenchService()
    service.set_resource_enabled(resource_id="builtin:data.map", enabled=False)
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "display_name": "HTTP Request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "display_name": "Map Result",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_runtime_session(None)
    assert started["status"] == "failed"
    assert started["runtime_session"]["session_id"] is None
    assert started["runtime_session"]["status"] == "diagnostic_blocked"
    assert started["runtime_session"]["execution_supported"] is False
    assert started["runtime_plan"] is None
    assert started["diagnostics"]["entries"][0]["category"] == "graph.node.resource_disabled"
    assert started["diagnostics"]["entries"][0]["stage_extension"]["rule"] == (
        "graph.node.resource_enabled"
    )


def test_runtime_health_exposes_host_session_capabilities_and_entrypoints() -> None:
    service = CompilationWorkbenchService()

    health = service.get_runtime_health()

    assert health["status"] == "ok"
    assert health["service"] == "weconduct-api"
    assert health["host_mode"] == "python_core"
    assert health["api_version"] == "0.4.0"
    assert health["workspace_state_version"] == 1
    assert health["workspace_session_id"].startswith("ws-")
    assert health["service_started_at"]
    assert health["capabilities"]["compiler_available"] is True
    assert health["capabilities"]["graph_workspace_available"] is True
    assert health["capabilities"]["runtime_available"] is True
    assert health["capabilities"]["debug_available"] is True
    assert health["entrypoints"]["snapshot"] == "/api/workbench/snapshot"
    assert health["entrypoints"]["compile_action"] == "/api/workbench/compile"
    assert health["entrypoints"]["graph_document"] == "/api/workbench/graph"
    assert health["entrypoints"]["runtime_prepare_action"] == "/api/workbench/runtime/prepare"
    assert health["entrypoints"]["debug_prepare_action"] == "/api/workbench/debug/prepare"
    assert health["entrypoints"]["host_info"] == "/api/host/info"


def test_workbench_snapshot_exposes_graph_workspace_entrypoint() -> None:
    service = CompilationWorkbenchService()

    snapshot = service.get_workbench_snapshot()

    assert snapshot["capabilities"]["graph_workspace_available"] is True
    assert snapshot["entrypoints"]["graph_document"] == "/api/workbench/graph"
    assert snapshot["entrypoints"]["graph_validate_action"] == "/api/workbench/graph/validate"
    assert snapshot["entrypoints"]["graph_compile_action"] == "/api/workbench/graph/compile"


def test_service_exposes_editable_graph_workspace_document() -> None:
    service = CompilationWorkbenchService()

    result = service.get_graph_document()

    assert result["graph_model"].graph_model_id == "graph:workspace"
    assert result["graph_model"].compilation_id is None
    assert result["graph_model"].graph_schema_version == "graph-v1"
    assert result["graph_model"].nodes == []
    assert result["graph_model"].edges == []
    assert result["view"]["is_editable"] is True
    assert result["view"]["authority_mode"] == "workspace_graph_draft"
    assert result["view"]["compile_source_authority"] == "graph_document"


def test_service_persists_graph_workspace_document_to_shared_store() -> None:
    store = InMemoryWorkspaceStateStore()
    first_service = CompilationWorkbenchService(state_store=store)
    second_service = CompilationWorkbenchService(state_store=store)
    graph_model = GraphModel(
        graph_model_id="graph:workspace",
        compilation_id=None,
        graph_schema_version="graph-v1",
        nodes=[
            GraphNode(
                node_id="node-1",
                lowered_kind="execution",
                source_anchor_ref="n1",
                expansion_role="action:request",
                display_name="HTTP Request",
                node_kind="http.request",
                position=GraphPosition(x=120, y=80),
                ports=[
                    GraphPort(
                        port_id="out-main",
                        direction="output",
                        relation_layer="data",
                        semantic_slot="out.result",
                    )
                ],
            )
        ],
        edges=[
            GraphEdge(
                edge_id="edge-1",
                relation_layer="data",
                from_node_id="node-1",
                to_node_id="node-2",
                from_port_id="out-main",
                to_port_id="in-main",
            )
        ],
    )

    save_result = first_service.save_graph_document(graph_model.model_dump())
    loaded_result = second_service.get_graph_document()

    assert save_result["status"] == "saved"
    assert save_result["view"]["graph_document_save_revision"] == 1
    assert save_result["view"]["graph_document_saved_at"]
    assert loaded_result["graph_model"].nodes[0].display_name == "HTTP Request"
    assert loaded_result["graph_model"].nodes[0].ports[0].port_id == "out-main"
    assert loaded_result["graph_model"].edges[0].from_port_id == "out-main"
    assert loaded_result["view"]["graph_document_save_revision"] == 1
    assert loaded_result["view"]["graph_document_saved_at"]


def test_service_rejects_graph_save_when_expected_revision_is_stale() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    store = InMemoryWorkspaceStateStore()
    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {},
                "compile_settings": {},
                "security_settings": {},
                "python_runtime_settings": {},
                "graph_settings": {
                    "save_conflict_policy": "strict",
                },
                "other_settings": {},
            }
        )
    )
    first_service = CompilationWorkbenchService(
        state_store=store,
        preferences_service=preferences_service,
    )
    second_service = CompilationWorkbenchService(
        state_store=store,
        preferences_service=preferences_service,
    )

    first_service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    try:
        second_service.save_graph_document(
            {
                "graph_model_id": "graph:workspace",
                "compilation_id": None,
                "graph_schema_version": "graph-v1",
                "nodes": [],
                "edges": [],
                "graph_effective_diagnostic_anchor_refs": [],
            },
            expected_graph_document_save_revision=0,
        )
    except ValueError as exc:
        assert str(exc) == (
            "graph document save revision conflict: expected 0, current 1"
        )
    else:
        raise AssertionError("expected stale save revision conflict")


def test_service_prefers_current_graph_when_graph_save_revision_is_stale_and_policy_allows() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    store = InMemoryWorkspaceStateStore()
    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {},
                "compile_settings": {},
                "security_settings": {},
                "python_runtime_settings": {},
                "graph_settings": {
                    "save_conflict_policy": "prefer_current_graph",
                },
                "other_settings": {},
            }
        )
    )
    first_service = CompilationWorkbenchService(
        state_store=store,
        preferences_service=preferences_service,
    )
    second_service = CompilationWorkbenchService(
        state_store=store,
        preferences_service=preferences_service,
    )

    first_service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-first",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-first",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "first"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    save_result = second_service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-second",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-second",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "second"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        },
        expected_graph_document_save_revision=0,
    )
    loaded_result = first_service.get_graph_document()

    assert save_result["status"] == "saved"
    assert save_result["view"]["graph_document_save_revision"] == 2
    assert loaded_result["graph_model"].nodes[0].node_id == "node-second"
    assert loaded_result["graph_model"].nodes[0].node_config["value"] == "second"


def test_service_graph_workspace_view_exposes_save_conflict_policy_preference() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {},
                "compile_settings": {},
                "security_settings": {},
                "python_runtime_settings": {},
                "graph_settings": {
                    "save_conflict_policy": "strict",
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    snapshot = service.get_workbench_snapshot()

    assert snapshot["graph_workspace"]["graph_preferences"]["save_conflict_policy"] == "strict"
    assert (
        snapshot["graph_workspace"]["preferences_state"]["graph_settings"]["save_conflict_policy"]
        == "active"
    )


def test_service_validates_graph_document_and_reports_missing_edge_targets() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                }
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "missing-node",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == "graph.edge.missing_target_node"
    assert validation_result["diagnostics"][0]["stage_extension"] == {
        "subject_ref": "edge-1",
        "action": "validated graph document",
        "rule": "graph.edge.target_node_exists",
        "result": "failed",
        "graph_ref": {
            "graph_model_id": "graph:workspace",
            "edge_id": "edge-1",
            "from_node_id": "node-1",
            "to_node_id": "missing-node",
            "from_port_id": "out",
            "to_port_id": "in",
        },
    }


def test_service_validates_graph_document_and_rejects_unsupported_observe_edges() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "observe-out",
                            "direction": "output",
                            "relation_layer": "observe",
                            "semantic_slot": "out.observe",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "observe-in",
                            "direction": "input",
                            "relation_layer": "observe",
                            "semantic_slot": "in.observe",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-observe",
                    "relation_layer": "observe",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "observe-out",
                    "to_port_id": "observe-in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == "graph.edge.observe_unsupported"


def test_service_validates_graph_document_as_valid_when_nodes_edges_and_ports_match() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "valid"
    assert validation_result["summary"]["error_count"] == 0
    assert validation_result["diagnostics"] == []


def test_service_validates_graph_document_as_invalid_when_node_resource_is_disabled() -> None:
    service = CompilationWorkbenchService()
    service.set_resource_enabled(resource_id="builtin:data.map", enabled=False)

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "transform:map",
                    "display_name": "Map Result",
                    "node_kind": "data.map",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == "graph.node.resource_disabled"
    assert validation_result["diagnostics"][0]["object_ref"] == "node-1"
    assert validation_result["diagnostics"][0]["stage_extension"] == {
        "subject_ref": "node-1",
        "action": "validated graph document",
        "rule": "graph.node.resource_enabled",
        "result": "failed",
        "graph_ref": {
            "graph_model_id": "graph:workspace",
            "node_id": "node-1",
            "node_kind": "data.map",
            "resource_id": "builtin:data.map",
            "resource_status": "disabled",
        },
    }


def test_service_validates_flow_graph_and_reports_unreachable_business_nodes() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "ports": [],
                },
                {
                    "node_id": "node-live",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-live",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "ok"},
                    "ports": [],
                },
                {
                    "node_id": "node-dead",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-dead",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "status", "value": "dead"},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-live",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-live",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "valid"
    assert validation_result["summary"]["error_count"] == 0
    assert validation_result["summary"]["warning_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == "graph.node.unreachable_in_flow_graph"
    assert validation_result["diagnostics"][0]["severity"] == "warning"
    assert validation_result["diagnostics"][0]["object_ref"] == "node-dead"
    assert validation_result["diagnostics"][0]["stage_extension"] == {
        "subject_ref": "node-dead",
        "action": "validated graph document",
        "rule": "graph.node.reachable_from_flow_start",
        "result": "failed",
        "graph_ref": {
            "graph_model_id": "graph:workspace",
            "node_id": "node-dead",
            "entry_node_ids": ["node-start"],
        },
    }


def test_p13_debugweb_project_captcha_branch_is_reachable() -> None:
    service = CompilationWorkbenchService()
    project_path = (
        Path(__file__).resolve().parents[3]
        / "docs"
        / "dev"
        / "phase-13"
        / "artifacts"
        / "p13_d_debugweb_test_project"
        / "p13_d_debugweb_test_project.weconduct.json"
    )

    opened = service.open_project(project_path=project_path)
    graph_document = opened["graph_document"].model_dump(mode="json")
    validation_result = service.validate_graph_document(graph_document)

    assert validation_result["status"] == "valid"
    unreachable_refs = {
        item["object_ref"]
        for item in validation_result["diagnostics"]
        if item.get("category") == "graph.node.unreachable_in_flow_graph"
    }
    assert "node-43503c860c7a" not in unreachable_refs
    assert "node-60cf3b303191" not in unreachable_refs


def test_p13_debugweb_project_progress_nodes_use_runtime_auth_token() -> None:
    service = CompilationWorkbenchService()
    project_path = (
        Path(__file__).resolve().parents[3]
        / "docs"
        / "dev"
        / "phase-13"
        / "artifacts"
        / "p13_d_debugweb_test_project"
        / "p13_d_debugweb_test_project.weconduct.json"
    )

    opened = service.open_project(project_path=project_path)
    graph_document = opened["graph_document"].model_dump(mode="json")
    progress_nodes = [
        node
        for node in graph_document["nodes"]
        if node.get("node_kind") == "http.request"
        and node.get("node_config", {}).get("url") == "${base_url}/api/progress"
    ]

    assert progress_nodes
    for node in progress_nodes:
        assert (
            node.get("node_config", {})
            .get("headers", {})
            .get("Authorization")
            == "Bearer ${auth_token}"
        )

def test_service_validates_graph_call_subgraph_and_reports_missing_subgraph_id() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == "graph.call_subgraph.subgraph_id_required"
    assert validation_result["diagnostics"][0]["object_ref"] == "node-call"
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call",
        "node_kind": "graph.call_subgraph",
    }


def test_service_validates_graph_call_subgraph_and_reports_missing_resource_or_invalid_mappings() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:subgraph",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_subgraph = service.save_subgraph_resource(resource_name="Validation Target")

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-missing",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-missing",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Missing Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": "subgraph_resource:missing",
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-call-invalid-inputs",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-invalid-inputs",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Invalid Inputs",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                        "inputs": [],
                    },
                    "ports": [],
                },
                {
                    "node_id": "node-call-invalid-outputs",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-invalid-outputs",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Invalid Outputs",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                        "outputs": "bad",
                    },
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 3
    categories = [item["category"] for item in validation_result["diagnostics"]]
    assert categories == [
        "graph.call_subgraph.subgraph_missing",
        "graph.call_subgraph.input_mapping_invalid",
        "graph.call_subgraph.output_mapping_invalid",
    ]
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-missing",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": "subgraph_resource:missing",
    }
    assert validation_result["diagnostics"][1]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-invalid-inputs",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "mapping_field": "inputs",
    }
    assert validation_result["diagnostics"][2]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-invalid-outputs",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "mapping_field": "outputs",
    }


def test_service_validates_flow_start_constraints_and_reports_multiple_entries() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "start-a",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start-a",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "ports": [],
                },
                {
                    "node_id": "start-b",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start-b",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == "graph.flow_start.invalid_entry_count"
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "entry_node_ids": ["start-a", "start-b"],
        "entry_count": 2,
    }


def test_service_validates_flow_start_constraints_and_reports_control_inputs() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "ports": [
                        {
                            "port_id": "boot",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        }
                    ],
                },
                {
                    "node_id": "source",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-source",
                    "expansion_role": "control:fork",
                    "node_kind": "control.parallel_fork",
                    "ports": [
                        {
                            "port_id": "branch:a",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:a",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-source-start",
                    "relation_layer": "control",
                    "from_node_id": "source",
                    "to_node_id": "start",
                    "from_port_id": "branch:a",
                    "to_port_id": "boot",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    categories = [item["category"] for item in validation_result["diagnostics"]]
    assert "graph.flow_start.control_input_forbidden" in categories


def test_service_validates_edge_relation_layer_matches_connected_ports() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-a",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-a",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.value",
                        }
                    ],
                },
                {
                    "node_id": "node-b",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-b",
                    "expansion_role": "transform:map",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-a-b",
                    "relation_layer": "control",
                    "from_node_id": "node-a",
                    "to_node_id": "node-b",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 2
    assert validation_result["diagnostics"][0]["category"] == "graph.node.parameter_blank_required"
    assert validation_result["diagnostics"][1]["category"] == "graph.edge.relation_layer_mismatch"
    assert validation_result["diagnostics"][1]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "edge_id": "edge-a-b",
        "from_node_id": "node-a",
        "to_node_id": "node-b",
        "from_port_id": "out",
        "to_port_id": "in",
        "edge_relation_layer": "control",
        "from_port_relation_layer": "data",
        "to_port_relation_layer": "data",
    }


def test_service_validates_input_port_max_connections() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "source-a",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-source-a",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.value",
                        }
                    ],
                },
                {
                    "node_id": "source-b",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-source-b",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.value",
                        }
                    ],
                },
                {
                    "node_id": "target",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-target",
                    "expansion_role": "transform:map",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                            "max_connections": 1,
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-a-target",
                    "relation_layer": "data",
                    "from_node_id": "source-a",
                    "to_node_id": "target",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-b-target",
                    "relation_layer": "data",
                    "from_node_id": "source-b",
                    "to_node_id": "target",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 3
    assert validation_result["diagnostics"][0]["category"] == "graph.node.parameter_blank_required"
    assert validation_result["diagnostics"][1]["category"] == "graph.node.parameter_blank_required"
    assert validation_result["diagnostics"][2]["category"] == "graph.port.max_connections_exceeded"
    assert validation_result["diagnostics"][2]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "target",
        "port_id": "in",
        "relation_layer": "data",
        "max_connections": 1,
        "connection_count": 2,
        "edge_ids": ["edge-a-target", "edge-b-target"],
    }


def test_service_validates_control_if_shape_and_requires_condition_source() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-if",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-if",
                    "expansion_role": "control:if",
                    "node_kind": "control.if",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "true",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.true",
                        },
                    ],
                    "node_config": {},
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    categories = {item["category"] for item in validation_result["diagnostics"]}
    assert "graph.control.if.missing_required_port" in categories
    assert "graph.control.if.condition_missing" in categories


def test_service_validates_parallel_fork_join_while_retry_and_failover_shapes() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch:a",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:a",
                        },
                    ],
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "ports": [
                        {
                            "port_id": "in:a",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:a",
                        }
                    ],
                    "node_config": {"mode": "quorum", "quorum": 3},
                },
                {
                    "node_id": "node-while",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-while",
                    "expansion_role": "control:while",
                    "node_kind": "control.while",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "repeat",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.repeat",
                        },
                        {
                            "port_id": "loop",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.loop",
                        },
                    ],
                    "node_config": {},
                },
                {
                    "node_id": "node-retry",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-retry",
                    "expansion_role": "control:retry",
                    "node_kind": "control.retry",
                    "ports": [],
                    "node_config": {"max_attempts": 0},
                },
                {
                    "node_id": "node-failover",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-failover",
                    "expansion_role": "control:failover",
                    "node_kind": "control.failover",
                    "ports": [
                        {
                            "port_id": "failed",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.failed",
                        }
                    ],
                    "node_config": {},
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    categories = {item["category"] for item in validation_result["diagnostics"]}
    assert "graph.control.parallel_fork.branch_count_invalid" in categories
    assert "graph.control.join.branch_count_invalid" in categories
    assert "graph.control.join.quorum_invalid" in categories
    assert "graph.control.while.missing_required_port" in categories
    assert "graph.control.while.condition_missing" in categories
    assert "graph.control.retry.max_attempts_invalid" in categories
    assert "graph.control.failover.branch_count_invalid" in categories


def test_service_get_graph_document_normalizes_parallel_fork_and_join_ports_from_branches() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "resource_language": "zh-CN",
                    "theme": "light",
                    "default_window_size": {"width": 1440, "height": 900},
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                    "preferences_auto_save": True,
                    "font_scale": 100,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                    "stop_on_first_error": True,
                    "emit_runtime_plan": True,
                    "emit_debug_plan": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": True,
                    "allow_file_access": True,
                    "allow_browser_executor": True,
                    "allow_local_network_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                    "capture_stdout_stderr": True,
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": True,
                    "show_disabled_resource_badge": True,
                    "snap_to_grid": True,
                    "grid_enabled": True,
                    "auto_open_node_on_drop": True,
                    "confirm_delete_node": True,
                    "show_inline_config_summary": True,
                },
                "other_settings": {
                    "workspace_draft_recovery_enabled": True,
                    "workspace_draft_recovery_ttl_minutes": 30,
                },
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch:legacy",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:legacy",
                        },
                    ],
                    "node_config": {
                        "branches": [
                            {"key": "alpha", "label": "Alpha"},
                            {"key": "beta", "label": "Beta"},
                            {"key": "gamma", "label": "Gamma"},
                        ]
                    },
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "ports": [
                        {
                            "port_id": "in:legacy",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:legacy",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                    "node_config": {
                        "branches": [
                            {"key": "alpha", "label": "Alpha"},
                            {"key": "beta", "label": "Beta"},
                            {"key": "gamma", "label": "Gamma"},
                        ],
                        "mode": "all",
                        "quorum": None,
                    },
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    graph_document = service.get_graph_document()
    graph_nodes = {
        node.node_id: node
        for node in graph_document["graph_model"].nodes
    }

    fork_ports = graph_nodes["node-fork"].ports
    assert [port.port_id for port in fork_ports] == [
        "in",
        "branch:alpha",
        "branch:beta",
        "branch:gamma",
    ]
    assert [port.display_name for port in fork_ports] == [
        None,
        "Alpha",
        "Beta",
        "Gamma",
    ]

    join_ports = graph_nodes["node-join"].ports
    assert [port.port_id for port in join_ports] == [
        "in:alpha",
        "in:beta",
        "in:gamma",
        "out",
    ]
    assert [port.display_name for port in join_ports] == [
        "Alpha",
        "Beta",
        "Gamma",
        None,
    ]


def test_service_validate_graph_document_uses_branches_as_port_source_of_truth() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "branch:legacy",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.branch:legacy",
                        },
                    ],
                    "node_config": {
                        "branches": [
                            {"key": "alpha", "label": "Alpha"},
                            {"key": "beta", "label": "Beta"},
                            {"key": "gamma", "label": "Gamma"},
                        ]
                    },
                },
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "ports": [
                        {
                            "port_id": "in:legacy",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:legacy",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                    "node_config": {
                        "branches": [
                            {"key": "alpha", "label": "Alpha"},
                            {"key": "beta", "label": "Beta"},
                            {"key": "gamma", "label": "Gamma"},
                        ],
                        "mode": "quorum",
                        "quorum": 2,
                    },
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "valid"
    assert validation_result["diagnostics"] == []


def test_service_normalize_graph_document_rebuilds_branch_ports_without_persisting() -> None:
    service = CompilationWorkbenchService()

    normalized = service.normalize_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "ports": [],
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                            {"key": "audit", "label": "Audit"},
                        ]
                    },
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    normalized_ports = normalized["graph_model"].nodes[0].ports
    assert [port.port_id for port in normalized_ports] == [
        "in",
        "branch:left",
        "branch:right",
        "branch:audit",
    ]

    persisted_graph = service.get_graph_document()["graph_model"]
    assert persisted_graph.nodes == []


def test_service_normalize_graph_document_reports_changed_flag_for_branch_port_rebuild() -> None:
    service = CompilationWorkbenchService()

    normalized = service.normalize_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-join",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-join",
                    "expansion_role": "control:join",
                    "node_kind": "control.join",
                    "ports": [
                        {
                            "port_id": "in:legacy",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.branch:legacy",
                        }
                    ],
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                        ],
                        "mode": "all",
                        "quorum": None,
                    },
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert normalized["status"] == "normalized"
    assert normalized["changed"] is True
    assert [port.port_id for port in normalized["graph_model"].nodes[0].ports] == [
        "in:left",
        "in:right",
        "out",
    ]


def test_service_get_graph_source_projection_normalizes_branch_ports_before_projection() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-fork",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-fork",
                    "expansion_role": "control:parallel_fork",
                    "node_kind": "control.parallel_fork",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        }
                    ],
                    "node_config": {
                        "branches": [
                            {"key": "left", "label": "Left"},
                            {"key": "right", "label": "Right"},
                        ]
                    },
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    projection = service.get_graph_source_projection_document(
        target_source_kind="native_flow",
    )

    assert projection["status"] == "ready"
    assert projection["source_kind"] == "graph_workspace"
    assert projection["source_text"] == '{"graph_model_id":"graph:workspace","compilation_id":null,"graph_schema_version":"graph-v1","nodes":[{"node_id":"node-fork","lowered_kind":"control","source_anchor_ref":"n-fork","expansion_role":"control:parallel_fork","display_name":null,"node_kind":"control.parallel_fork","position":null,"ports":[{"port_id":"in","direction":"input","relation_layer":"control","semantic_slot":"in.control","display_name":null,"max_connections":null},{"port_id":"branch:left","direction":"output","relation_layer":"control","semantic_slot":"out.branch:left","display_name":"Left","max_connections":null},{"port_id":"branch:right","direction":"output","relation_layer":"control","semantic_slot":"out.branch:right","display_name":"Right","max_connections":null}],"node_config":{"branches":[{"key":"left","label":"Left"},{"key":"right","label":"Right"}]}}],"edges":[],"viewport":null,"root_metadata":{},"graph_effective_diagnostic_anchor_refs":[]}'


def test_service_validates_control_switch_requires_input_and_case_outputs() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-switch",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-switch",
                    "expansion_role": "control:switch",
                    "node_kind": "control.switch",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        }
                    ],
                    "node_config": {},
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    categories = {item["category"] for item in validation_result["diagnostics"]}
    assert "graph.control.switch.case_count_invalid" in categories
    assert "graph.control.switch.selector_missing" in categories


def test_service_validates_graph_call_subgraph_and_reports_disabled_subgraph_resource() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:subgraph",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_subgraph = service.save_subgraph_resource(resource_name="Disabled Validation Target")
    service.set_resource_enabled(
        resource_id=saved_subgraph["resource"]["resource_id"],
        enabled=False,
    )

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-disabled",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-disabled",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Disabled Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == (
        "graph.call_subgraph.subgraph_disabled"
    )
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-disabled",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "resource_id": saved_subgraph["resource"]["resource_id"],
        "resource_status": "disabled",
    }


def test_service_validates_graph_call_subgraph_and_reports_missing_required_schema_input() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_subgraph = service.save_subgraph_resource(resource_name="Schema Required Input Target")

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-required",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-required",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Required Input Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                        "inputs": {},
                        "outputs": {"message": "result_message"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == (
        "graph.call_subgraph.input_mapping_missing_required"
    )
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-required",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "missing_required_inputs": ["incoming"],
    }


def test_service_validates_builtin_component_required_parameters_before_runtime() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-shot",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-shot",
                    "expansion_role": "action:screenshot",
                    "node_kind": "browser.screenshot",
                    "node_config": {"path": ""},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-set-variable",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-set-variable",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "", "value": "demo"},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "in:value",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.value",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
                {
                    "node_id": "node-python",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-python",
                    "expansion_role": "action:run_python",
                    "node_kind": "python.run",
                    "node_config": {"code": ""},
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "control",
                            "semantic_slot": "in.control",
                        },
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        },
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-start-shot",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-shot",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-shot-set",
                    "relation_layer": "control",
                    "from_node_id": "node-shot",
                    "to_node_id": "node-set-variable",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
                {
                    "edge_id": "edge-set-python",
                    "relation_layer": "control",
                    "from_node_id": "node-set-variable",
                    "to_node_id": "node-python",
                    "from_port_id": "out",
                    "to_port_id": "in",
                },
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    categories = {item["category"] for item in validation_result["diagnostics"]}
    assert "graph.node.parameter_blank_required" in categories
    assert any(
        item["stage_extension"]["graph_ref"].get("node_id") == "node-shot"
        and item["stage_extension"]["graph_ref"].get("parameter_name") == "path"
        for item in validation_result["diagnostics"]
        if item["category"] == "graph.node.parameter_blank_required"
    )
    assert any(
        item["stage_extension"]["graph_ref"].get("node_id") == "node-set-variable"
        and item["stage_extension"]["graph_ref"].get("parameter_name") == "name"
        for item in validation_result["diagnostics"]
        if item["category"] == "graph.node.parameter_blank_required"
    )
    assert any(
        item["stage_extension"]["graph_ref"].get("node_id") == "node-python"
        and item["stage_extension"]["graph_ref"].get("parameter_name") == "code"
        for item in validation_result["diagnostics"]
        if item["category"] == "graph.node.parameter_blank_required"
    )

    compile_result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-shot",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-shot",
                    "expansion_role": "action:screenshot",
                    "node_kind": "browser.screenshot",
                    "node_config": {"path": ""},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert compile_result["status"] == "failed"
    assert any(
        item.category == "graph.node.parameter_blank_required"
        for item in compile_result["outcome"].diagnostic_catalog.entries
    )


def test_service_validates_required_parameters_declared_by_parameter_schema() -> None:
    service = CompilationWorkbenchService()

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-file-write",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-file-write",
                    "expansion_role": "action:write_text_file",
                    "node_kind": "file.write_text_file",
                    "node_config": {"path": "", "content": "hello", "encoding": "utf-8"},
                    "ports": [],
                },
                {
                    "node_id": "node-excel-write",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-excel-write",
                    "expansion_role": "action:write_excel_file",
                    "node_kind": "excel.write_file",
                    "node_config": {"path": "", "sheet_name": "Sheet1", "rows": []},
                    "ports": [],
                },
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert any(
        item["stage_extension"]["graph_ref"].get("node_id") == "node-file-write"
        and item["stage_extension"]["graph_ref"].get("parameter_name") == "path"
        for item in validation_result["diagnostics"]
        if item["category"] == "graph.node.parameter_blank_required"
    )
    assert any(
        item["stage_extension"]["graph_ref"].get("node_id") == "node-excel-write"
        and item["stage_extension"]["graph_ref"].get("parameter_name") == "path"
        for item in validation_result["diagnostics"]
        if item["category"] == "graph.node.parameter_blank_required"
    )

    compile_result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-file-write",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-file-write",
                    "expansion_role": "action:write_text_file",
                    "node_kind": "file.write_text_file",
                    "node_config": {"path": "", "content": "hello", "encoding": "utf-8"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert compile_result["status"] == "failed"
    assert any(
        item.category == "graph.node.parameter_blank_required"
        and item.stage_extension.get("graph_ref", {}).get("parameter_name") == "path"
        for item in compile_result["outcome"].diagnostic_catalog.entries
    )


def test_service_validates_graph_call_subgraph_and_reports_unknown_schema_output_mapping() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-output",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": False},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_subgraph = service.save_subgraph_resource(resource_name="Schema Unknown Output Target")

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-output",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-output",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Unknown Output Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                        "inputs": {"incoming": "hello"},
                        "outputs": {"unknown_output": "result_message"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == (
        "graph.call_subgraph.output_mapping_unknown_output"
    )
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-output",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "unknown_outputs": ["unknown_output"],
    }


def test_service_validates_graph_call_subgraph_and_reports_string_schema_input_type_mismatch() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-string",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "incoming": {"type": "string", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_subgraph = service.save_subgraph_resource(resource_name="Schema String Type Target")

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-string-type",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-string-type",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call String Type Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                        "inputs": {"incoming": ["not-a-string"]},
                        "outputs": {"message": "result_message"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == (
        "graph.call_subgraph.input_mapping_type_mismatch"
    )
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-string-type",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "invalid_typed_inputs": ["incoming"],
    }


def test_service_validates_graph_call_subgraph_and_reports_number_schema_input_type_mismatch() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:child-schema-number",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "root_metadata": {
                "input_schema": {
                    "count": {"type": "number", "required": True},
                },
                "output_schema": {
                    "message": {"type": "string"},
                },
            },
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved_subgraph = service.save_subgraph_resource(resource_name="Schema Number Type Target")

    validation_result = service.validate_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-call-number-type",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-call-number-type",
                    "expansion_role": "action:call_subgraph",
                    "display_name": "Call Number Type Subgraph",
                    "node_kind": "graph.call_subgraph",
                    "node_config": {
                        "subgraph_id": saved_subgraph["resource"]["resource_id"],
                        "inputs": {"count": "not-a-number"},
                        "outputs": {"message": "result_message"},
                    },
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert validation_result["status"] == "invalid"
    assert validation_result["summary"]["error_count"] == 1
    assert validation_result["diagnostics"][0]["category"] == (
        "graph.call_subgraph.input_mapping_type_mismatch"
    )
    assert validation_result["diagnostics"][0]["stage_extension"]["graph_ref"] == {
        "graph_model_id": "graph:workspace",
        "node_id": "node-call-number-type",
        "node_kind": "graph.call_subgraph",
        "subgraph_id": saved_subgraph["resource"]["resource_id"],
        "invalid_typed_inputs": ["count"],
    }


def test_service_compile_graph_document_blocks_disabled_node_resources() -> None:
    service = CompilationWorkbenchService()
    service.set_resource_enabled(resource_id="builtin:data.map", enabled=False)

    result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "transform:map",
                    "display_name": "Map Result",
                    "node_kind": "data.map",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert result["status"] == "failed"
    assert result["view"]["status"] == "failed"
    assert result["view"]["primary_diagnostic"]["category"] == "graph.node.resource_disabled"
    assert result["outcome"].diagnostic_catalog.entries[0].stage_extension["rule"] == (
        "graph.node.resource_enabled"
    )


def test_service_compile_graph_document_returns_failed_envelope_when_graph_is_invalid() -> None:
    service = CompilationWorkbenchService()

    result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                }
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "missing-node",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert result["status"] == "failed"
    assert result["view"]["status"] == "failed"
    assert result["view"]["stage_overview"]["terminal_stage"] == "validate"
    assert result["view"]["primary_diagnostic"]["category"] == "graph.edge.missing_target_node"
    assert result["outcome"].diagnostic_catalog.entries[0].stage_extension == {
        "subject_ref": "edge-1",
        "action": "validated graph document",
        "rule": "graph.edge.target_node_exists",
        "result": "failed",
        "graph_ref": {
            "graph_model_id": "graph:workspace",
            "edge_id": "edge-1",
            "from_node_id": "node-1",
            "to_node_id": "missing-node",
            "from_port_id": "out",
            "to_port_id": "in",
        },
    }
    assert result["outcome"].graph_model is not None
    assert result["outcome"].graph_model.graph_model_id == "graph:workspace"


def test_service_compile_graph_document_records_failed_graph_validation_in_snapshot() -> None:
    service = CompilationWorkbenchService()

    result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                }
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "missing-node",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    snapshot = service.get_workbench_snapshot()

    assert result["status"] == "failed"
    assert snapshot["workbench"]["compile_counter"] == 1
    assert snapshot["project"]["last_compile_status"] == "failed"
    assert snapshot["last_compile"]["status"] == "failed"
    assert snapshot["last_compile"]["source_kind"] == "graph_workspace"
    assert snapshot["last_compile"]["entry_document"] == "graph:workspace"
    assert snapshot["last_compile"]["request_origin"] == "graph_document"
    assert snapshot["last_compile"]["requested_graph_model_id"] == "graph:workspace"
    assert snapshot["last_compile"]["primary_diagnostic"]["category"] == "graph.edge.missing_target_node"
    assert snapshot["compile_history"][0] == snapshot["last_compile"]


def test_service_compile_graph_document_uses_graph_workspace_as_formal_compile_source() -> None:
    service = CompilationWorkbenchService()

    result = service.compile_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "display_name": "HTTP Request",
                    "node_kind": "http.request",
                    "position": {"x": 120, "y": 80},
                    "ports": [
                        {
                            "port_id": "out-main",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.result",
                        }
                    ],
                    "node_config": {"method": "GET"},
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "display_name": "Map Result",
                    "node_kind": "data.map",
                    "position": {"x": 360, "y": 80},
                    "ports": [
                        {
                            "port_id": "in-main",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                    "node_config": {"mode": "map"},
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out-main",
                    "to_port_id": "in-main",
                    "edge_state": "draft",
                }
            ],
            "viewport": {"x": 0, "y": 0, "zoom": 1.1},
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert result["status"] == "succeeded"
    assert result["request"]["source_kind"] == "graph_workspace"
    assert result["request"]["source"]["kind"] == "graph_workspace"
    assert result["request"]["request_origin"] == "graph_document"
    assert result["request"]["requested_graph_model_id"] == "graph:workspace"
    assert result["view"]["status"] == "succeeded"
    assert result["view"]["graph_stats"]["node_count"] == 2
    assert result["view"]["graph_stats"]["edge_count"] == 1
    assert result["outcome"].graph_model is not None
    assert result["outcome"].graph_model.graph_model_id == "graph:workspace"
    assert result["outcome"].graph_model.nodes[0].node_id == "node-1"
    assert result["outcome"].graph_model.nodes[0].position is not None
    assert result["outcome"].graph_model.nodes[0].position.x == 120
    assert result["outcome"].graph_model.nodes[0].ports[0].port_id == "out-main"
    assert result["outcome"].graph_model.nodes[0].node_config == {"method": "GET"}
    assert result["outcome"].graph_model.edges[0].relation_layer == "data"
    assert result["outcome"].graph_model.edges[0].from_port_id == "out-main"
    assert result["outcome"].graph_model.edges[0].to_port_id == "in-main"
    assert result["outcome"].graph_model.edges[0].edge_state == "draft"

    snapshot = service.get_workbench_snapshot()

    assert snapshot["last_compile"]["source_kind"] == "graph_workspace"
    assert snapshot["last_compile"]["request_origin"] == "graph_document"
    assert snapshot["last_compile"]["requested_graph_model_id"] == "graph:workspace"
    assert snapshot["compile_history"][0]["request_origin"] == "graph_document"


def test_service_compile_graph_document_uses_saved_workspace_graph_when_payload_is_omitted() -> None:
    service = CompilationWorkbenchService()
    saved_graph = {
        "graph_model_id": "graph:workspace",
        "compilation_id": None,
        "graph_schema_version": "graph-v1",
        "nodes": [
            {
                "node_id": "node-1",
                "lowered_kind": "execution",
                "source_anchor_ref": "n1",
                "expansion_role": "action:request",
                "node_kind": "http.request",
                "ports": [
                    {
                        "port_id": "out",
                        "direction": "output",
                        "relation_layer": "data",
                        "semantic_slot": "out.default",
                    }
                ],
            },
            {
                "node_id": "node-2",
                "lowered_kind": "execution",
                "source_anchor_ref": "n2",
                "expansion_role": "transform:map",
                "node_kind": "data.map",
                "ports": [
                    {
                        "port_id": "in",
                        "direction": "input",
                        "relation_layer": "data",
                        "semantic_slot": "in.default",
                    }
                ],
            },
        ],
        "edges": [
            {
                "edge_id": "edge-1",
                "relation_layer": "data",
                "from_node_id": "node-1",
                "to_node_id": "node-2",
                "from_port_id": "out",
                "to_port_id": "in",
            }
        ],
        "graph_effective_diagnostic_anchor_refs": [],
    }
    service.save_graph_document(saved_graph)

    result = service.compile_graph_document(None)

    assert result["status"] == "succeeded"
    assert result["request"]["entry_document"] == "graph:workspace"
    assert result["request"]["source_kind"] == "graph_workspace"
    assert result["request"]["source"]["kind"] == "graph_workspace"
    assert result["request"]["request_origin"] == "graph_document"
    assert result["request"]["requested_graph_model_id"] == "graph:workspace"
    assert result["request"]["requested_graph_save_revision"] == 1
    assert result["request"]["requested_graph_saved_at"]
    assert result["view"]["graph_stats"]["node_count"] == 2
    assert result["view"]["graph_stats"]["edge_count"] == 1
    assert '"graph_model_id":"graph:workspace"' in result["request"]["source"]["source_text"]
    assert '"from_port_id":"out"' in result["request"]["source"]["source_text"]


def test_service_snapshot_exposes_preferences_state() -> None:
    service = CompilationWorkbenchService()

    snapshot = service.get_workbench_snapshot()

    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["default_project_directory"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["language"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["resource_language"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["theme"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["default_window_size"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["startup_action"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["preferences_auto_save"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["program_settings"]["font_scale"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["compile_settings"]["default_source_kind"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["compile_settings"]["diagnostic_level"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["compile_settings"]["stop_on_first_error"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["compile_settings"]["emit_runtime_plan"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["compile_settings"]["emit_debug_plan"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["security_settings"]["confirm_high_risk_actions"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["security_settings"]["allow_file_access"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["security_settings"]["allow_browser_executor"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["security_settings"]["allow_local_network_access"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["python_runtime_settings"]["timeout_seconds"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["python_runtime_settings"]["python_executable_path"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["python_runtime_settings"]["sandbox_mode"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["python_runtime_settings"]["capture_stdout_stderr"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["graph_settings"]["auto_open_node_on_drop"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["graph_settings"]["confirm_delete_node"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["graph_settings"]["show_inline_config_summary"] == "active"
    assert snapshot["graph_workspace"]["preferences_state"]["other_settings"]["workspace_draft_recovery_enabled"] == "stored_only"
    assert snapshot["graph_workspace"]["preferences_state"]["other_settings"]["workspace_draft_recovery_ttl_minutes"] == "stored_only"


def test_service_graph_workspace_view_exposes_graph_preferences() -> None:
    from weconduct.application.preferences_service import PreferencesService
    from weconduct.application.preferences_store import InMemoryPreferencesStore

    preferences_service = PreferencesService(
        preferences_store=InMemoryPreferencesStore(
            {
                "preferences_file_version": 1,
                "program_settings": {
                    "language": "zh-CN",
                    "theme": "light",
                    "startup_action": "restore_last_workspace",
                    "default_project_directory": None,
                    "recent_project_limit": 10,
                },
                "compile_settings": {
                    "default_source_kind": "graph_workspace",
                    "diagnostic_level": "error",
                    "block_on_disabled_components": True,
                    "allow_degraded_compile": True,
                },
                "security_settings": {
                    "confirm_high_risk_actions": True,
                    "allow_external_programs": False,
                    "allow_file_access": True,
                },
                "python_runtime_settings": {
                    "python_executable_path": None,
                    "timeout_seconds": 60,
                    "sandbox_mode": "restricted",
                },
                "graph_settings": {
                    "auto_sync_mode": "responsive",
                    "show_node_id_on_node": False,
                    "show_disabled_resource_badge": False,
                    "snap_to_grid": False,
                    "grid_enabled": False,
                },
                "other_settings": {},
            }
        )
    )
    service = CompilationWorkbenchService(preferences_service=preferences_service)

    snapshot = service.get_workbench_snapshot()

    assert snapshot["graph_workspace"]["graph_preferences"] == {
        "auto_sync_mode": "responsive",
        "show_node_id_on_node": False,
        "show_disabled_resource_badge": False,
        "snap_to_grid": False,
        "grid_enabled": False,
        "auto_open_node_on_drop": True,
        "confirm_delete_node": True,
        "show_inline_config_summary": True,
        "save_conflict_policy": "prefer_current_graph",
    }


def test_workbench_snapshot_exposes_graph_workspace_saved_state_and_compile_relation() -> None:
    service = CompilationWorkbenchService()

    initial_snapshot = service.get_workbench_snapshot()
    assert initial_snapshot["graph_workspace"]["graph_model_id"] == "graph:workspace"
    assert initial_snapshot["graph_workspace"]["graph_document_save_revision"] == 0
    assert initial_snapshot["graph_workspace"]["graph_document_saved_at"] is None
    assert initial_snapshot["graph_workspace"]["last_compile_matches_saved_graph"] is None

    first_save = service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    first_saved_at = first_save["view"]["graph_document_saved_at"]

    compile_result = service.compile_graph_document(None)
    compiled_snapshot = service.get_workbench_snapshot()

    assert compile_result["request"]["requested_graph_save_revision"] == 1
    assert compile_result["request"]["requested_graph_saved_at"] == first_saved_at
    assert compiled_snapshot["graph_workspace"]["graph_document_save_revision"] == 1
    assert compiled_snapshot["graph_workspace"]["graph_document_saved_at"] == first_saved_at
    assert compiled_snapshot["graph_workspace"]["last_compiled_graph_save_revision"] == 1
    assert compiled_snapshot["graph_workspace"]["last_compile_matches_saved_graph"] is True

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    updated_snapshot = service.get_workbench_snapshot()

    assert updated_snapshot["graph_workspace"]["graph_document_save_revision"] == 2
    assert updated_snapshot["graph_workspace"]["last_compiled_graph_save_revision"] == 1
    assert updated_snapshot["graph_workspace"]["last_compile_matches_saved_graph"] is False


def test_service_compile_empty_saved_graph_workspace_records_failed_compiler_validation_in_snapshot() -> None:
    service = CompilationWorkbenchService()

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    result = service.compile_graph_document(None)
    snapshot = service.get_workbench_snapshot()

    assert result["status"] == "failed"
    assert result["view"]["primary_diagnostic"]["category"] == "source.empty"
    assert result["request"]["requested_graph_save_revision"] == 1
    assert result["request"]["source_kind"] == "graph_workspace"
    assert result["request"]["source"]["kind"] == "graph_workspace"
    assert snapshot["workbench"]["compile_counter"] == 1
    assert snapshot["last_compile"]["status"] == "failed"
    assert snapshot["last_compile"]["source_kind"] == "graph_workspace"
    assert snapshot["last_compile"]["primary_diagnostic"]["category"] == "source.empty"
    assert snapshot["last_compile"]["requested_graph_save_revision"] == 1
    assert snapshot["graph_workspace"]["last_compiled_graph_save_revision"] == 1
    assert snapshot["graph_workspace"]["last_compile_matches_saved_graph"] is True


def test_service_prepares_runtime_session_from_saved_graph_workspace() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "display_name": "HTTP Request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "display_name": "Map",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    result = service.prepare_runtime_session(None)

    assert result["status"] == "ready"
    assert result["request"]["request_origin"] == "saved_graph_document"
    assert result["request"]["requested_graph_model_id"] == "graph:workspace"
    assert result["request"]["requested_graph_save_revision"] == 1
    assert result["request"]["requested_graph_saved_at"]
    assert result["runtime_session"]["execution_supported"] is False
    assert result["runtime_session"]["status"] == "prepared"
    assert result["runtime_plan"]["graph_model_id"] == "graph:workspace"
    assert result["runtime_plan"]["node_count"] == 2
    assert result["runtime_plan"]["edge_count"] == 1
    assert result["runtime_plan"]["start_node_ids"] == ["node-1"]
    assert result["runtime_plan"]["terminal_node_ids"] == ["node-2"]
    assert result["runtime_plan"]["executable_nodes"][0]["node_id"] == "node-1"
    assert result["runtime_plan"]["executable_nodes"][0]["outgoing_edge_ids"] == ["edge-1"]
    assert result["runtime_plan"]["executable_nodes"][1]["incoming_edge_ids"] == ["edge-1"]
    assert result["diagnostics"]["highest_severity"] == "info"
    assert result["diagnostics"]["entries"][0]["stage"] == "parse"
    assert result["runtime_session"]["debug_snapshot"]["scheduler_mode"] == "legacy_sequence"
    assert result["runtime_session"]["debug_snapshot"]["token_queue"] == []
    assert result["runtime_session"]["debug_snapshot"]["join_buffers"] == {}


def test_service_prepares_debug_session_and_links_invalid_graph_diagnostics() -> None:
    service = CompilationWorkbenchService()

    result = service.prepare_debug_session(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                }
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "missing-node",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    assert result["status"] == "failed"
    assert result["request"]["request_origin"] == "graph_document_payload"
    assert result["debug_session"]["resume_supported"] is False
    assert result["debug_session"]["breakpoint_slots"] == []
    assert result["stage_timeline"][2]["stage"] == "validate"
    assert result["stage_timeline"][2]["status"] == "failed"
    assert result["object_index"]["graph_model_id"] == "graph:workspace"
    assert result["object_index"]["nodes"][0]["node_id"] == "node-1"
    assert result["object_index"]["edges"][0]["edge_id"] == "edge-1"
    assert result["diagnostic_links"][0]["category"] == "graph.edge.missing_target_node"
    assert result["diagnostic_links"][0]["graph_ref"]["edge_id"] == "edge-1"
    assert result["diagnostic_links"][0]["subject_ref"] == "edge-1"


def test_service_can_start_and_query_debug_session_from_saved_graph_workspace() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "node_kind": "http.request",
                    "ports": [
                        {
                            "port_id": "out",
                            "direction": "output",
                            "relation_layer": "data",
                            "semantic_slot": "out.default",
                        }
                    ],
                },
                {
                    "node_id": "node-2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n2",
                    "expansion_role": "transform:map",
                    "node_kind": "data.map",
                    "ports": [
                        {
                            "port_id": "in",
                            "direction": "input",
                            "relation_layer": "data",
                            "semantic_slot": "in.default",
                        }
                    ],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-1",
                    "relation_layer": "data",
                    "from_node_id": "node-1",
                    "to_node_id": "node-2",
                    "from_port_id": "out",
                    "to_port_id": "in",
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_debug_session(None)
    session = service.get_debug_session(session_id=started["debug_session"]["session_id"])
    sessions = service.list_debug_sessions()

    assert started["status"] == "started"
    assert started["request"]["request_origin"] == "saved_graph_document"
    assert started["debug_session"]["status"] == "prepared"
    assert started["debug_session"]["resume_supported"] is False
    assert started["debug_session"]["breakpoint_slots"] == []
    assert started["object_index"]["graph_model_id"] == "graph:workspace"
    assert started["stage_timeline"][-1]["stage"] == "emit"
    assert started["runtime_preview"]["scheduler_mode"] == "legacy_sequence"
    assert started["runtime_preview"]["token_queue"] == []
    assert started["runtime_preview"]["join_buffers"] == {}
    assert session["debug_session"]["session_id"] == started["debug_session"]["session_id"]
    assert session["debug_session"]["status"] == "prepared"
    assert session["runtime_preview"]["scheduler_mode"] == "legacy_sequence"
    assert sessions["sessions"][0]["session_id"] == started["debug_session"]["session_id"]
    assert sessions["sessions"][0]["graph_model_id"] == "graph:workspace"


def test_service_debug_session_and_lists_expose_runtime_preview_summary() -> None:
    service = CompilationWorkbenchService()
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-start",
                    "expansion_role": "flow:start",
                    "node_kind": "flow.start",
                    "node_config": {},
                    "ports": [
                        {
                            "port_id": "out-control",
                            "direction": "output",
                            "relation_layer": "control",
                            "semantic_slot": "out.control",
                        }
                    ],
                },
                {
                    "node_id": "node-next",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-next",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "debug_value", "value": 1},
                    "ports": [],
                },
            ],
            "edges": [
                {
                    "edge_id": "edge-control",
                    "relation_layer": "control",
                    "from_node_id": "node-start",
                    "to_node_id": "node-next",
                    "from_port_id": "out-control",
                    "to_port_id": None,
                }
            ],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    started = service.start_debug_session(None)
    session = service.get_debug_session(session_id=started["debug_session"]["session_id"])
    sessions = service.list_debug_sessions()

    assert started["debug_session"]["status"] == "prepared"
    assert session["runtime_preview_summary"]["scheduler_mode"] == "flow_graph"
    assert session["runtime_preview_summary"]["queued_node_count"] == 0
    assert session["runtime_preview_summary"]["executed_node_count"] == 0
    assert session["runtime_preview_summary"]["join_buffer_count"] == 0
    assert session["runtime_preview_summary"]["retry_state_count"] == 0
    assert session["runtime_preview_summary"]["current_node_id"] is None
    assert sessions["sessions"][0]["scheduler_mode"] == "flow_graph"
    assert sessions["sessions"][0]["queued_node_count"] == 0


def test_workbench_snapshot_reflects_most_recent_compile_summary() -> None:
    service = CompilationWorkbenchService()

    service.compile_source(
        source_kind="native_flow",
        entry_document="examples/recent.json",
        source_text='{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}',
    )
    snapshot = service.get_workbench_snapshot()

    assert snapshot["workbench"]["compile_counter"] == 1
    assert snapshot["project"]["project_status"] == "ready"
    assert snapshot["project"]["last_compile_status"] == "succeeded"
    assert snapshot["project"]["last_compile_request_sequence"] == 1
    assert snapshot["last_compile"]["status"] == "succeeded"
    assert snapshot["last_compile"]["request_sequence"] == 1
    assert isinstance(snapshot["last_compile"]["compiled_at"], str)
    assert snapshot["last_compile"]["compiled_at"]
    assert snapshot["last_compile"]["source_kind"] == "native_flow"
    assert snapshot["last_compile"]["entry_document"] == "examples/recent.json"
    assert snapshot["last_compile"]["graph_stats"]["node_count"] == 1
    assert snapshot["last_compile"]["stage_overview"]["succeeded_stage_count"] == 6
    assert snapshot["last_compile"]["stage_overview"]["terminal_stage"] == "emit"
    assert snapshot["last_compile"]["diagnostic_summary"]["highest_severity"] == "info"
    assert snapshot["last_compile"]["primary_diagnostic"]["stage"] == "parse"
    assert snapshot["last_compile"]["stage_cards"][-1]["stage"] == "emit"
    assert snapshot["last_compile"]["duration_ms"] is not None
    assert isinstance(snapshot["last_compile"]["duration_ms"], int)
    assert snapshot["last_compile"]["duration_ms"] >= 0


def test_workbench_snapshot_preserves_last_compile_failure_summary() -> None:
    service = CompilationWorkbenchService()

    service.compile_source(
        source_kind="native_flow",
        entry_document="examples/empty.json",
        source_text='{"nodes":[]}',
    )
    snapshot = service.get_workbench_snapshot()

    assert snapshot["workbench"]["compile_counter"] == 1
    assert snapshot["project"]["last_compile_status"] == "failed"
    assert snapshot["project"]["last_compile_request_sequence"] == 1
    assert snapshot["last_compile"]["status"] == "failed"
    assert snapshot["last_compile"]["request_sequence"] == 1
    assert snapshot["last_compile"]["compiled_at"]
    assert snapshot["last_compile"]["stage_overview"]["failed_stage_count"] == 1
    assert snapshot["last_compile"]["diagnostic_summary"]["highest_severity"] == "fatal"
    assert snapshot["last_compile"]["primary_diagnostic"]["category"] == "source.empty"
    assert snapshot["last_compile"]["duration_ms"] is not None
    assert isinstance(snapshot["last_compile"]["duration_ms"], int)
    assert snapshot["last_compile"]["duration_ms"] >= 0


def test_workbench_snapshot_exposes_recent_compile_history_with_latest_first() -> None:
    service = CompilationWorkbenchService()

    service.compile_source(
        source_kind="native_flow",
        entry_document="examples/first.json",
        source_text='{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}',
    )
    service.compile_source(
        source_kind="native_flow",
        entry_document="examples/second.json",
        source_text='{"nodes":[]}',
    )

    snapshot = service.get_workbench_snapshot()

    assert len(snapshot["compile_history"]) == 2
    assert snapshot["compile_history"][0]["request_sequence"] == 2
    assert snapshot["compile_history"][1]["request_sequence"] == 1
    assert snapshot["compile_history"][0]["compiled_at"]
    assert snapshot["compile_history"][1]["compiled_at"]
    assert snapshot["compile_history"][0]["entry_document"] == "examples/second.json"
    assert snapshot["compile_history"][0]["status"] == "failed"
    assert snapshot["compile_history"][1]["entry_document"] == "examples/first.json"
    assert snapshot["compile_history"][1]["status"] == "succeeded"
    assert snapshot["compile_history"][0]["duration_ms"] is not None
    assert snapshot["compile_history"][1]["duration_ms"] is not None
    assert isinstance(snapshot["compile_history"][0]["duration_ms"], int)
    assert isinstance(snapshot["compile_history"][1]["duration_ms"], int)
    assert snapshot["compile_history"][0]["duration_ms"] >= 0
    assert snapshot["compile_history"][1]["duration_ms"] >= 0
    assert snapshot["compile_history"][0] == snapshot["last_compile"]


def test_workbench_snapshot_compile_history_keeps_only_latest_five_items() -> None:
    service = CompilationWorkbenchService()

    for index in range(6):
        service.compile_source(
            source_kind="native_flow",
            entry_document=f"examples/run-{index}.json",
            source_text='{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}',
        )

    snapshot = service.get_workbench_snapshot()
    history_documents = [item["entry_document"] for item in snapshot["compile_history"]]

    assert len(snapshot["compile_history"]) == 5
    assert history_documents == [
        "examples/run-5.json",
        "examples/run-4.json",
        "examples/run-3.json",
        "examples/run-2.json",
        "examples/run-1.json",
    ]


def test_workbench_snapshot_compile_counter_counts_all_compile_requests() -> None:
    service = CompilationWorkbenchService()

    initial_snapshot = service.get_workbench_snapshot()

    service.compile_source(
        source_kind="native_flow",
        entry_document="examples/success.json",
        source_text='{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}',
    )
    service.compile_source(
        source_kind="native_flow",
        entry_document="examples/failure.json",
        source_text='{"nodes":[]}',
    )
    service.compile_source(
        source_kind="unknown_flow",
        entry_document="examples/unsupported.txt",
        source_text="ignored",
    )

    snapshot = service.get_workbench_snapshot()

    assert initial_snapshot["workbench"]["workspace_session_id"] == snapshot["workbench"]["workspace_session_id"]
    assert initial_snapshot["workbench"]["service_started_at"] == snapshot["workbench"]["service_started_at"]
    assert snapshot["workbench"]["compile_counter"] == 3
    assert snapshot["project"]["last_compile_status"] == "unsupported"
    assert snapshot["project"]["last_compile_request_sequence"] == 3
    assert snapshot["last_compile"]["request_sequence"] == 3


def test_services_share_workspace_state_when_using_same_store() -> None:
    store = InMemoryWorkspaceStateStore()
    first_service = CompilationWorkbenchService(state_store=store)
    second_service = CompilationWorkbenchService(state_store=store)

    first_service.compile_source(
        source_kind="native_flow",
        entry_document="examples/shared.json",
        source_text='{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}',
    )

    second_snapshot = second_service.get_workbench_snapshot()

    assert second_snapshot["workbench"]["compile_counter"] == 1
    assert second_snapshot["last_compile"]["entry_document"] == "examples/shared.json"
    assert second_snapshot["compile_history"][0]["entry_document"] == "examples/shared.json"


class _BarrierFileWorkspaceStateStore(FileWorkspaceStateStore):
    def __init__(self, path: Path) -> None:
        super().__init__(path)
        self._load_barrier: threading.Barrier | None = None
        self._barrier_armed = False
        self._barrier_lock = threading.Lock()

    def arm_next_load(self, barrier: threading.Barrier) -> None:
        with self._barrier_lock:
            self._load_barrier = barrier
            self._barrier_armed = True

    def load(self) -> dict | None:
        state = super().load()
        barrier: threading.Barrier | None = None
        with self._barrier_lock:
            if self._barrier_armed:
                barrier = self._load_barrier
                self._barrier_armed = False
        if barrier is not None:
            barrier.wait(timeout=5)
        return state


def test_file_backed_services_keep_both_compile_records_under_concurrent_compile_requests(
    tmp_path,
) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))

    first_store = _BarrierFileWorkspaceStateStore(state_file)
    second_store = _BarrierFileWorkspaceStateStore(state_file)
    first_service = CompilationWorkbenchService(state_store=first_store)
    second_service = CompilationWorkbenchService(state_store=second_store)

    shared_barrier = threading.Barrier(2)
    first_store.arm_next_load(shared_barrier)
    second_store.arm_next_load(shared_barrier)

    errors: list[BaseException] = []

    def _run_compile(service: CompilationWorkbenchService, entry_document: str) -> None:
        try:
            service.compile_source(
                source_kind="native_flow",
                entry_document=entry_document,
                source_text=(
                    '{"nodes":[{"id":"n1","role":"action",'
                    '"capability_domain":"http","action_kind":"request"}]}'
                ),
            )
        except BaseException as exc:  # pragma: no cover - test should surface real failure
            errors.append(exc)

    first_thread = threading.Thread(
        target=_run_compile,
        args=(first_service, "examples/concurrent-a.json"),
    )
    second_thread = threading.Thread(
        target=_run_compile,
        args=(second_service, "examples/concurrent-b.json"),
    )
    first_thread.start()
    second_thread.start()
    first_thread.join(timeout=10)
    second_thread.join(timeout=10)

    assert errors == []
    assert first_thread.is_alive() is False
    assert second_thread.is_alive() is False

    snapshot = CompilationWorkbenchService(
        state_store=FileWorkspaceStateStore(state_file)
    ).get_workbench_snapshot()

    assert snapshot["workbench"]["compile_counter"] == 2
    assert len(snapshot["compile_history"]) == 2
    assert {
        item["entry_document"]
        for item in snapshot["compile_history"]
    } == {"examples/concurrent-a.json", "examples/concurrent-b.json"}


def test_service_uses_store_initialized_workspace_state() -> None:
    store = InMemoryWorkspaceStateStore()
    seeded_state = {
        "workspace_state_version": 1,
        "workbench": {
            "host_mode": "python_core",
            "api_version": "phase1",
            "workspace_session_id": "ws-seeded123456",
            "service_started_at": "2026-06-13T08:00:00+00:00",
            "compile_counter": 2,
        },
        "last_compile": {
            "status": "failed",
            "request_sequence": 2,
            "compiled_at": "2026-06-13T08:10:00+00:00",
            "source_kind": "native_flow",
            "entry_document": "examples/seeded.json",
            "stage_cards": [],
            "stage_overview": {
                "total_stage_count": 6,
                "succeeded_stage_count": 1,
                "failed_stage_count": 1,
                "terminal_stage": "validate",
            },
            "diagnostic_summary": {
                "total_count": 1,
                "highest_severity": "fatal",
            },
            "primary_diagnostic": {
                "stage": "validate",
                "category": "source.empty",
                "severity": "fatal",
                "message": "seeded failure",
            },
            "graph_stats": {
                "graph_model_id": None,
                "node_count": 0,
                "edge_count": 0,
                "effective_diagnostic_anchor_count": 0,
            },
        },
        "compile_history": [
            {
                "status": "failed",
                "request_sequence": 2,
                "compiled_at": "2026-06-13T08:10:00+00:00",
                "source_kind": "native_flow",
                "entry_document": "examples/seeded.json",
                "stage_cards": [],
                "stage_overview": {
                    "total_stage_count": 6,
                    "succeeded_stage_count": 1,
                    "failed_stage_count": 1,
                    "terminal_stage": "validate",
                },
                "diagnostic_summary": {
                    "total_count": 1,
                    "highest_severity": "fatal",
                },
                "primary_diagnostic": {
                    "stage": "validate",
                    "category": "source.empty",
                    "severity": "fatal",
                    "message": "seeded failure",
                },
                "graph_stats": {
                    "graph_model_id": None,
                    "node_count": 0,
                    "edge_count": 0,
                    "effective_diagnostic_anchor_count": 0,
                },
            }
        ],
    }
    store.save(seeded_state)

    service = CompilationWorkbenchService(state_store=store)
    snapshot = service.get_workbench_snapshot()

    assert snapshot["workbench"]["workspace_session_id"] == "ws-seeded123456"
    assert snapshot["workbench"]["service_started_at"] == "2026-06-13T08:00:00+00:00"
    assert snapshot["workbench"]["compile_counter"] == 2
    assert snapshot["project"]["last_compile_request_sequence"] == 2
    assert snapshot["last_compile"]["request_sequence"] == 2
    assert snapshot["last_compile"]["compiled_at"] == "2026-06-13T08:10:00+00:00"
    assert snapshot["last_compile"]["duration_ms"] is None
    assert snapshot["last_compile"]["entry_document"] == "examples/seeded.json"
    assert snapshot["compile_history"][0]["duration_ms"] is None
    assert snapshot["compile_history"][0]["entry_document"] == "examples/seeded.json"


def test_file_workspace_state_store_returns_none_when_file_missing(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"

    store = FileWorkspaceStateStore(state_file)

    assert store.load() is None


def test_file_workspace_state_store_persists_state_to_json_file(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    store = FileWorkspaceStateStore(state_file)
    state = {
        "workspace_state_version": 1,
        "workbench": {
            "host_mode": "python_core",
            "api_version": "phase1",
            "workspace_session_id": "ws-file123456",
            "service_started_at": "2026-06-13T09:00:00+00:00",
            "compile_counter": 1,
        },
        "last_compile": None,
        "compile_history": [],
    }

    store.save(state)
    loaded_state = store.load()

    assert state_file.exists() is True
    assert loaded_state == state


def test_workbench_snapshot_exposes_workspace_state_version() -> None:
    service = CompilationWorkbenchService()

    snapshot = service.get_workbench_snapshot()

    assert snapshot["workbench"]["workspace_state_version"] == 1


def test_service_can_restore_workspace_state_from_file_store(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    first_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))

    first_service.compile_source(
        source_kind="native_flow",
        entry_document="examples/file-backed.json",
        source_text='{"nodes":[{"id":"n1","role":"action","capability_domain":"http","action_kind":"request"}]}',
    )

    second_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))
    snapshot = second_service.get_workbench_snapshot()

    assert snapshot["workbench"]["workspace_state_version"] == 1
    assert snapshot["project"]["has_persisted_workspace_state"] is True
    assert snapshot["project"]["last_compile_status"] == "succeeded"
    assert snapshot["workbench"]["compile_counter"] == 1
    assert snapshot["last_compile"]["entry_document"] == "examples/file-backed.json"
    assert snapshot["compile_history"][0]["entry_document"] == "examples/file-backed.json"


def test_service_compatibly_loads_legacy_workspace_state_without_pending_recovery_key(
    tmp_path,
) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    state_file.parent.mkdir(parents=True, exist_ok=True)
    state_file.write_text(
        json.dumps(
            {
                "workspace_state_version": 1,
                "workbench": {
                    "host_mode": "python_core",
                    "api_version": "phase1",
                    "workspace_session_id": "ws-legacy123456",
                    "service_started_at": "2026-06-17T10:00:00+00:00",
                    "compile_counter": 0,
                },
                "project": {
                    "project_id": "legacy-project",
                    "project_name": "Legacy Project",
                    "project_schema_version": "project-v1",
                    "project_status": "ready",
                    "workspace_root": r"I:\\WeConduct Object\\WeConduct",
                    "source_of_truth": "graph_document",
                    "main_graph_document_id": "graph:workspace",
                    "resource_registry_revision": 0,
                },
                "project_runtime": {
                    "project_file_path": None,
                    "is_dirty": False,
                },
                "last_compile": None,
                "compile_history": [],
                "recent_projects": [],
                "resource_registry": [],
                "editor_history": {
                    "undo_stack": [],
                    "redo_stack": [],
                },
                "execution_history": {
                    "runtime_runs": [],
                    "debug_sessions": [],
                },
                "runtime_sessions": [],
                "debug_sessions": [],
                "graph_document": {
                    "graph_model_id": "graph:workspace",
                    "compilation_id": None,
                    "graph_schema_version": "graph-v1",
                    "nodes": [],
                    "edges": [],
                    "graph_effective_diagnostic_anchor_refs": [],
                },
                "graph_document_meta": {
                    "save_revision": 0,
                    "saved_at": None,
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))
    snapshot = service.get_workbench_snapshot()

    assert snapshot["workbench"]["workspace_session_id"] == "ws-legacy123456"
    assert snapshot["project"]["pending_recovery"] is None
    assert snapshot["project"]["is_dirty"] is False


def test_service_exposes_pending_recovery_and_can_restore_dirty_workspace_from_file_store(
    tmp_path,
) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    project_path = tmp_path / "recovery-project.weconduct.json"
    first_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))

    first_service.create_project(project_name="Recovery Project")
    first_service.save_project_as(project_path=project_path)
    first_service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "draft-node",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-draft",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "draft"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    restored_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))
    pending_snapshot = restored_service.get_workbench_snapshot()

    assert pending_snapshot["project"]["is_dirty"] is False
    assert pending_snapshot["project"]["pending_recovery"]["status"] == "pending"
    assert (
        pending_snapshot["project"]["pending_recovery"]["project_file_path"]
        == str(project_path.resolve())
    )
    assert pending_snapshot["graph_workspace"]["node_count"] == 0

    restore_result = restored_service.restore_pending_recovery()
    restored_snapshot = restored_service.get_workbench_snapshot()

    assert restore_result["status"] == "restored"
    assert restore_result["project"]["pending_recovery"] is None
    assert restore_result["graph_document"].nodes[0].node_id == "draft-node"
    assert restored_snapshot["project"]["project_file_path"] == str(project_path.resolve())
    assert restored_snapshot["project"]["is_dirty"] is True
    assert restored_snapshot["project"]["pending_recovery"] is None
    assert restored_snapshot["graph_workspace"]["node_count"] == 1


def test_service_restored_workspace_snapshot_contains_full_graph_document_state(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    project_path = tmp_path / "full-restore.weconduct.json"
    first_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))

    first_service.create_project(project_name="Full Restore Project")
    first_service.save_project_as(project_path=project_path)
    first_service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-restored",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-restored",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "username", "value": "alice"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    first_service.save_project()

    restored_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))
    snapshot = restored_service.get_workbench_snapshot()
    graph_document = restored_service.get_graph_document()
    project_document = restored_service.get_project_document()

    assert snapshot["project"]["project_file_path"] == str(project_path.resolve())
    assert snapshot["project"]["pending_recovery"] is None
    assert snapshot["graph_workspace"]["node_count"] == 1
    assert graph_document["graph_model"].nodes[0].node_id == "node-restored"
    assert project_document["graph_workspace"]["node_count"] == 1


def test_service_can_discard_pending_recovery_from_file_store(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    project_path = tmp_path / "discard-project.weconduct.json"
    first_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))

    first_service.create_project(project_name="Discard Project")
    first_service.save_project_as(project_path=project_path)
    first_service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "discard-node",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-discard",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "discard"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    restored_service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))
    discard_result = restored_service.discard_pending_recovery()
    discarded_snapshot = restored_service.get_workbench_snapshot()

    assert discard_result["status"] == "discarded"
    assert discard_result["project"]["pending_recovery"] is None
    assert discarded_snapshot["project"]["pending_recovery"] is None
    assert discarded_snapshot["project"]["is_dirty"] is False
    assert discarded_snapshot["graph_workspace"]["node_count"] == 0


def test_service_save_project_does_not_convert_in_process_dirty_project_to_pending_recovery(
    tmp_path,
) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    project_path = tmp_path / "active-project.weconduct.json"
    service = CompilationWorkbenchService(state_store=FileWorkspaceStateStore(state_file))

    service.create_project(project_name="Active Project")
    service.save_project_as(project_path=project_path)
    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n1",
                    "expansion_role": "action:request",
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    result = service.save_project()
    project_payload = json.loads(project_path.read_text(encoding="utf-8"))
    storage_root = project_path.parent / f"{project_path.stem}.data"
    graph_payload = json.loads(
        (storage_root / "graphs" / "workspace.graph.json").read_text(encoding="utf-8")
    )
    snapshot = service.get_workbench_snapshot()

    assert result["status"] == "saved"
    assert project_payload["project_file_schema_version"] == 2
    assert graph_payload["nodes"][0]["node_id"] == "node-1"
    assert snapshot["project"]["pending_recovery"] is None
    assert snapshot["graph_workspace"]["node_count"] == 1


def test_service_save_project_persists_provided_graph_document_payload(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "payload-save.weconduct.json"

    service.create_project(project_name="Payload Save")
    service.save_project_as(project_path=project_path)

    save_result = service.save_project(
        graph_document_payload={
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-inline",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-inline",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "username", "value": "alice"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    project_payload = json.loads(project_path.read_text(encoding="utf-8"))
    storage_root = project_path.parent / f"{project_path.stem}.data"
    graph_payload = json.loads(
        (storage_root / "graphs" / "workspace.graph.json").read_text(encoding="utf-8")
    )
    snapshot = service.get_workbench_snapshot()

    assert save_result["status"] == "saved"
    assert save_result["graph_document"].nodes[0].node_id == "node-inline"
    assert project_payload["project_file_schema_version"] == 2
    assert graph_payload["nodes"][0]["node_id"] == "node-inline"
    assert graph_payload["nodes"][0]["node_config"]["value"] == "alice"
    assert snapshot["graph_workspace"]["graph_document_save_revision"] == 1
    assert snapshot["project"]["is_dirty"] is False


def test_service_save_project_as_persists_provided_graph_document_payload(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "payload-save-as.weconduct.json"

    service.create_project(project_name="Payload Save As")

    save_result = service.save_project_as(
        project_path=project_path,
        graph_document_payload={
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-save-as",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-save-as",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "password", "value": "secret"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        },
    )

    project_payload = json.loads(project_path.read_text(encoding="utf-8"))
    storage_root = project_path.parent / f"{project_path.stem}.data"
    graph_payload = json.loads(
        (storage_root / "graphs" / "workspace.graph.json").read_text(encoding="utf-8")
    )
    snapshot = service.get_workbench_snapshot()

    assert save_result["status"] == "saved"
    assert save_result["graph_document"].nodes[0].node_id == "node-save-as"
    assert project_payload["project_file_schema_version"] == 2
    assert graph_payload["nodes"][0]["node_id"] == "node-save-as"
    assert graph_payload["nodes"][0]["node_config"]["value"] == "secret"
    assert snapshot["graph_workspace"]["graph_document_save_revision"] == 1
    assert snapshot["project"]["project_file_path"] == str(project_path.resolve())
    assert snapshot["project"]["is_dirty"] is False


def test_service_save_graph_document_updates_open_project_file_payload(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "graph-save-sync.weconduct.json"

    service.create_project(project_name="Graph Save Sync")
    service.save_project_as(project_path=project_path)

    graph_save_result = service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-project-sync",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-project-sync",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "message", "value": "from-graph-save"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    project_payload = json.loads(project_path.read_text(encoding="utf-8"))
    storage_root = project_path.parent / f"{project_path.stem}.data"
    graph_payload = json.loads(
        (storage_root / "graphs" / "workspace.graph.json").read_text(encoding="utf-8")
    )

    assert graph_save_result["status"] == "saved"
    assert graph_save_result["view"]["graph_document_save_revision"] == 1
    assert project_payload["project_file_schema_version"] == 2
    assert graph_payload["nodes"][0]["node_id"] == "node-project-sync"
    assert graph_payload["nodes"][0]["node_config"]["value"] == "from-graph-save"
    assert project_payload["graph_document_meta"]["save_revision"] == 1


def test_save_project_writes_split_project_storage_layout(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "demo" / "demo.weconduct.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-start",
                    "lowered_kind": "control",
                    "source_anchor_ref": "n-node-start",
                    "expansion_role": "flow.start",
                    "display_name": "流程入口",
                    "node_kind": "flow.start",
                    "node_config": {"initial_variables": {"username": "demo"}},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )

    service.save_custom_node_graph_resource(resource_name="登录流程")
    result = service.save_project_as(project_path=str(project_path))

    storage_root = project_path.parent / f"{project_path.stem}.data"
    assert result["status"] == "saved"
    assert project_path.exists()
    assert (storage_root / "graphs" / "workspace.graph.json").exists()
    assert (storage_root / "resources" / "index.json").exists()
    assert (storage_root / "resource-overrides.json").exists()

    payload = json.loads(project_path.read_text(encoding="utf-8"))
    assert payload["project_file_schema_version"] == 2
    assert "builtin_resource_refs" in payload
    assert "project_resource_refs" in payload
    assert "resource_registry" not in payload
    assert payload["project"]["main_graph_path"] == f"{project_path.stem}.data/graphs/workspace.graph.json"


def test_save_project_writes_project_resources_into_resource_directories(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "demo.weconduct.json"

    service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    saved = service.save_custom_node_graph_resource(resource_name="登录流程")
    service.save_project_as(project_path=str(project_path))

    storage_root = project_path.parent / f"{project_path.stem}.data"
    index_payload = json.loads(
        (storage_root / "resources" / "index.json").read_text(encoding="utf-8")
    )
    resource_ref = next(
        item for item in index_payload["resources"] if item["resource_id"] == saved["resource"]["resource_id"]
    )
    manifest_path = project_path.parent / Path(resource_ref["manifest_path"])
    graph_path = project_path.parent / Path(resource_ref["graph_path"])

    assert manifest_path.exists()
    assert graph_path.exists()

    manifest_payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    graph_payload = json.loads(graph_path.read_text(encoding="utf-8"))

    assert manifest_payload["resource_id"] == saved["resource"]["resource_id"]
    assert manifest_payload["resource_type"] == "custom_node_graph"
    assert graph_payload["graph_model_id"] == saved["resource"]["source_graph_document_id"]


def test_save_project_extracts_enabled_and_tags_into_resource_overrides(tmp_path: Path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "demo.weconduct.json"

    saved = service.save_custom_node_graph_resource(resource_name="登录流程")
    service.set_resource_enabled(resource_id=saved["resource"]["resource_id"], enabled=False)
    service.update_resource_tags(
        resource_id=saved["resource"]["resource_id"],
        tags=["project", "auth", "favorite"],
    )
    service.save_project_as(project_path=str(project_path))

    storage_root = project_path.parent / f"{project_path.stem}.data"
    overrides_payload = json.loads(
        (storage_root / "resource-overrides.json").read_text(encoding="utf-8")
    )
    item = overrides_payload["resources"][saved["resource"]["resource_id"]]

    assert item["enabled"] is False
    assert sorted(item["tags"]) == ["auth", "favorite", "project"]


def test_open_legacy_v1_project_file_migrates_into_split_project_state(tmp_path: Path) -> None:
    legacy_project_path = tmp_path / "legacy.weconduct.json"
    legacy_payload = {
        "project_file_schema_version": 1,
        "saved_at": "2026-06-18T00:00:00Z",
        "project": {
            "project_id": "legacy-project",
            "project_name": "Legacy Project",
            "project_schema_version": "project-v1",
            "project_status": "ready",
            "workspace_root": str(tmp_path),
            "source_of_truth": "graph_document",
            "main_graph_document_id": "graph:workspace",
            "resource_registry_revision": 0,
        },
        "resource_registry": [],
        "editor_history": {"undo_stack": [], "redo_stack": []},
        "execution_history": {"runtime_runs": [], "debug_sessions": []},
        "graph_document": {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        },
        "graph_document_meta": {"save_revision": 0, "saved_at": None},
    }
    legacy_project_path.write_text(
        json.dumps(legacy_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    service = CompilationWorkbenchService()
    opened = service.open_project(project_path=str(legacy_project_path))

    assert opened["status"] == "opened"
    assert opened["project"]["project_file_path"] == str(legacy_project_path)
    assert opened["project"]["project_schema_version"] in {"project-v1", "project-v2"}
    assert opened["graph_document"].graph_model_id == "graph:workspace"


def test_open_split_project_file_restores_graph_and_project_resources(tmp_path: Path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "demo.weconduct.json"

    service.save_custom_node_graph_resource(resource_name="登录流程")
    service.save_project_as(project_path=str(project_path))

    reopened = CompilationWorkbenchService()
    opened = reopened.open_project(project_path=str(project_path))
    registry = reopened.get_resource_registry_document()["resources"]

    assert opened["status"] == "opened"
    assert any(item["resource_type"] == "custom_node_graph" for item in registry)
    assert opened["graph_document"].graph_model_id == "graph:workspace"


def test_service_can_save_project_then_save_graph_document_without_revision_drift(tmp_path) -> None:
    service = CompilationWorkbenchService()
    project_path = tmp_path / "revision-drift.weconduct.json"

    service.create_project(project_name="Revision Drift")
    service.save_project_as(project_path=project_path)

    save_project_result = service.save_project(
        graph_document_payload={
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-v1",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-v1",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "username", "value": "alice"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        }
    )
    snapshot_after_project_save = service.get_workbench_snapshot()

    graph_save_result = service.save_graph_document(
        {
            "graph_model_id": "graph:workspace",
            "compilation_id": None,
            "graph_schema_version": "graph-v1",
            "nodes": [
                {
                    "node_id": "node-v2",
                    "lowered_kind": "execution",
                    "source_anchor_ref": "n-v2",
                    "expansion_role": "action:set_variable",
                    "node_kind": "data.set_variable",
                    "node_config": {"name": "username", "value": "bob"},
                    "ports": [],
                }
            ],
            "edges": [],
            "graph_effective_diagnostic_anchor_refs": [],
        },
        expected_graph_document_save_revision=snapshot_after_project_save["graph_workspace"][
            "graph_document_save_revision"
        ],
    )

    project_payload = json.loads(project_path.read_text(encoding="utf-8"))
    storage_root = project_path.parent / f"{project_path.stem}.data"
    graph_payload = json.loads(
        (storage_root / "graphs" / "workspace.graph.json").read_text(encoding="utf-8")
    )

    assert save_project_result["status"] == "saved"
    assert snapshot_after_project_save["graph_workspace"]["graph_document_save_revision"] == 1
    assert graph_save_result["view"]["graph_document_save_revision"] == 2
    assert project_payload["project_file_schema_version"] == 2
    assert graph_payload["nodes"][0]["node_id"] == "node-v2"
    assert project_payload["graph_document_meta"]["save_revision"] == 2


def test_file_workspace_state_store_rejects_missing_workspace_state_version(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    state_file.parent.mkdir(parents=True, exist_ok=True)
    state_file.write_text(
        """
{
  "workbench": {},
  "last_compile": null,
  "compile_history": []
}
""".strip(),
        encoding="utf-8",
    )

    store = FileWorkspaceStateStore(state_file)

    try:
        store.load()
    except ValueError as exc:
        assert "workspace_state_version" in str(exc)
    else:
        raise AssertionError("expected ValueError for missing workspace_state_version")


def test_file_workspace_state_store_rejects_missing_workbench_key(tmp_path) -> None:
    state_file = tmp_path / "workspace" / "state.json"
    state_file.parent.mkdir(parents=True, exist_ok=True)
    state_file.write_text(
        """
{
  "workspace_state_version": 1,
  "last_compile": null,
  "compile_history": []
}
""".strip(),
        encoding="utf-8",
    )

    store = FileWorkspaceStateStore(state_file)

    try:
        store.load()
    except ValueError as exc:
        assert "workbench" in str(exc)
    else:
        raise AssertionError("expected ValueError for missing workbench")
